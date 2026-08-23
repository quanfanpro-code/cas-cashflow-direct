from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from cashflow_direct.classification import load_rule_pack
from cashflow_direct.pipeline import (
    confirm_cash_scope,
    run_classification,
    run_preflight,
)


def _write_single_sided_detail_fixture(path: Path) -> None:
    rules = load_rule_pack(Path(__file__).resolve().parents[1])
    workbook = Workbook()
    detail = workbook.active
    detail.title = "明细"
    detail.append(["日期", "凭证字", "凭证号", "分录行号", "摘要", "科目编码", "科目名称", "借方", "贷方", None, "流量金额（原币）", None, "主表项目名称"])
    detail.append(["2026/2/27", "记", 231, 11, "财务应付", "2202.03", "应付账款_财务", 58972968.30, None, None, 58972968.30, None, "购建固定资产、无形资产和其他长期资产支付的现金"])
    detail.append(["2026/2/27", "记", 231, 13, "5-票据红冲", "1121.01", "应收票据_银行承兑汇票", None, 48972968.30, None, -48972968.30, None, "购建固定资产、无形资产和其他长期资产支付的现金"])
    detail.append(["2026/1/31", "记", 176, 6, "计提2026.1月工资", "2211.01", "应付职工薪酬_工资", None, 926170.61, None, 926170.61, None, "购买商品、接受劳务支付的现金"])
    statement = workbook.create_sheet("现流表")
    statement.append(["单位名称：测试主体", None, "未审数"])
    statement.append(["项目", "行次", "2026年1-6月"])
    opening = 10926170.61
    row = 3
    for item in rules.statement_items:
        value = 0
        if item.item_id == "CASH-OPENING":
            value = opening
        elif item.item_id == "CASH-CLOSING":
            value = 0
        statement.cell(row, 1, item.name)
        statement.cell(row, 2, item.display_order)
        statement.cell(row, 3, value)
        row += 1
    workbook.save(path)


class SingleSidedDetailRegressionTests(unittest.TestCase):
    def test_single_sided_detail_without_confirmed_cash_proxy_requests_cleanup(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "单边现流明细.xlsx"
            _write_single_sided_detail_fixture(source)
            preflight = run_preflight(
                [source],
                ("100000", "50000", "5000"),
                output_parent=root,
                statement_path=source,
            )
            confirm_cash_scope(preflight.run_dir, {})
            classified = run_classification(preflight.run_dir)
            state_path = preflight.run_dir / "计算留痕数据" / "运行状态.json"
            state = json.loads(state_path.read_text(encoding="utf-8-sig"))
            self.assertEqual("待用户清洗现金分录", classified.status)
            self.assertEqual(0, classified.component_count)
            self.assertEqual("waiting_cash_row_cleanup", state["stage"])
            self.assertTrue(state["cash_row_cleanup_requests"])
            self.assertTrue((preflight.run_dir / "现金分录清洗请求.md").is_file())


if __name__ == "__main__":
    unittest.main()
