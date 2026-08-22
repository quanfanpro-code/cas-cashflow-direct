from __future__ import annotations

import json
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from cashflow_direct.pipeline import (
    confirm_cash_scope,
    finalize_run,
    run_classification,
    run_preflight,
)
from tests.fixture_factory import mark_dictionary_complete, write_end_to_end_case
from tests.test_pipeline import _complete_all_pending_ai


def test_final_workbook_exposes_the_complete_decision_chain_per_effective_cashflow_segment() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        preflight = run_preflight(
            write_end_to_end_case(root),
            ("1000000", "750000", "50000"),
            output_parent=root,
        )
        confirm_cash_scope(preflight.run_dir, preflight.recommended_cash_decisions)
        mark_dictionary_complete(preflight.run_dir)
        run_classification(preflight.run_dir)
        _complete_all_pending_ai(preflight.run_dir, root)
        final = finalize_run(preflight.run_dir)
        state = json.loads(
            (
                preflight.run_dir / "计算留痕数据" / "运行状态.json"
            ).read_text(encoding="utf-8-sig")
        )

        workbook = load_workbook(final.workbook_path, data_only=False)
        try:
            assert "原表与系统决定差异" in workbook.sheetnames
            assert "科目语义词典" in workbook.sheetnames
            assert "同类检查" in workbook.sheetnames
            trace = workbook["全量分类留痕"]
            headers = [cell.value for cell in trace[1]]
            expected_order = [
                "日期",
                "凭证字",
                "凭证号",
                "本行摘要",
                "本行科目路径",
                "原始完整科目路径",
                "本行完整对方科目路径",
                "原现流项目",
                "系统候选项目",
                "判断理由",
                "摘要来源质量",
                "完整路径来源质量",
                "两个来源是否独立",
                "证据质量说明",
                "证据得分",
                "单笔金额",
                "同类累计金额",
                "有效重要性层级",
                "强制检查",
                "唯一动作",
                "异常",
                "AI复核过程",
                "当前决定形成过程",
                "最终决定项目",
                "复核状态",
                "评分版本",
                "动作表版本",
            ]
            positions = [headers.index(header) for header in expected_order]
            assert positions == sorted(positions)
            assert trace.max_row == 3
            original_item_column = headers.index("原现流项目") + 1
            assert all(
                trace.cell(row, original_item_column).value != "原项目为空"
                for row in range(2, trace.max_row + 1)
            )
            assert "现金账户路径" in headers
            assert "现金账户范围状态" in headers
            assert "对方科目范围状态" in headers
            for header in ("来源文件", "来源工作表", "来源行号", "来源单元格"):
                column_index = headers.index(header) + 1
                assert any(
                    dimension.hidden
                    and int(dimension.min or 0) <= column_index <= int(dimension.max or 0)
                    for dimension in trace.column_dimensions.values()
                )
            assert all(
                trace.cell(row, headers.index("评分版本") + 1).value
                == state["versions"]["scoring"]
                for row in range(2, trace.max_row + 1)
            )
            action_column = headers.index("唯一动作") + 1
            assert all(
                trace.cell(row, action_column).value
                not in {
                    "automatic_keep",
                    "automatic_change",
                    "human_decision",
                    "low_amount_human_batch",
                }
                for row in range(2, trace.max_row + 1)
            )

            difference = workbook["原表与系统决定差异"]
            difference_headers = [cell.value for cell in difference[1]]
            for header in (
                "差异形成原因",
                "打分逻辑描述及打分结果",
                "独立来源1",
                "独立来源2",
            ):
                assert header in difference_headers
            assert "主表项目名称" not in difference_headers
            assert "审定现流表项目" in difference_headers
        finally:
            workbook.close()
