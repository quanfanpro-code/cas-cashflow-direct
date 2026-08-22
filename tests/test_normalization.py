from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from cashflow_direct.models import SourceLocator
from cashflow_direct.normalization import (
    NormalizationResult,
    RowExclusion,
    normalize_dataset,
    subtotal_exclusion_warning,
)
from cashflow_direct.semantic_mapping import DatasetMapping, infer_dataset_mapping
from cashflow_direct.workbook_structure import scan_workbook
from tests.fixture_factory import break_dimension, write_all_input_types


class NormalizationTests(unittest.TestCase):
    def test_optional_source_fields_preserve_blank_zero_and_negative_amounts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "原始字段.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "日期",
                    "凭证字",
                    "凭证号",
                    "摘要",
                    "科目编码",
                    "科目名称",
                    "借方",
                    "贷方",
                    "流量金额",
                    "主表项目",
                ]
            )
            sheet.append(
                [
                    "2026-01-01",
                    "记",
                    1,
                    "匿名红冲",
                    "1002.01",
                    "银行存款",
                    0,
                    None,
                    -12.34,
                    "销售商品、提供劳务收到的现金",
                ]
            )
            workbook.save(path)
            mapping = infer_dataset_mapping(scan_workbook(path))
            self.assertIsInstance(mapping, DatasetMapping)

            entry = normalize_dataset(path, "FRAW", mapping).entries[0]

            self.assertEqual("记", entry.voucher_word)
            self.assertEqual("1002.01", entry.account_code)
            self.assertEqual(0, entry.source_debit_cent)
            self.assertIsNone(entry.source_credit_cent)
            self.assertEqual(-1_234, entry.source_flow_amount_cent)

    def test_five_input_shapes_normalize_without_template_numbers(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            files = write_all_input_types(Path(tmp))
            results = []
            for index, path in enumerate(files, start=1):
                mapping = infer_dataset_mapping(scan_workbook(path))
                self.assertIsInstance(mapping, DatasetMapping, path.name)
                results.append(normalize_dataset(path, f"F{index}", mapping))
            self.assertEqual([4, 5, 6, 4, 4], [len(result.entries) for result in results])
            self.assertTrue(results[0].profile.matched_counterparty)
            self.assertTrue(results[1].profile.has_flow_item)
            self.assertTrue(results[2].profile.split_duplication_risk)
            self.assertEqual(frozenset({"cash", "counterpart"}), results[3].profile.retained_side_values)
            self.assertTrue(results[4].profile.summary_only)
            for result in results:
                self.assertEqual(
                    result.rows_read,
                    len(result.entries) + len(result.exclusions) + len(result.errors),
                )
                for entry in result.entries:
                    self.assertTrue(entry.source.file_id)
                    self.assertTrue(entry.source.sheet_name)
                    self.assertGreater(entry.source.row_start, 0)
                    self.assertTrue(entry.source.cell_range)

    def test_invalid_money_is_a_located_error_not_silent_zero(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "坏金额.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["发生额方向", "金额", "摘要", "现流项目"])
            sheet.append(["借", "无法识别", "匿名事项", "收到其他经营现金"])
            workbook.save(path)
            mapping = infer_dataset_mapping(scan_workbook(path))
            self.assertIsInstance(mapping, DatasetMapping)
            result = normalize_dataset(path, "FBAD", mapping)
            self.assertEqual(0, len(result.entries))
            self.assertEqual(1, len(result.errors))
            self.assertIn("B2", result.errors[0].source.cell_range)
            self.assertIn("金额", result.errors[0].message)

    def test_subtotal_rows_without_date_and_voucher_are_excluded_with_trace(self) -> None:
        # 鼎弘式：小计行日期与凭证号双空 → 剔除留痕，数据行保留
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "鼎弘式小计.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["日期", "凭证号", "摘要", "科目", "借方", "贷方", "流量金额", "现流项目"])
            sheet.append([None, None, None, None, 0, 0, 15593144.77, "经营活动现金流入小计"])
            sheet.append(["2025-02-12", "记-5", "匿名结息", "财务费用", -85708.43, None, 85708.43, "收到其他经营现金"])
            workbook.save(path)
            mapping = infer_dataset_mapping(scan_workbook(path))
            self.assertIsInstance(mapping, DatasetMapping)
            result = normalize_dataset(path, "FSUB", mapping)
            self.assertEqual(1, len(result.entries))
            self.assertEqual(1, len(result.exclusions))
            self.assertEqual("subtotal_row", result.exclusions[0].discard_reason)
            self.assertIn("A2", result.exclusions[0].source.cell_range)

    def test_subtotal_exclusion_warning_flags_only_abnormal_share(self) -> None:
        def make_result(rows_read: int, subtotal_count: int) -> NormalizationResult:
            exclusions = tuple(
                RowExclusion(SourceLocator("F", "S", i, i, f"A{i}"), "subtotal_row")
                for i in range(subtotal_count)
            )
            return NormalizationResult((), None, exclusions, (), rows_read)

        flagged = subtotal_exclusion_warning(make_result(8, 1))  # 12.5% > 10%
        self.assertIsNotNone(flagged)
        self.assertIn("1/8", str(flagged["message"]))
        quiet = subtotal_exclusion_warning(make_result(12, 1))  # 8.3% <= 10%
        self.assertIsNone(quiet)

    def test_normalize_dataset_tolerates_broken_dimension(self) -> None:
        # dimension 损坏的文件仍应读出全部列与金额
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "坏维度明细.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["日期", "凭证号", "摘要", "科目", "借方", "贷方"])
            sheet.append(["2026-01-01", "记-1", "匿名收款", "银行存款", 100, None])
            workbook.save(path)
            break_dimension(path)
            mapping = infer_dataset_mapping(scan_workbook(path))
            self.assertIsInstance(mapping, DatasetMapping)
            result = normalize_dataset(path, "FBD", mapping)
            self.assertEqual(1, len(result.entries))
            self.assertEqual(100_00, result.entries[0].debit_cent)

    def test_blank_summary_is_kept_as_illegal_input_and_warned(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "摘要为空.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["发生额方向", "金额", "摘要", "科目名称", "现流项目"])
            sheet.append(["贷", 100, None, "应付账款_供应商", "购买商品、接受劳务支付的现金"])
            workbook.save(path)
            mapping = infer_dataset_mapping(scan_workbook(path))
            self.assertIsInstance(mapping, DatasetMapping)

            result = normalize_dataset(path, "F-EMPTY-SUMMARY", mapping)

            self.assertEqual(1, len(result.entries))
            self.assertEqual(10_000, result.entries[0].credit_cent)
            self.assertIn("summary_empty", result.entries[0].input_issues)
            self.assertTrue(any("摘要为空" in item.message for item in result.warnings))

    def test_blank_account_path_is_kept_and_is_not_reconstructed(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "路径为空.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["发生额方向", "金额", "摘要", "科目名称", "现流项目"])
            sheet.append(["贷", 100, "支付货款", None, "购买商品、接受劳务支付的现金"])
            workbook.save(path)
            mapping = infer_dataset_mapping(scan_workbook(path))
            self.assertIsInstance(mapping, DatasetMapping)

            result = normalize_dataset(path, "F-EMPTY-PATH", mapping)

            self.assertEqual(1, len(result.entries))
            self.assertEqual("", result.entries[0].account_name)
            self.assertEqual("", result.entries[0].counterpart_name)
            self.assertIn("account_path_empty", result.entries[0].input_issues)
            self.assertTrue(any("对方科目路径为空" in item.message for item in result.warnings))

    def test_only_real_merged_cells_inherit_summary(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "合并与普通空白.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["日期", "凭证号", "摘要", "科目", "借方", "贷方"])
            sheet.append(["2026-01-01", "记-1", "发放工资", "银行存款", None, 100])
            sheet.append([None, None, None, "应付职工薪酬_工资", 100, None])
            sheet.append(["2026-01-02", "记-2", None, "银行存款", None, 50])
            sheet.merge_cells("A2:A3")
            sheet.merge_cells("B2:B3")
            sheet.merge_cells("C2:C3")
            workbook.save(path)
            mapping = infer_dataset_mapping(scan_workbook(path))
            self.assertIsInstance(mapping, DatasetMapping)

            result = normalize_dataset(path, "F-MERGED", mapping)

            self.assertEqual("发放工资", result.entries[1].summary)
            self.assertEqual("", result.entries[2].summary)
            self.assertNotIn("summary_empty", result.entries[1].input_issues)
            self.assertIn("summary_empty", result.entries[2].input_issues)

    def test_merged_and_blank_voucher_fields_stay_in_one_voucher(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "合并凭证.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["日期", "凭证号", "摘要", "科目", "借方", "贷方"])
            sheet.append(["2026-01-01", "记-1", "合计回款不是合计行", "银行存款", 100, None])
            sheet.append([None, None, "合计回款不是合计行", "主营业务收入", None, 100])
            sheet.merge_cells("A2:A3")
            sheet.merge_cells("B2:B3")
            workbook.save(path)
            mapping = infer_dataset_mapping(scan_workbook(path))
            self.assertIsInstance(mapping, DatasetMapping)
            result = normalize_dataset(path, "FM", mapping)
            self.assertEqual(2, len(result.entries))
            self.assertEqual(1, len({item.voucher_key for item in result.entries}))
            self.assertEqual(0, len(result.errors))


if __name__ == "__main__":
    unittest.main()
