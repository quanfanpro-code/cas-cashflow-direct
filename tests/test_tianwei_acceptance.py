from __future__ import annotations

import hashlib
import json
import os
import re
from collections import Counter
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from cashflow_direct.classification import load_rule_pack
from cashflow_direct.pipeline import (
    confirm_account_mapping,
    confirm_cash_scope,
    confirm_mapping,
    finalize_run,
    import_ai_results,
    import_dictionary_results,
    import_summary_results,
    run_classification,
    run_preflight,
    scan_accounts,
    supplement_cash_balances,
)
from cashflow_direct.statement import ExistingStatementResult, detect_statement_sheets
from cashflow_direct.workbook_output import SHEET_NAMES, USE_SYSTEM_RECOMMENDATION


ALLOWED_SCORES = {0, 10, 20, 25, 35, 45, 50, 55, 70, 90, None}


def _real_materiality() -> tuple[Decimal, Decimal, Decimal]:
    raw = os.environ.get("CAS_CASHFLOW_REAL_MATERIALITY")
    if not raw:
        pytest.skip("未设置真实验收重要性参数")
    values = tuple(Decimal(value.strip()) for value in raw.split(","))
    assert len(values) == 3, "真实验收重要性参数应依次提供整体、实际执行和明显微小金额"
    return values


def _real_fx_amount() -> Decimal:
    raw = os.environ.get("CAS_CASHFLOW_REAL_FX")
    if raw is None:
        pytest.skip("未设置经用户确认的真实汇率变动金额")
    return Decimal(raw)


def _real_level1_mappings() -> dict[str, str]:
    raw_path = os.environ.get("CAS_CASHFLOW_REAL_ACCOUNT_MAPPINGS")
    if not raw_path:
        pytest.skip("未设置经用户确认的真实一级科目映射文件")
    path = Path(raw_path)
    assert path.is_file()
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    assert isinstance(payload, dict) and payload
    return {str(key): str(value) for key, value in payload.items()}


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _state(run_dir: Path) -> dict[str, object]:
    return json.loads(
        (run_dir / "计算留痕数据" / "运行状态.json").read_text(
            encoding="utf-8-sig"
        )
    )


def _statement_file(paths: tuple[Path, ...], rules) -> Path:
    candidates = []
    for path in paths:
        results = detect_statement_sheets(path, rules)
        if any(isinstance(item, ExistingStatementResult) for item in results.values()):
            candidates.append(path)
    assert len(candidates) == 1, "两个输入中必须且只能识别出一份客户现有正表"
    return candidates[0]


def _complete_mapping(run_dir: Path) -> None:
    for _ in range(20):
        state = _state(run_dir)
        questions = state.get("mapping_questions", ())
        if not questions:
            return
        choices = {
            f"{item['file_id']}:{item.get('sheet', '')}:{item['role']}": item[
                "recommended"
            ]
            for item in questions
            if item.get("kind") != "statement"
        }
        assert len(choices) == len(questions), "指定正表后不应再出现正表选择问题"
        confirm_mapping(run_dir, choices)
    raise AssertionError("字段映射未能在有限轮次内完成")


def _confirm_real_level1_accounts(run_dir: Path) -> set[str]:
    state = _state(run_dir)
    pending = {
        str(item["original_level1"])
        for item in state.get("account_mapping_records", ())
        if item.get("status") != "confirmed"
    }
    if pending:
        mappings = _real_level1_mappings()
        assert pending <= set(mappings), "映射文件必须覆盖本次全部待确认一级科目"
        confirm_account_mapping(
            run_dir,
            {name: mappings[name] for name in pending},
        )
    confirmed = _state(run_dir)
    assert not confirmed.get("account_mapping_questions")
    assert all(
        item.get("status") == "confirmed"
        for item in confirmed.get("account_mapping_records", ())
    )
    return pending


def _complete_dictionary(run_dir: Path, result_path: Path) -> None:
    result = scan_accounts(run_dir)
    if not result.get("missing"):
        return
    state = _state(run_dir)
    expected = {
        task["task_id"]: task["account"]
        for task in state["account_dictionary"]["tasks"]
    }
    records = tuple(
        json.loads(line)
        for line in result_path.read_text(encoding="utf-8-sig").splitlines()
        if line.strip()
    )
    selected = tuple(record for record in records if record.get("task_id") in expected)
    assert {record["task_id"] for record in selected} == set(expected)
    assert all(record["account"] == expected[record["task_id"]] for record in selected)
    allowed_fields = {"task_id", "account", "node_concepts", "relations"}
    assert all(set(record) <= allowed_fields for record in selected), (
        "科目路径Agent只能补节点概念和父子关系，不能沿用旧项目、质量或置信度答案"
    )
    matched_path = run_dir / "科目语义判断结果_匹配当前输入.jsonl"
    matched_path.write_text(
        "".join(json.dumps(record, ensure_ascii=False) + "\n" for record in selected),
        encoding="utf-8-sig",
    )
    imported = import_dictionary_results(run_dir, matched_path)
    assert imported["status"] == "科目语义已导入"


def _complete_summary_semantics(run_dir: Path, result_path: Path | None) -> None:
    state = _state(run_dir)
    expected = {
        task["task_id"]: task["summary"]
        for task in state["summary_semantics"]["tasks"]
    }
    if not expected:
        return
    assert result_path is not None and result_path.is_file(), (
        "固定规则仍有未决语言槽位时，必须提供当前Agent形成的受限槽位结果"
    )
    records = tuple(
        json.loads(line)
        for line in result_path.read_text(encoding="utf-8-sig").splitlines()
        if line.strip()
    )
    selected = tuple(record for record in records if record.get("task_id") in expected)
    assert {record["task_id"] for record in selected} == set(expected)
    assert all(record["summary"] == expected[record["task_id"]] for record in selected)
    forbidden = {
        "item_id",
        "candidate_item_ids",
        "quality",
        "score",
        "confidence",
        "decision_action",
    }
    assert all(not forbidden.intersection(record) for record in selected)
    matched_path = run_dir / "摘要语义判断结果_匹配当前输入.jsonl"
    matched_path.write_text(
        "".join(json.dumps(record, ensure_ascii=False) + "\n" for record in selected),
        encoding="utf-8-sig",
    )
    imported = import_summary_results(run_dir, matched_path)
    assert imported["status"] == "摘要语义已导入"


def _real_cash_scope_decisions(state: dict[str, object]) -> dict[str, str]:
    raw_path = os.environ.get("CAS_CASHFLOW_REAL_CASH_SCOPE_DECISIONS")
    if not raw_path:
        pytest.skip("未设置经用户逐项确认的真实现金范围决定文件")
    path = Path(raw_path)
    assert path.is_file()
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    assert isinstance(payload, dict)
    decisions = {str(key): str(value) for key, value in payload.items()}
    expected_keys = {
        str(item["account_key"])
        for item in state["cash_scope_proposal"]["candidates"]
    }
    assert set(decisions) == expected_keys, "现金范围决定必须逐项覆盖本次全部候选账户"
    assert set(decisions.values()) <= {"include", "exclude"}
    return decisions


def _component_signatures(state: dict[str, object]) -> Counter[tuple[object, ...]]:
    return Counter(
        (
            item["voucher_date"],
            str(item["voucher_no"]),
            item["summary"],
            int(item["cash_delta_cent"]),
            tuple(item["counterpart_accounts"]),
            item["original_item_text"],
        )
        for item in state["components"]
    )


def test_real_files_complete_the_generic_pipeline() -> None:
    raw_directory = os.environ.get("CAS_CASHFLOW_REAL_DIR")
    if not raw_directory:
        pytest.skip("未设置真实验收目录")
    ai_results_value = os.environ.get("CAS_CASHFLOW_REAL_AI_RESULTS")
    if not ai_results_value:
        pytest.skip("未设置结构化AI复核结果或技术失败注入文件")
    ai_results_paths = tuple(Path(value) for value in ai_results_value.split(os.pathsep))
    assert all(path.is_file() for path in ai_results_paths)
    dictionary_results_value = os.environ.get(
        "CAS_CASHFLOW_REAL_DICTIONARY_RESULTS"
    )
    if not dictionary_results_value:
        pytest.skip("未设置已经完成的科目语义判断结果")
    dictionary_results_path = Path(dictionary_results_value)
    assert dictionary_results_path.is_file()
    summary_results_value = os.environ.get("CAS_CASHFLOW_REAL_SUMMARY_SLOT_RESULTS")
    summary_results_path = Path(summary_results_value) if summary_results_value else None
    case_dir = Path(raw_directory)
    detail_value = os.environ.get("CAS_CASHFLOW_REAL_DETAIL")
    statement_value = os.environ.get("CAS_CASHFLOW_REAL_STATEMENT")
    if detail_value and statement_value:
        inputs = (Path(detail_value), Path(statement_value))
        assert all(path.is_file() for path in inputs)
    else:
        inputs = tuple(sorted(case_dir.glob("*.xlsx")))
        assert len(inputs) == 2, "目录存在其他Excel时必须显式指定明细和正表"
    before_hashes = {str(path.resolve()): _sha256(path) for path in inputs}

    rules = load_rule_pack(Path(__file__).resolve().parents[1])
    statement_path = _statement_file(inputs, rules)
    output_parent = Path(
        os.environ.get(
            "CAS_CASHFLOW_REAL_OUTPUT",
            str(case_dir.parent / "真实验收输出"),
        )
    )
    materiality = _real_materiality()
    preflight = run_preflight(
        inputs,
        materiality,
        output_parent,
        statement_path,
    )
    run_dir = preflight.run_dir
    confirmed_mappings = _real_level1_mappings()
    fx_amount = _real_fx_amount()
    (run_dir / "用户确认参数.md").write_text(
        "# 真实验收参数\n\n"
        f"1. 财务报表整体重要性：{materiality[0]}元。\n"
        f"2. 实际执行的重要性：{materiality[1]}元。\n"
        f"3. 明显微小错报临界值：{materiality[2]}元。\n"
        "4. 现金及现金等价物期初、期末余额取客户现有现金流量表正表。\n"
        "5. 现金范围由用户逐项确认文件提供，不由测试自动推断。\n"
        "6. 每个候选账户必须有纳入或不纳入决定。\n"
        "7. 没有公司特殊规则。\n"
        f"8. 经用户确认的汇率变动金额：{fx_amount}元。\n"
        + "".join(
            f"{index}. 一级科目映射：{original}→{standard}。\n"
            for index, (original, standard) in enumerate(
                sorted(confirmed_mappings.items()), 9
            )
        ),
        encoding="utf-8-sig",
    )
    _complete_mapping(run_dir)

    state = _state(run_dir)
    pending_level1 = {
        str(item["original_level1"]): tuple(item["candidate_standard_names"])
        for item in state["account_mapping_records"]
        if item["status"] != "confirmed"
    }
    assert set(pending_level1) <= set(confirmed_mappings)
    assert all(
        confirmed_mappings[original] in candidates
        for original, candidates in pending_level1.items()
    )
    assert _confirm_real_level1_accounts(run_dir) == set(pending_level1)
    state = _state(run_dir)
    assert len(state["entries"]) > 0
    mapped_roles = state["mappings"][0]["roles"]
    assert "account_name" in mapped_roles
    assert any(
        re.search(r"[_/\\>|：:]", str(item["account_name"]))
        for item in state["entries"]
        if item["account_name"]
    ), "真实明细存在分层科目时必须保留完整路径"

    cash_choices = _real_cash_scope_decisions(state)
    confirm_cash_scope(run_dir, cash_choices)
    _complete_dictionary(run_dir, dictionary_results_path)

    state = _state(run_dir)
    existing = next(
        item
        for item in detect_statement_sheets(statement_path, rules).values()
        if isinstance(item, ExistingStatementResult)
    )
    opening = existing.values["CASH-OPENING"]
    closing = existing.values["CASH-CLOSING"]
    assert opening is not None and closing is not None
    supplement_cash_balances(
        run_dir,
        Decimal(opening) / 100,
        Decimal(closing) / 100,
        fx_amount,
        "用户确认：期初期末取客户现有正表，汇率影响取外部确认参数",
    )

    first_classification = run_classification(run_dir)
    if first_classification.status == "待摘要语义确认":
        _complete_summary_semantics(run_dir, summary_results_path)
        run_classification(run_dir)
    state = _state(run_dir)
    components = {item["component_id"]: item for item in state["components"]}
    allocations: dict[str, int] = {}
    allocation_pairs = set()
    for item in state["source_allocations"]:
        pair = (item["component_id"], item["entry_id"])
        assert pair not in allocation_pairs
        allocation_pairs.add(pair)
        allocations[item["component_id"]] = (
            allocations.get(item["component_id"], 0) + item["allocated_cent"]
        )
    assert all(
        allocations.get(component_id, 0) == component["cash_delta_cent"]
        for component_id, component in components.items()
    )
    assert all(
        item["evidence_score"] in ALLOWED_SCORES or item["excluded"]
        for item in state["decisions"]
    )
    assert all(item["decision_action"] for item in state["decisions"])
    assert "materiality_group_confirmation_requests" not in state
    assert "materiality_potential_group_warnings" not in state
    assert all(
        item["sources_independent"]
        for item in state["decisions"]
        if item["evidence_score"] in {70, 90}
    )
    summary_results = state["summary_semantics"]["results"]
    assert summary_results
    assert any(item["candidate_item_ids"] for item in summary_results)
    if len(summary_results) >= 100:
        assert any(
            item["evidence_score"] in {70, 90}
            for item in state["decisions"]
            if not item["excluded"]
        ), "大样本真实验收没有任何70分或90分决定，应先复核摘要语义结果，不能继续出具通过结论"

    for ai_results_path in ai_results_paths:
        imported = import_ai_results(run_dir, ai_results_path)
    assert imported.missing_count == 0
    final = finalize_run(run_dir)
    state = _state(run_dir)

    assert final.workbook_path.is_file()
    assert final.overall_status == "待完成人工确认"
    assert int(state["classification_summary"]["ai_tasks_missing"]) == 0
    unresolved = tuple(
        item for item in state["decisions"] if not item["resolved"] and not item["excluded"]
    )
    assert unresolved, "真实验收不得用自动化夹具冒充人工决定"
    decisions_by_component = {
        item["component_id"]: item for item in state["decisions"]
    }
    no_preferred_item = tuple(
        item
        for item in state["review_batches"]
        if not str(item.get("proposed_item_code") or "").strip()
    )
    assert all(
        all(
            decisions_by_component[component_id]["candidate_status"]
            == "invalid_input"
            for component_id in item["component_ids"]
        )
        and item["alternative_item_codes"]
        for item in no_preferred_item
    ), "只有摘要和原项目均无法形成首选的非法输入，才允许提示人工改选具体项目"

    workbook = load_workbook(final.workbook_path, data_only=False, read_only=False)
    difference_business_rows = 0
    try:
        assert tuple(workbook.sheetnames) == SHEET_NAMES
        assert "可靠同类组批量处理" not in workbook.sheetnames
        review = workbook["重要待复核事项"]
        review_headers = [cell.value for cell in review[1]]
        review_data_end_row = len(state["review_batches"]) + 1
        review_data_rows = range(2, review_data_end_row + 1)
        assert review_data_end_row > 1
        assert review.max_row >= review_data_end_row
        for header in (
            "日期",
            "本行摘要",
            "本行完整对方科目路径",
            "原项目标准化结果",
            "系统候选项目",
            "证据得分",
            "单笔重要性层级",
            "唯一动作",
            "人工确认项目",
            "人工依据",
            "处理人",
            "处理时间",
            "人工处理状态",
        ):
            assert header in review_headers
        assert "来源文件" not in review_headers
        included_count_column = review_headers.index("包含笔数(技术)") + 1
        assert sum(
            int(review.cell(row, included_count_column).value or 0)
            for row in review_data_rows
        ) == len(unresolved)
        debit_column = review_headers.index("借方") + 1
        credit_column = review_headers.index("贷方") + 1
        amount_column = review_headers.index("单笔金额") + 1
        allocated_column = review_headers.index("本行分配现金变化") + 1
        option_column = review_headers.index("人工可选标准项目") + 1
        assert all(
            review.cell(row, amount_column).data_type == "n"
            and {
                review.cell(row, column).number_format
                for column in (debit_column, credit_column, allocated_column)
                if review.cell(row, column).data_type == "n"
            }
            == {review.cell(row, amount_column).number_format}
            for row in review_data_rows
        )
        assert all(
            USE_SYSTEM_RECOMMENDATION
            not in str(review.cell(row, option_column).value or "")
            for row in review_data_rows
        )
        validations = review.data_validations.dataValidation
        choice_column = get_column_letter(review_headers.index("人工确认项目") + 1)
        for row, batch in zip(review_data_rows, state["review_batches"], strict=True):
            matching = [
                validation
                for validation in validations
                if f"{choice_column}{row}" in validation.sqref
            ]
            if batch.get("follows_component_id"):
                assert not matching
                assert review.cell(row, option_column).value == (
                    "随基础项目自动确定（无需重复选择）"
                )
                continue
            assert len(matching) == 1
            helper_range = str(matching[0].formula1).split("!")[-1].replace("'", "")
            first_cell = helper_range.split(":", 1)[0].replace("$", "")
            assert review[first_cell].value == USE_SYSTEM_RECOMMENDATION

        differences = workbook["原表与系统决定差异"]
        difference_headers = [cell.value for cell in differences[1]]
        result_column = difference_headers.index("审定现流表项目") + 1
        score_column = difference_headers.index("打分逻辑描述及打分结果") + 1
        assert all(
            not str(differences.cell(row, result_column).value or "").startswith("等待人工复核")
            for row in range(2, differences.max_row + 1)
        ), "待人工事项不是系统已决定差异，不得进入差异明细"
        assert all(
            "合计0分" not in str(differences.cell(row, score_column).value or "")
            for row in range(2, differences.max_row + 1)
            if differences.cell(row, result_column).value == "不进入正表"
        ), "内部划转不得显示为按0分修改项目"
        difference_business_rows = sum(
            bool(differences.cell(row, result_column).value)
            for row in range(2, differences.max_row + 1)
        )
        if state["internal_transfers"]:
            assert any(
                differences.cell(row, result_column).value == "不进入正表"
                and "不适用：内部划转"
                in str(differences.cell(row, score_column).value or "")
                for row in range(2, differences.max_row + 1)
            ), "存在内部划转时，差异明细必须明确说明分类评分不适用"

        trace = workbook["全量分类留痕"]
        trace_headers = [cell.value for cell in trace[1]]
        assert trace.max_row - 1 == len(state["source_allocations"])
        component_column = trace_headers.index("业务组成编号(技术)") + 1
        original_column = trace_headers.index("原现流项目") + 1
        original_by_component = {
            item["component_id"]: str(item.get("original_item_text") or "")
            for item in state["components"]
        }
        for row in range(2, trace.max_row + 1):
            component_id = str(trace.cell(row, component_column).value)
            expected_original = original_by_component[component_id]
            actual_original = str(trace.cell(row, original_column).value or "")
            assert actual_original == (expected_original or "原项目为空")

        ai_sheet = workbook["AI复核记录"]
        ai_headers = [cell.value for cell in ai_sheet[1]]
        required_ai_headers = (
            "review_round",
            "reviewer_id",
            "model_id",
            "reviewed_at",
            "prior_result_difference",
        )
        if set(required_ai_headers) <= set(ai_headers):
            ai_indexes = {header: ai_headers.index(header) + 1 for header in required_ai_headers}
            component_index = ai_headers.index("component_id") + 1
            reviewers_by_component: dict[str, dict[str, str]] = {}
            for row in range(2, ai_sheet.max_row + 1):
                assert all(
                    str(ai_sheet.cell(row, ai_indexes[header]).value or "").strip()
                    for header in required_ai_headers
                )
                component_id = str(ai_sheet.cell(row, component_index).value)
                review_round = str(ai_sheet.cell(row, ai_indexes["review_round"]).value)
                if review_round in {"A", "B"}:
                    reviewers_by_component.setdefault(component_id, {})[review_round] = str(
                        ai_sheet.cell(row, ai_indexes["reviewer_id"]).value
                    )
            assert all(
                rounds.get("A") != rounds.get("B")
                for rounds in reviewers_by_component.values()
                if {"A", "B"} <= set(rounds)
            )
        else:
            assert {"阶段", "task_id", "attempt", "status"} <= set(ai_headers)
            task_index = ai_headers.index("task_id") + 1
            attempt_index = ai_headers.index("attempt") + 1
            status_index = ai_headers.index("status") + 1
            attempts_by_task: dict[str, set[int]] = {}
            final_status_by_task: dict[str, str] = {}
            for row in range(2, ai_sheet.max_row + 1):
                task_id = str(ai_sheet.cell(row, task_index).value or "")
                attempt = int(ai_sheet.cell(row, attempt_index).value or 0)
                status = str(ai_sheet.cell(row, status_index).value or "")
                assert task_id and attempt in {1, 2, 3}
                attempts_by_task.setdefault(task_id, set()).add(attempt)
                if attempt == 3:
                    final_status_by_task[task_id] = status
            assert attempts_by_task
            assert all(attempts == {1, 2, 3} for attempts in attempts_by_task.values())
            assert set(final_status_by_task) == set(attempts_by_task)
            assert set(final_status_by_task.values()) == {"无有效结果"}

        cash_scope = workbook["现金范围与现金流量表与货币资金变动的勾稽核对"]
        scope_rows = tuple(
            (str(cash_scope.cell(row, 1).value), str(cash_scope.cell(row, 2).value))
            for row in range(2, cash_scope.max_row + 1)
            if cash_scope.cell(row, 1).value
        )
        expected_scope_labels = {
            "纳入" if decision == "include" else "不纳入"
            for decision in cash_choices.values()
        }
        assert expected_scope_labels <= {decision for _, decision in scope_rows}
    finally:
        workbook.close()

    cached = load_workbook(final.workbook_path, data_only=True, read_only=True)
    try:
        for sheet in cached.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    _ = cell.value
    finally:
        cached.close()

    after_hashes = {str(path.resolve()): _sha256(path) for path in inputs}
    assert after_hashes == before_hashes

    manifest = {
        "run_dir": str(run_dir),
        "workbook_path": str(final.workbook_path),
        "input_hashes_before": before_hashes,
        "input_hashes_after": after_hashes,
        "component_count": len(state["components"]),
        "source_entry_count": len(state["entries"]),
        "summary_count": len(summary_results),
        "summary_agent_task_count": len(state["summary_semantics"]["tasks"]),
        "summary_candidate_count": sum(
            bool(item["candidate_item_ids"]) for item in summary_results
        ),
        "summary_quality_distribution": dict(
            Counter(str(item["quality"]) for item in summary_results)
        ),
        "score_distribution": dict(
            Counter(
                str(item["evidence_score"])
                for item in state["decisions"]
                if not item["excluded"]
            )
        ),
        "score_70_or_90_count": sum(
            item["evidence_score"] in {70, 90}
            for item in state["decisions"]
            if not item["excluded"]
        ),
        "automatic_change_count": sum(
            item["decision_action"] == "automatic_change"
            for item in state["decisions"]
        ),
        "difference_business_rows": difference_business_rows,
        "workbook_sheets": list(SHEET_NAMES),
        "reconciliation_difference_cent": state["reconciliation"]["difference_cent"],
        "overall_status": final.overall_status,
        "versions": state["versions"],
        "pending_manual_decisions": len(unresolved),
        "manual_gate": "未冒充人工决定；须由用户完成重要待复核事项后才能进入最终可使用",
    }
    (run_dir / "真实验收记录.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8-sig",
    )


def test_manual_filtered_detail_has_only_the_user_confirmed_omissions() -> None:
    original_value = os.environ.get("CAS_CASHFLOW_REAL_DETAIL")
    filtered_value = os.environ.get("CAS_CASHFLOW_REAL_DETAIL_FILTERED")
    statement_value = os.environ.get("CAS_CASHFLOW_REAL_STATEMENT")
    dictionary_value = os.environ.get("CAS_CASHFLOW_REAL_DICTIONARY_RESULTS")
    summary_value = os.environ.get("CAS_CASHFLOW_REAL_SUMMARY_SLOT_RESULTS")
    expected_omissions_value = os.environ.get(
        "CAS_CASHFLOW_REAL_FILTERED_EXPECTED_OMISSIONS"
    )
    output_value = os.environ.get("CAS_CASHFLOW_REAL_OUTPUT")
    if not all(
        (
            original_value,
            filtered_value,
            statement_value,
            dictionary_value,
            expected_omissions_value,
            output_value,
        )
    ):
        pytest.skip("未设置两份等价明细的比较参数")

    original_path = Path(original_value)
    filtered_path = Path(filtered_value)
    statement_path = Path(statement_value)
    dictionary_path = Path(dictionary_value)
    summary_path = Path(summary_value) if summary_value else None
    expected_omissions_path = Path(expected_omissions_value)
    assert all(
        path.is_file()
        for path in (
            original_path,
            filtered_path,
            statement_path,
            dictionary_path,
            expected_omissions_path,
        )
    )
    before_hashes = {
        original_path: _sha256(original_path),
        filtered_path: _sha256(filtered_path),
    }

    rules = load_rule_pack(Path(__file__).resolve().parents[1])
    existing = next(
        item
        for item in detect_statement_sheets(statement_path, rules).values()
        if isinstance(item, ExistingStatementResult)
    )

    def classified_state(detail_path: Path) -> dict[str, object]:
        preflight = run_preflight(
            (detail_path, statement_path),
            _real_materiality(),
            Path(output_value),
            statement_path,
        )
        _complete_mapping(preflight.run_dir)
        _confirm_real_level1_accounts(preflight.run_dir)
        current = _state(preflight.run_dir)
        choices = _real_cash_scope_decisions(current)
        confirm_cash_scope(preflight.run_dir, choices)
        _complete_dictionary(preflight.run_dir, dictionary_path)
        supplement_cash_balances(
            preflight.run_dir,
            Decimal(existing.values["CASH-OPENING"]) / 100,
            Decimal(existing.values["CASH-CLOSING"]) / 100,
            _real_fx_amount(),
            "用户确认：期初期末取客户现有正表，汇率影响取外部确认参数",
        )
        first_classification = run_classification(preflight.run_dir)
        if first_classification.status == "待摘要语义确认":
            _complete_summary_semantics(preflight.run_dir, summary_path)
            run_classification(preflight.run_dir)
        return _state(preflight.run_dir)

    original_state = classified_state(original_path)
    filtered_state = classified_state(filtered_path)
    original_signatures = _component_signatures(original_state)
    filtered_signatures = _component_signatures(filtered_state)
    omissions_payload = json.loads(
        expected_omissions_path.read_text(encoding="utf-8-sig")
    )
    expected_omissions = Counter(
        {
            (
                str(item["voucher_date"]),
                str(item["voucher_no"]),
                str(item["summary"]),
                int(item["cash_delta_cent"]),
                tuple(str(account) for account in item["counterpart_accounts"]),
                str(item["original_item_text"]),
            ): int(item.get("count", 1))
            for item in omissions_payload["expected_missing_components"]
        }
    )

    def serializable(counter: Counter[tuple[object, ...]]) -> list[dict[str, object]]:
        return [
            {
                "voucher_date": signature[0],
                "voucher_no": signature[1],
                "summary": signature[2],
                "cash_delta_cent": signature[3],
                "counterpart_accounts": list(signature[4]),
                "original_item_text": signature[5],
                "count": count,
            }
            for signature, count in sorted(counter.items(), key=lambda item: repr(item[0]))
        ]

    comparison_record = {
        "original_file": str(original_path.resolve()),
        "filtered_file": str(filtered_path.resolve()),
        "original_sha256": before_hashes[original_path],
        "filtered_sha256": before_hashes[filtered_path],
        "original_component_count": len(original_state["components"]),
        "filtered_component_count": len(filtered_state["components"]),
        "unexpected_filtered_components": serializable(
            filtered_signatures - original_signatures
        ),
        "actual_missing_components": serializable(
            original_signatures - filtered_signatures
        ),
        "user_confirmed_missing_components": serializable(expected_omissions),
    }
    comparison_path = Path(output_value) / "手工筛选表通用差异验收.json"
    comparison_path.parent.mkdir(parents=True, exist_ok=True)
    comparison_path.write_text(
        json.dumps(comparison_record, ensure_ascii=False, indent=2),
        encoding="utf-8-sig",
    )

    assert not filtered_signatures - original_signatures
    assert original_signatures - filtered_signatures == expected_omissions
    for current in (original_state, filtered_state):
        allocated: dict[str, int] = {}
        for item in current["source_allocations"]:
            component_id = item["component_id"]
            allocated[component_id] = (
                allocated.get(component_id, 0) + item["allocated_cent"]
            )
        assert all(
            allocated.get(item["component_id"], 0) == item["cash_delta_cent"]
            for item in current["components"]
        )
    assert all(_sha256(path) == digest for path, digest in before_hashes.items())
