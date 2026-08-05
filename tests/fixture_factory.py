from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook
import xlsxwriter

from cashflow_direct.models import CashflowComponent, NormalizedEntry, SourceLocator


def write_complex_header_fixture(path: Path, header_row: int, label_side: str) -> None:
    """生成匿名的合并多行表头，不包含任何真实客户信息。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "随机页签甲" if header_row < 5 else "无规则名称乙"
    sheet.cell(1, 1, "匿名序时账导出")
    sheet.cell(header_row - 1, 1, "金额单位：元")
    sheet.merge_cells(start_row=header_row, start_column=1, end_row=header_row, end_column=2)
    sheet.cell(header_row, 1, "凭证信息")
    sheet.merge_cells(start_row=header_row, start_column=5, end_row=header_row, end_column=6)
    sheet.cell(header_row, 5, "发生额")
    sheet.cell(header_row + 1, 1, "凭证日期")
    sheet.cell(header_row + 1, 2, "凭证号")
    sheet.cell(header_row + 1, 3, "业务摘要")
    sheet.cell(header_row + 1, 4, "科目名称")
    sheet.cell(header_row + 1, 5, "借方发生额")
    sheet.cell(header_row + 1, 6, "贷方发生额")
    sheet.cell(header_row + 1, 7, "现金流量项目")
    data_row = header_row + 2
    sheet.append([])
    values = ["2026-01-03", "记-1", "匿名销售回款", "银行存款", 100, None, "销售商品收到的现金"]
    if label_side == "counterpart":
        values[3] = "应收账款"
    for column, value in enumerate(values, 1):
        sheet.cell(data_row, column, value)
    sheet.cell(data_row + 1, 1, "2026-01-04")
    sheet.cell(data_row + 1, 2, "记-2")
    sheet.cell(data_row + 1, 3, "匿名采购付款")
    sheet.cell(data_row + 1, 4, "原材料")
    sheet.cell(data_row + 1, 6, 80)
    sheet.cell(data_row + 1, 7, "购买商品支付的现金")
    workbook.save(path)


def write_ambiguous_money_fixture(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    headers = ["日期", "凭证号", "摘要", "科目", "借方发生额", "借方金额", "贷方发生额", "现流项目"]
    sheet.append(headers)
    sheet.append(["2026-01-03", "记-1", "匿名事项", "银行存款", 100, 100, None, "其他经营流入"])
    workbook.save(path)


def write_hostile_header_fixture(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "页签_7F3"
    sheet.cell(1, 1, "匿名导出文件")
    sheet.cell(2, 1, "金额单位：人民币元")
    sheet.merge_cells("A5:G5")
    sheet.cell(5, 1, "凭证及现金流数据")
    sheet.merge_cells("A6:B6")
    sheet.cell(6, 1, "凭证信息")
    sheet.merge_cells("E6:F6")
    sheet.cell(6, 5, "发生额")
    for column, value in enumerate(
        ["日期", "凭证编号", "摘要", "会计科目", "借方", "贷方", "主表项目"], 1
    ):
        sheet.cell(7, column, value)
    sheet.append([])
    for column, value in enumerate(
        ["2026-02-01", "转-1", "匿名收款", "银行存款", 20, None, "销售商品收到的现金"], 1
    ):
        sheet.cell(8, column, value)
    for column, value in enumerate(
        ["日期", "凭证编号", "摘要", "会计科目", "借方", "贷方", "主表项目"], 1
    ):
        sheet.cell(10, column, value)
    sheet.column_dimensions["H"].hidden = True
    sheet.cell(7, 8, "内部辅助列")
    workbook.save(path)


def _write_table(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "匿名数据"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def write_all_input_types(root: Path) -> tuple[Path, ...]:
    """生成五类结构等效匿名输入，名称和金额均为虚构。"""
    common = ["日期", "凭证号", "摘要", "科目", "对方科目", "借方", "贷方"]
    first = root / "类型甲.xlsx"
    _write_table(
        first,
        common,
        [
            ["2026-01-01", "记-1", "匿名收款", "银行存款", "应收款项", 120, None],
            ["2026-01-01", "记-1", "匿名收款", "应收款项", "银行存款", None, 120],
            ["2026-01-02", "记-2", "匿名付款", "应付款项", "银行存款", 40, None],
            ["2026-01-02", "记-2", "匿名付款", "银行存款", "应付款项", None, 40],
        ],
    )

    second = root / "类型乙.xlsx"
    _write_table(
        second,
        ["日期", "凭证号", "摘要", "科目", "借方", "贷方", "现流项目"],
        [
            ["2026-01-03", "记-3", "匿名销售", "银行存款", 90, None, "销售商品收到的现金"],
            ["2026-01-03", "记-3", "匿名销售", "主营业务收入", None, 90, "销售商品收到的现金"],
            ["2026-01-04", "记-4", "匿名采购", "原材料", 30, None, "购买商品支付的现金"],
            ["2026-01-04", "记-4", "匿名采购", "银行存款", None, 30, "购买商品支付的现金"],
            ["2026-01-05", "记-5", "匿名费用", "管理费用", 10, None, "支付其他经营现金"],
        ],
    )

    third = root / "类型丙.xlsx"
    _write_table(
        third,
        common + ["现流项目"],
        [
            ["2026-01-06", "记-6", "匿名收款", "银行存款", "应收款项", 70, None, "销售商品收到的现金"],
            ["2026-01-06", "记-6", "匿名收款", "应收款项", "银行存款", None, 70, "销售商品收到的现金"],
            ["2026-01-07", "记-7", "匿名付款", "应付款项", "银行存款", 25, None, "购买商品支付的现金"],
            ["2026-01-07", "记-7", "匿名付款", "银行存款", "应付款项", None, 25, "购买商品支付的现金"],
            ["2026-01-08", "记-8", "匿名往来", "银行存款", "其他往来", 15, None, "收到其他经营现金"],
            ["2026-01-08", "记-8", "匿名往来", "其他往来", "银行存款", None, 15, "收到其他经营现金"],
            ["日期", "凭证号", "摘要", "科目", "对方科目", "借方", "贷方", "现流项目"],
        ],
    )

    fourth = root / "类型丁.xlsx"
    _write_table(
        fourth,
        ["日期", "凭证号", "摘要", "科目", "发生额方向", "流量金额", "现流项目"],
        [
            ["2026-01-09", "记-9", "匿名收款", "银行存款", "借", 55, "收到其他经营现金"],
            ["2026-01-10", "记-10", "匿名付款", "应付款项", "借", 22, "购买商品支付的现金"],
            ["2026-01-11", "记-11", "匿名付款", "银行存款", "贷", 12, "支付其他经营现金"],
            ["2026-01-12", "记-12", "匿名收款", "其他往来", "贷", 8, "收到其他经营现金"],
        ],
    )

    fifth = root / "类型戊.xlsx"
    _write_table(
        fifth,
        ["发生额方向", "金额", "摘要", "现流项目"],
        [
            ["借", 18, "匿名收款一", "收到其他经营现金"],
            ["贷", 9, "匿名付款一", "支付其他经营现金"],
            ["借", -2, "匿名退款一", "收到其他经营现金"],
            ["贷", 3, "匿名付款二", "支付其他经营现金"],
        ],
    )
    return first, second, third, fourth, fifth


def _component_entry(
    row: int,
    voucher: str,
    account: str,
    debit_cent: int = 0,
    credit_cent: int = 0,
    item: str = "",
    flow_amount_cent: int = 0,
    retained_side: str = "unknown",
    counterpart_name: str = "",
    summary: str = "匿名业务",
) -> NormalizedEntry:
    return NormalizedEntry(
        entry_id=f"E{row}",
        source=SourceLocator("FSYN", "匿名页", row, row, f"A{row}:H{row}"),
        voucher_key=voucher,
        voucher_date="2026-03-01",
        voucher_no=voucher,
        summary=summary,
        account_name=account,
        counterpart_name=counterpart_name,
        debit_cent=debit_cent,
        credit_cent=credit_cent,
        flow_amount_cent=flow_amount_cent,
        original_flow_item=item,
        label_side=retained_side if item else "unknown",
        retained_side=retained_side,
    )


def component_entries(case: str) -> tuple[NormalizedEntry, ...]:
    cases = {
        "internal_and_external": (
            _component_entry(1, "V1", "1002 银行存款甲", credit_cent=100_000, retained_side="cash"),
            _component_entry(2, "V1", "1002 银行存款乙", debit_cent=130_000, retained_side="cash"),
            _component_entry(3, "V1", "主营业务收入", credit_cent=30_000, retained_side="counterpart"),
        ),
        "pure_internal": (
            _component_entry(4, "V2", "1002 银行存款甲", credit_cent=60_000, retained_side="cash"),
            _component_entry(5, "V2", "1002 银行存款乙", debit_cent=60_000, retained_side="cash"),
        ),
        "multi_project_receipt": (
            _component_entry(6, "V3", "1002 银行存款", debit_cent=9_052_530, retained_side="cash"),
            _component_entry(7, "V3", "往来科目甲", credit_cent=5_000_000, item="项目甲", retained_side="counterpart"),
            _component_entry(8, "V3", "往来科目乙", credit_cent=4_052_530, item="项目乙", retained_side="counterpart"),
        ),
        "one_sided_counterpart": (
            _component_entry(9, "V4", "应收款项", credit_cent=12_000, item="项目甲", flow_amount_cent=12_000, retained_side="counterpart"),
        ),
        "one_sided_counterpart_without_flow_amount": (
            _component_entry(
                28,
                "V15",
                "应收款项",
                credit_cent=12_000,
                item="销售商品、提供劳务收到的现金",
                retained_side="counterpart",
                counterpart_name="1002 银行存款",
            ),
        ),
        "one_sided_cash": (
            _component_entry(10, "V5", "1002 银行存款", debit_cent=14_000, item="项目甲", flow_amount_cent=14_000, retained_side="cash"),
        ),
        "summary_only": (
            _component_entry(11, "V6", "", debit_cent=16_000, item="项目甲", flow_amount_cent=16_000),
        ),
        "summary_only_counterpart_direction": (
            _component_entry(
                17,
                "V10",
                "",
                debit_cent=16_000,
                item="支付其他与经营活动有关的现金",
                flow_amount_cent=16_000,
            ),
        ),
        "flow_amount_differs": (
            _component_entry(12, "V7", "1002 银行存款", debit_cent=20_000, item="项目甲", flow_amount_cent=18_000, retained_side="cash"),
        ),
        "split_label_duplication": (
            _component_entry(13, "V8", "1002 银行存款", debit_cent=25_000, item="项目甲", retained_side="cash"),
            _component_entry(14, "V8", "应收款项", credit_cent=25_000, item="项目甲", retained_side="counterpart"),
        ),
        "unbalanced_cash_fact": (
            _component_entry(15, "V9", "1002 银行存款", debit_cent=30_000, retained_side="cash"),
            _component_entry(16, "V9", "其他科目", credit_cent=29_000, retained_side="counterpart"),
        ),
        "receipt_and_fee": (
            _component_entry(18, "V11", "1002 银行存款甲", debit_cent=1_000_000, retained_side="cash"),
            _component_entry(19, "V11", "主营业务收入", credit_cent=1_000_000, item="销售商品收到的现金", retained_side="counterpart"),
            _component_entry(20, "V11", "财务费用", debit_cent=500, item="支付其他与经营活动有关的现金", retained_side="counterpart"),
            _component_entry(21, "V11", "1002 银行存款乙", credit_cent=500, retained_side="cash"),
        ),
        "explicit_internal_transfer": (
            _component_entry(22, "V12", "1002 银行存款甲", credit_cent=60_000, retained_side="cash", counterpart_name="1002 银行存款乙"),
            _component_entry(23, "V12", "1002 银行存款乙", debit_cent=60_000, retained_side="cash", counterpart_name="1002 银行存款甲"),
        ),
        "principal_and_interest": (
            _component_entry(24, "V13", "1002 银行存款", credit_cent=110_000, retained_side="cash", summary="偿还本金及利息"),
            _component_entry(25, "V13", "短期借款", debit_cent=100_000, retained_side="counterpart", summary="偿还本金及利息"),
            _component_entry(26, "V13", "应付利息", debit_cent=10_000, retained_side="counterpart", summary="偿还本金及利息"),
        ),
        "purchase_with_input_vat": (
            _component_entry(29, "V16", "1002 银行存款", credit_cent=113_000, retained_side="cash", summary="采购原材料"),
            _component_entry(30, "V16", "原材料", debit_cent=100_000, retained_side="counterpart", summary="采购原材料"),
            _component_entry(31, "V16", "应交税费-应交增值税（进项税额）", debit_cent=13_000, retained_side="counterpart", summary="采购原材料"),
        ),
        "one_sided_internal_transfer": (
            _component_entry(
                27,
                "V14",
                "1002 银行存款甲",
                credit_cent=60_000,
                flow_amount_cent=60_000,
                retained_side="cash",
                counterpart_name="1002 银行存款乙",
            ),
        ),
    }
    return cases[case]


def cashflow_component(
    summary: str,
    cash_delta_cent: int,
    counterpart_accounts: tuple[str, ...] = ("普通往来科目",),
    original_item_text: str = "",
    anomalies: tuple[str, ...] = (),
    evidence_strength: str = "strong",
    component_id: str = "C-SYN",
) -> CashflowComponent:
    return CashflowComponent(
        component_id=component_id,
        voucher_key=f"V-{component_id}",
        summary=summary,
        cash_delta_cent=cash_delta_cent,
        counterpart_accounts=counterpart_accounts,
        original_item_text=original_item_text,
        source_keys=(f"E-{component_id}",),
        anomalies=anomalies,
        evidence_strength=evidence_strength,
    )


def ai_case(
    component_id: str,
    amount_cent: int,
    weak: bool,
    anomaly: bool,
    *,
    labeled: bool = True,
    summary_pattern: str = "普通往来",
    alternatives: tuple[str, ...] = ("CFO-03", "CFI-05"),
):
    from cashflow_direct.ai_review import build_ai_task
    from cashflow_direct.models import ClassificationDecision, UnresolvedDecision

    component = cashflow_component(
        summary="匿名往来事项",
        cash_delta_cent=amount_cent,
        original_item_text="收到其他经营现金" if labeled else "",
        anomalies=("direction_anomaly",) if anomaly else (),
        evidence_strength="weak" if weak else "strong",
        component_id=component_id,
    )
    decision = ClassificationDecision(
        component_id=component_id,
        system_item_id="CFO-03",
        system_item_name="收到其他与经营活动有关的现金",
        normal_direction="inflow",
        matched_rule_id="CFO-03-FALLBACK" if weak else "CFO-03-CURRENT",
        reason="匿名规则证据",
        evidence_level="low" if weak else "high",
    )
    task = build_ai_task(component, decision)
    unresolved = UnresolvedDecision(
        component_id=component_id,
        cash_delta_cent=amount_cent,
        cash_direction="inflow" if amount_cent >= 0 else "outflow",
        original_item=component.original_item_text,
        system_item_id=decision.system_item_id,
        adjudication_status="unresolved",
        counterpart_group="普通往来科目",
        summary_pattern=summary_pattern,
        alternative_item_ids=alternatives,
        reason="证据仍不足",
    )
    return SimpleNamespace(component=component, decision=decision, task=task, unresolved=unresolved)


def duplicate_components(
    amount_cent: int,
    *,
    first_date: str = "2026-04-01",
    second_date: str = "2026-04-01",
    first_voucher: str = "记-20",
    second_voucher: str = "记-20",
    first_summary: str = "匿名往来 收款",
    second_summary: str = "匿名往来，收款。",
    second_amount_cent: int | None = None,
    same_file: bool = False,
) -> tuple[CashflowComponent, CashflowComponent]:
    second_amount = amount_cent if second_amount_cent is None else second_amount_cent
    return (
        CashflowComponent(
            component_id="DUP-A",
            voucher_key="V-DUP-A",
            summary=first_summary,
            cash_delta_cent=amount_cent,
            counterpart_accounts=("普通往来科目",),
            original_item_text="收到其他经营现金",
            source_keys=("F1:E1",),
            voucher_date=first_date,
            voucher_no=first_voucher,
            source_file_ids=("F1",),
        ),
        CashflowComponent(
            component_id="DUP-B",
            voucher_key="V-DUP-B",
            summary=second_summary,
            cash_delta_cent=second_amount,
            counterpart_accounts=("普通往来科目",),
            original_item_text="收到其他经营现金",
            source_keys=(("F1:E2",) if same_file else ("F2:E1",)),
            voucher_date=second_date,
            voucher_no=second_voucher,
            source_file_ids=(("F1",) if same_file else ("F2",)),
        ),
    )


def classified_components():
    from cashflow_direct.classification import classify_all, load_rule_pack

    root = Path(__file__).resolve().parents[1]
    rules = load_rule_pack(root)
    components = (
        cashflow_component("销售商品收到货款", 1_000_000, component_id="S1"),
        cashflow_component("采购原材料付款", -400_000, component_id="S2"),
        cashflow_component("收回长期股权投资", 300_000, component_id="S3"),
        cashflow_component("取得银行借款", 200_000, component_id="S4"),
    )
    return SimpleNamespace(components=components, decisions=classify_all(components, rules), rules=rules)


def write_existing_statement_fixture(
    path: Path,
    header_row: int,
    with_custom_rows: bool,
    *,
    include_unknown: bool = False,
) -> None:
    from cashflow_direct.classification import load_rule_pack

    root = Path(__file__).resolve().parents[1]
    rules = load_rule_pack(root)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "报表页_随机"
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    sheet.cell(1, 1, "现金流量表")
    sheet.cell(2, 1, "金额单位：万元")
    for column, value in enumerate(["项目", "行次", "本期金额", "上期金额"], 1):
        sheet.cell(header_row, column, value)
    row = header_row + 1
    for item in rules.statement_items:
        name = "销售商品和提供劳务收到的现金" if item.item_id == "CFO-01" else item.name
        current = 0 if item.item_id == "CFO-02" else item.display_order / 100
        prior = None if item.item_id == "CFO-02" else item.display_order / 200
        sheet.cell(row, 1, name)
        sheet.cell(row, 2, item.display_order)
        sheet.cell(row, 3, current)
        sheet.cell(row, 4, prior)
        row += 1
        if with_custom_rows and item.item_id == "CFO-03":
            sheet.cell(row, 1, "收到匿名专项经营款")
            sheet.cell(row, 3, 0.01)
            row += 1
    if include_unknown:
        sheet.cell(row, 1, "无法归属的客户自定义总额")
        sheet.cell(row, 3, 1)
    workbook.save(path)


def workbook_model(review_batches: int, duplicate_groups: int):
    from cashflow_direct.duplicates import DuplicateGroup
    from cashflow_direct.models import ReviewBatch
    from cashflow_direct.statement import aggregate_statement
    from cashflow_direct.workbook_output import WorkbookModel

    case = classified_components()
    statement = aggregate_statement(case.components, case.decisions, case.rules)
    reviews = tuple(
        ReviewBatch(
            batch_id=f"REV-{index + 1}",
            component_ids=(f"RC-{index + 1}",),
            proposed_item_code="CFO-07",
            alternative_item_codes=("CFI-09",),
            worst_case_impact_cent=10_000 + index,
            reason="匿名重大剩余事项",
            baseline_statement_amount_cent=10_000 + index,
            cash_delta_cent=-(10_000 + index),
        )
        for index in range(review_batches)
    )
    duplicates = tuple(
        DuplicateGroup(
            group_id=f"DUP-{index + 1}",
            component_ids=(f"DA-{index + 1}", f"DB-{index + 1}"),
            component_amounts_cent=(20_000 + index, 20_000 + index),
            signature=("2026-01-01", "记-1", "匿名", "inflow", str(20_000 + index), "CFO-03"),
            default_decision="keep",
            worst_case_impact_cent=20_000 + index,
            blocks_manual_completion=True,
            item_id="CFO-03",
            baseline_statement_amount_cent=20_000 + index,
        )
        for index in range(duplicate_groups)
    )
    trace_rows = tuple(
        {
            "component_id": component.component_id,
            "summary": component.summary,
            "cash_delta_cent": component.cash_delta_cent,
            "system_item_id": decision.system_item_id,
            "reason": decision.reason,
        }
        for component, decision in zip(case.components, case.decisions, strict=True)
    )
    return WorkbookModel(
        statement=statement,
        rules=case.rules,
        comparison=None,
        review_batches=reviews,
        duplicate_groups=duplicates,
        ai_records=(),
        cash_scope_rows=({"科目":"1002 银行存款","决定":"纳入"},),
        reconciliation=None,
        trace_rows=trace_rows,
        mapping_rows=({"文件":"匿名输入.xlsx","工作表":"匿名数据","字段":"摘要","来源列":"C"},),
        overall_status="待处理重大事项" if reviews or duplicates else "自动计算完成",
    )


def write_end_to_end_case(
    root: Path,
    include_existing_statement: bool = False,
    *,
    include_cash_balances: bool = True,
) -> tuple[Path, ...]:
    transaction = root / "匿名序时账.xlsx"
    workbook = Workbook()
    journal = workbook.active
    journal.title = "随机序时数据"
    journal.append(["日期", "凭证号", "摘要", "科目", "借方", "贷方", "现流项目"])
    journal.append(["2026-01-01", "记-1", "匿名销售收款", "1002 银行存款", 100, None, "销售商品收到的现金"])
    journal.append(["2026-01-01", "记-1", "匿名销售收款", "主营业务收入", None, 100, "销售商品收到的现金"])
    journal.append(["2026-01-02", "记-2", "匿名采购付款", "原材料", 40, None, "购买商品支付的现金"])
    journal.append(["2026-01-02", "记-2", "匿名采购付款", "1002 银行存款", None, 40, "购买商品支付的现金"])
    if include_cash_balances:
        balance = workbook.create_sheet("现金余额资料")
        balance.append(["项目", "金额"])
        balance.append(["期初现金及现金等价物余额", 1000])
        balance.append(["期末现金及现金等价物余额", 1060])
        balance.append(["汇率变动对现金及现金等价物的影响", 0])
    workbook.save(transaction)
    inputs: list[Path] = [transaction]
    if include_existing_statement:
        existing = root / "匿名客户正表.xlsx"
        write_existing_statement_fixture(existing, 7, True)
        inputs.append(existing)
    return tuple(inputs)


def write_large_case(root: Path, row_count: int) -> tuple[Path, int]:
    """流式生成匿名大样例，并用整数分独立累计预期现金变化。"""
    if row_count <= 0 or row_count % 2:
        raise ValueError("大型样例行数必须是正偶数")
    path = root / "匿名十万行序时账.xlsx"
    workbook = xlsxwriter.Workbook(str(path), {"constant_memory": True})
    journal = workbook.add_worksheet("匿名全量序时数据")
    headers = ["日期", "凭证号", "摘要", "科目", "借方", "贷方", "现流项目"]
    for column, header in enumerate(headers):
        journal.write(0, column, header)
    patterns = (
        (100, "匿名销售收款", "销售商品收到的现金", "主营业务收入"),
        (-80, "匿名采购付款", "购买商品支付的现金", "原材料"),
        (20, "匿名税费返还", "收到的税费返还", "其他应收税款"),
        (-30, "匿名职工付款", "支付给职工的现金", "应付职工薪酬"),
        (-10, "匿名缴纳税费", "支付的各项税费", "应交税费"),
        (50, "匿名收回投资", "收回投资收到的现金", "长期股权投资"),
        (-40, "匿名购建设备", "购建固定资产支付的现金", "固定资产"),
        (60, "匿名取得借款", "取得借款收到的现金", "短期借款"),
        (-25, "匿名偿还借款", "偿还债务支付的现金", "长期借款"),
        (15, "匿名供应商退回采购款", "购买商品支付的现金", "应付账款"),
        (0, "匿名账户内部划转", "", "1001 库存现金"),
    )
    expected_cash_delta = 0
    output_row = 1
    for voucher_index in range(row_count // 2):
        amount_cent, summary, item, counterpart = patterns[voucher_index % len(patterns)]
        date_text = f"2026-{voucher_index % 12 + 1:02d}-{voucher_index % 28 + 1:02d}"
        voucher_no = f"记-{voucher_index + 1}"
        if amount_cent == 0:
            amount_yuan = 1
            first = [date_text, voucher_no, summary, "1002 银行存款", None, amount_yuan, item]
            second = [date_text, voucher_no, summary, counterpart, amount_yuan, None, item]
        elif amount_cent > 0:
            amount_yuan = amount_cent / 100
            first = [date_text, voucher_no, summary, "1002 银行存款", amount_yuan, None, item]
            second = [date_text, voucher_no, summary, counterpart, None, amount_yuan, item]
            expected_cash_delta += amount_cent
        else:
            amount_yuan = abs(amount_cent) / 100
            first = [date_text, voucher_no, summary, counterpart, amount_yuan, None, item]
            second = [date_text, voucher_no, summary, "1002 银行存款", None, amount_yuan, item]
            expected_cash_delta += amount_cent
        for record in (first, second):
            for column, value in enumerate(record):
                journal.write(output_row, column, value)
            output_row += 1

    opening_cent = 10_000_000
    balance = workbook.add_worksheet("现金余额资料")
    balance.write_row(0, 0, ["项目", "金额"])
    balance.write_row(1, 0, ["期初现金及现金等价物余额", opening_cent / 100])
    balance.write_row(2, 0, ["期末现金及现金等价物余额", (opening_cent + expected_cash_delta) / 100])
    balance.write_row(3, 0, ["汇率变动对现金及现金等价物的影响", 0])
    workbook.close()
    return path, expected_cash_delta


def write_ai_end_to_end_case(root: Path) -> Path:
    path = root / "匿名弱证据明细.xlsx"
    workbook = Workbook()
    detail = workbook.active
    detail.title = "匿名流量明细"
    detail.append(["发生额方向", "金额", "摘要", "现流项目"])
    detail.append(["借", 800000, "匿名往来款", "收到其他经营现金"])
    balance = workbook.create_sheet("现金余额资料")
    balance.append(["项目", "金额"])
    balance.append(["期初现金及现金等价物余额", 1000000])
    balance.append(["期末现金及现金等价物余额", 1800000])
    balance.append(["汇率变动对现金及现金等价物的影响", 0])
    workbook.save(path)
    return path


def write_ai_batch_case(root: Path, count: int = 26) -> Path:
    path = root / "多批匿名弱证据明细.xlsx"
    workbook = Workbook()
    detail = workbook.active
    detail.title = "匿名流量明细"
    detail.append(["发生额方向", "金额", "摘要", "现流项目"])
    for index in range(count):
        detail.append(["借", 800000, f"匿名往来款{index + 1}", "收到其他经营现金"])
    balance = workbook.create_sheet("现金余额资料")
    balance.append(["项目", "金额"])
    balance.append(["期初现金及现金等价物余额", 1000000])
    balance.append(["期末现金及现金等价物余额", 1000000 + count * 800000])
    balance.append(["汇率变动对现金及现金等价物的影响", 0])
    workbook.save(path)
    return path
