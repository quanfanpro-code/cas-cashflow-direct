from __future__ import annotations

import json
import sqlite3
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook
from openpyxl import load_workbook

import cashflow_direct.pipeline as pipeline_module
from cashflow_direct.pipeline import (
    confirm_company_notes,
    confirm_mapping,
    confirm_cash_scope,
    finalize_run,
    import_ai_results,
    import_dictionary_results,
    import_summary_results,
    run_classification,
    run_preflight,
    scan_accounts,
    supplement_cash_balances,
)
from tests.fixture_factory import (
    mark_dictionary_complete,
    write_ai_batch_case,
    write_ai_end_to_end_case,
    write_ambiguous_money_fixture,
    write_detail_plus_statement_fixture,
    write_end_to_end_case,
    write_existing_statement_fixture,
)
from cashflow_direct.models import CashflowComponent, ClassificationDecision
from cashflow_direct.summary_semantics import (
    analyze_summary,
    build_summary_agent_task,
    load_summary_rules,
)


ROOT = Path(__file__).resolve().parents[1]


class SummarySemanticsPipelineTests(unittest.TestCase):
    def test_summary_import_rejects_agent_accounting_decision_fields(self) -> None:
        unresolved = analyze_summary(
            "代甲方向乙方转处理尾款",
            load_summary_rules(ROOT),
        )
        task = build_summary_agent_task(unresolved)
        self.assertIsNotNone(task)
        state = {
            "summary_semantics": {
                "tasks": [task],
                "results": [pipeline_module._summary_result_to_dict(unresolved)],
                "missing_ids": [task["task_id"]],
            },
            "summary_semantics_completed": False,
        }

        with tempfile.TemporaryDirectory() as tmp:
            result_path = Path(tmp) / "摘要补槽位.jsonl"
            result_path.write_text(
                json.dumps(
                    {
                        "task_id": task["task_id"],
                        "summary": unresolved.summary,
                        "item_id": "CFO-07",
                        "spans": [],
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8-sig",
            )
            with (
                patch.object(pipeline_module, "_load_state", return_value=state),
                patch.object(pipeline_module, "_assert_inputs_unchanged"),
                patch.object(pipeline_module, "_save_state"),
            ):
                with self.assertRaisesRegex(ValueError, "不得返回"):
                    import_summary_results(Path(tmp), result_path)


def test_dictionary_display_uses_the_confirmed_full_path_result() -> None:
    confirmed = {
        "应付账款_应付设备款": {
            "semantic": "设备购建形成的应付款",
            "item_id": "CFI-06",
        }
    }

    result = pipeline_module._dictionary_display_result(
        confirmed,
        "应付账款_应付设备款",
        "客户应付账款_应付设备款",
    )

    assert result["semantic"] == "设备购建形成的应付款"
    assert result["item_id"] == "CFI-06"


def test_pipeline_builds_two_blind_followups_when_first_ai_cannot_confirm_change() -> None:
    component = CashflowComponent(
        component_id="CMP-FOLLOWUP",
        voucher_key="V-1",
        summary="支付设备尾款",
        cash_delta_cent=-100_000,
        counterpart_accounts=("应付账款_应付设备款",),
    )
    decision = ClassificationDecision(
        component_id=component.component_id,
        system_item_id="CFI-06",
        system_item_name="购建固定资产、无形资产和其他长期资产支付的现金",
        normal_direction="outflow",
        matched_rule_id="TEST",
        reason="首轮未确认保留原项目",
        evidence_level="medium",
        candidate_item_ids=("CFI-06", "CFO-04"),
        decision_action="ai_double_followup_review",
        ai_review_policy="valid_original_retention",
        resolved=False,
    )

    tasks = pipeline_module._build_pending_ai_followups(
        (decision,), (component,), (), ()
    )

    assert len(tasks) == 2
    assert "独立复核A" in tasks[0].context
    assert "独立复核B" in tasks[1].context


def test_pipeline_builds_only_round_c_after_two_blind_reviews_disagree() -> None:
    component = CashflowComponent(
        component_id="CMP-THIRD",
        voucher_key="V-2",
        summary="收到往来款",
        cash_delta_cent=100_000,
        counterpart_accounts=("其他应付款_往来款",),
    )
    decision = ClassificationDecision(
        component_id=component.component_id,
        system_item_id="CFO-03",
        system_item_name="收到其他与经营活动有关的现金",
        normal_direction="inflow",
        matched_rule_id="TEST",
        reason="前两份互盲结果不一致",
        evidence_level="weak",
        candidate_item_ids=("CFO-01", "CFO-03"),
        decision_action="ai_third_review",
        ai_review_policy="blank_low_majority",
        resolved=False,
    )

    tasks = pipeline_module._build_pending_ai_followups(
        (decision,), (component,), (), ()
    )

    assert len(tasks) == 1
    assert "独立复核C" in tasks[0].context
    assert "不得查看独立复核A、B结果" in tasks[0].context


def test_overall_materiality_blank_score_90_ai_fill_is_not_forced_back_to_human() -> None:
    decision = ClassificationDecision(
        component_id="CMP-M3-90",
        system_item_id="CFO-01",
        system_item_name="销售商品、提供劳务收到的现金",
        normal_direction="inflow",
        matched_rule_id="TEST",
        reason="90分且AI确认",
        evidence_level="strong",
        evidence_score=90,
        original_item_state="blank",
        decision_action="automatic_fill",
        decision_source="ai_reviewed_system_decision",
        resolved=True,
    )

    assert pipeline_module._requires_overall_manual_override(decision) is False


def test_overall_materiality_blank_score_90_ai_fill_is_not_listed_for_manual_review() -> None:
    decision = ClassificationDecision(
        component_id="CMP-M3-BLANK-90",
        system_item_id="CFO-01",
        system_item_name="销售商品、提供劳务收到的现金",
        normal_direction="inflow",
        matched_rule_id="AI",
        reason="两个独立强来源经AI确认",
        evidence_level="strong",
        decision_source="ai_reviewed_system_decision",
        resolved=True,
        evidence_score=90,
        original_item_state="blank",
        decision_action="automatic_fill",
    )

    assert pipeline_module._needs_overall_manual_listing(
        decision, 2_200_000_00, 2_200_000_00
    ) is False


def test_overall_materiality_other_nonmanual_decisions_still_require_human() -> None:
    decision = ClassificationDecision(
        component_id="CMP-M3-OTHER",
        system_item_id="CFO-01",
        system_item_name="销售商品、提供劳务收到的现金",
        normal_direction="inflow",
        matched_rule_id="TEST",
        reason="不属于90分空白项目例外",
        evidence_level="strong",
        evidence_score=70,
        original_item_state="blank",
        decision_action="automatic_fill",
        decision_source="ai_reviewed_system_decision",
        resolved=True,
    )

    assert pipeline_module._requires_overall_manual_override(decision) is True


def _structured_ai_payload(
    task: dict[str, object],
    *,
    summary_quality: str = "strong",
    account_quality: str = "strong",
    candidate_item_id: str | None = None,
) -> dict[str, object]:
    context = str(task["context"])
    summary = context.split("摘要原文：", 1)[1].split("；", 1)[0]
    account_path = context.split("完整对方科目路径：", 1)[1].split("；", 1)[0]
    candidate = candidate_item_id or str(task["candidate_item_ids"][0])
    review_round = (
        "A"
        if "；独立复核A：" in context
        else "B"
        if "；独立复核B：" in context
        else "C"
        if "；独立复核C：" in context
        else "single"
    )
    return {
        "task_id": task["task_id"],
        "component_id": task["component_id"],
        "summary": {
            "candidate_item_id": candidate,
            "quality": summary_quality,
            "basis_text": summary,
            "classification_facts": [f"summary:{summary}"],
            "conflict": False,
        },
        "account_path": {
            "candidate_item_id": candidate,
            "quality": account_quality,
            "basis_text": account_path,
            "classification_facts": [f"account:{account_path}"],
            "conflict": False,
        },
        "sources_independent": True,
        "business_conflict": False,
        "direction_status": "compatible",
        "reason": "只重新解释本行摘要和完整对方科目路径",
        "alternative_item_ids": [],
        "note_ids": [],
        "review_round": review_round,
        "reviewer_id": f"test-reviewer-{review_round}-{task['task_id']}",
        "model_id": "test-model",
        "reviewed_at": "2026-08-21T00:00:00+08:00",
        "prior_result_difference": (
            "互盲复核，未查看另一结果"
            if review_round in {"A", "B", "C"}
            else "第二轮重新核对原文结构与第一次解释的差异"
            if review_round == "second"
            else "首轮复核，无前序结果"
        ),
    }


def _complete_all_pending_ai(run_dir: Path, root: Path) -> None:
    state_path = run_dir / "计算留痕数据" / "运行状态.json"
    for round_no in range(1, 6):
        state = json.loads(state_path.read_text(encoding="utf-8-sig"))
        completed_ids = {
            item["task_id"]
            for item in state.get("structured_ai_validation", {}).get(
                "valid_results", ()
            )
        }
        pending = [
            task for task in state.get("ai_tasks", ()) if task["task_id"] not in completed_ids
        ]
        if not pending:
            return
        result_path = root / f"自动完成AI复核_{round_no}.jsonl"
        result_path.write_text(
            "".join(
                json.dumps(_structured_ai_payload(task), ensure_ascii=False) + "\n"
                for task in pending
            ),
            encoding="utf-8-sig",
        )
        import_ai_results(run_dir, result_path)
    raise AssertionError("测试用AI复核在五轮内仍未完成")


class PipelineTests(unittest.TestCase):
    def test_flow_detail_without_confirmed_cash_leg_requests_row_cleanup(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "现流明细加排除账户完整分录.xlsx"
            workbook = Workbook()
            journal = workbook.active
            journal.title = "现流明细"
            journal.append(
                [
                    "日期", "凭证字", "凭证号", "摘要", "科目编码", "科目",
                    "借方", "贷方", "流量金额", "现流项目",
                ]
            )
            journal.append(
                [
                    "2026/1/1", "记", "1", "支付货款", "2202.01", "应付账款_货款",
                    1000, None, 1000, "购买商品、接受劳务支付的现金",
                ]
            )
            journal.append(
                [
                    "2026/1/2", "记", "2", "转入理财户", "1101.02", "交易性金融资产_一般户",
                    500, None, 500, "收到其他与投资活动有关的现金",
                ]
            )
            journal.append(
                [
                    "2026/1/3", "记", "3", "保证金账户收款", "1012.01", "其他货币资金_保证金",
                    300, None, 300, "收到其他与经营活动有关的现金",
                ]
            )
            balances = workbook.create_sheet("现金余额资料")
            balances.append(["项目", "金额"])
            balances.append(["期初现金及现金等价物余额", 1000])
            balances.append(["汇率变动对现金及现金等价物的影响", 0])
            balances.append(["期末现金及现金等价物余额", 0])
            workbook.save(source)
            workbook.close()

            preflight = run_preflight([source], ("100000", "50000", "5000"), root)
            preflight_state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(
                    encoding="utf-8-sig"
                )
            )
            confirm_cash_scope(
                preflight.run_dir,
                {
                    item["account_key"]: "exclude"
                    for item in preflight_state["cash_scope_proposal"]["candidates"]
                },
            )
            mark_dictionary_complete(preflight.run_dir)

            result = run_classification(preflight.run_dir)
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(
                    encoding="utf-8-sig"
                )
            )

            self.assertEqual("待用户清洗现金分录", result.status)
            self.assertEqual(0, result.component_count)
            self.assertEqual("waiting_cash_row_cleanup", state["stage"])
            self.assertTrue(state["cash_row_cleanup_requests"])
            self.assertTrue((preflight.run_dir / "现金分录清洗请求.md").is_file())

    def test_standard_original_flow_item_cannot_bypass_cash_row_cleanup(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "现金分录不明确.xlsx"
            workbook = Workbook()
            journal = workbook.active
            journal.title = "现流明细"
            journal.append(
                [
                    "日期",
                    "凭证字",
                    "凭证号",
                    "摘要",
                    "科目编码",
                    "科目",
                    "借方",
                    "贷方",
                    "流量金额",
                    "现流项目",
                ]
            )
            journal.append(
                [
                    "2026/1/1",
                    "记",
                    "1",
                    "票据背书",
                    "1121",
                    "应收票据",
                    None,
                    1000,
                    1000,
                    "销售商品、提供劳务收到的现金",
                ]
            )
            balances = workbook.create_sheet("现金余额资料")
            balances.append(["项目", "金额"])
            balances.append(["期初现金及现金等价物余额", 0])
            balances.append(["汇率变动对现金及现金等价物的影响", 0])
            balances.append(["期末现金及现金等价物余额", 0])
            workbook.save(source)
            workbook.close()

            preflight = run_preflight([source], ("100000", "50000", "5000"), root)
            confirm_cash_scope(
                preflight.run_dir, preflight.recommended_cash_decisions
            )

            result = run_classification(preflight.run_dir)

            self.assertEqual("待用户清洗现金分录", result.status)
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(
                    encoding="utf-8-sig"
                )
            )
            self.assertEqual("waiting_cash_row_cleanup", state["stage"])
            self.assertTrue(state["cash_row_cleanup_requests"])
            self.assertTrue((preflight.run_dir / "现金分录清洗请求.md").is_file())

    def test_cash_scope_sheet_lists_confirmed_included_and_excluded_accounts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "两个现金账户.xlsx"
            source_workbook = Workbook()
            journal = source_workbook.active
            journal.title = "序时账"
            journal.append(
                [
                    "日期",
                    "凭证字",
                    "凭证号",
                    "摘要",
                    "科目编码",
                    "科目",
                    "借方",
                    "贷方",
                    "流量金额",
                    "现流项目",
                ]
            )
            journal.append(
                [
                    "2026/1/1",
                    "记",
                    "1",
                    "支付材料采购款",
                    "1002.01",
                    "银行存款_一般户",
                    None,
                    1000,
                    1000,
                    "购买商品、接受劳务支付的现金",
                ]
            )
            journal.append(
                [
                    "2026/1/1",
                    "记",
                    "1",
                    "支付材料采购款",
                    "2202.01",
                    "应付账款_应付商品款",
                    1000,
                    None,
                    1000,
                    "购买商品、接受劳务支付的现金",
                ]
            )
            journal.append(
                [
                    "2026/1/2",
                    "记",
                    "2",
                    "保证金账户付款",
                    "1002.02",
                    "银行存款_保证金账户",
                    None,
                    500,
                    500,
                    "支付其他与经营活动有关的现金",
                ]
            )
            journal.append(
                [
                    "2026/1/2",
                    "记",
                    "2",
                    "保证金账户付款",
                    "2241.01",
                    "其他应付款_保证金",
                    500,
                    None,
                    500,
                    "支付其他与经营活动有关的现金",
                ]
            )
            balances = source_workbook.create_sheet("现金余额资料")
            balances.append(["项目", "金额"])
            balances.append(["期初现金及现金等价物余额", 1000])
            balances.append(["汇率变动对现金及现金等价物的影响", 0])
            balances.append(["期末现金及现金等价物余额", 0])
            source_workbook.save(source)
            source_workbook.close()
            preflight = run_preflight(
                [source],
                ("1000000", "750000", "50000"),
                output_parent=root,
            )
            state_path = preflight.run_dir / "计算留痕数据" / "运行状态.json"
            state = json.loads(state_path.read_text(encoding="utf-8-sig"))
            decisions = {}
            excluded_key = ""
            for candidate in state["cash_scope_proposal"]["candidates"]:
                key = candidate["account_key"]
                if any("保证金" in name for name in candidate["account_names"]):
                    decisions[key] = "exclude"
                    excluded_key = key
                else:
                    decisions[key] = "include"
            confirm_cash_scope(preflight.run_dir, decisions)
            mark_dictionary_complete(preflight.run_dir)
            classified = run_classification(preflight.run_dir)
            if classified.ai_tasks_missing:
                tasks = json.loads(state_path.read_text(encoding="utf-8-sig"))[
                    "ai_tasks"
                ]
                ai_path = root / "AI结果.jsonl"
                ai_path.write_text(
                    "\n".join(
                        json.dumps(_structured_ai_payload(task), ensure_ascii=False)
                        for task in tasks
                    )
                    + "\n",
                    encoding="utf-8-sig",
                )
                import_ai_results(preflight.run_dir, ai_path)
            final = finalize_run(preflight.run_dir)
            workbook = load_workbook(final.workbook_path, data_only=False)
            try:
                sheet = workbook[
                    "现金范围与现金流量表与货币资金变动的勾稽核对"
                ]
                rows = {
                    str(sheet.cell(row, 1).value): sheet.cell(row, 2).value
                    for row in range(2, sheet.max_row + 1)
                }
                self.assertIn("不纳入", rows.values())
                excluded_names = next(
                    names
                    for key, names in json.loads(state_path.read_text(encoding="utf-8-sig"))[
                        "cash_scope"
                    ]["account_names_by_key"]
                    if key == excluded_key
                )
                self.assertIn("、".join(excluded_names), rows)
            finally:
                workbook.close()

    def test_original_and_auto_item_difference_is_exported_without_affecting_result(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "原项目与自动判定不一致.xlsx"
            workbook = Workbook()
            detail = workbook.active
            detail.title = "序时账"
            detail.append(
                [
                    "日期",
                    "凭证字",
                    "凭证号",
                    "摘要",
                    "科目编码",
                    "科目名称",
                    "借方",
                    "贷方",
                    "主表项目名称",
                ]
            )
            detail.append(
                ["2026/1/1", "记", "1", "支付税收滞纳金", "1002", "银行存款", None, 100, None]
            )
            detail.append(
                [
                    "2026/1/1",
                    "记",
                    "1",
                    "支付税收滞纳金",
                    "6711",
                    "营业外支出_罚款、滞纳金",
                    100,
                    None,
                    "客户自定义项目",
                ]
            )
            balances = workbook.create_sheet("现金余额资料")
            balances.append(["项目", "金额"])
            balances.append(["期初现金及现金等价物余额", 1_000])
            balances.append(["汇率变动影响", 0])
            balances.append(["期末现金及现金等价物余额", 900])
            workbook.save(source)
            workbook.close()

            preflight = run_preflight(
                [source], ("1000000", "750000", "50000"), output_parent=root
            )
            confirm_cash_scope(preflight.run_dir, preflight.recommended_cash_decisions)
            supplement_cash_balances(
                preflight.run_dir, "1000", "900", "0", "测试现金余额"
            )
            # 本用例不涉及词典机制，直接标记齐备（门禁通行）
            mark_dictionary_complete(preflight.run_dir)
            run_classification(preflight.run_dir)
            _complete_all_pending_ai(preflight.run_dir, root)
            final = finalize_run(preflight.run_dir)

            self.assertEqual("最终可使用", final.overall_status)
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(
                    encoding="utf-8-sig"
                )
            )
            self.assertEqual(0, state["reconciliation"]["difference_cent"])
            workbook = load_workbook(final.workbook_path, data_only=False)
            try:
                sheet = workbook["原表与系统决定差异"]
                headers = [cell.value for cell in sheet[1]]
                row = {
                    header: sheet.cell(2, column).value
                    for column, header in enumerate(headers, start=1)
                }
                self.assertEqual(2, sheet.max_row)
                self.assertEqual("记", row["凭证字"])
                self.assertEqual("1", str(row["凭证号"]))
                self.assertEqual("6711", str(row["科目编码"]))
                self.assertEqual("原项目无法标准化", row["原项目标准化结果"])
                self.assertEqual("支付其他与经营活动有关的现金", row["审定现流表项目"])
                self.assertEqual(100, row["借方"])
                self.assertIsNone(row["流量金额（原币）"])
                self.assertEqual(
                    "证据得分90分；金额档位为低于明显微小错报临界值；符合自动修改条件。",
                    row["差异形成原因"],
                )
            finally:
                workbook.close()

    def test_unresolved_consistency_group_does_not_duplicate_human_review_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "待人工业务组.xlsx"
            workbook = Workbook()
            detail = workbook.active
            detail.title = "明细"
            detail.append(
                ["日期", "凭证字", "凭证号", "分录号", "摘要", "科目编码", "科目", "对方科目", "借方", "贷方", "流量金额", "现流项目"]
            )
            detail.append(
                ["2026/6/15", "记", "70", "2", "支付款项", "2001.03", "短期借款_财务", "银行存款", None, 60_000, 60_000, None]
            )
            detail.append(
                ["2026/6/15", "记", "70", "3", "支付款项", "2211.03", "应付职工薪酬_财务", "银行存款", None, 60_000, 60_000, None]
            )
            balances = workbook.create_sheet("现金余额资料")
            balances.append(["项目", "金额"])
            balances.append(["期初现金及现金等价物余额", 0])
            balances.append(["汇率变动影响", 0])
            balances.append(["期末现金及现金等价物余额", 120_000])
            workbook.save(source)
            workbook.close()

            preflight = run_preflight([source], ("100000", "50000", "5000"), root)
            confirm_cash_scope(
                preflight.run_dir, preflight.recommended_cash_decisions
            )
            supplement_cash_balances(preflight.run_dir, "0", "120000", "0", "测试现金余额")
            # 本用例不涉及词典机制，直接标记齐备（门禁通行）
            mark_dictionary_complete(preflight.run_dir)
            run_classification(preflight.run_dir)
            state_path = preflight.run_dir / "计算留痕数据" / "运行状态.json"
            state = json.loads(state_path.read_text(encoding="utf-8-sig"))
            result_path = root / "结构化逐条结果.jsonl"
            result_path.write_text(
                "".join(
                    json.dumps(_structured_ai_payload(task), ensure_ascii=False)
                    + "\n"
                    for task in state["ai_tasks"]
                ),
                encoding="utf-8-sig",
            )
            import_ai_results(preflight.run_dir, result_path)
            state = json.loads(state_path.read_text(encoding="utf-8-sig"))
            self.assertEqual([], state.get("consistency_groups", []))

    def test_same_summary_with_different_counterpart_paths_never_forms_one_ai_group(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "同一业务组.xlsx"
            workbook = Workbook()
            detail = workbook.active
            detail.title = "明细"
            detail.append(
                ["日期", "凭证字", "凭证号", "分录号", "摘要", "科目编码", "科目", "对方科目", "借方", "贷方", "流量金额", "现流项目"]
            )
            detail.append(
                ["2026/6/15", "记", "70", "2", "支付款项", "2001.03", "短期借款_财务", "银行存款", None, 60_000, 60_000, None]
            )
            detail.append(
                ["2026/6/15", "记", "71", "3", "支付款项", "2211.03", "应付职工薪酬_财务", "银行存款", None, 60_000, 60_000, None]
            )
            balances = workbook.create_sheet("现金余额资料")
            balances.append(["项目", "金额"])
            balances.append(["期初现金及现金等价物余额", 0])
            balances.append(["汇率变动影响", 0])
            balances.append(["期末现金及现金等价物余额", 120_000])
            workbook.save(source)
            workbook.close()

            preflight = run_preflight(
                [source], ("100000", "50000", "5000"), root
            )
            confirm_cash_scope(
                preflight.run_dir, preflight.recommended_cash_decisions
            )
            supplement_cash_balances(
                preflight.run_dir, "0", "120000", "0", "测试现金余额"
            )
            # 本用例不涉及词典机制，直接标记齐备（门禁通行）
            mark_dictionary_complete(preflight.run_dir)
            classified = run_classification(preflight.run_dir)
            # 两笔按各自完整路径分别形成A、B两份互盲任务，不能因摘要相同而合并。
            self.assertEqual(4, classified.ai_tasks_missing)
            state_path = preflight.run_dir / "计算留痕数据" / "运行状态.json"
            state = json.loads(state_path.read_text(encoding="utf-8-sig"))

            result_path = root / "结构化逐条结果.jsonl"
            result_path.write_text(
                "".join(
                    json.dumps(
                        {
                            **_structured_ai_payload(
                                task,
                                candidate_item_id=task[
                                    "account_path_candidate_item_ids"
                                ][0],
                            ),
                            "sources_independent": False,
                            "summary": {
                                "candidate_item_id": "",
                                "candidate_item_ids": [],
                                "quality": "invalid",
                                "basis_text": "",
                                "classification_facts": [],
                                "conflict": False,
                            },
                        },
                        ensure_ascii=False,
                    )
                    + "\n"
                    for task in state["ai_tasks"]
                ),
                encoding="utf-8-sig",
            )
            imported = import_ai_results(preflight.run_dir, result_path)
            validation_state = json.loads(
                state_path.read_text(encoding="utf-8-sig")
            )["structured_ai_validation"]
            self.assertEqual(
                [],
                validation_state["invalid_ids"],
                [
                    (
                        task["task_id"],
                        task["candidate_item_ids"],
                        task["account_path_candidate_item_ids"],
                        task["context"],
                    )
                    for task in state["ai_tasks"]
                ],
            )
            self.assertEqual([], validation_state["missing_ids"])
            self.assertEqual("AI 已完成", imported.status)
            state = json.loads(state_path.read_text(encoding="utf-8-sig"))
            # 相同摘要不能跨越不同的完整对方科目路径拼成一个证据组。
            self.assertEqual([], state.get("consistency_groups", []))
            self.assertEqual(
                2,
                len({item["system_item_id"] for item in state["decisions"]}),
            )

            final = finalize_run(preflight.run_dir)
            self.assertEqual("最终可使用", final.overall_status)

    def test_statement_path_is_honored_and_failure_stops(self) -> None:
        # 指定的文件识别不出正表 → 报错停下（不静默降级为编制模式）
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (journal,) = write_end_to_end_case(root)
            bogus = root / "指定但不是正表.xlsx"
            workbook = Workbook()
            workbook.active.append(["不是", "正表", "内容"])
            workbook.save(bogus)
            with self.assertRaises(ValueError):
                run_preflight(
                    [journal, bogus], ("1000000", "750000", "50000"), statement_path=bogus
                )

    def test_statement_path_success_skips_auto_detection_for_others(self) -> None:
        # 指定合法正表：该文件走正表识别并登记，其余文件只按明细处理
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            journal, existing = write_end_to_end_case(root, include_existing_statement=True)
            preflight = run_preflight(
                [journal, existing],
                ("1000000", "750000", "50000"),
                output_parent=root,
                statement_path=existing,
            )
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(
                    encoding="utf-8-sig"
                )
            )
            self.assertEqual(str(existing), state["existing_statement_path"])
            self.assertIn(str(state["files"][0]["file_id"]), state["evidence_profiles"])
            profile = state["evidence_profiles"][str(state["files"][0]["file_id"])]
            self.assertTrue(profile["has_flow_item"])
            self.assertTrue(profile["full_voucher"])
            self.assertEqual(340_000, state["cash_balances"]["opening_cent"])
            self.assertEqual(350_000, state["cash_balances"]["closing_cent"])
            self.assertEqual(320_000, state["cash_balances"]["fx_cent"])

    def test_statement_path_outside_inputs_raises(self) -> None:
        # 指定的正表文件不在已选输入中 → 必须报错，不得静默跳过
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (journal,) = write_end_to_end_case(root)
            missing = root / "未选中的文件.xlsx"
            with self.assertRaises(ValueError):
                run_preflight(
                    [journal], ("1000000", "750000", "50000"), statement_path=missing
                )

    def test_trace_marks_direction_source(self) -> None:
        # 全量分类留痕表含"现金方向依据"列，序时账明细（无流量金额列）标"借贷差额"
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            inputs = write_end_to_end_case(root)
            preflight = run_preflight(
                inputs, ("1000000", "750000", "50000"), output_parent=root
            )
            confirm_cash_scope(preflight.run_dir, preflight.recommended_cash_decisions)
            mark_dictionary_complete(preflight.run_dir)
            run_classification(preflight.run_dir)
            _complete_all_pending_ai(preflight.run_dir, root)
            final = finalize_run(preflight.run_dir)
            workbook = load_workbook(final.workbook_path, data_only=False)
            try:
                sheet = workbook["全量分类留痕"]
                headers = [cell.value for cell in sheet[1]]
                self.assertIn("现金方向依据", headers)
                column = headers.index("现金方向依据") + 1
                values = {
                    sheet.cell(row, column).value for row in range(2, sheet.max_row + 1)
                }
                self.assertIn("借贷差额", values)
            finally:
                workbook.close()

    def test_multiple_statement_sheet_names_survive_preflight_for_user_action(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "多张客户正表.xlsx"
            write_existing_statement_fixture(source, header_row=7, with_custom_rows=False)
            workbook = load_workbook(source)
            duplicate = workbook.copy_worksheet(workbook.worksheets[0])
            duplicate.title = "另一张现金流量表"
            workbook.save(source)
            workbook.close()

            preflight = run_preflight(
                [source], ("1000000", "750000", "50000"), output_parent=root
            )
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(
                    encoding="utf-8-sig"
                )
            )
            question = state["mapping_questions"][0]
            self.assertEqual("statement_sheet", question["role"])
            self.assertIn("报表页_随机", question["message"])
            self.assertIn("另一张现金流量表", question["message"])
            with self.assertRaisesRegex(RuntimeError, "报表页_随机.*另一张现金流量表"):
                confirm_mapping(preflight.run_dir, {})

    def test_ai_result_batches_accumulate_without_resubmitting_completed_ids(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            preflight = run_preflight(
                [write_ai_batch_case(root)],
                ("100000000", "75000000", "50000"),
                root,
            )
            confirm_cash_scope(
                preflight.run_dir, preflight.recommended_cash_decisions
            )
            mark_dictionary_complete(preflight.run_dir)
            classified = run_classification(preflight.run_dir)
            self.assertEqual(52, classified.ai_tasks_missing)
            state_path = preflight.run_dir / "计算留痕数据" / "运行状态.json"
            tasks = json.loads(state_path.read_text(encoding="utf-8-sig"))["ai_tasks"]

            def write_results(path: Path, selected: list[dict[str, object]]) -> None:
                path.write_text(
                    "".join(
                        json.dumps(
                            _structured_ai_payload(item),
                            ensure_ascii=False,
                        ) + "\n"
                        for item in selected
                    ),
                    encoding="utf-8-sig",
                )

            first_path = root / "第一批结果.jsonl"
            write_results(first_path, tasks[:25])
            first = import_ai_results(preflight.run_dir, first_path)
            self.assertEqual(25, first.valid_count)
            self.assertEqual(27, first.missing_count)

            second_path = root / "第二批结果.jsonl"
            write_results(second_path, tasks[25:])
            second = import_ai_results(preflight.run_dir, second_path)
            self.assertEqual(52, second.valid_count)
            self.assertEqual(0, second.missing_count)
            self.assertEqual("AI 已完成", second.status)

            conflicting_path = root / "冲突结果.jsonl"
            conflicting_path.write_text(
                json.dumps(
                    _structured_ai_payload(
                        tasks[0], summary_quality="medium"
                    ),
                    ensure_ascii=False,
                ) + "\n",
                encoding="utf-8-sig",
            )
            with self.assertRaisesRegex(ValueError, "冲突"):
                import_ai_results(preflight.run_dir, conflicting_path)

    def test_preflight_reads_every_data_sheet_and_keeps_row_errors(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "按月序时账.xlsx"
            workbook = Workbook()
            for index, title in enumerate(("一月", "二月")):
                sheet = workbook.active if index == 0 else workbook.create_sheet()
                sheet.title = title
                sheet.append(["日期", "凭证号", "摘要", "科目", "借方", "贷方"])
                amount = 100 if index == 0 else "坏金额"
                sheet.append([f"2026-0{index + 1}-01", "记-1", "匿名收款", "银行存款", amount, None])
            workbook.save(source)
            preflight = run_preflight(
                [source], ("1000000", "750000", "50000"), output_parent=root
            )
            self.assertEqual(1, preflight.source_entry_count)
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(
                    encoding="utf-8-sig"
                )
            )
            self.assertEqual({"一月", "二月"}, {item["sheet"] for item in state["mappings"]})
            self.assertTrue(any(item["kind"] == "错误" for item in state["normalization_issues"]))
            confirm_cash_scope(preflight.run_dir, preflight.recommended_cash_decisions)
            with self.assertRaisesRegex(RuntimeError, "补充期初、期末现金余额和汇率影响"):
                run_classification(preflight.run_dir)
            supplement_cash_balances(preflight.run_dir, "0", "100", "0", "匿名余额资料")
            run_classification(preflight.run_dir)
            final = finalize_run(preflight.run_dir)
            self.assertEqual("诊断材料，不可作为最终表", final.overall_status)

    def test_cash_balance_sheet_in_ten_thousand_yuan_is_scaled(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "万元余额.xlsx"
            workbook = Workbook()
            journal = workbook.active
            journal.append(["日期", "凭证号", "摘要", "科目", "借方", "贷方"])
            journal.append(["2026-01-01", "记-1", "匿名收款", "银行存款", 100, None])
            journal.append(["2026-01-01", "记-1", "匿名收款", "主营业务收入", None, 100])
            balance = workbook.create_sheet("现金余额资料")
            balance.append(["金额单位：万元", None])
            balance.append(["期初现金及现金等价物余额", 1])
            balance.append(["期末现金及现金等价物余额", 1.01])
            balance.append(["汇率变动对现金及现金等价物的影响", 0])
            workbook.save(source)
            preflight = run_preflight(
                [source], ("1000000", "750000", "50000"), output_parent=root
            )
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(
                    encoding="utf-8-sig"
                )
            )
            self.assertEqual(1_000_000, state["cash_balances"]["opening_cent"])
            self.assertEqual(1_010_000, state["cash_balances"]["closing_cent"])

    def test_compile_and_verify_flow_produces_workbook_and_baseline(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            inputs = write_end_to_end_case(root, include_existing_statement=True)
            preflight = run_preflight(inputs, ("1000000", "750000", "50000"), output_parent=root)
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(
                    encoding="utf-8-sig"
                )
            )
            statement_question = next(
                question
                for question in state["mapping_questions"]
                if question.get("kind") == "statement"
            )
            confirm_mapping(
                preflight.run_dir,
                {
                    f"{statement_question['file_id']}:statement:{statement_question['sheet']}": "use"
                },
            )
            supplement_cash_balances(preflight.run_dir, "1000", "1060", "0", "匿名余额资料")
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(
                    encoding="utf-8-sig"
                )
            )
            confirm_cash_scope(preflight.run_dir, state["recommended_cash_decisions"])
            mark_dictionary_complete(preflight.run_dir)
            classified = run_classification(preflight.run_dir)
            self.assertEqual(0, classified.ai_tasks_missing)
            final = finalize_run(preflight.run_dir)
            self.assertTrue(final.workbook_path.is_file())
            self.assertTrue((final.run_dir / "计算留痕数据" / "计算留痕.sqlite3").is_file())
            self.assertEqual("最终可使用", final.overall_status)
            workbook = load_workbook(final.workbook_path, data_only=False)
            try:
                self.assertIsNotNone(workbook["现金流量表正表"]["C4"].value)
            finally:
                workbook.close()

    def test_resume_does_not_duplicate_completed_components(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            preflight = run_preflight(
                write_end_to_end_case(root),
                ("1000000", "750000", "50000"),
                output_parent=root,
            )
            confirm_cash_scope(preflight.run_dir, preflight.recommended_cash_decisions)
            first = run_classification(preflight.run_dir)
            second = run_classification(preflight.run_dir)
            self.assertEqual(first.component_count, second.component_count)
            self.assertEqual(first.component_hash, second.component_hash)

    def test_finalize_skips_partial_file_left_by_prior_failure(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            preflight = run_preflight(
                write_end_to_end_case(root),
                ("1000000", "750000", "50000"),
                output_parent=root,
            )
            confirm_cash_scope(preflight.run_dir, preflight.recommended_cash_decisions)
            mark_dictionary_complete(preflight.run_dir)
            run_classification(preflight.run_dir)
            _complete_all_pending_ai(preflight.run_dir, root)
            partial = preflight.run_dir / "现金流量表正表及复核底稿_生成中.xlsx"
            partial.write_bytes("模拟中断残留".encode("utf-8"))
            final = finalize_run(preflight.run_dir)
            self.assertEqual("现金流量表正表及复核底稿_重建2.xlsx", final.workbook_path.name)
            self.assertEqual("模拟中断残留".encode("utf-8"), partial.read_bytes())

    def test_cash_confirmation_and_input_hash_are_hard_boundaries(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            inputs = write_end_to_end_case(root)
            preflight = run_preflight(inputs, ("1000000", "750000", "50000"), root)
            with self.assertRaisesRegex(RuntimeError, "确认现金范围"):
                run_classification(preflight.run_dir)
            inputs[0].write_bytes(inputs[0].read_bytes() + b"changed")
            with self.assertRaisesRegex(RuntimeError, "输入文件已被修改.*新运行目录"):
                confirm_cash_scope(preflight.run_dir, preflight.recommended_cash_decisions)

    def test_missing_balances_block_classify_until_supplemented(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            preflight = run_preflight(
                write_end_to_end_case(root, include_cash_balances=False),
                ("1000000", "750000", "50000"),
                root,
            )
            confirm_cash_scope(preflight.run_dir, preflight.recommended_cash_decisions)
            with self.assertRaisesRegex(RuntimeError, "补充期初、期末现金余额和汇率影响"):
                run_classification(preflight.run_dir)
            supplement_cash_balances(preflight.run_dir, "1000", "1060", "0", "客户盖章现金余额表")
            mark_dictionary_complete(preflight.run_dir)
            run_classification(preflight.run_dir)
            _complete_all_pending_ai(preflight.run_dir, root)
            final = finalize_run(preflight.run_dir)
            self.assertEqual("最终可使用", final.overall_status)

    def test_missing_balances_can_be_supplemented_without_new_run(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            preflight = run_preflight(
                write_end_to_end_case(root, include_cash_balances=False),
                ("1000000", "750000", "50000"),
                root,
            )
            confirm_cash_scope(preflight.run_dir, preflight.recommended_cash_decisions)
            with self.assertRaisesRegex(RuntimeError, "补充期初、期末现金余额和汇率影响"):
                run_classification(preflight.run_dir)
            supplement_cash_balances(
                preflight.run_dir, "1000", "1060", "0", "客户盖章现金余额表"
            )
            mark_dictionary_complete(preflight.run_dir)
            run_classification(preflight.run_dir)
            _complete_all_pending_ai(preflight.run_dir, root)
            final = finalize_run(preflight.run_dir)
            self.assertEqual("最终可使用", final.overall_status)

    def test_mapping_confirmation_is_applied_without_restarting_run(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "歧义明细.xlsx"
            write_ambiguous_money_fixture(source)
            preflight = run_preflight([source], ("1000000", "750000", "50000"), root)
            self.assertEqual(1, preflight.mapping_question_count)
            state_path = preflight.run_dir / "计算留痕数据" / "运行状态.json"
            state = json.loads(state_path.read_text(encoding="utf-8-sig"))
            question = state["mapping_questions"][0]
            with self.assertRaisesRegex(RuntimeError, "字段映射"):
                confirm_cash_scope(preflight.run_dir, {})
            confirm_mapping(
                preflight.run_dir,
                {f"{question['file_id']}:{question['role']}": question["recommended"]},
            )
            updated = json.loads(state_path.read_text(encoding="utf-8-sig"))
            self.assertEqual([], updated["mapping_questions"])
            self.assertEqual(1, len(updated["entries"]))

    def test_same_file_detail_and_statement_prompts_and_reconciles(self) -> None:
        # 同一文件含明细+正表：不带 statement-path → 明细不丢，且产生疑似正表提问；确认 use 后生成核对报告
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "明细加正表.xlsx"
            write_detail_plus_statement_fixture(source)
            preflight = run_preflight(
                [source], ("1000000", "750000", "50000"), output_parent=root
            )
            self.assertGreater(preflight.source_entry_count, 0)
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(
                    encoding="utf-8-sig"
                )
            )
            statement_questions = [
                question
                for question in state["mapping_questions"]
                if question.get("kind") == "statement"
            ]
            self.assertEqual(1, len(statement_questions))
            key = (
                f"{statement_questions[0]['file_id']}:statement:"
                f"{statement_questions[0]['sheet']}"
            )
            confirm_mapping(preflight.run_dir, {key: "use"})
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(encoding="utf-8-sig")
            )
            self.assertIn("opening_cent", state["cash_balances"])
            confirm_cash_scope(preflight.run_dir, state["recommended_cash_decisions"])
            mark_dictionary_complete(preflight.run_dir)
            run_classification(preflight.run_dir)
            _complete_all_pending_ai(preflight.run_dir, root)
            final = finalize_run(preflight.run_dir)
            workbook = load_workbook(final.workbook_path, data_only=False)
            try:
                self.assertIsNotNone(workbook["正表核对报告"]["A2"].value)
            finally:
                workbook.close()

    def test_same_file_with_statement_path_keeps_detail(self) -> None:
        # 同一文件带 --statement-path 指向自身 → 明细仍被读取，且正表路径已登记
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "明细加正表.xlsx"
            write_detail_plus_statement_fixture(source)
            preflight = run_preflight(
                [source],
                ("1000000", "750000", "50000"),
                output_parent=root,
                statement_path=source,
            )
            self.assertGreater(preflight.source_entry_count, 0)
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(
                    encoding="utf-8-sig"
                )
            )
            self.assertEqual(str(source), state["existing_statement_path"])

    def test_unconfirmed_statement_blocks_final_usable(self) -> None:
        # 疑似正表未纳入核对（确认 ignore）→ 不允许"最终可使用"
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "明细加正表.xlsx"
            write_detail_plus_statement_fixture(source)
            preflight = run_preflight(
                [source], ("1000000", "750000", "50000"), output_parent=root
            )
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(
                    encoding="utf-8-sig"
                )
            )
            statement_question = next(
                question
                for question in state["mapping_questions"]
                if question.get("kind") == "statement"
            )
            key = (
                f"{statement_question['file_id']}:statement:"
                f"{statement_question['sheet']}"
            )
            confirm_mapping(preflight.run_dir, {key: "ignore"})
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(
                    encoding="utf-8-sig"
                )
            )
            confirm_cash_scope(preflight.run_dir, state["recommended_cash_decisions"])
            supplement_cash_balances(preflight.run_dir, "0", "60", "0", "匿名余额资料")
            mark_dictionary_complete(preflight.run_dir)
            run_classification(preflight.run_dir)
            _complete_all_pending_ai(preflight.run_dir, root)
            final = finalize_run(preflight.run_dir)
            self.assertNotEqual("最终可使用", final.overall_status)

    def test_structured_ai_cannot_invent_candidate_and_system_keeps_final_control(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = write_ai_end_to_end_case(root)
            preflight = run_preflight([source], ("1000000", "750000", "50000"), root)
            confirm_cash_scope(
                preflight.run_dir, preflight.recommended_cash_decisions
            )
            mark_dictionary_complete(preflight.run_dir)
            classified = run_classification(preflight.run_dir)
            self.assertEqual(2, classified.ai_tasks_missing)
            request_path = preflight.run_dir / "计算留痕数据" / "AI复核请求_第01批.jsonl"
            self.assertTrue(request_path.is_file())
            request_rows = tuple(
                json.loads(line)
                for line in request_path.read_text(encoding="utf-8-sig").splitlines()
            )
            self.assertTrue(request_rows)
            forbidden_fields = {
                "original_item",
                "system_item_id",
                "rule_evidence",
                "candidate_item_ids",
                "summary_candidate_item_ids",
                "account_path_candidate_item_ids",
            }
            self.assertTrue(
                all(not forbidden_fields.intersection(row) for row in request_rows)
            )
            state_path = preflight.run_dir / "计算留痕数据" / "运行状态.json"
            state = json.loads(state_path.read_text(encoding="utf-8-sig"))
            initial_decision = state["decisions"][0]
            self.assertEqual("", initial_decision["system_item_id"])
            self.assertEqual("no_candidate", initial_decision["candidate_status"])
            tasks = state["ai_tasks"]
            invented_result = root / "AI越界结果.jsonl"
            invented_result.write_text(
                json.dumps(
                    _structured_ai_payload(
                        tasks[0], candidate_item_id="CFO-04"
                    ),
                    ensure_ascii=False,
                )
                + "\n",
                encoding="utf-8-sig",
            )
            rejected = import_ai_results(preflight.run_dir, invented_result)
            self.assertEqual("AI 未完成", rejected.status)

            valid_result = root / "AI结构化结果.jsonl"
            valid_result.write_text(
                "".join(
                    json.dumps(_structured_ai_payload(task), ensure_ascii=False) + "\n"
                    for task in tasks
                ),
                encoding="utf-8-sig",
            )
            completed = import_ai_results(preflight.run_dir, valid_result)
            self.assertEqual("AI 已完成", completed.status)
            repeated = import_ai_results(preflight.run_dir, valid_result)
            self.assertEqual("AI 已完成", repeated.status)
            self.assertEqual(0, repeated.missing_count)
            final = finalize_run(preflight.run_dir)
            self.assertEqual("最终可使用", final.overall_status)
            state = json.loads(state_path.read_text(encoding="utf-8-sig"))
            self.assertEqual("CFO-01", state["decisions"][0]["system_item_id"])
            self.assertEqual(90, state["decisions"][0]["evidence_score"])
            workbook = load_workbook(final.workbook_path, data_only=False)
            try:
                ai_sheet = workbook["AI复核记录"]
                self.assertEqual(5, ai_sheet.max_row)
                ai_headers = [cell.value for cell in ai_sheet[1]]
                stage_column = ai_headers.index("阶段") + 1
                self.assertCountEqual(
                    [
                        "分类AI技术失败",
                        "分类AI技术失败",
                        "分类AI有效结果",
                        "分类AI有效结果",
                    ],
                    [
                        ai_sheet.cell(row, stage_column).value
                        for row in range(2, ai_sheet.max_row + 1)
                    ],
                )
                ai_headers = [cell.value for cell in ai_sheet[1]]
                for header in (
                    "阶段",
                    "review_round",
                    "reviewer_id",
                    "model_id",
                    "reviewed_at",
                    "prior_result_difference",
                ):
                    self.assertIn(header, ai_headers)
                trace_headers = [cell.value for cell in workbook["全量分类留痕"][1]]
                for header in (
                    "原现流项目",
                    "本行完整对方科目路径",
                    "系统候选项目",
                    "摘要来源质量",
                    "来源文件",
                    "来源工作表",
                    "来源单元格",
                ):
                    self.assertIn(header, trace_headers)
            finally:
                workbook.close()
            connection = sqlite3.connect(preflight.run_dir / "计算留痕数据" / "计算留痕.sqlite3")
            try:
                self.assertEqual(2, connection.execute("SELECT COUNT(*) FROM ai_result").fetchone()[0])
                final_payload = json.loads(
                    connection.execute(
                        "SELECT payload_json FROM classification_decision WHERE record_id = ?",
                        (state["decisions"][0]["component_id"],),
                    ).fetchone()[0]
                )
                self.assertEqual("CFO-01", final_payload["system_item_id"])
            finally:
                connection.close()

    def test_low_confidence_two_ai_consensus_fills_without_human_review(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            preflight = run_preflight(
                [write_ai_end_to_end_case(root)],
                ("1000000", "750000", "50000"),
                root,
            )
            confirm_cash_scope(
                preflight.run_dir, preflight.recommended_cash_decisions
            )
            mark_dictionary_complete(preflight.run_dir)
            run_classification(preflight.run_dir)
            state_path = preflight.run_dir / "计算留痕数据" / "运行状态.json"
            tasks = json.loads(state_path.read_text(encoding="utf-8-sig"))["ai_tasks"]
            self.assertEqual(2, len(tasks))
            result_path = root / "两份互盲结果.jsonl"
            result_path.write_text(
                "".join(
                    json.dumps(
                        _structured_ai_payload(
                            task,
                            summary_quality="weak",
                            account_quality="weak",
                        ),
                        ensure_ascii=False,
                    )
                    + "\n"
                    for task in tasks
                ),
                encoding="utf-8-sig",
            )
            imported = import_ai_results(preflight.run_dir, result_path)
            self.assertEqual("AI 已完成", imported.status)
            state = json.loads(state_path.read_text(encoding="utf-8-sig"))
            self.assertTrue(state["decisions"][0]["resolved"])
            self.assertEqual("automatic_fill", state["decisions"][0]["decision_action"])
            final = finalize_run(preflight.run_dir)
            self.assertEqual("最终可使用", final.overall_status)
            workbook = load_workbook(final.workbook_path, data_only=False)
            try:
                review = workbook["重要待复核事项"]
                headers = [cell.value for cell in review[1]]
                self.assertEqual(
                    "本期无重大剩余不确定事项，无需人工复核。",
                    review.cell(2, 1).value,
                )
                main = workbook["现金流量表正表"]
                proposed_row = next(
                    row
                    for row in range(4, main.max_row + 1)
                    if main.cell(row, 2).value
                    == "销售商品、提供劳务收到的现金"
                )
                self.assertNotEqual(0, main.cell(proposed_row, 4).value)
                self.assertNotIn("来源文件", headers)
            finally:
                workbook.close()

    def test_classify_requires_balances_before_running(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            preflight = run_preflight(
                write_end_to_end_case(root, include_cash_balances=False),
                ("1000000", "750000", "50000"),
                root,
            )
            confirm_cash_scope(preflight.run_dir, preflight.recommended_cash_decisions)
            with self.assertRaisesRegex(RuntimeError, "补充期初、期末现金余额和汇率影响"):
                run_classification(preflight.run_dir)
            supplement_cash_balances(preflight.run_dir, "1000", "1060", "0", "客户盖章现金余额表")
            mark_dictionary_complete(preflight.run_dir)
            classified = run_classification(preflight.run_dir)
            self.assertEqual(0, classified.ai_tasks_missing)
            self.assertEqual("consistency_completed", classified.status)

    def test_rough_reconciliation_is_only_diagnostic_before_cash_row_cleanup(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "单边明细.xlsx"
            workbook = Workbook()
            detail = workbook.active
            detail.title = "单边数据"
            detail.append(["日期", "凭证号", "摘要", "科目", "借方", "贷方", "流量金额", "现流项目"])
            detail.append(["2026-01-01", "记-1", "匿名付款", "应付账款", 100, None, 100, "购买商品、接受劳务支付的现金"])
            balance = workbook.create_sheet("现金余额资料")
            balance.append(["项目", "金额"])
            balance.append(["期初现金及现金等价物余额", 200])
            balance.append(["期末现金及现金等价物余额", 100])
            balance.append(["汇率变动对现金及现金等价物的影响", 0])
            workbook.save(source)
            preflight = run_preflight([source], ("1000000", "750000", "50000"), root)
            confirm_cash_scope(
                preflight.run_dir, preflight.recommended_cash_decisions
            )
            mark_dictionary_complete(preflight.run_dir)
            classified = run_classification(preflight.run_dir)
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(encoding="utf-8-sig")
            )
            self.assertEqual("待用户清洗现金分录", classified.status)
            self.assertEqual(0, classified.component_count)
            self.assertEqual("waiting_cash_row_cleanup", state["stage"])
            rough = state["rough_reconciliation"]
            self.assertTrue(rough["applicable"])
            self.assertEqual("相符", rough["status"])
            self.assertEqual(-10_000, rough["detail_sum_cent"])
            trace = preflight.run_dir / "计算留痕数据" / "粗勾稽留痕.jsonl"
            self.assertTrue(trace.is_file())
            self.assertIn("detail_sum_cent", trace.read_text(encoding="utf-8-sig"))

    def test_classify_blocked_until_dictionary_completed(self) -> None:
        # 含未知明细对方科目时，classify 必须先停在"待科目语义确认"；导入后放行（Task 5）
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "含明细科目.xlsx"
            workbook = Workbook()
            journal = workbook.active
            journal.title = "序时账"
            journal.append(["日期", "凭证字", "凭证号", "摘要", "科目编码", "科目", "借方", "贷方", "流量金额", "现流项目"])
            journal.append(["2026/1/1", "记", "1", "支付设备款", "1002", "银行存款", None, 100000, None, "购建固定资产支付的现金"])
            journal.append(["2026/1/1", "记", "1", "支付设备款", "2202", "应付账款_全新明细段XYZ", 100000, None, None, "购建固定资产支付的现金"])
            balances = workbook.create_sheet("现金余额资料")
            balances.append(["项目", "金额"])
            balances.append(["期初现金及现金等价物余额", 0])
            balances.append(["汇率变动对现金及现金等价物的影响", 0])
            balances.append(["期末现金及现金等价物余额", 100000])
            workbook.save(source)
            workbook.close()

            preflight = run_preflight([source], ("100000", "50000", "5000"), root)
            confirm_cash_scope(preflight.run_dir, dict(preflight.recommended_cash_decisions))
            result = run_classification(preflight.run_dir)
            self.assertEqual("待科目语义确认", result.status)
            self.assertGreater(result.ai_tasks_missing, 0)

            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(encoding="utf-8-sig")
            )
            tasks = state["account_dictionary"]["tasks"]
            self.assertTrue(
                all(
                    not {
                        "sample_contexts",
                        "summary",
                        "original_item",
                        "amount",
                        "cash_delta_cent",
                    }.intersection(task)
                    for task in tasks
                )
            )
            batch_files = list((preflight.run_dir / "计算留痕数据").glob("科目语义待判断_*.jsonl"))
            self.assertTrue(batch_files)
            result_path = root / "词典结果.jsonl"
            result_path.write_text(
                "".join(
                    json.dumps(
                        {
                            "task_id": task["task_id"],
                            "account": task["account"],
                            "semantic": "测试语义",
                            "item_id": "CFI-06",
                            "confidence": "high",
                            "basis": "知识库第13行：测试依据",
                            "standard_basis": (
                                "《企业会计准则第31号——现金流量表》第十三条："
                                "购建固定资产、无形资产和其他长期资产支付的现金（CFI-06）"
                            ),
                            "classification_facts": ["object:设备", "purpose:购建长期资产"],
                        },
                        ensure_ascii=False,
                    )
                    + "\n"
                    for task in tasks
                ),
                encoding="utf-8-sig",
            )
            import_result = import_dictionary_results(preflight.run_dir, result_path)
            self.assertEqual("科目语义已导入", import_result["status"])
            second = run_classification(preflight.run_dir)
            self.assertNotEqual("待摘要语义确认", second.status)
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(encoding="utf-8-sig")
            )
            self.assertEqual([], state["summary_semantics"]["tasks"])
            summary_result = state["summary_semantics"]["results"][0]
            self.assertEqual("rule_complete", summary_result["status"])
            self.assertEqual(["CFI-06"], summary_result["candidate_item_ids"])
            self.assertEqual(25, summary_result["quality"])
            self.assertEqual("科目语义词典说明.md", (preflight.run_dir / "科目语义词典说明.md").name)
            self.assertTrue((preflight.run_dir / "科目语义词典说明.md").is_file())

    def test_company_notes_gate_and_injection(self) -> None:
        # 传入 --notes 后，未 confirm-notes 前 scan-accounts 停在待确认；确认后注意事项进入词典批次 context（Task 5B）
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "含明细科目.xlsx"
            workbook = Workbook()
            journal = workbook.active
            journal.title = "序时账"
            journal.append(["日期", "凭证字", "凭证号", "摘要", "科目编码", "科目", "借方", "贷方", "流量金额", "现流项目"])
            journal.append(["2026/1/1", "记", "1", "支付设备款", "1002", "银行存款", None, 100000, None, "购建固定资产支付的现金"])
            journal.append(["2026/1/1", "记", "1", "支付设备款", "2202", "应付账款_全新明细段XYZ", 100000, None, None, "购建固定资产支付的现金"])
            balances = workbook.create_sheet("现金余额资料")
            balances.append(["项目", "金额"])
            balances.append(["期初现金及现金等价物余额", 0])
            balances.append(["汇率变动对现金及现金等价物的影响", 0])
            balances.append(["期末现金及现金等价物余额", 100000])
            workbook.save(source)
            workbook.close()

            notes_text = "该公司把应付设备款做到其他应付款"
            preflight = run_preflight(
                [source], ("100000", "50000", "5000"), root, notes=notes_text
            )
            confirm_cash_scope(preflight.run_dir, dict(preflight.recommended_cash_decisions))
            blocked = scan_accounts(preflight.run_dir)
            self.assertEqual("待确认公司特殊规则", blocked["status"])

            confirmed = confirm_company_notes(
                preflight.run_dir,
                [
                    {
                        "note_id": "NOTE-01",
                        "内容": "该公司把应付设备款做到其他应付款",
                        "涉及科目或词": ["其他应付款", "设备款"],
                        "建议处理": "其他应付款中的设备采购支出按购建固定资产判断",
                        "依据": "管理层核算习惯说明",
                    }
                ],
            )
            self.assertEqual("completed", confirmed.status)

            scan = scan_accounts(preflight.run_dir)
            self.assertEqual("待科目语义确认", scan["status"])
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(encoding="utf-8-sig")
            )
            task = state["account_dictionary"]["tasks"][0]
            self.assertIn("company_notes", task)
            self.assertTrue(task["company_notes"])

    def test_confirm_notes_without_raw_text(self) -> None:
        # 复核修复：无 --notes 文本时也可登记口述规则；缺省状态"采用"、自动生成 NOTE 编号
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = _write_ledger_case(root)
            preflight = run_preflight([source], ("100000", "50000", "5000"), root)

            confirmed = confirm_company_notes(
                preflight.run_dir,
                [
                    {
                        "内容": "电费押金走其他应收",
                        "涉及科目或词": ["押金"],
                        "适用完整路径": ["其他应收款_电费押金"],
                        "规则类型": "退款或反向冲减",
                    }
                ],
            )

            self.assertEqual("completed", confirmed.status)
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(encoding="utf-8-sig")
            )
            self.assertEqual("NOTE-01", state["company_notes"][0]["note_id"])
            self.assertEqual("采用", state["company_notes"][0]["状态"])
            self.assertEqual(
                ["其他应收款_电费押金"],
                state["company_notes"][0]["适用完整路径"],
            )
            self.assertEqual(
                "退款或反向冲减", state["company_notes"][0]["规则类型"]
            )
            note = state["company_notes"][0]
            self.assertEqual(1, note["规则版本"])
            self.assertEqual(str(state["run_id"]), note["运行编号"])
            self.assertIn("适用主体", note)
            self.assertIn("适用期间", note)
            self.assertIn("确认时间", note)
            self.assertTrue(
                (preflight.run_dir / "计算留痕数据" / "公司规则登记.json").is_file()
            )

    def test_company_note_ids_must_use_note_number_format_and_be_unique(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = _write_ledger_case(root)
            preflight = run_preflight([source], ("100000", "50000", "5000"), root)

            with self.assertRaisesRegex(ValueError, "NOTE-xx"):
                confirm_company_notes(
                    preflight.run_dir,
                    [{"note_id": "规则-01", "内容": "押金走其他应收"}],
                )
            with self.assertRaisesRegex(ValueError, "重复"):
                confirm_company_notes(
                    preflight.run_dir,
                    [
                        {"note_id": "NOTE-01", "内容": "押金走其他应收"},
                        {"note_id": "NOTE-01", "内容": "设备款走其他应付"},
                    ],
                )

    def test_disabling_a_company_note_invalidates_all_dependent_results(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = _write_ledger_case(root)
            preflight = run_preflight([source], ("100000", "50000", "5000"), root)
            confirm_company_notes(
                preflight.run_dir,
                [
                    {
                        "note_id": "NOTE-01",
                        "内容": "设备款实际为职工垫付报销",
                        "涉及科目或词": ["全新明细段XYZ"],
                    }
                ],
            )
            confirm_cash_scope(
                preflight.run_dir, dict(preflight.recommended_cash_decisions)
            )
            self.assertEqual("待科目语义确认", scan_accounts(preflight.run_dir)["status"])

            confirm_company_notes(
                preflight.run_dir,
                [
                    {
                        "note_id": "NOTE-01",
                        "内容": "设备款实际为职工垫付报销",
                        "涉及科目或词": ["全新明细段XYZ"],
                        "状态": "已停用",
                    }
                ],
            )
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(
                    encoding="utf-8-sig"
                )
            )
            self.assertNotIn("account_dictionary", state)
            self.assertFalse(state["account_dictionary_completed"])
            self.assertIn("NOTE-01", str(state["note_dependency_rebuild"]))
            self.assertEqual(2, state["company_notes"][0]["规则版本"])
            self.assertEqual("cash_scope_confirmed", state["stage"])
            self.assertEqual("待科目语义确认", scan_accounts(preflight.run_dir)["status"])
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(
                    encoding="utf-8-sig"
                )
            )
            self.assertNotIn("company_notes", state["account_dictionary"]["tasks"][0])

    def test_conflicted_notes_are_never_injected(self) -> None:
        # 复核修复："冲突未采用"规则不得进入科目语义任务上下文与 AI 任务上下文
        from cashflow_direct.ai_review import build_ai_task
        from cashflow_direct.models import CashflowComponent, ClassificationDecision

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = _write_ledger_case(root)
            preflight = run_preflight([source], ("100000", "50000", "5000"), root)
            confirm_company_notes(
                preflight.run_dir,
                [
                    {"note_id": "NOTE-01", "内容": "采用规则甲", "涉及科目或词": ["全新明细段XYZ"]},
                    {"note_id": "NOTE-02", "内容": "冲突规则乙", "涉及科目或词": ["全新明细段XYZ"], "状态": "冲突未采用"},
                ],
            )
            confirm_cash_scope(preflight.run_dir, dict(preflight.recommended_cash_decisions))
            scan = scan_accounts(preflight.run_dir)

            self.assertEqual("待科目语义确认", scan["status"])
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(encoding="utf-8-sig")
            )
            notes_in_tasks = str(
                [task.get("company_notes") for task in state["account_dictionary"]["tasks"]]
            )
            self.assertIn("采用规则甲", notes_in_tasks)
            self.assertNotIn("冲突规则乙", notes_in_tasks)

            component = CashflowComponent(
                component_id="C1", voucher_key="V1", summary="支付全新明细段XYZ款项",
                cash_delta_cent=-100, counterpart_accounts=("应付账款_全新明细段XYZ",),
            )
            decision = ClassificationDecision(
                component_id="C1", system_item_id="CFI-06",
                system_item_name="购建固定资产、无形资产和其他长期资产支付的现金",
                normal_direction="outflow", matched_rule_id="TEST",
                reason="测试", evidence_level="medium",
            )
            ai_task = build_ai_task(component, decision, state["company_notes"])
            self.assertNotIn("采用规则甲", ai_task.context)
            self.assertNotIn("冲突规则乙", ai_task.context)

    def test_dictionary_import_rejects_misaligned_account(self) -> None:
        # 复核修复：结果行 account 与任务 account 不一致 → 该行无效并列入 missing（防张冠李戴）
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = _write_ledger_case(root)
            preflight = run_preflight([source], ("100000", "50000", "5000"), root)
            confirm_cash_scope(preflight.run_dir, dict(preflight.recommended_cash_decisions))
            scan_accounts(preflight.run_dir)
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(encoding="utf-8-sig")
            )
            task = state["account_dictionary"]["tasks"][0]
            result_path = root / "词典结果.jsonl"
            result_path.write_text(
                json.dumps({
                    "task_id": task["task_id"],
                    "account": "张冠李戴段",
                    "semantic": "测试语义", "item_id": "CFI-06",
                    "confidence": "high", "basis": "知识库第1行：测试依据",
                    "classification_facts": ["object:设备", "purpose:购建长期资产"],
                }, ensure_ascii=False) + "\n",
                encoding="utf-8-sig",
            )

            result = import_dictionary_results(preflight.run_dir, result_path)

            self.assertEqual("AI 未完成", result["status"])
            self.assertIn(task["task_id"], result["missing_ids"])

    def test_dictionary_candidate_requires_item_specific_standard_basis(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = _write_ledger_case(root)
            preflight = run_preflight([source], ("100000", "50000", "5000"), root)
            confirm_cash_scope(preflight.run_dir, dict(preflight.recommended_cash_decisions))
            scan_accounts(preflight.run_dir)
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(
                    encoding="utf-8-sig"
                )
            )
            task = state["account_dictionary"]["tasks"][0]
            result_path = root / "词典结果.jsonl"
            payload = {
                "task_id": task["task_id"],
                "account": task["account"],
                "semantic": "设备采购对价",
                "item_id": "CFI-06",
                "confidence": "high",
                "basis": f"完整科目路径“{task['account']}”显示为设备采购款",
                "classification_facts": ["object:设备", "purpose:购建长期资产"],
            }
            result_path.write_text(
                json.dumps(payload, ensure_ascii=False) + "\n", encoding="utf-8-sig"
            )

            rejected = import_dictionary_results(preflight.run_dir, result_path)

            self.assertEqual("AI 未完成", rejected["status"])
            payload["standard_basis"] = (
                "《企业会计准则第31号——现金流量表》第十三条："
                "购建固定资产、无形资产和其他长期资产支付的现金（CFI-06）"
            )
            result_path.write_text(
                json.dumps(payload, ensure_ascii=False) + "\n", encoding="utf-8-sig"
            )

            accepted = import_dictionary_results(preflight.run_dir, result_path)

            self.assertEqual("科目语义已导入", accepted["status"])

    def test_dictionary_import_rejects_unknown_note_id(self) -> None:
        # 复核修复：结果行 note_id 不在已登记"采用"清单 → 无效并列入 missing
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = _write_ledger_case(root)
            preflight = run_preflight([source], ("100000", "50000", "5000"), root)
            confirm_cash_scope(preflight.run_dir, dict(preflight.recommended_cash_decisions))
            scan_accounts(preflight.run_dir)
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(encoding="utf-8-sig")
            )
            task = state["account_dictionary"]["tasks"][0]
            result_path = root / "词典结果.jsonl"
            result_path.write_text(
                json.dumps({
                    "task_id": task["task_id"], "account": task["account"],
                    "semantic": "测试语义", "item_id": "CFI-06",
                    "confidence": "high", "basis": "依据公司特殊规则：NOTE-99",
                    "classification_facts": ["object:设备", "purpose:购建长期资产"],
                    "note_id": "NOTE-99",
                }, ensure_ascii=False) + "\n",
                encoding="utf-8-sig",
            )

            result = import_dictionary_results(preflight.run_dir, result_path)

            self.assertEqual("AI 未完成", result["status"])
            self.assertIn(task["task_id"], result["missing_ids"])

    def test_scan_accounts_forces_task_for_adopted_note_segment(self) -> None:
        # 通用词典先正常生效；用户确认的公司特殊规则命中后，再强制生成专属确认任务。
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = _write_ledger_case(root, account_name="其他应付款_设备款")
            preflight = run_preflight([source], ("100000", "50000", "5000"), root)
            confirm_cash_scope(preflight.run_dir, dict(preflight.recommended_cash_decisions))

            first = scan_accounts(preflight.run_dir)
            self.assertEqual("科目语义已齐备", first["status"])

            confirm_company_notes(
                preflight.run_dir,
                [{"note_id": "NOTE-01", "内容": "设备款实际为职工垫付报销", "涉及科目或词": ["设备款"]}],
            )
            second = scan_accounts(preflight.run_dir)

            self.assertEqual("待科目语义确认", second["status"])
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(encoding="utf-8-sig")
            )
            accounts = [task["account"] for task in state["account_dictionary"]["tasks"]]
            self.assertIn("其他应付款_设备款", accounts)
            self.assertTrue(
                all("通用业务语义" in task["instruction"] for task in state["account_dictionary"]["tasks"])
            )
            # 设计 3.1.3：任务上下文必须注明对应 NOTE 编号，答题方才能在结果里留痕
            task_notes = [
                note
                for task in state["account_dictionary"]["tasks"]
                for note in task.get("company_notes", ())
            ]
            self.assertTrue(any("NOTE-01" in note for note in task_notes))
            # 生成了待确认任务后，主流程必须重新等待词典导入，不得沿用"已齐备"标志
            self.assertFalse(state.get("account_dictionary_completed", False))

    def test_dictionary_note_id_flows_into_classification_reason(self) -> None:
        # 复核修复：导入条目带 NOTE 编号，分类命中后理由须留痕"依据公司特殊规则：NOTE-01"
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = _write_ledger_case(root)
            preflight = run_preflight([source], ("100000", "50000", "5000"), root)
            confirm_company_notes(
                preflight.run_dir,
                [{"note_id": "NOTE-01", "内容": "全新明细段XYZ是设备采购款", "涉及科目或词": ["全新明细段XYZ"]}],
            )
            confirm_cash_scope(preflight.run_dir, dict(preflight.recommended_cash_decisions))
            scan_accounts(preflight.run_dir)
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(encoding="utf-8-sig")
            )
            task = state["account_dictionary"]["tasks"][0]
            result_path = root / "词典结果.jsonl"
            result_path.write_text(
                json.dumps({
                    "task_id": task["task_id"], "account": task["account"],
                    "semantic": "设备采购对价", "item_id": "CFI-06",
                    "confidence": "high", "basis": "依据公司特殊规则：NOTE-01",
                    "standard_basis": (
                        "《企业会计准则第31号——现金流量表》第十三条："
                        "购建固定资产、无形资产和其他长期资产支付的现金（CFI-06）"
                    ),
                    "note_id": "NOTE-01",
                    "classification_facts": ["object:设备", "purpose:购建长期资产"],
                }, ensure_ascii=False) + "\n",
                encoding="utf-8-sig",
            )
            imported = import_dictionary_results(preflight.run_dir, result_path)
            self.assertEqual("科目语义已导入", imported["status"])

            mark_dictionary_complete(preflight.run_dir)
            run_classification(preflight.run_dir)

            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(encoding="utf-8-sig")
            )
            reasons = json.dumps(state["decisions"], ensure_ascii=False)
            self.assertIn("依据公司特殊规则：NOTE-01", reasons)

    def test_dictionary_import_keeps_direction_specific_items(self) -> None:
        """同一完整路径的流入、流出项目必须从导入结果完整进入运行状态。"""
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = _write_ledger_case(root, "其他应收款_日常往来")
            preflight = run_preflight([source], ("100000", "50000", "5000"), root)
            confirm_cash_scope(preflight.run_dir, dict(preflight.recommended_cash_decisions))
            scan_accounts(preflight.run_dir)
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(encoding="utf-8-sig")
            )
            task = state["account_dictionary"]["tasks"][0]
            result_path = root / "词典结果.jsonl"
            result_path.write_text(
                json.dumps({
                    "task_id": task["task_id"],
                    "account": task["account"],
                    "semantic": "日常经营往来款",
                    "item_id": "",
                    "inflow_item_id": "CFO-03",
                    "outflow_item_id": "CFO-07",
                    "confidence": "medium",
                    "basis": f"完整科目路径“{task['account']}”显示为日常经营往来",
                    "standard_basis": (
                        "《企业会计准则第31号——现金流量表》第十条："
                        "收到其他与经营活动有关的现金（CFO-03）；"
                        "支付其他与经营活动有关的现金（CFO-07）"
                    ),
                    "classification_facts": ["object:经营往来"],
                }, ensure_ascii=False) + "\n",
                encoding="utf-8-sig",
            )

            imported = import_dictionary_results(preflight.run_dir, result_path)

            self.assertEqual("科目语义已导入", imported["status"])
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(encoding="utf-8-sig")
            )
            entry = state["account_dictionary"]["valid_results"][0]
            self.assertEqual("CFO-03", entry["inflow_item_id"])
            self.assertEqual("CFO-07", entry["outflow_item_id"])

    def test_scan_accounts_does_not_treat_generic_fallback_as_full_path_judgment(self) -> None:
        """通用词典的泛化科目段不能替代完整路径语义确认。"""
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = _write_ledger_case(root, "应交税费_应交增值税_进项税")
            preflight = run_preflight([source], ("100000", "50000", "5000"), root)
            confirm_cash_scope(preflight.run_dir, dict(preflight.recommended_cash_decisions))

            scan = scan_accounts(preflight.run_dir)

            self.assertEqual("待科目语义确认", scan["status"])
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(encoding="utf-8-sig")
            )
            self.assertIn(
                "应交税费_应交增值税_进项税",
                {task["account"] for task in state["account_dictionary"]["tasks"]},
            )

    def test_scan_accounts_ignores_a_separate_non_cash_voucher(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = _write_ledger_case(root, "应付账款_全新现金业务XYZ")
            workbook = load_workbook(source)
            try:
                journal = workbook["序时账"]
                journal.append(
                    [
                        "2026/1/2",
                        "记",
                        "2",
                        "计提折旧",
                        "5101",
                        "制造费用_全新非现金业务XYZ",
                        500,
                        None,
                        None,
                        "",
                    ]
                )
                journal.append(
                    [
                        "2026/1/2",
                        "记",
                        "2",
                        "计提折旧",
                        "1602",
                        "累计折旧_机器设备",
                        None,
                        500,
                        None,
                        "",
                    ]
                )
                workbook.save(source)
            finally:
                workbook.close()
            preflight = run_preflight(
                [source], ("100000", "50000", "5000"), root
            )
            confirm_cash_scope(
                preflight.run_dir, dict(preflight.recommended_cash_decisions)
            )

            scan = scan_accounts(preflight.run_dir)
            state = json.loads(
                (preflight.run_dir / "计算留痕数据" / "运行状态.json").read_text(
                    encoding="utf-8-sig"
                )
            )

            self.assertEqual("待科目语义确认", scan["status"])
            self.assertEqual(
                {"应付账款_全新现金业务XYZ"},
                {task["account"] for task in state["account_dictionary"]["tasks"]},
            )

    def test_scan_accounts_accepts_builtin_segment_with_decisive_item(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = _write_ledger_case(root, "应付账款_应付设备款_往来款")
            preflight = run_preflight([source], ("100000", "50000", "5000"), root)
            confirm_cash_scope(preflight.run_dir, dict(preflight.recommended_cash_decisions))

            scan = scan_accounts(preflight.run_dir)

            self.assertEqual("科目语义已齐备", scan["status"])
            self.assertEqual(0, scan["missing"])


def _write_ledger_case(root: Path, account_name: str = "应付账款_全新明细段XYZ") -> Path:
    """最小运行态夹具：一张序时账（含指定对方科目明细段）+ 现金余额资料。"""
    source = root / "最小夹具.xlsx"
    workbook = Workbook()
    journal = workbook.active
    journal.title = "序时账"
    journal.append(["日期", "凭证字", "凭证号", "摘要", "科目编码", "科目", "借方", "贷方", "流量金额", "现流项目"])
    journal.append(["2026/1/1", "记", "1", "支付设备款", "1002", "银行存款", None, 100000, None, "购建固定资产支付的现金"])
    journal.append(["2026/1/1", "记", "1", "支付设备款", "2202", account_name, 100000, None, None, "购建固定资产支付的现金"])
    balances = workbook.create_sheet("现金余额资料")
    balances.append(["项目", "金额"])
    balances.append(["期初现金及现金等价物余额", 0])
    balances.append(["汇率变动对现金及现金等价物的影响", 0])
    balances.append(["期末现金及现金等价物余额", 100000])
    workbook.save(source)
    workbook.close()
    return source


def test_preflight_accepts_explicit_paths(tmp_path, monkeypatch):
    """--paths 给定中文路径时不再弹窗（A1，Task 11）。"""
    from cashflow_direct.cli import main
    from tests.fixture_factory import write_end_to_end_case

    inputs = write_end_to_end_case(tmp_path)
    monkeypatch.setattr(
        "cashflow_direct.cli.choose_input_files",
        lambda: (_ for _ in ()).throw(AssertionError("不应弹窗")),
    )
    code = main(
        [
            "preflight",
            "--overall", "1000000", "--performance", "500000", "--trivial", "50000",
            "--paths", str(inputs[0]),
            "--output-parent", str(tmp_path / "输出"),
        ]
    )
    assert code == 0


if __name__ == "__main__":
    unittest.main()
