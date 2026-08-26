from collections import Counter
import json
import tempfile
from pathlib import Path

from openpyxl import Workbook
import pytest

from cashflow_direct.account_mapping import (
    AccountMappingRecord,
    StandardAccount,
    apply_account_mapping,
    build_account_mappings,
    load_standard_accounts,
    resolve_account_mappings,
    standardize_component_accounts,
    standardize_entries,
)
from cashflow_direct.components import discover_cash_scope
from cashflow_direct.models import CashflowComponent, NormalizedEntry, SourceLocator
from cashflow_direct.classification import (
    classify_all,
    load_rule_pack,
)
from cashflow_direct.pipeline import confirm_account_mapping, run_preflight


PROJECT_ROOT = Path(__file__).resolve().parents[1]


def test_standard_account_baseline_has_all_201_rows_and_metadata() -> None:
    accounts = load_standard_accounts(PROJECT_ROOT)

    assert len(accounts) == 201
    assert len({item.standard_name for item in accounts}) == 201
    assert Counter(item.category for item in accounts) == {
        "资产类": 99,
        "负债类": 39,
        "共同类": 5,
        "所有者权益类": 10,
        "成本类": 10,
        "损益类": 38,
    }
    tax = next(item for item in accounts if item.standard_name == "税金及附加")
    assert "营业税金及附加" in tax.aliases
    assert tax.optional_codes
    assert tax.source_locations
    assert tax.report_note


def test_exact_name_alias_and_code_prefix_are_uniquely_mapped() -> None:
    accounts = load_standard_accounts(PROJECT_ROOT)

    records = build_account_mappings(
        (
            "应收账款_客户甲",
            "营业税金及附加_环保税",
            "1002 银行存款_一般户",
        ),
        accounts,
    )

    mapped = {item.original_level1: item for item in records}
    assert (mapped["银行存款"].standard_level1, mapped["银行存款"].status) == (
        "银行存款",
        "confirmed",
    )
    assert (mapped["应收账款"].standard_level1, mapped["应收账款"].status) == (
        "应收账款",
        "confirmed",
    )
    assert (mapped["营业税金及附加"].standard_level1, mapped["营业税金及附加"].status) == (
        "税金及附加",
        "confirmed",
    )


def test_unknown_or_ambiguous_name_is_never_guessed() -> None:
    baseline = (
        StandardAccount(1, "资产类", "甲科目", ("共同旧称",), ("1001",), ("S1-1",), "甲"),
        StandardAccount(2, "负债类", "乙科目", ("共同旧称",), ("2001",), ("S2-1",), "乙"),
    )

    records = build_account_mappings(("共同旧称_明细", "客户自定义科目_明细"), baseline)

    assert records[0].status == "ambiguous"
    assert records[0].candidate_standard_names == ("甲科目", "乙科目")
    assert records[0].standard_level1 == ""
    assert records[1].status == "unmapped"
    assert records[1].candidate_standard_names == ()
    assert records[1].standard_level1 == ""


def test_similar_and_custom_level1_names_are_never_auto_confirmed() -> None:
    accounts = load_standard_accounts(PROJECT_ROOT)

    records = build_account_mappings(("研发支出费", "客户自定义成本"), accounts)
    by_name = {item.original_level1: item for item in records}

    assert "研发支出" in by_name["研发支出费"].candidate_standard_names
    assert not by_name["客户自定义成本"].standard_level1
    assert {item.status for item in records} == {"unmapped"}
    assert all(not item.standard_level1 for item in records)


def test_mapping_only_replaces_level1_and_preserves_detail_path() -> None:
    record = AccountMappingRecord(
        original_level1="营业税金及附加",
        standard_level1="税金及附加",
        status="confirmed",
        candidate_standard_names=("税金及附加",),
        basis="别名唯一命中",
    )

    assert apply_account_mapping("营业税金及附加_环保税_本年", {record.original_level1: record}) == (
        "税金及附加_环保税_本年",
        True,
    )
    assert apply_account_mapping("客户自定义科目_明细", {record.original_level1: record}) == (
        "客户自定义科目_明细",
        False,
    )


def test_every_unknown_name_must_be_mapped_before_processing_can_continue() -> None:
    accounts = load_standard_accounts(PROJECT_ROOT)
    pending = build_account_mappings(("研发支出费_项目甲", "客户自定义成本_检验费"), accounts)

    with pytest.raises(ValueError, match="必须映射至201条基线"):
        resolve_account_mappings(
            pending,
            {"研发支出费": "研发支出", "客户自定义成本": "manual"},
            accounts,
        )


def test_confirmed_mapping_can_be_corrected_before_downstream_processing() -> None:
    accounts = load_standard_accounts(PROJECT_ROOT)
    confirmed = resolve_account_mappings(
        build_account_mappings(("客户自定义成本",), accounts),
        {"客户自定义成本": "生产成本"},
        accounts,
    )

    corrected = resolve_account_mappings(
        confirmed, {"客户自定义成本": "管理费用"}, accounts
    )

    assert corrected[0].standard_level1 == "管理费用"
    assert corrected[0].basis == "用户确认"


def test_component_keeps_raw_path_but_classifies_with_standard_path() -> None:
    component = CashflowComponent(
        component_id="C1",
        voucher_key="V1",
        summary="支付环保税",
        cash_delta_cent=-100,
        counterpart_accounts=("营业税金及附加_环保税",),
    )
    record = AccountMappingRecord(
        original_level1="营业税金及附加",
        standard_level1="税金及附加",
        status="confirmed",
        candidate_standard_names=("税金及附加",),
        basis="别名唯一命中",
    )

    standardized = standardize_component_accounts(component, {record.original_level1: record})

    assert standardized.counterpart_accounts == ("税金及附加_环保税",)
    assert standardized.original_counterpart_accounts == ("营业税金及附加_环保税",)
    assert standardized.account_mapping_status == "confirmed"


def test_preflight_stops_at_account_mapping_and_persists_user_confirmation() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        source = root / "客户明细.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "序时账"
        sheet.append(["日期", "凭证号", "摘要", "科目", "借方", "贷方", "现流项目"])
        sheet.append(["2026-01-01", "记-1", "支付检验费", "银行存款_一般户", None, 100, "支付其他经营现金"])
        sheet.append(["2026-01-01", "记-1", "支付检验费", "客户自定义成本_检验费", 100, None, "支付其他经营现金"])
        workbook.save(source)

        preflight = run_preflight((source,), (2_200_000, 1_100_000, 55_000), root / "输出")
        assert preflight.status == "waiting_account_mapping"
        state_path = preflight.run_dir / "计算留痕数据" / "运行状态.json"
        state = json.loads(state_path.read_text(encoding="utf-8-sig"))
        pending = {
            item["original_level1"]: item
            for item in state["account_mapping_records"]
            if item["status"] != "confirmed"
        }
        assert set(pending) == {"客户自定义成本"}

        result = confirm_account_mapping(preflight.run_dir, {"客户自定义成本": "生产成本"})
        assert result.status == "completed"
        state = json.loads(state_path.read_text(encoding="utf-8-sig"))
        quality = next(
            item for item in state["account_mapping_records"]
            if item["original_level1"] == "客户自定义成本"
        )
        assert quality["standard_level1"] == "生产成本"
        assert quality["basis"] == "用户确认"
        assert state["stage"] == "waiting_cash_scope"

        from cashflow_direct.pipeline import confirm_cash_scope

        confirm_cash_scope(preflight.run_dir, state["recommended_cash_decisions"])
        with pytest.raises(RuntimeError, match="新建运行目录"):
            confirm_account_mapping(preflight.run_dir, {"客户自定义成本": "管理费用"})


def test_cash_scope_is_created_only_after_all_level1_mappings_are_confirmed() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        source = root / "客户明细.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "序时账"
        sheet.append(["日期", "凭证号", "摘要", "科目", "借方", "贷方", "现流项目"])
        sheet.append(["2026-01-01", "记-1", "收到货款", "客户银行款_一般户", 100, None, "销售商品收到现金"])
        sheet.append(["2026-01-01", "记-1", "收到货款", "应收账款_客户甲", None, 100, "销售商品收到现金"])
        workbook.save(source)

        preflight = run_preflight((source,), (2_200_000, 1_100_000, 55_000), root / "输出")
        state_path = preflight.run_dir / "计算留痕数据" / "运行状态.json"
        state = json.loads(state_path.read_text(encoding="utf-8-sig"))

        assert preflight.status == "waiting_account_mapping"
        assert "cash_scope_proposal" not in state
        assert "recommended_cash_decisions" not in state

        confirm_account_mapping(preflight.run_dir, {"客户银行款": "银行存款"})
        state = json.loads(state_path.read_text(encoding="utf-8-sig"))
        candidates = state["cash_scope_proposal"]["candidates"]

        assert state["stage"] == "waiting_cash_scope"
        assert len(candidates) == 1
        assert candidates[0]["account_names"] == ["客户银行款_一般户"]
        assert set(state["recommended_cash_decisions"]) == {
            candidates[0]["account_key"]
        }


def test_counterpart_placeholder_is_traced_as_missing_instead_of_mapped() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        source = root / "含对方科目占位提示的序时账.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "序时账"
        sheet.append(
            ["日期", "凭证号", "摘要", "科目名称", "借方发生额", "贷方发生额", "对方科目"]
        )
        sheet.append(
            [
                "2026-01-01",
                "记-1",
                "收到销售款",
                "银行存款_一般户",
                100,
                None,
                "主营业务收入",
            ]
        )
        sheet.append(
            [
                "2026-01-02",
                "记-2",
                "一分钱异常",
                "营业外支出_其他",
                0.01,
                None,
                "未找到匹配",
            ]
        )
        workbook.save(source)

        preflight = run_preflight(
            (source,), (250_000, 125_000, 25_000), root / "输出"
        )
        state_path = preflight.run_dir / "计算留痕数据" / "运行状态.json"
        state = json.loads(state_path.read_text(encoding="utf-8-sig"))

        assert preflight.status == "waiting_cash_scope"
        assert "未找到匹配" not in {
            item["original_level1"] for item in state["account_mapping_records"]
        }
        placeholder_entry = next(
            item for item in state["entries"] if item["summary"] == "一分钱异常"
        )
        assert placeholder_entry["counterpart_name"] == ""
        assert any(
            issue["kind"] == "警告" and "未找到匹配" in issue["message"]
            for issue in state["normalization_issues"]
        )


def test_cash_scope_keeps_subaccounts_distinct_after_level1_mapping_is_inherited() -> None:
    baseline = load_standard_accounts(PROJECT_ROOT)
    records = resolve_account_mappings(
        build_account_mappings(("客户银行款_一般户", "客户银行款_保证金户"), baseline),
        {"客户银行款": "银行存款"},
        baseline,
    )
    source = SourceLocator("F1", "序时账", 2, 2, "A2:H2")
    entries = (
        NormalizedEntry(
            "E1", source, "V1", "2026-01-01", "1", "转款",
            "客户银行款_一般户", "", 100, 0, 0, "", account_code="1002.01"
        ),
        NormalizedEntry(
            "E2", source, "V1", "2026-01-01", "1", "转款",
            "客户银行款_保证金户", "", 0, 100, 0, "", account_code="1002.02"
        ),
    )

    standardized = standardize_entries(
        entries, {item.original_level1: item for item in records}
    )
    proposal = discover_cash_scope(standardized)

    assert len(records) == 1
    assert records[0].original_level1 == "客户银行款"
    assert tuple(item.account_name for item in standardized) == (
        "银行存款_一般户",
        "银行存款_保证金户",
    )
    assert {item.account_key for item in proposal.candidates} == {
        "1002.01",
        "1002.02",
    }
    assert {item.account_names for item in proposal.candidates} == {
        ("客户银行款_一般户",),
        ("客户银行款_保证金户",),
    }


def test_standardized_level1_recomputes_cash_side_in_both_directions() -> None:
    baseline = load_standard_accounts(PROJECT_ROOT)
    records = resolve_account_mappings(
        build_account_mappings(
            ("客户现金折扣_明细", "客户银行款_一般户"), baseline
        ),
        {
            "客户现金折扣": "财务费用",
            "客户银行款": "银行存款",
        },
        baseline,
    )
    source = SourceLocator("F1", "序时账", 2, 2, "A2:H2")
    entries = (
        NormalizedEntry(
            "E1", source, "V1", "2026-01-01", "1", "折扣",
            "客户现金折扣_明细", "", 100, 0, 100, "项目甲",
            retained_side="cash", label_side="cash",
        ),
        NormalizedEntry(
            "E2", source, "V2", "2026-01-01", "2", "收款",
            "客户银行款_一般户", "", 100, 0, 100, "项目乙",
            retained_side="counterpart", label_side="counterpart",
        ),
    )

    standardized = standardize_entries(
        entries, {item.original_level1: item for item in records}
    )

    assert tuple(item.retained_side for item in standardized) == (
        "counterpart",
        "cash",
    )
    assert tuple(item.label_side for item in standardized) == (
        "counterpart",
        "cash",
    )


def test_mapping_decisions_reject_detail_account_keys() -> None:
    baseline = load_standard_accounts(PROJECT_ROOT)
    records = build_account_mappings(("客户银行款_一般户",), baseline)

    with pytest.raises(ValueError, match="只接受客户一级科目"):
        resolve_account_mappings(
            records,
            {
                "客户银行款": "银行存款",
                "客户银行款_一般户": "银行存款",
            },
            baseline,
        )


def test_unconfirmed_mapping_cannot_enter_classification_routing() -> None:
    component = CashflowComponent(
        component_id="C-MANUAL",
        voucher_key="V-MANUAL",
        summary="销售商品收到货款",
        cash_delta_cent=100_000,
        counterpart_accounts=("客户自定义成本_检验费",),
        original_item_text="销售商品、提供劳务收到的现金",
        original_counterpart_accounts=("客户自定义成本_检验费",),
        account_mapping_status="manual",
    )
    rules = load_rule_pack(PROJECT_ROOT)

    with pytest.raises(ValueError, match="一级科目映射未全部确认"):
        classify_all((component,), rules)
