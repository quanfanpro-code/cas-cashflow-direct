from __future__ import annotations

import json
import sqlite3
import tempfile
from dataclasses import asdict
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

import cashflow_direct.pipeline as pipeline_module
from cashflow_direct.component_structure_ai import build_structure_ai_tasks
from cashflow_direct.versions import current_versions
from cashflow_direct.pipeline import (
    confirm_cash_scope,
    finalize_run,
    import_ai_results,
    run_classification,
    run_preflight,
)


def test_ai_task_becomes_terminal_only_after_three_explicit_technical_failures() -> None:
    pipeline = pipeline_module
    state: dict[str, object] = {}
    payload = ({"task_id": "AI-1"},)

    first = pipeline._register_ai_technical_attempts(
        state, payload, ("AI-1",), ()
    )
    second = pipeline._register_ai_technical_attempts(
        state, payload, ("AI-1",), ()
    )
    missing_only = pipeline._register_ai_technical_attempts(state, (), (), ())
    third = pipeline._register_ai_technical_attempts(
        state, payload, ("AI-1",), ()
    )

    assert first == set()
    assert second == set()
    assert missing_only == set()
    assert third == {"AI-1"}
    assert state["ai_task_attempts"] == {"AI-1": 3}
    assert [item["status"] for item in state["ai_technical_failure_log"]] == [
        "技术失败，可重试",
        "技术失败，可重试",
        "无有效结果",
    ]


def test_component_structure_confirmation_accepts_only_a_listed_complete_choice(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    state: dict[str, object] = {
        "run_id": "RUN-1",
        "component_structure_requests": [
            {
                "voucher_key": "V-1",
                "candidate_entry_id_combinations": [
                    ["E-1", "E-3"],
                    ["E-2", "E-3"],
                ],
            }
        ],
    }
    monkeypatch.setattr(pipeline_module, "_load_state", lambda _path: state)
    monkeypatch.setattr(pipeline_module, "_assert_inputs_unchanged", lambda _state: None)
    monkeypatch.setattr(pipeline_module, "_save_state", lambda _path, _state: None)

    with pytest.raises(ValueError, match="既有候选组合"):
        pipeline_module.confirm_component_structure(
            tmp_path, {"V-1": ("E-9",)}
        )

    result = pipeline_module.confirm_component_structure(
        tmp_path, {"V-1": ("E-2", "E-3")}
    )

    assert result.status == "completed"
    assert state["component_structure_selections"] == {
        "V-1": ["E-2", "E-3"]
    }
    assert "component_structure_requests" not in state


def test_component_structure_ai_low_first_result_gets_second_review_then_finishes(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    request = {
        "voucher_key": "V-1",
        "cash_delta_cent": -10_000,
        "materiality_level": "M1",
        "candidate_entry_id_combinations": [
            ["E-1", "E-3"],
            ["E-2", "E-3"],
        ],
        "candidate_details": ["组合1", "组合2"],
    }
    first_task = build_structure_ai_tasks(request, "M1", ("single",))[0]
    state: dict[str, object] = {
        "run_id": "RUN-1",
        "component_structure_requests": [request],
        "component_structure_ai_tasks": [asdict(first_task)],
        "component_structure_ai_results": [],
    }
    monkeypatch.setattr(pipeline_module, "_load_state", lambda _path: state)
    monkeypatch.setattr(pipeline_module, "_assert_inputs_unchanged", lambda _state: None)
    monkeypatch.setattr(pipeline_module, "_save_state", lambda _path, _state: None)
    (tmp_path / "计算留痕数据").mkdir()

    first_path = tmp_path / "first.jsonl"
    first_path.write_text(
        json.dumps(
            {
                "task_id": first_task.task_id,
                "voucher_key": "V-1",
                "review_round": "single",
                "selected_entry_ids": ["E-1", "E-3"],
                "confidence": "low",
                "reason": "存在两个金额均守恒的组合",
                "reviewer_id": "R-1",
                "model_id": "M-1",
                "reviewed_at": "2026-08-22T18:00:00+08:00",
            },
            ensure_ascii=False,
        )
        + "\n",
        encoding="utf-8-sig",
    )
    first = pipeline_module.import_component_structure_ai_results(
        tmp_path, first_path
    )
    assert first["status"] == "待AI继续判断业务组成"
    second_task = pipeline_module._structure_ai_task_from_dict(
        state["component_structure_ai_tasks"][-1]
    )

    second_path = tmp_path / "second.jsonl"
    second_path.write_text(
        json.dumps(
            {
                "task_id": second_task.task_id,
                "voucher_key": "V-1",
                "review_round": "second",
                "selected_entry_ids": ["E-1", "E-3"],
                "confidence": "high",
                "reason": "复核后该组合金额守恒且与现有摘要关系最清楚",
                "reviewer_id": "R-2",
                "model_id": "M-2",
                "reviewed_at": "2026-08-22T18:05:00+08:00",
            },
            ensure_ascii=False,
        )
        + "\n",
        encoding="utf-8-sig",
    )
    completed = pipeline_module.import_component_structure_ai_results(
        tmp_path, second_path
    )

    assert completed["status"] == "业务组成AI判断已完成"
    assert state["component_structure_selections"] == {
        "V-1": ["E-1", "E-3"]
    }
    assert state["component_structure_selection_basis"] == {
        "V-1": "existing_evidence"
    }
    assert state["stage"] == "component_structure_confirmed"


def test_refund_confirmation_api_has_been_removed() -> None:
    assert "confirm_reversal_patterns" not in dir(pipeline_module)
from cashflow_direct.models import CashflowComponent, ClassificationDecision
from tests.fixture_factory import mark_dictionary_complete, write_end_to_end_case


def test_evidence_record_keeps_every_forced_check_that_can_change_the_route() -> None:
    decision = ClassificationDecision(
        component_id="CF-1",
        system_item_id="CFO-01",
        system_item_name="销售商品收到的现金",
        normal_direction="inflow",
        matched_rule_id="TEST",
        reason="构造测试",
        evidence_level="strong",
        company_rule_conflict=True,
        vat_base_missing=True,
        vat_base_component_id="VAT-BASE-1",
        vat_relation_status="unique",
        net_item_facts_missing=True,
    )

    payload = pipeline_module._evidence_assessment_payload(decision)

    assert payload["company_rule_conflict"] is True
    assert payload["vat_base_missing"] is True
    assert payload["vat_base_component_id"] == "VAT-BASE-1"
    assert payload["vat_relation_status"] == "unique"
    assert payload["net_item_facts_missing"] is True
    assert "new_reversal_pattern" not in payload


def test_pipeline_refreshes_vat_after_base_decision_changes() -> None:
    state = {
        "components": [
            asdict(
                CashflowComponent(
                    "BASE",
                    "V-1",
                    "支付材料款",
                    -100,
                    ("应付账款_材料供应商",),
                )
            ),
            asdict(
                CashflowComponent(
                    "VAT",
                    "V-1",
                    "支付材料进项税",
                    -13,
                    ("应交税费_应交增值税_进项税额",),
                )
            ),
        ],
        "source_allocations": [
            {"component_id": "BASE", "entry_id": "CASH-1", "allocated_cent": -100},
            {"component_id": "VAT", "entry_id": "CASH-1", "allocated_cent": -13},
        ],
    }
    decisions = (
        ClassificationDecision(
            "BASE",
            "CFO-04",
            "购买商品、接受劳务支付的现金",
            "outflow",
            "TEST",
            "基础项目已落定",
            "strong",
            resolved=True,
            decision_action="automatic_change",
        ),
        ClassificationDecision(
            "VAT",
            "CFO-06",
            "支付的各项税费",
            "outflow",
            "TEST",
            "等待基础项目",
            "strong",
            resolved=False,
            vat_base_missing=True,
        ),
    )

    refreshed = pipeline_module._refresh_vat_companion_decisions(state, decisions)

    assert refreshed[1].system_item_id == "CFO-04"
    assert refreshed[1].resolved is True
    assert refreshed[1].vat_base_component_id == "BASE"


def test_final_ai_records_include_valid_results_and_technical_failure_terminal_states() -> None:
    state = {
        "structured_ai_validation": {
            "valid_results": [{"task_id": "AI-OK", "status": "有效"}]
        },
        "ai_technical_failure_log": [
            {"task_id": "AI-FAIL", "attempt": 3, "status": "无有效结果"}
        ],
        "component_structure_ai_results": [
            {"task_id": "STRUCTURE-OK", "status": "有效"}
        ],
        "component_structure_ai_technical_failure_log": [
            {
                "task_id": "STRUCTURE-FAIL",
                "attempt": 3,
                "status": "无有效结果",
            }
        ],
    }

    records = pipeline_module._ai_records_from_state(state)

    assert {(item["阶段"], item["task_id"], item["status"]) for item in records} == {
        ("分类AI有效结果", "AI-OK", "有效"),
        ("分类AI技术失败", "AI-FAIL", "无有效结果"),
        ("业务组成结构AI有效结果", "STRUCTURE-OK", "有效"),
        ("业务组成结构AI技术失败", "STRUCTURE-FAIL", "无有效结果"),
    }


def test_removed_refund_state_cannot_create_a_completion_gate() -> None:
    pipeline_module._assert_agent_gates_closed(
        {"stage": "cash_scope_confirmed", "历史退款记录": [{"component_id": "CF-1"}]}
    )


def test_removed_group_has_no_separate_confirmation_command() -> None:
    from cashflow_direct.cli import build_parser

    with pytest.raises(SystemExit):
        build_parser().parse_args(
            ["confirm-materiality-groups", "--run-dir", "构造运行目录"]
        )


def test_state_save_retries_when_windows_temporarily_locks_target(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    original_replace = Path.replace
    attempts = 0

    def temporarily_locked(source: Path, target: Path) -> Path:
        nonlocal attempts
        attempts += 1
        if attempts == 1:
            raise PermissionError(5, "Windows临时拒绝访问", str(target))
        return original_replace(source, target)

    monkeypatch.setattr(Path, "replace", temporarily_locked)
    with tempfile.TemporaryDirectory() as tmp:
        run_dir = Path(tmp) / "构造运行目录"
        pipeline_module._save_state(run_dir, {"stage": "test"})

        assert attempts == 2
        assert json.loads(
            (
                run_dir / "计算留痕数据" / "运行状态.json"
            ).read_text(encoding="utf-8-sig")
        ) == {"stage": "test"}


def test_state_save_raises_after_ten_persistent_windows_lock_failures(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    attempts = 0

    def persistently_locked(source: Path, target: Path) -> Path:
        nonlocal attempts
        attempts += 1
        raise PermissionError(5, "Windows持续拒绝访问", str(target))

    monkeypatch.setattr(Path, "replace", persistently_locked)
    monkeypatch.setattr(pipeline_module, "sleep", lambda _seconds: None)
    with tempfile.TemporaryDirectory() as tmp:
        with pytest.raises(PermissionError, match="Windows持续拒绝访问"):
            pipeline_module._save_state(Path(tmp) / "构造运行目录", {"stage": "test"})

    assert attempts == 10


def test_state_save_does_not_retry_non_permission_file_errors(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    attempts = 0

    def other_file_error(source: Path, target: Path) -> Path:
        nonlocal attempts
        attempts += 1
        raise OSError(123, "其他文件错误", str(target))

    monkeypatch.setattr(Path, "replace", other_file_error)
    with tempfile.TemporaryDirectory() as tmp:
        with pytest.raises(OSError, match="其他文件错误"):
            pipeline_module._save_state(Path(tmp) / "构造运行目录", {"stage": "test"})

    assert attempts == 1


def _write_ai_routed_case(root: Path) -> Path:
    path = root / "构造AI分流.xlsx"
    workbook = Workbook()
    detail = workbook.active
    detail.title = "序时账"
    detail.append(["日期", "凭证号", "摘要", "科目", "借方", "贷方", "现流项目"])
    detail.append(
        [
            "2026-01-01",
            "记-1",
            "销售回款",
            "1002 银行存款",
            5_000,
            None,
            "收到其他与经营活动有关的现金",
        ]
    )
    detail.append(
        [
            "2026-01-01",
            "记-1",
            "销售回款",
            "主营业务收入",
            None,
            5_000,
            "收到其他与经营活动有关的现金",
        ]
    )
    balances = workbook.create_sheet("现金余额资料")
    balances.append(["项目", "金额"])
    balances.append(["期初现金及现金等价物余额", 0])
    balances.append(["期末现金及现金等价物余额", 5_000])
    balances.append(["汇率变动对现金及现金等价物的影响", 0])
    workbook.save(path)
    workbook.close()
    return path


def _write_illegal_summary_case(root: Path) -> Path:
    path = root / "构造空摘要.xlsx"
    workbook = Workbook()
    detail = workbook.active
    detail.title = "序时账"
    detail.append(["日期", "凭证号", "摘要", "科目", "借方", "贷方", "现流项目"])
    detail.append(["2026-01-01", "记-1", None, "1002 银行存款", 100, None, None])
    detail.append(["2026-01-01", "记-1", None, "主营业务收入", None, 100, None])
    balances = workbook.create_sheet("现金余额资料")
    balances.append(["项目", "金额"])
    balances.append(["期初现金及现金等价物余额", 0])
    balances.append(["期末现金及现金等价物余额", 100])
    balances.append(["汇率变动对现金及现金等价物的影响", 0])
    workbook.save(path)
    workbook.close()
    return path


def _write_split_vat_case(root: Path) -> Path:
    path = root / "构造价税拆分.xlsx"
    workbook = Workbook()
    detail = workbook.active
    detail.title = "序时账"
    detail.append(
        ["日期", "凭证号", "摘要", "科目", "借方", "贷方", "流量金额", "现流项目"]
    )
    detail.append(
        [
            "2026-01-01",
            "记-1",
            "支付材料含税款",
            "1002 银行存款",
            None,
            113,
            None,
            "购买商品、接受劳务支付的现金",
        ]
    )
    detail.append(
        [
            "2026-01-01",
            "记-1",
            "支付材料款",
            "应付账款_应付材料款",
            100,
            None,
            100,
            "购买商品、接受劳务支付的现金",
        ]
    )
    detail.append(
        [
            "2026-01-01",
            "记-1",
            "支付材料进项税",
            "应交税费_应交增值税_进项税额",
            13,
            None,
            13,
            "支付的各项税费",
        ]
    )
    balances = workbook.create_sheet("现金余额资料")
    balances.append(["项目", "金额"])
    balances.append(["期初现金及现金等价物余额", 113])
    balances.append(["期末现金及现金等价物余额", 0])
    balances.append(["汇率变动对现金及现金等价物的影响", 0])
    workbook.save(path)
    workbook.close()
    return path


def _write_split_vat_ai_case(root: Path) -> Path:
    path = root / "构造价税待AI.xlsx"
    workbook = Workbook()
    detail = workbook.active
    detail.title = "序时账"
    detail.append(
        ["日期", "凭证号", "摘要", "科目", "借方", "贷方", "流量金额", "现流项目"]
    )
    detail.append(
        [
            "2026-01-01",
            "记-2",
            "支付生产线设备含税款",
            "1002 银行存款",
            None,
            11_300,
            None,
            "支付其他与经营活动有关的现金",
        ]
    )
    detail.append(
        [
            "2026-01-01",
            "记-2",
            "支付自有设备购置款用于生产线建设",
            "在建工程_生产线设备",
            10_000,
            None,
            10_000,
            "支付其他与经营活动有关的现金",
        ]
    )
    detail.append(
        [
            "2026-01-01",
            "记-2",
            "支付设备进项税",
            "应交税费_应交增值税_进项税额",
            1_300,
            None,
            1_300,
            "支付的各项税费",
        ]
    )
    balances = workbook.create_sheet("现金余额资料")
    balances.append(["项目", "金额"])
    balances.append(["期初现金及现金等价物余额", 11_300])
    balances.append(["期末现金及现金等价物余额", 0])
    balances.append(["汇率变动对现金及现金等价物的影响", 0])
    workbook.save(path)
    workbook.close()
    return path


def test_run_classification_persists_the_automatic_fill_route() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        preflight = run_preflight(
            write_end_to_end_case(root),
            ("1000000", "750000", "50000"),
            output_parent=root,
        )
        confirm_cash_scope(preflight.run_dir, preflight.recommended_cash_decisions)
        mark_dictionary_complete(preflight.run_dir)

        result = run_classification(preflight.run_dir)

        state_path = preflight.run_dir / "计算留痕数据" / "运行状态.json"
        state = json.loads(state_path.read_text(encoding="utf-8-sig"))
        assert result.status == "consistency_completed"
        assert state["ai_tasks"] == []
        assert state["materiality_assessments"]
        assert state["source_allocations"]
        assert {
            decision["decision_action"] for decision in state["decisions"]
        } == {"automatic_fill"}
        assert {
            decision["materiality_level"] for decision in state["decisions"]
        } == {"M0"}
        assert all(decision["resolved"] for decision in state["decisions"])
        connection = sqlite3.connect(
            preflight.run_dir / "计算留痕数据" / "计算留痕.sqlite3"
        )
        try:
            counts = {
                table: connection.execute(
                    f'SELECT COUNT(*) FROM "{table}"'
                ).fetchone()[0]
                for table in (
                    "source_allocation",
                    "evidence_assessment",
                    "materiality_assessment",
                    "decision_route",
                    "run_version",
                )
            }
        finally:
            connection.close()
        assert counts["source_allocation"] == len(state["source_allocations"])
        assert counts["evidence_assessment"] == len(state["decisions"])
        assert counts["materiality_assessment"] == len(state["decisions"])
        assert counts["decision_route"] == len(state["decisions"])
        assert counts["run_version"] == len(
            current_versions(Path(__file__).resolve().parents[1])
        )


def test_pipeline_uses_source_allocations_for_split_vat_companion() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        preflight = run_preflight(
            [_write_split_vat_case(root)],
            ("100000", "50000", "5000"),
            output_parent=root,
        )
        confirm_cash_scope(preflight.run_dir, preflight.recommended_cash_decisions)
        mark_dictionary_complete(preflight.run_dir)

        run_classification(preflight.run_dir)

        state = json.loads(
            (
                preflight.run_dir / "计算留痕数据" / "运行状态.json"
            ).read_text(encoding="utf-8-sig")
        )
        decisions = {item["component_id"]: item for item in state["decisions"]}
        vat = next(
            item
            for item in decisions.values()
            if item["vat_relation_status"] == "unique"
        )
        base = decisions[vat["vat_base_component_id"]]

        assert vat["system_item_id"] == base["system_item_id"] == "CFO-04"
        assert vat["decision_action"] == "vat_follow_base"
        assert vat["vat_base_missing"] is False
        assert vat["component_id"] not in {
            task["component_id"] for task in state["ai_tasks"]
        }


def test_pipeline_refreshes_split_vat_after_ai_resolves_the_base() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        preflight = run_preflight(
            [_write_split_vat_ai_case(root)],
            ("100000", "50000", "5000"),
            output_parent=root,
        )
        confirm_cash_scope(preflight.run_dir, preflight.recommended_cash_decisions)
        mark_dictionary_complete(preflight.run_dir)

        classified = run_classification(preflight.run_dir)
        state_path = preflight.run_dir / "计算留痕数据" / "运行状态.json"
        state = json.loads(state_path.read_text(encoding="utf-8-sig"))
        assert classified.status == "waiting_ai"
        assert len(state["ai_tasks"]) == 1
        task = state["ai_tasks"][0]
        vat_before = next(
            item for item in state["decisions"] if item["vat_relation_status"] == "unique"
        )
        assert vat_before["resolved"] is False
        assert task["component_id"] == vat_before["vat_base_component_id"]
        assert "CFI-06" in task["candidate_item_ids"], task
        assert "支付自有设备购置款用于生产线建设" in task["context"], task["context"]
        assert "在建工程_生产线设备" in task["context"], task["context"]

        result_path = root / "价税基础AI结果.jsonl"
        result_path.write_text(
            json.dumps(
                {
                    "task_id": task["task_id"],
                    "component_id": task["component_id"],
                    "summary": {
                        "candidate_item_id": "CFI-06",
                        "quality": "strong",
                        "basis_text": "支付自有设备购置款用于生产线建设",
                        "classification_facts": ["purpose:long_asset_construction"],
                        "conflict": False,
                    },
                    "account_path": {
                        "candidate_item_id": "CFI-06",
                        "quality": "strong",
                        "basis_text": "在建工程_生产线设备",
                        "classification_facts": ["account:construction_in_progress"],
                        "conflict": False,
                    },
                    "sources_independent": True,
                    "business_conflict": False,
                    "direction_status": "compatible",
                    "reason": "摘要和完整路径均唯一指向自有长期资产购建",
                    "alternative_item_ids": [],
                    "note_ids": [],
                    "review_round": "single",
                    "reviewer_id": "test-reviewer",
                    "model_id": "test-model",
                    "reviewed_at": "2026-08-24T12:00:00+08:00",
                    "prior_result_difference": "首轮复核，无前序结果",
                },
                ensure_ascii=False,
            )
            + "\n",
            encoding="utf-8-sig",
        )

        imported = import_ai_results(preflight.run_dir, result_path)
        state = json.loads(state_path.read_text(encoding="utf-8-sig"))
        vat_after = next(
            item for item in state["decisions"] if item["vat_relation_status"] == "unique"
        )

        assert imported.status == "AI 已完成", state["structured_ai_validation"]
        assert vat_after["system_item_id"] == "CFI-06"
        assert vat_after["resolved"] is True
        assert vat_after["decision_action"] == "vat_follow_base"


def test_pipeline_builds_one_manual_choice_for_pending_base_and_vat() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        preflight = run_preflight(
            [_write_split_vat_ai_case(root)],
            ("10000", "5000", "500"),
            output_parent=root,
        )
        confirm_cash_scope(preflight.run_dir, preflight.recommended_cash_decisions)
        mark_dictionary_complete(preflight.run_dir)

        classified = run_classification(preflight.run_dir)
        assert classified.status == "waiting_human"
        final = finalize_run(preflight.run_dir)
        state = json.loads(
            (
                preflight.run_dir / "计算留痕数据" / "运行状态.json"
            ).read_text(encoding="utf-8-sig")
        )
        dependent = next(
            batch for batch in state["review_batches"] if batch["follows_component_id"]
        )

        assert dependent["follows_component_id"] in {
            component_id
            for batch in state["review_batches"]
            for component_id in batch["component_ids"]
        }
        workbook = load_workbook(final.workbook_path, data_only=False)
        try:
            review = workbook["重要待复核事项"]
            headers = [cell.value for cell in review[1]]
            manual_column = headers.index("人工确认项目") + 1
            dependent_row = next(
                row
                for row, batch in enumerate(state["review_batches"], 2)
                if batch["follows_component_id"]
            )
            self_choice = review.cell(dependent_row, manual_column)
            assert self_choice.data_type == "f"
            assert self_choice.protection.locked is True
        finally:
            workbook.close()


def test_pipeline_imports_structured_ai_sources_then_system_recalculates_action() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        preflight = run_preflight(
            [_write_ai_routed_case(root)],
            ("100000", "50000", "5000"),
            output_parent=root,
        )
        confirm_cash_scope(preflight.run_dir, preflight.recommended_cash_decisions)
        mark_dictionary_complete(preflight.run_dir)
        classified = run_classification(preflight.run_dir)
        assert classified.status == "waiting_ai"

        state_path = preflight.run_dir / "计算留痕数据" / "运行状态.json"
        state = json.loads(state_path.read_text(encoding="utf-8-sig"))
        task = state["ai_tasks"][0]
        result_path = root / "AI结构化结果.jsonl"
        result_path.write_text(
            json.dumps(
                {
                    "task_id": task["task_id"],
                    "component_id": task["component_id"],
                    "summary": {
                        "candidate_item_id": "CFO-01",
                        "quality": "strong",
                        "basis_text": "销售回款",
                        "classification_facts": ["action:sale_collection"],
                        "conflict": False,
                    },
                    "account_path": {
                        "candidate_item_id": "CFO-01",
                        "quality": "strong",
                        "basis_text": "主营业务收入",
                        "classification_facts": ["account:main_revenue"],
                        "conflict": False,
                    },
                    "sources_independent": True,
                    "business_conflict": False,
                    "direction_status": "compatible",
                    "reason": "只重新解释摘要和完整路径",
                    "alternative_item_ids": [],
                    "note_ids": [],
                    "review_round": "single",
                    "reviewer_id": "test-reviewer-single",
                    "model_id": "test-model",
                    "reviewed_at": "2026-08-21T00:00:00+08:00",
                    "prior_result_difference": "首轮复核，无前序结果",
                },
                ensure_ascii=False,
            )
            + "\n",
            encoding="utf-8-sig",
        )

        imported = import_ai_results(preflight.run_dir, result_path)

        assert imported.status == "AI 已完成"
        state = json.loads(state_path.read_text(encoding="utf-8-sig"))
        decision = state["decisions"][0]
        assert decision["system_item_id"] == "CFO-01"
        assert decision["evidence_score"] == 90
        assert decision["decision_action"] == "automatic_change"
        assert decision["resolved"] is True
        connection = sqlite3.connect(
            preflight.run_dir / "计算留痕数据" / "计算留痕.sqlite3"
        )
        try:
            stored_decision = json.loads(
                connection.execute(
                    "SELECT payload_json FROM classification_decision WHERE record_id = ?",
                    (decision["component_id"],),
                ).fetchone()[0]
            )
            stored_route = json.loads(
                connection.execute(
                    "SELECT payload_json FROM decision_route WHERE record_id = ?",
                    (decision["component_id"],),
                ).fetchone()[0]
            )
        finally:
            connection.close()
        assert stored_decision["evidence_score"] == 90
        assert stored_route["action"] == "automatic_change"


def test_three_invalid_ai_submissions_clear_the_queue_and_use_the_cell_failure_exit() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        preflight = run_preflight(
            [_write_ai_routed_case(root)],
            ("100000", "50000", "5000"),
            output_parent=root,
        )
        confirm_cash_scope(preflight.run_dir, preflight.recommended_cash_decisions)
        mark_dictionary_complete(preflight.run_dir)
        classified = run_classification(preflight.run_dir)
        assert classified.status == "waiting_ai"
        state_path = preflight.run_dir / "计算留痕数据" / "运行状态.json"
        task_id = json.loads(state_path.read_text(encoding="utf-8-sig"))["ai_tasks"][0][
            "task_id"
        ]

        for attempt in range(1, 4):
            invalid_path = root / f"无效AI结果_{attempt}.jsonl"
            invalid_path.write_text(
                json.dumps({"task_id": task_id}, ensure_ascii=False) + "\n",
                encoding="utf-8-sig",
            )
            result = import_ai_results(preflight.run_dir, invalid_path)

        state = json.loads(state_path.read_text(encoding="utf-8-sig"))
        assert result.missing_count == 0
        assert state["classification_summary"]["ai_tasks_missing"] == 0
        assert state["ai_terminal_failure_ids"] == [task_id]
        assert state["ai_technical_failure_log"][-1]["status"] == "无有效结果"
        assert state["decisions"][0]["decision_action"] == "automatic_keep"
        assert state["decisions"][0]["resolved"] is True


def test_three_missing_ai_submissions_also_clear_the_queue_and_reach_a_terminal_exit() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        preflight = run_preflight(
            [_write_ai_routed_case(root)],
            ("100000", "50000", "5000"),
            output_parent=root,
        )
        confirm_cash_scope(preflight.run_dir, preflight.recommended_cash_decisions)
        mark_dictionary_complete(preflight.run_dir)
        classified = run_classification(preflight.run_dir)
        assert classified.status == "waiting_ai"
        state_path = preflight.run_dir / "计算留痕数据" / "运行状态.json"
        task_id = json.loads(state_path.read_text(encoding="utf-8-sig"))["ai_tasks"][0][
            "task_id"
        ]

        for attempt in range(1, 4):
            missing_path = root / f"漏答AI结果_{attempt}.jsonl"
            missing_path.write_text("", encoding="utf-8-sig")
            result = import_ai_results(preflight.run_dir, missing_path)

        state = json.loads(state_path.read_text(encoding="utf-8-sig"))
        assert result.missing_count == 0
        assert state["classification_summary"]["ai_tasks_missing"] == 0
        assert state["ai_terminal_failure_ids"] == [task_id]
        assert state["ai_task_attempts"] == {task_id: 3}
        assert state["ai_technical_failure_log"][-1]["status"] == "无有效结果"


def test_pipeline_restores_valid_original_when_ai_does_not_prove_change() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        preflight = run_preflight(
            [_write_ai_routed_case(root)],
            ("100000", "50000", "5000"),
            output_parent=root,
        )
        confirm_cash_scope(preflight.run_dir, preflight.recommended_cash_decisions)
        mark_dictionary_complete(preflight.run_dir)
        run_classification(preflight.run_dir)
        state_path = preflight.run_dir / "计算留痕数据" / "运行状态.json"
        state = json.loads(state_path.read_text(encoding="utf-8-sig"))
        first_task = state["ai_tasks"][0]

        def payload(task, *, review_round, account_quality):
            return {
                "task_id": task["task_id"],
                "component_id": task["component_id"],
                "summary": {
                    "candidate_item_id": "CFO-01",
                    "quality": "strong",
                    "basis_text": "销售回款",
                    "classification_facts": ["action:sale_collection"],
                    "conflict": False,
                },
                "account_path": {
                    "candidate_item_id": "CFO-01",
                    "quality": account_quality,
                    "basis_text": "主营业务收入",
                    "classification_facts": ["account:main_revenue"],
                    "conflict": False,
                },
                "sources_independent": True,
                "business_conflict": False,
                "direction_status": "compatible",
                "reason": "只重新解释摘要和完整路径",
                "alternative_item_ids": [],
                "note_ids": [],
                "review_round": review_round,
                "reviewer_id": f"test-reviewer-{review_round}",
                "model_id": "test-model",
                "reviewed_at": "2026-08-21T00:00:00+08:00",
                "prior_result_difference": (
                    "首轮复核，无前序结果"
                    if review_round == "single"
                    else "第二轮重新核对完整路径的业务属性"
                ),
            }

        first_path = root / "AI首轮结果.jsonl"
        first_path.write_text(
            json.dumps(payload(first_task, review_round="single", account_quality="weak"), ensure_ascii=False) + "\n",
            encoding="utf-8-sig",
        )
        first = import_ai_results(preflight.run_dir, first_path)

        assert first.status == "AI 已完成"
        state = json.loads(state_path.read_text(encoding="utf-8-sig"))
        assert state["decisions"][0]["resolved"] is True
        assert state["decisions"][0]["decision_action"] == "automatic_keep"
        assert state["decisions"][0]["system_item_id"] == "CFO-03"


def test_old_or_changed_scoring_version_cannot_continue_in_the_same_run() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        preflight = run_preflight(
            write_end_to_end_case(root),
            ("1000000", "750000", "50000"),
            output_parent=root,
        )
        state_path = preflight.run_dir / "计算留痕数据" / "运行状态.json"
        state = json.loads(state_path.read_text(encoding="utf-8-sig"))
        assert {
            "schema",
            "scoring",
            "action_matrix",
            "account_mapping",
            "company_notes",
            "rule_pack",
            "account_dictionary",
        } <= set(state["versions"])
        state["versions"]["scoring"] = "旧评分版本"
        state_path.write_text(
            json.dumps(state, ensure_ascii=False), encoding="utf-8-sig"
        )

        with pytest.raises(RuntimeError, match="旧运行目录.*新建运行目录"):
            confirm_cash_scope(
                preflight.run_dir, preflight.recommended_cash_decisions
            )


def test_illegal_blank_summary_enters_the_same_workbook_for_user_decision() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        preflight = run_preflight(
            [_write_illegal_summary_case(root)],
            ("100000", "50000", "5000"),
            output_parent=root,
        )
        confirm_cash_scope(preflight.run_dir, preflight.recommended_cash_decisions)
        mark_dictionary_complete(preflight.run_dir)

        classified = run_classification(preflight.run_dir)

        state = json.loads(
            (
                preflight.run_dir / "计算留痕数据" / "运行状态.json"
            ).read_text(encoding="utf-8-sig")
        )
        assert classified.status == "waiting_human"
        assert len(state["components"]) == 1
        assert state["components"][0]["cash_delta_cent"] == 10_000
        assert "summary_empty" in state["components"][0]["anomalies"]
        assert state["decisions"][0]["decision_action"] == "isolate_invalid_input"
        assert state["decisions"][0]["resolved"] is False
        assert state["ai_tasks"] == []
        assert any(
            issue["kind"] == "警告" and "摘要为空" in issue["message"]
            for issue in state["normalization_issues"]
        )

        final = finalize_run(preflight.run_dir)
        assert final.workbook_path.is_file()
        assert final.overall_status == "待完成人工确认"
        state = json.loads(
            (
                preflight.run_dir / "计算留痕数据" / "运行状态.json"
            ).read_text(encoding="utf-8-sig")
        )
        assert state["review_batches"]
        assert state["components"][0]["component_id"] in state["review_batches"][0][
            "component_ids"
        ]
        assert not any(
            "非法输入" in error for error in state["final_readiness"]["errors"]
        )
