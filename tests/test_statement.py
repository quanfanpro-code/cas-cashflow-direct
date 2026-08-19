from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from cashflow_direct.models import ClassificationDecision
from cashflow_direct.classification import load_rule_pack
from cashflow_direct.semantic_mapping import MappingQuestion
from cashflow_direct.statement import (
    ExistingStatementResult,
    aggregate_statement,
    compare_statement,
    detect_statement_sheets,
    parse_existing_statement,
    reconcile_cash,
)
from tests.fixture_factory import (
    cashflow_component,
    classified_components,
    write_detail_plus_statement_fixture,
    write_existing_statement_fixture,
)


ROOT = Path(__file__).resolve().parents[1]


def _single_decision_case(
    component_id: str,
    cash_delta_cent: int,
    item_id: str,
    item_name: str,
    normal_direction: str,
):
    rules = load_rule_pack(ROOT)
    component = cashflow_component(
        "匿名净额事项", cash_delta_cent, component_id=component_id
    )
    decision = ClassificationDecision(
        component_id=component_id,
        system_item_id=item_id,
        system_item_name=item_name,
        normal_direction=normal_direction,
        matched_rule_id="MANUAL-LABEL",
        reason="匿名测试决策",
        evidence_level="high",
    )
    return rules, (component,), (decision,)


class StatementTests(unittest.TestCase):
    def test_leaf_subtotals_and_net_cash_reconcile(self) -> None:
        case = classified_components()
        result = aggregate_statement(case.components, case.decisions, case.rules)
        self.assertEqual(35, len(result.values))
        self.assertEqual(result.values["CFO-IN"] - result.values["CFO-OUT"], result.values["CFO-NET"])
        expected = (
            result.values["CFO-NET"]
            + result.values["CFI-NET"]
            + result.values["CFF-NET"]
            + result.values["FX"]
        )
        self.assertEqual(expected, result.values["NET-CASH"])
        self.assertEqual(("S1",), result.support_component_ids["CFO-01"])
        self.assertIsNone(result.prior_values["CFO-01"])

    def test_disposal_net_negative_moves_to_other_investing_outflow(self) -> None:
        # 应用指南三(二)3：处置长期资产收回的现金净额为负数时，
        # 在"支付其他与投资活动有关的现金"项目中反映
        rules, components, decisions = _single_decision_case(
            "NEG-DISP",
            -50_000,
            "CFI-03",
            "处置固定资产、无形资产和其他长期资产收回的现金净额",
            "inflow",
        )

        result = aggregate_statement(components, decisions, rules)

        self.assertEqual(0, result.values["CFI-03"])
        self.assertEqual(50_000, result.values["CFI-09"])
        self.assertEqual(("NEG-DISP",), result.support_component_ids["CFI-09"])
        self.assertEqual((), result.support_component_ids["CFI-03"])

    def test_acquire_subsidiary_net_negative_moves_to_other_investing_inflow(self) -> None:
        # 应用指南三(二)5：取得子公司支付的现金净额为负数时，
        # 在"收到其他与投资活动有关的现金"项目中反映
        rules, components, decisions = _single_decision_case(
            "NEG-ACQ",
            80_000,
            "CFI-08",
            "取得子公司及其他营业单位支付的现金净额",
            "outflow",
        )

        result = aggregate_statement(components, decisions, rules)

        self.assertEqual(0, result.values["CFI-08"])
        self.assertEqual(80_000, result.values["CFI-05"])
        self.assertEqual(("NEG-ACQ",), result.support_component_ids["CFI-05"])
        self.assertEqual((), result.support_component_ids["CFI-08"])

    def test_dispose_subsidiary_net_negative_moves_to_other_investing_outflow(self) -> None:
        # 应用指南三(二)4：处置子公司及其他营业单位收到的现金净额为负数时，
        # 填列至"支付其他与投资活动有关的现金"项目
        rules, components, decisions = _single_decision_case(
            "NEG-DISP-SUB",
            -30_000,
            "CFI-04",
            "处置子公司及其他营业单位收到的现金净额",
            "inflow",
        )

        result = aggregate_statement(components, decisions, rules)

        self.assertEqual(0, result.values["CFI-04"])
        self.assertEqual(30_000, result.values["CFI-09"])
        self.assertEqual(("NEG-DISP-SUB",), result.support_component_ids["CFI-09"])
        self.assertEqual((), result.support_component_ids["CFI-04"])

    def test_positive_net_items_are_not_migrated(self) -> None:
        # 回归保护：净额为正时保持原项目，不发生迁移
        rules, components, decisions = _single_decision_case(
            "POS-DISP",
            60_000,
            "CFI-03",
            "处置固定资产、无形资产和其他长期资产收回的现金净额",
            "inflow",
        )

        result = aggregate_statement(components, decisions, rules)

        self.assertEqual(60_000, result.values["CFI-03"])
        self.assertEqual(0, result.values["CFI-09"])
        self.assertEqual(("POS-DISP",), result.support_component_ids["CFI-03"])

    def test_opening_and_fx_are_injected_into_closing_cash_formula(self) -> None:
        case = classified_components()
        result = aggregate_statement(
            case.components,
            case.decisions,
            case.rules,
            opening_cent=1_000_000,
            fx_cent=12_300,
        )
        self.assertEqual(1_000_000, result.values["CASH-OPENING"])
        self.assertEqual(12_300, result.values["FX"])
        self.assertEqual(
            result.values["CASH-OPENING"] + result.values["NET-CASH"],
            result.values["CASH-CLOSING"],
        )

    def test_custom_rows_map_to_standard_parent_without_double_counting(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "客户正表.xlsx"
            write_existing_statement_fixture(path, header_row=7, with_custom_rows=True)
            case = classified_components()
            existing = parse_existing_statement(path, case.rules)
            self.assertNotIsInstance(existing, MappingQuestion)
            self.assertEqual(35, len(existing.values))
            self.assertEqual("CFO-03", existing.custom_rows[0].parent_item_id)
            self.assertEqual(existing.values["CFO-03"], existing.standardized_values["CFO-03"])
            self.assertEqual(0, existing.values["CFO-02"])
            self.assertIsNone(existing.prior_values["CFO-02"])

    def test_existing_comparison_has_each_standard_row_and_source_support(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "客户正表.xlsx"
            write_existing_statement_fixture(path, header_row=11, with_custom_rows=False)
            case = classified_components()
            existing = parse_existing_statement(path, case.rules)
            computed = aggregate_statement(case.components, case.decisions, case.rules)
            comparison = compare_statement(existing, computed)
            self.assertEqual(35, len(comparison.rows))
            row = next(item for item in comparison.rows if item.item_id == "CFO-01")
            self.assertEqual(computed.values["CFO-01"] - existing.values["CFO-01"], row.difference_cent)
            self.assertEqual(("S1",), row.support_component_ids)

    def test_existing_statement_scans_all_sheets_and_accepts_year_headers(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "多页客户正表.xlsx"
            write_existing_statement_fixture(path, header_row=7, with_custom_rows=False)
            workbook = load_workbook(path)
            statement_sheet = workbook.worksheets[0]
            statement_sheet["C7"] = "本年金额"
            statement_sheet["D7"] = "上年金额"
            cover = workbook.create_sheet("封面", 0)
            cover["A1"] = "审计资料封面"
            workbook.save(path)
            workbook.close()

            result = parse_existing_statement(path, classified_components().rules)
            self.assertNotIsInstance(result, MappingQuestion)
            self.assertEqual("报表页_随机", result.sheet_name)
            self.assertEqual(35, len(result.values))

    def test_multiple_statement_sheets_return_an_ambiguity_question(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "重复正表.xlsx"
            write_existing_statement_fixture(path, header_row=7, with_custom_rows=False)
            workbook = load_workbook(path)
            duplicate = workbook.copy_worksheet(workbook.worksheets[0])
            duplicate.title = "另一张现金流量表"
            workbook.save(path)
            workbook.close()

            result = parse_existing_statement(path, classified_components().rules)
            self.assertIsInstance(result, MappingQuestion)
            self.assertEqual("statement_sheet", result.role)

    def test_unmapped_total_returns_question_instead_of_guessing(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "歧义正表.xlsx"
            write_existing_statement_fixture(path, 5, False, include_unknown=True)
            result = parse_existing_statement(path, classified_components().rules)
            self.assertIsInstance(result, MappingQuestion)

    def test_existing_statement_accepts_period_range_header(self) -> None:
        # 清平式表头：项目 | 行次 | 2026年1-6月（无"本期/本年"字样）
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "清平式正表.xlsx"
            rules = classified_components().rules
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["单位名称：匿名公司", None, "未审数"])
            sheet.append(["项目", "行次", "2026年1-6月"])
            row_index = 3
            for item in rules.statement_items:
                sheet.cell(row_index, 1, item.name)
                sheet.cell(row_index, 2, item.display_order)
                sheet.cell(row_index, 3, item.display_order / 100)
                row_index += 1
            workbook.save(path)
            result = parse_existing_statement(path, rules)
            self.assertIsInstance(result, ExistingStatementResult)
            self.assertEqual(1, result.values["CFO-01"])  # display_order=1 → 0.01元 → 1分

    def test_sequential_integer_column_never_wins_amount(self) -> None:
        # 序号列 1,2,3…（数值占比100%）不得抢金额列
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "带序号正表.xlsx"
            rules = classified_components().rules
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["序号", "项目", "金额"])
            row_index = 2
            for number, item in enumerate(rules.statement_items, 1):
                sheet.cell(row_index, 1, number)
                sheet.cell(row_index, 2, item.name)
                sheet.cell(row_index, 3, item.display_order / 100)
                row_index += 1
            workbook.save(path)
            result = parse_existing_statement(path, rules)
            self.assertIsInstance(result, ExistingStatementResult)
            self.assertEqual(1, result.values["CFO-01"])  # 金额列选中，序号列排除

    def test_statement_tie_prefers_period_words_then_asks(self) -> None:
        rules = classified_components().rules
        with tempfile.TemporaryDirectory() as tmp:
            # 并列且无"本期/本年/金额"字样 → 提问
            ambiguous = Path(tmp) / "并列无字样.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["项目", "2026年1-6月", "2026年1-5月"])
            row_index = 2
            for item in rules.statement_items:
                sheet.cell(row_index, 1, item.name)
                sheet.cell(row_index, 2, item.display_order / 100)
                sheet.cell(row_index, 3, item.display_order / 100)
                row_index += 1
            workbook.save(ambiguous)
            result = parse_existing_statement(ambiguous, rules)
            self.assertIsInstance(result, MappingQuestion)
            self.assertEqual("statement_header", result.role)
            self.assertIn("并列", str(result.sample_values[0]))

            # 并列时含"本期"字样的列胜出
            tied = Path(tmp) / "并列有本期.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["项目", "累计数", "本期数"])
            row_index = 2
            for item in rules.statement_items:
                sheet.cell(row_index, 1, item.name)
                sheet.cell(row_index, 2, item.display_order / 10)
                sheet.cell(row_index, 3, item.display_order / 100)
                row_index += 1
            workbook.save(tied)
            result = parse_existing_statement(tied, rules)
            self.assertIsInstance(result, ExistingStatementResult)
            self.assertEqual(1, result.values["CFO-01"])  # 本期数列 0.01 → 1分，非累计列 10分

    def test_low_hit_rate_sheet_is_not_statement(self) -> None:
        # 35 标准行 + 40 行"其中"自定义行 → 命中率 46.7% < 50% → 不当正表
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "低命中率.xlsx"
            rules = classified_components().rules
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["项目", "本期金额"])
            row_index = 2
            for item in rules.statement_items:
                sheet.cell(row_index, 1, item.name)
                sheet.cell(row_index, 2, item.display_order / 100)
                row_index += 1
            for index in range(40):
                sheet.cell(row_index, 1, f"其中匿名明细{index + 1}")
                sheet.cell(row_index, 2, 0.01)
                row_index += 1
            workbook.save(path)
            result = parse_existing_statement(path, rules)
            self.assertNotIsInstance(result, ExistingStatementResult)

    def test_section_heading_rows_without_amount_are_skipped(self) -> None:
        # 清平式正表带节标题行（金额为空）→ 跳过，不影响整表识别
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "带节标题正表.xlsx"
            rules = classified_components().rules
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["项目", "行次", "2026年1-6月"])
            row_index = 3
            sheet.cell(row_index, 1, "一、经营活动产生的现金流量：")
            row_index += 1
            for item in rules.statement_items:
                sheet.cell(row_index, 1, item.name)
                sheet.cell(row_index, 2, item.display_order)
                sheet.cell(row_index, 3, item.display_order / 100)
                row_index += 1
            sheet.cell(row_index, 1, "经营活动现金流入小计")
            sheet.cell(row_index, 3, 123.45)
            workbook.save(path)
            result = parse_existing_statement(path, rules)
            self.assertIsInstance(result, ExistingStatementResult)
            self.assertEqual(1, result.values["CFO-01"])
            self.assertEqual(35, len(result.values))

    def test_ordinal_prefixed_item_names_still_match(self) -> None:
        # 清平式：净额类项目带"四、五、六"序数前缀也要能匹配
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "序数前缀正表.xlsx"
            rules = classified_components().rules
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["项目", "2026年1-6月"])
            row_index = 2
            for item in rules.statement_items:
                name = item.name
                if item.item_id == "FX":
                    name = "四、" + item.name
                elif item.item_id == "NET-CASH":
                    name = "五、" + item.name
                elif item.item_id == "CASH-CLOSING":
                    name = "六、" + item.name
                sheet.cell(row_index, 1, name)
                sheet.cell(row_index, 2, item.display_order / 100)
                row_index += 1
            workbook.save(path)
            result = parse_existing_statement(path, rules)
            self.assertIsInstance(result, ExistingStatementResult)
            self.assertEqual(35, len(result.values))

    def test_hit_rate_gate_uses_ordinal_fallback(self) -> None:
        # 全部项目名带序数前缀时，命中率统计也要用兜底匹配，不得误拒
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "全序数正表.xlsx"
            rules = classified_components().rules
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["项目", "本期金额"])
            ordinals = "一二三四五六七八九十"
            row_index = 2
            for number, item in enumerate(rules.statement_items):
                sheet.cell(row_index, 1, f"{ordinals[number % 10]}、" + item.name)
                sheet.cell(row_index, 2, item.display_order / 100)
                row_index += 1
            workbook.save(path)
            result = parse_existing_statement(path, rules)
            self.assertIsInstance(result, ExistingStatementResult)
            self.assertEqual(35, len(result.values))

    def test_two_char_ordinal_prefix_still_matches(self) -> None:
        # "十一、"这类两位序数前缀也要能匹配（锁定既有正则行为）
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "两位序数正表.xlsx"
            rules = classified_components().rules
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["项目", "2026年1-6月"])
            row_index = 2
            for item in rules.statement_items:
                name = item.name
                if item.item_id == "FX":
                    name = "十一、" + item.name
                sheet.cell(row_index, 1, name)
                sheet.cell(row_index, 2, item.display_order / 100)
                row_index += 1
            workbook.save(path)
            result = parse_existing_statement(path, rules)
            self.assertIsInstance(result, ExistingStatementResult)
            self.assertEqual(35, len(result.values))

    def test_low_hit_rate_reports_accurate_reason(self) -> None:
        # 命中率不足时如实说明"匹配率过低"，而不是笼统的"未找到项目列"
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "低命中率.xlsx"
            rules = classified_components().rules
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["项目", "本期金额"])
            row_index = 2
            for item in rules.statement_items:
                sheet.cell(row_index, 1, item.name)
                sheet.cell(row_index, 2, item.display_order / 100)
                row_index += 1
            for index in range(40):
                sheet.cell(row_index, 1, f"其中匿名明细{index + 1}")
                sheet.cell(row_index, 2, 0.01)
                row_index += 1
            workbook.save(path)
            result = parse_existing_statement(path, rules)
            self.assertIsInstance(result, MappingQuestion)
            self.assertIn("匹配率", str(result.sample_values[0]))

    def test_zero_hit_sheet_with_project_column_is_not_statement(self) -> None:
        # 项目列零命中 → 不当正表，通用提示，不产生"匹配率"提问
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "零命中.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["项目", "本期金额"])
            sheet.append(["完全无关的内容", 1])
            sheet.append(["另一个无关行", 2])
            workbook.save(path)
            result = parse_existing_statement(path, classified_components().rules)
            self.assertIsInstance(result, MappingQuestion)
            self.assertNotIn("匹配率", str(result.sample_values[0]))

    def test_detect_statement_sheets_returns_per_sheet_results(self) -> None:
        # 明细页不是正表 → None；正表页唯一命中 → ExistingStatementResult
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "明细加正表.xlsx"
            write_detail_plus_statement_fixture(path)
            detected = detect_statement_sheets(path, classified_components().rules)
            self.assertIsNone(detected.get("随机明细"))
            self.assertIsInstance(detected.get("报表页_随机"), ExistingStatementResult)

    def test_detect_statement_sheets_reports_multiple_statement_sheets(self) -> None:
        # 多正表页 → 每个命中 sheet 都返回结果，由调用方做歧义判断
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "重复正表.xlsx"
            write_existing_statement_fixture(path, header_row=7, with_custom_rows=False)
            workbook = load_workbook(path)
            duplicate = workbook.copy_worksheet(workbook.worksheets[0])
            duplicate.title = "另一张现金流量表"
            workbook.save(path)
            workbook.close()
            detected = detect_statement_sheets(path, classified_components().rules)
            hits = [
                item for item in detected.values() if isinstance(item, ExistingStatementResult)
            ]
            self.assertEqual(2, len(hits))

    def test_cash_reconciliation_never_plugs_missing_fx(self) -> None:
        case = classified_components()
        statement = aggregate_statement(case.components, case.decisions, case.rules)
        incomplete = reconcile_cash(statement, opening_cent=1_000, closing_cent=2_000, fx_cent=None)
        self.assertEqual("现金流量表与货币资金变动的勾稽核对：未完成", incomplete.status)
        self.assertIsNone(incomplete.fx_cent)
        fx = 100
        net = statement.values["CFO-NET"] + statement.values["CFI-NET"] + statement.values["CFF-NET"] + fx
        completed = reconcile_cash(statement, opening_cent=1_000, closing_cent=1_000 + net, fx_cent=fx)
        self.assertEqual("现金流量表与货币资金变动的勾稽核对：相符", completed.status)
        self.assertEqual(0, completed.difference_cent)


def test_fuzzy_item_name_matching(tmp_path):
    """正表措辞变体（"收到税收返还"）能模糊匹配到 CFO-02，不再直接失败（A2，Task 12）。"""
    rules = load_rule_pack(ROOT)
    path = Path(tmp_path) / "措辞变体正表.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "现流表"
    ws.append(["项目", "行次", "本期金额", "上期金额"])
    for item in rules.statement_items:
        name = "收到的税收返还" if item.item_id == "CFO-02" else item.name
        ws.append([name, item.display_order, 0, None])
    wb.save(path)
    result = parse_existing_statement(path, rules)
    assert isinstance(result, ExistingStatementResult)
    assert result.values["CFO-02"] == 0


def test_multi_year_column_selected_by_reference_years(tmp_path):
    """多时间列正表：结合明细日期年份选定本期列（2025），不再并列直停（A3，Task 13）。"""
    rules = load_rule_pack(ROOT)
    path = Path(tmp_path) / "多时间列正表.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "现流表"
    ws.append(["项目", "行次", "2024年度金额", "2025年度金额"])
    for item in rules.statement_items:
        ws.append([item.name, item.display_order, 100, 200])
    wb.save(path)
    result = parse_existing_statement(path, rules, reference_years=frozenset({2025}))
    assert isinstance(result, ExistingStatementResult)
    # 选 2025 列，200 元=20000 分（结果以分计）
    assert result.values["CFO-01"] == 20000


if __name__ == "__main__":
    unittest.main()
