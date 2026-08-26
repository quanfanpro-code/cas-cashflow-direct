from __future__ import annotations

import tempfile
from dataclasses import replace
from pathlib import Path

from openpyxl import load_workbook

from cashflow_direct.models import ReviewBatch
from cashflow_direct.workbook_output import (
    USE_SYSTEM_FALLBACK,
    build_output_workbook,
    calculate_manual_adjustments,
)
from tests.fixture_factory import workbook_model


def trace_rows() -> tuple[dict[str, object], ...]:
    common = {
        "业务组成编号(技术)": "RC-1",
        "日期": "2025-03-22",
        "凭证字": "记",
        "凭证号": "146",
        "本行完整对方科目路径": "主营业务收入",
        "标准一级科目": "主营业务收入",
        "现金账户路径": "银行存款_民生银行成都分行2385",
        "现金方向依据": "借贷差额",
        "原项目标准化结果": "原项目为空",
        "系统候选项目": "销售商品、提供劳务收到的现金",
        "判断理由": "低于实际执行重要性，系统兜底",
        "摘要来源质量": "中等证据25分",
        "完整路径来源质量": "强证据45分",
        "证据得分": 70,
    }
    return (
        {**common, "本行摘要": "收3.21日营业款", "借方": 34_352.90, "贷方": 0.0, "流量金额（原币）": 34_352.90, "本行分配现金变化": 34_352.90},
        {**common, "本行摘要": "收3.22日营业款", "借方": 97_631.43, "贷方": 0.0, "流量金额（原币）": 97_631.43, "本行分配现金变化": 97_631.43},
    )


def fallback_batch() -> ReviewBatch:
    return ReviewBatch(
        batch_id="FALLBACK-1",
        component_ids=("RC-1",),
        proposed_item_code="CFO-01",
        alternative_item_codes=("CFO-03",),
        worst_case_impact_cent=13_198_433,
        reason="低于实际执行重要性，按较高分独立来源兜底；同分优先完整科目路径来源",
        baseline_statement_amount_cent=13_198_433,
        cash_delta_cent=13_198_433,
        baseline_item_code="CFO-01",
        fallback_source="account_path",
        fallback_step="source_preferred",
    )


def test_manual_sheets_only_show_real_rows_and_each_amount_is_numeric() -> None:
    base = workbook_model(1, 0)
    important = replace(base.review_batches[0], component_ids=("RC-1",), cash_delta_cent=13_198_433)
    model = replace(base, review_batches=(important,), low_amount_fallback_batches=(fallback_batch(),), trace_rows=trace_rows())

    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "真实明细.xlsx"
        build_output_workbook(model, path)
        workbook = load_workbook(path, data_only=False)
        try:
            assert len(workbook.sheetnames) == 13
            assert "低金额系统兜底明细" in workbook.sheetnames
            assert "低金额批量处理" not in workbook.sheetnames
            for sheet_name in ("重要待复核事项", "低金额系统兜底明细"):
                sheet = workbook[sheet_name]
                headers = [cell.value for cell in sheet[1]]
                for removed in ("行类型", "批次编号(技术)", "批次最不利影响金额", "批次现金变化金额", "人工处理状态", "人工可选标准项目"):
                    assert removed not in headers
                business_col = headers.index("同一业务序号") + 1
                business_rows = [row for row in range(2, sheet.max_row + 1) if sheet.cell(row, business_col).value]
                assert business_rows == [2, 3]
                amount_header = "本行分配现金变化" if sheet_name == "重要待复核事项" else "本行金额"
                amount_col = headers.index(amount_header) + 1
                assert [sheet.cell(row, amount_col).value for row in business_rows] == [34_352.90, 97_631.43]
                assert all(isinstance(sheet.cell(row, amount_col).value, (int, float)) for row in business_rows)
                text = "".join(str(cell.value or "") for row in sheet.iter_rows() for cell in row)
                assert "批次主行" not in text
                assert "明细随批次主行生效" not in text
        finally:
            workbook.close()


def test_fallback_default_is_valid_and_one_business_choice_updates_all_real_rows() -> None:
    model = replace(workbook_model(0, 0), low_amount_fallback_batches=(fallback_batch(),), trace_rows=trace_rows())

    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "兜底改选.xlsx"
        build_output_workbook(model, path)
        workbook = load_workbook(path, data_only=False)
        try:
            sheet = workbook["低金额系统兜底明细"]
            headers = [cell.value for cell in sheet[1]]
            choice_col = headers.index("人工改选项目") + 1
            final_col = headers.index("最终采用项目") + 1
            first_choice = sheet.cell(2, choice_col)
            second_choice = sheet.cell(3, choice_col)
            assert first_choice.value == "采用系统兜底项目"
            assert first_choice.protection.locked is False
            assert second_choice.data_type == "f"
            assert first_choice.coordinate in second_choice.value
            assert sheet.cell(2, final_col).data_type == "f"
            assert sheet.cell(3, final_col).data_type == "f"
            assert sheet.data_validations.dataValidation
        finally:
            workbook.close()


def test_important_businesses_are_contiguous_without_synthetic_rows() -> None:
    base = workbook_model(2, 0)
    first = replace(base.review_batches[0], component_ids=("RC-1",))
    second = replace(base.review_batches[1], component_ids=("RC-2",))
    second_row = {
        **trace_rows()[0],
        "业务组成编号(技术)": "RC-2",
        "本行摘要": "第二项业务",
        "本行分配现金变化": -50.0,
    }
    model = replace(
        base,
        review_batches=(first, second),
        trace_rows=(*trace_rows(), second_row),
    )

    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "同一业务连续.xlsx"
        build_output_workbook(model, path)
        workbook = load_workbook(path, data_only=False)
        try:
            sheet = workbook["重要待复核事项"]
            headers = [cell.value for cell in sheet[1]]
            business_col = headers.index("同一业务序号") + 1
            choice_col = headers.index("人工确认项目") + 1
            assert [sheet.cell(row, business_col).value for row in (2, 3, 4)] == [
                "业务1",
                "业务1",
                "业务2",
            ]
            assert sheet.cell(3, choice_col).data_type == "f"
            assert sheet.cell(2, choice_col).coordinate in sheet.cell(3, choice_col).value
            assert sheet.cell(4, choice_col).data_type != "f"
        finally:
            workbook.close()


def test_fallback_manual_change_moves_amount_without_changing_cash_total() -> None:
    model = replace(
        workbook_model(0, 0),
        low_amount_fallback_batches=(fallback_batch(),),
    )
    assert calculate_manual_adjustments(model, {}, {}, {}) == {}
    assert calculate_manual_adjustments(
        model,
        {},
        {},
        {"FALLBACK-1": USE_SYSTEM_FALLBACK},
    ) == {}
    assert calculate_manual_adjustments(
        model,
        {},
        {},
        {"FALLBACK-1": "CFO-03"},
    ) == {"CFO-01": -13_198_433, "CFO-03": 13_198_433}
