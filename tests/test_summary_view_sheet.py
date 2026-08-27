# -*- coding: utf-8 -*-
"""分类汇总视图（第13张表）的行为测试：组合汇总、校验行、纯展示无输入区。"""

from __future__ import annotations

import tempfile
from dataclasses import replace
from pathlib import Path

from openpyxl import load_workbook

from cashflow_direct.workbook_output import build_output_workbook
from tests.fixture_factory import workbook_model

SALES = "销售商品、提供劳务收到的现金"
PURCHASE = "购买商品、接受劳务支付的现金"


def _view_trace_rows() -> tuple[dict[str, object], ...]:
    def row(item: str, account: str, amount: float) -> dict[str, object]:
        return {
            "业务组成编号(技术)": f"RC-{item}-{account}-{amount}",
            "日期": "2026-01-02",
            "凭证字": "记",
            "凭证号": "1",
            "本行摘要": "匿名业务",
            "本行完整对方科目路径": account,
            "标准一级科目": account,
            "现金账户路径": "银行存款_一般户",
            "借方": max(amount, 0.0),
            "贷方": max(-amount, 0.0),
            "本行分配现金变化": amount,
            "现金方向依据": "借贷差额",
            "原项目标准化结果": item,
            "系统候选项目": item,
            "最终决定项目": item,
            "判断理由": "匿名",
        }

    return (
        row(SALES, "主营业务收入", 100.0),
        row(SALES, "主营业务收入", 200.0),
        row(PURCHASE, "原材料", -50.0),
        row(PURCHASE, "原材料", 30.0),
        row("等待人工复核", "其他应付款", -70.0),
        row("明确排除", "银行存款", -5.0),
    )


def _build(model) -> tuple[object, object]:
    """返回（公式视图， 缓存值视图）。openpyxl 已在内存解析，临时文件删除不影响断言。"""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "汇总视图.xlsx"
        build_output_workbook(model, path)
        return load_workbook(path), load_workbook(path, data_only=True)


def _view_rows(sheet) -> list[tuple[int, tuple[object, ...]]]:
    return [
        (record[0].row, tuple(cell.value for cell in record[:5]))
        for record in sheet.iter_rows(min_row=4)
        if record[0].value
    ]


def test_workbook_has_summary_view_as_the_thirteenth_sheet() -> None:
    model = replace(workbook_model(0, 0), trace_rows=_view_trace_rows())
    workbook, _ = _build(model)
    try:
        assert len(workbook.sheetnames) == 14
        assert workbook.sheetnames[12] == "分类汇总视图"
        assert workbook.sheetnames[13] == "输入识别与字段映射"
    finally:
        workbook.close()


def test_summary_view_combos_match_independent_recount() -> None:
    model = replace(workbook_model(0, 0), trace_rows=_view_trace_rows())
    _, values = _build(model)
    try:
        sheet = values["分类汇总视图"]
        combos = [
            record
            for _, record in _view_rows(sheet)
            if record[0] in (SALES, PURCHASE) and record[1] in ("主营业务收入", "原材料")
        ]
        assert combos == [
            (SALES, "主营业务收入", "流入", 2, 300.0),
            (PURCHASE, "原材料", "流出", 1, -50.0),
            (PURCHASE, "原材料", "流入", 1, 30.0),
        ]
    finally:
        values.close()


def test_summary_view_check_rows_cover_every_leaf_item() -> None:
    model = replace(workbook_model(0, 0), trace_rows=_view_trace_rows())
    formulas, values = _build(model)
    try:
        check_rows = [
            (row_number, record)
            for row_number, record in _view_rows(values["分类汇总视图"])
            if record[1] == "校验：与留痕按项目合计差额"
        ]
        assert len(check_rows) == 22
        formula_sheet = formulas["分类汇总视图"]
        for row_number, record in check_rows:
            assert record[4] == 0
            assert "SUMIFS" in str(formula_sheet.cell(row_number, 5).value)
    finally:
        formulas.close()
        values.close()


def test_summary_view_is_display_only_without_input_cells() -> None:
    model = replace(workbook_model(0, 0), trace_rows=_view_trace_rows())
    workbook, _ = _build(model)
    try:
        sheet = workbook["分类汇总视图"]
        assert not sheet.data_validations.dataValidation
        for record in sheet.iter_rows():
            for cell in record:
                assert cell.protection.locked
        text = "".join(str(cell.value or "") for row in sheet.iter_rows() for cell in row)
        assert "纯展示" in text
        assert "复核义务" in text
    finally:
        workbook.close()
