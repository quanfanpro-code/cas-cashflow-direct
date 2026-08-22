from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from cashflow_direct.semantic_mapping import (
    DatasetMapping,
    MappingQuestion,
    _is_strict_date,
    _strict_date_share,
    infer_dataset_mapping,
    infer_dataset_mappings,
)
from cashflow_direct.workbook_structure import find_header_bands, scan_workbook
from tests.fixture_factory import (
    break_dimension,
    write_ambiguous_money_fixture,
    write_complex_header_fixture,
    write_hostile_header_fixture,
)


class StructureAndMappingTests(unittest.TestCase):
    def test_voucher_word_and_account_code_are_mapped_separately(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "独立编码字段.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                ["日期", "凭证字", "凭证号", "摘要", "科目编码", "科目名称", "借方", "贷方"]
            )
            sheet.append(
                ["2026-01-01", "记", 1, "匿名收款", "1002.01", "银行存款", 100, None]
            )
            workbook.save(path)

            mapping = infer_dataset_mapping(scan_workbook(path))

            self.assertIsInstance(mapping, DatasetMapping)
            self.assertEqual(2, mapping.role_to_column["voucher_word"].column_index)
            self.assertEqual(3, mapping.role_to_column["voucher_no"].column_index)
            self.assertEqual(5, mapping.role_to_column["account_code"].column_index)
            self.assertEqual(6, mapping.role_to_column["account_name"].column_index)

    def test_every_data_sheet_gets_its_own_mapping(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "按月序时账.xlsx"
            workbook = Workbook()
            for index, title in enumerate(("一月", "二月")):
                sheet = workbook.active if index == 0 else workbook.create_sheet()
                sheet.title = title
                sheet.append(["日期", "凭证号", "摘要", "科目", "借方", "贷方"])
                sheet.append([f"2026-0{index + 1}-01", "记-1", "匿名收款", "银行存款", 100, None])
            workbook.save(path)
            mappings = infer_dataset_mappings(scan_workbook(path))
            self.assertEqual(
                {"一月", "二月"},
                {item.sheet_name for item in mappings if isinstance(item, DatasetMapping)},
            )

    def test_merged_multiline_header_has_same_roles_at_different_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            first = Path(tmp) / "结构甲.xlsx"
            second = Path(tmp) / "结构乙.xlsx"
            write_complex_header_fixture(first, header_row=2, label_side="cash")
            write_complex_header_fixture(second, header_row=9, label_side="counterpart")
            snapshot_a = scan_workbook(first)
            snapshot_b = scan_workbook(second)
            mapping_a = infer_dataset_mapping(snapshot_a)
            mapping_b = infer_dataset_mapping(snapshot_b)
            self.assertIsInstance(mapping_a, DatasetMapping)
            self.assertIsInstance(mapping_b, DatasetMapping)
            required = {
                "voucher_date",
                "voucher_no",
                "summary",
                "account_name",
                "debit",
                "credit",
                "flow_item",
            }
            self.assertTrue(required.issubset(mapping_a.role_to_column))
            self.assertTrue(required.issubset(mapping_b.role_to_column))
            self.assertNotEqual(mapping_a.header_row_start, mapping_b.header_row_start)
            self.assertTrue(snapshot_a.sheets[0].merged_ranges)
            self.assertTrue(find_header_bands(snapshot_a))
            self.assertEqual("E", mapping_a.role_to_column["debit"].column_letter)
            self.assertIn("发生额", mapping_a.role_to_column["debit"].header_path)

    def test_close_semantic_candidates_return_question_instead_of_guessing(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "歧义.xlsx"
            write_ambiguous_money_fixture(path)
            result = infer_dataset_mapping(scan_workbook(path))
            self.assertIsInstance(result, MappingQuestion)
            self.assertEqual("debit", result.role)
            self.assertEqual(3, len(result.sample_values))
            self.assertTrue(result.recommended.header_path)
            self.assertTrue(result.alternatives)

    def test_strict_date_helpers_reject_non_dates_and_accept_common_formats(self) -> None:
        self.assertTrue(_is_strict_date("2026-02-26"))
        self.assertTrue(_is_strict_date("2026/02/06"))
        self.assertTrue(_is_strict_date("2026.2.6"))
        self.assertTrue(_is_strict_date("2026年2月6日"))
        self.assertTrue(_is_strict_date("2026-02-26 10:30:00"))
        self.assertFalse(_is_strict_date("-139311.41"))
        self.assertFalse(_is_strict_date(12345))
        self.assertFalse(_is_strict_date(None))
        share = _strict_date_share(("2026-02-26", "2026-03-26", "-100"))
        self.assertAlmostEqual(2 / 3, share)

    def test_date_column_recognized_by_data_format_without_header_term(self) -> None:
        # 表头“账期”不是词典日期词，但值全是日期 → 仍应识别为凭证日期
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "账期列.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["账期", "凭证编号", "摘要", "科目名称", "借方发生额", "贷方发生额"])
            sheet.append(["2026-02-26", "记-1", "匿名收款", "银行存款", 100, None])
            sheet.append(["2026-03-26", "记-2", "匿名付款", "管理费用", None, 50])
            workbook.save(path)
            mapping = infer_dataset_mapping(scan_workbook(path))
            self.assertIsInstance(mapping, DatasetMapping)
            self.assertIn("voucher_date", mapping.role_to_column)
            self.assertEqual("A", mapping.role_to_column["voucher_date"].column_letter)

    def test_dinghong_headers_map_flow_amount_and_item(self) -> None:
        # 鼎弘式明细表头：分配金额 → flow_amount；现金流量表项 → flow_item，编码列不得抢答
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "鼎弘式明细.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append([])  # 第1行空，表头在第2行
            sheet.append(
                ["现金流量表项编码", "现金流量表项", "日期", "凭证号", "摘要", "科目", "辅助核算", "借方", "贷方", "分配金额"]
            )
            sheet.append(
                ["1113", "收到的其他与经营活动有关的现金", "2025-02-12", "记-5", "匿名结息", "财务费用", None, -100, None, 100]
            )
            workbook.save(path)
            mapping = infer_dataset_mapping(scan_workbook(path))
            self.assertIsInstance(mapping, DatasetMapping)
            roles = mapping.role_to_column
            self.assertIn("flow_amount", roles)
            self.assertIn("flow_item", roles)
            self.assertEqual(10, roles["flow_amount"].column_index)
            self.assertEqual(2, roles["flow_item"].column_index)

    def test_scan_workbook_tolerates_broken_dimension(self) -> None:
        # 导出工具写坏 dimension（ref="A1"）时仍应读出完整表头行
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "坏维度.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["日期", "凭证号", "摘要", "科目", "借方", "贷方"])
            sheet.append(["2026-01-01", "记-1", "匿名收款", "银行存款", 100, None])
            workbook.save(path)
            break_dimension(path)
            snapshot = scan_workbook(path)
            self.assertGreaterEqual(len(snapshot.sheets[0].rows[0]), 6)
            self.assertEqual("日期", snapshot.sheets[0].rows[0][0])

    def test_level1_account_and_detail_account_ask_before_choosing(self) -> None:
        # 导出列表式：一级科目与科目名称并存 → account_name 歧义提问，不静默选明细科目
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "导出列表式.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                ["记账日期", "凭证号", "摘要", "一级科目", "科目编码", "科目名称", "借方", "贷方", "主表项目"]
            )
            sheet.append(
                ["2026-06-30", "00377", "匿名保费", "银行存款", "1002.01", "人民币", 40250, None, "销售商品、提供劳务收到的现金"]
            )
            workbook.save(path)
            result = infer_dataset_mapping(scan_workbook(path))
            self.assertIsInstance(result, MappingQuestion)
            self.assertEqual("account_name", result.role)

    def test_hierarchical_account_samples_prefer_the_complete_path_column(self) -> None:
        # 同时存在一级科目和明细路径时，层级化样本能够证明“科目名称”是完整路径。
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "分层科目导出.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                ["记账日期", "凭证号", "摘要", "一级科目", "科目编码", "科目名称", "借方", "贷方", "主表项目"]
            )
            sheet.append(
                ["2026-06-30", "00377", "匿名付款", "应交税费", "2221.01", "应交税费_税费明细", 40250, None, "支付的各项税费"]
            )
            sheet.append(
                ["2026-06-30", "00378", "匿名收款", "银行存款", "1002.01", "银行存款_基本账户", None, 40250, "收到其他与经营活动有关的现金"]
            )
            workbook.save(path)

            result = infer_dataset_mapping(scan_workbook(path))

            self.assertIsInstance(result, DatasetMapping)
            self.assertEqual(6, result.role_to_column["account_name"].column_index)

    def test_three_level_header_and_hostile_layout_keep_full_structure(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "恶劣结构.xlsx"
            write_hostile_header_fixture(path)
            snapshot = scan_workbook(path)
            result = infer_dataset_mapping(snapshot)
            self.assertIsInstance(result, DatasetMapping)
            self.assertEqual(5, result.header_row_start)
            self.assertEqual(7, result.header_row_end)
            self.assertEqual((8,), snapshot.sheets[0].hidden_columns)
            self.assertEqual(
                ("凭证及现金流数据", "发生额", "借方"),
                result.role_to_column["debit"].header_path,
            )


if __name__ == "__main__":
    unittest.main()
