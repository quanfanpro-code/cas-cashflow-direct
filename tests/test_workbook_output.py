from __future__ import annotations

import tempfile
import unittest
import zipfile
from dataclasses import replace
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

from cashflow_direct.workbook_output import (
    REVIEW_HEADERS,
    USE_SYSTEM_RECOMMENDATION,
    build_output_workbook,
    calculate_manual_adjustments,
    manual_adjustment_formula,
    validate_output_workbook,
)
from cashflow_direct.statement import (
    ExistingStatementResult,
    ReconciliationResult,
    compare_statement,
)
from cashflow_direct.models import ReviewBatch
from tests.fixture_factory import workbook_model


def test_review_sheet_exposes_level1_mapping_candidates() -> None:
    assert "一级科目映射候选" not in REVIEW_HEADERS
    assert "一级科目映射依据" not in REVIEW_HEADERS


EXPECTED_SHEETS = [
    "使用说明与状态",
    "现金流量表正表",
    "正表核对报告",
    "重要待复核事项",
    "低金额系统兜底明细",
    "疑似重复事项",
    "AI复核记录",
    "原表与系统决定差异",
    "现金范围与现金流量表与货币资金变动的勾稽核对",
    "全量分类留痕",
    "科目语义词典",
    "同类检查",
    "分类汇总视图",
    "输入识别与字段映射",
]


def _column_is_hidden(sheet, one_based_index: int) -> bool:
    return any(
        bool(dimension.hidden)
        and int(dimension.min or 0) <= one_based_index <= int(dimension.max or 0)
        for dimension in sheet.column_dimensions.values()
    )


class WorkbookOutputTests(unittest.TestCase):
    def test_vat_review_row_follows_the_base_without_a_second_manual_choice(self) -> None:
        base_batch = ReviewBatch(
            "REV-BASE",
            ("CMP-BASE",),
            "CFI-06",
            ("CFO-07",),
            10_000,
            "基础项目待人工决定",
            baseline_statement_amount_cent=-10_000,
            cash_delta_cent=-10_000,
            baseline_item_code="CFO-07",
        )
        vat_batch = ReviewBatch(
            "REV-VAT",
            ("CMP-VAT",),
            "CFO-06",
            ("CFI-06", "CFO-07"),
            1_300,
            "增值税随基础项目决定",
            baseline_statement_amount_cent=-1_300,
            cash_delta_cent=-1_300,
            baseline_item_code="CFO-06",
            mandatory=True,
            follows_component_id="CMP-BASE",
        )
        model = replace(
            workbook_model(0, 0),
            review_batches=(base_batch, vat_batch),
            trace_rows=(
                {"业务组成编号(技术)": "CMP-BASE", "最终决定项目": "等待人工复核"},
                {"业务组成编号(技术)": "CMP-VAT", "最终决定项目": "等待人工复核"},
            ),
        )
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "增值税随基础人工选择.xlsx"
            build_output_workbook(model, path)
            workbook = load_workbook(path, data_only=False)
            try:
                review = workbook["重要待复核事项"]
                headers = [cell.value for cell in review[1]]
                manual_column = headers.index("人工确认项目") + 1
                system_column = headers.index("系统项目(技术)") + 1
                status_column = REVIEW_HEADERS.index("人工处理状态") + 1
                base_choice = review.cell(2, manual_column)
                vat_choice = review.cell(3, manual_column)
                vat_status = review.cell(3, status_column)

                self.assertEqual("n", base_choice.data_type)
                self.assertEqual("f", vat_choice.data_type)
                self.assertIn(base_choice.coordinate, vat_choice.value)
                self.assertIn(review.cell(2, system_column).coordinate, vat_choice.value)
                self.assertIn("随基础项目待定", vat_status.value)
                self.assertIn("随基础项目完成", vat_status.value)
                validated_cells = " ".join(
                    str(validation.sqref)
                    for validation in review.data_validations.dataValidation
                )
                self.assertIn(base_choice.coordinate, validated_cells)
                self.assertNotIn(vat_choice.coordinate, validated_cells)
                self.assertTrue(vat_choice.protection.locked)
                self.assertNotIn("人工可选标准项目", headers)
                self.assertNotIn("人工处理状态", headers)
            finally:
                workbook.close()
            validation = validate_output_workbook(path, model)
            self.assertTrue(validation.valid, validation.errors)

    def test_status_sheet_displays_selected_automatic_change_threshold(self) -> None:
        model = replace(workbook_model(0, 0), automatic_change_threshold=55)
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "改判阈值.xlsx"
            build_output_workbook(model, path)
            workbook = load_workbook(path, data_only=False)
            try:
                status = workbook["使用说明与状态"]
                self.assertEqual("本次自动修改最低证据分", status["A6"].value)
                self.assertEqual("55分（客户选择；70为默认推荐）", status["B6"].value)
            finally:
                workbook.close()

    def test_status_sheet_displays_45_as_selected_threshold(self) -> None:
        model = replace(workbook_model(0, 0), automatic_change_threshold=45)
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "45分档说明.xlsx"
            build_output_workbook(model, path)
            workbook = load_workbook(path, data_only=False)
            try:
                self.assertEqual(
                    "45分（客户选择；70为默认推荐）",
                    workbook["使用说明与状态"]["B6"].value,
                )
            finally:
                workbook.close()

    def test_final_workbook_uses_date_only_and_human_decision_is_the_only_manual_gate(self) -> None:
        trace_row = {
            "业务组成编号(技术)": "RC-1",
            "日期": "2026-01-01T00:00:00",
            "本行摘要": "匿名业务",
            "本行完整对方科目路径": "应付账款_设备款",
            "标准一级科目": "应付账款",
            "现金账户路径": "银行存款_一般户",
            "原项目标准化结果": "支付其他与经营活动有关的现金",
            "系统候选项目": "购建固定资产、无形资产和其他长期资产支付的现金",
            "判断理由": "设备购置",
            "单笔金额": 100.0,
            "异常": "未发现异常",
        }
        difference_row = {
            "日期": "2026-01-01T00:00:00",
            "原项目标准化结果": "支付其他与经营活动有关的现金",
            "审定现流表项目": "购建固定资产、无形资产和其他长期资产支付的现金",
        }
        model = replace(
            workbook_model(1, 0),
            trace_rows=(trace_row,),
            difference_rows=(difference_row,),
        )
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "人工门禁与日期.xlsx"
            build_output_workbook(model, path)
            workbook = load_workbook(path, data_only=False)
            try:
                review = workbook["重要待复核事项"]
                review_headers = [cell.value for cell in review[1]]
                self.assertEqual("2026-01-01", review.cell(2, review_headers.index("日期") + 1).value)
                self.assertNotIn("已决定项目", review_headers)
                status_formula = review.cell(
                    2, REVIEW_HEADERS.index("人工处理状态") + 1
                ).value
                self.assertIn("人工确认项目", review_headers)
                self.assertNotIn("明确排除原因", review_headers)
                for optional in ("人工依据", "处理人", "处理时间"):
                    self.assertNotIn(
                        f"{get_column_letter(review_headers.index(optional) + 1)}2",
                        status_formula,
                    )
                self.assertEqual("2026-01-01", workbook["原表与系统决定差异"]["A2"].value)
                trace = workbook["全量分类留痕"]
                trace_headers = [cell.value for cell in trace[1]]
                self.assertEqual("2026-01-01", trace.cell(2, trace_headers.index("日期") + 1).value)
            finally:
                workbook.close()

    def test_review_sheet_contains_only_human_decision_facts_and_hides_optional_audit_fields(self) -> None:
        model = workbook_model(1, 0)
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "精简人工复核.xlsx"
            build_output_workbook(model, path)
            workbook = load_workbook(path, data_only=False)
            try:
                sheet = workbook["重要待复核事项"]
                headers = [cell.value for cell in sheet[1]]
                for removed in (
                    "来源文件", "来源工作表", "来源行号", "来源单元格",
                    "本行科目路径", "原始一级科目", "原始科目编码",
                    "原始完整科目路径", "中间层级", "末级明细", "映射状态",
                    "一级科目映射候选", "一级科目映射依据", "现金账户范围状态",
                    "原现流项目", "组成明细", "AI复核过程", "当前决定形成过程",
                    "已决定项目", "人工可选标准项目", "人工处理状态",
                    "行类型", "批次最不利影响金额", "批次现金变化金额",
                    "批次编号(技术)",
                ):
                    self.assertNotIn(removed, headers)
                for visible in (
                    "日期", "凭证号", "本行摘要", "本行完整对方科目路径",
                    "标准一级科目", "现金账户路径", "原项目标准化结果",
                    "系统候选项目", "判断理由", "单笔金额", "异常",
                    "同一业务序号", "人工确认项目",
                ):
                    self.assertIn(visible, headers)
                for hidden in (
                    "摘要来源质量", "完整路径来源质量", "两个来源是否独立",
                    "证据质量说明", "证据得分", "单笔重要性层级",
                    "强制检查", "唯一动作", "人工依据", "外部资料位置",
                    "处理人", "处理时间",
                ):
                    index = headers.index(hidden) + 1
                    self.assertTrue(_column_is_hidden(sheet, index))
            finally:
                workbook.close()

    def test_trace_sheet_hides_process_fields_and_has_no_artificial_human_decision_column(self) -> None:
        trace_row = {
            "日期": "2026-01-01",
            "原始完整科目路径": "应付账款_设备款",
            "本行完整对方科目路径": "应付账款_设备款",
            "中间层级": "设备款",
            "末级明细": "设备款",
            "映射状态": "已确认",
            "一级科目映射候选": "应付账款",
            "一级科目映射依据": "名称一致",
            "现金方向依据": "现金贷方",
            "原现流项目": "支付其他与经营活动有关的现金",
            "系统候选项目": "购建固定资产、无形资产和其他长期资产支付的现金",
            "判断理由": "设备购置",
            "摘要来源质量": "强（45分）",
            "完整路径来源质量": "中（25分）",
            "两个来源是否独立": "是",
            "证据质量说明": "合计70分",
            "证据得分": "70",
            "单笔金额": "100",
            "强制检查": "无",
            "异常": "未发现异常",
            "AI复核过程": "详细过程",
            "本行分配现金变化": -100,
            "组成明细": "详细组成",
            "评分版本": "S1",
            "动作表版本": "A1",
            "当前决定形成过程": "两次独立AI复核结果一致；系统重算后自动保留原项目。",
            "最终决定项目": "支付其他与经营活动有关的现金",
            "人工决定": "未经过人工决定",
        }
        model = replace(workbook_model(0, 0), trace_rows=(trace_row,))
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "精简留痕.xlsx"
            build_output_workbook(model, path)
            workbook = load_workbook(path, data_only=False)
            try:
                sheet = workbook["全量分类留痕"]
                headers = [cell.value for cell in sheet[1]]
                self.assertNotIn("人工决定", headers)
                self.assertFalse(_column_is_hidden(sheet, headers.index("原始完整科目路径") + 1))
                for hidden in (
                    "本行完整对方科目路径", "中间层级", "末级明细", "映射状态",
                    "一级科目映射候选", "一级科目映射依据", "现金方向依据",
                    "原现流项目", "系统候选项目", "判断理由", "摘要来源质量",
                    "完整路径来源质量", "两个来源是否独立", "证据质量说明",
                    "证据得分", "单笔金额", "强制检查", "异常",
                    "AI复核过程", "本行分配现金变化", "组成明细", "评分版本", "动作表版本",
                ):
                    self.assertTrue(_column_is_hidden(sheet, headers.index(hidden) + 1))
            finally:
                workbook.close()

    def test_pending_trace_result_follows_the_human_choice_in_the_same_workbook(self) -> None:
        model = replace(
            workbook_model(1, 0),
            trace_rows=(
                {
                    "业务组成编号(技术)": "RC-1",
                    "本行摘要": "待人工判断业务",
                    "最终决定项目": "等待人工复核",
                    "复核状态": "等待人工复核",
                },
            ),
        )
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "人工结果联动留痕.xlsx"
            build_output_workbook(model, path)
            workbook = load_workbook(path, data_only=False)
            try:
                trace = workbook["全量分类留痕"]
                headers = [cell.value for cell in trace[1]]
                formula = trace.cell(
                    2, headers.index("最终决定项目") + 1
                ).value
                self.assertIn("重要待复核事项", formula)
                self.assertIn("$AD2", formula)
                review = workbook["重要待复核事项"]
                review_headers = [cell.value for cell in review[1]]
                self.assertEqual(
                    "RC-1",
                    review.cell(
                        2, review_headers.index("业务组成编号(技术)") + 1
                    ).value,
                )
            finally:
                workbook.close()

    def test_trace_final_decision_is_editable_for_every_row_with_all_leaf_items(self) -> None:
        base = workbook_model(1, 0)
        review = replace(base.review_batches[0], baseline_item_code="CFO-07")
        model = replace(
            base,
            review_batches=(review,),
            trace_rows=(
                {
                    "业务组成编号(技术)": "DONE-1",
                    "本行分配现金变化": 100.0,
                    "最终决定项目": "销售商品、提供劳务收到的现金",
                },
                {
                    "业务组成编号(技术)": "RC-1",
                    "本行分配现金变化": -100.0,
                    "最终决定项目": "等待人工复核",
                },
            ),
        )
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "全量留痕逐行人工改选.xlsx"
            build_output_workbook(model, path)
            workbook = load_workbook(path, data_only=False)
            try:
                trace = workbook["全量分类留痕"]
                headers = [cell.value for cell in trace[1]]
                final_column = headers.index("最终决定项目") + 1
                final_range = (
                    f"{get_column_letter(final_column)}2:"
                    f"{get_column_letter(final_column)}3"
                )
                validation = next(
                    item
                    for item in trace.data_validations.dataValidation
                    if final_range in str(item.sqref)
                )
                source = str(validation.formula1).lstrip("=").replace("$", "")
                min_col, min_row, max_col, max_row = range_boundaries(source)
                options = (
                    [
                        trace.cell(min_row, column).value
                        for column in range(min_col, max_col + 1)
                    ]
                    if min_row == max_row
                    else [
                        trace.cell(row, min_col).value
                        for row in range(min_row, max_row + 1)
                    ]
                )
                expected_options = [
                    item.name
                    for item in sorted(
                        (item for item in model.rules.statement_items if item.is_leaf),
                        key=lambda item: item.display_order,
                    )
                ] + ["明确排除"]

                self.assertEqual(expected_options, options)
                self.assertIn(validation.errorStyle, (None, "stop"))
                self.assertEqual("项目无效", validation.errorTitle)
                self.assertFalse(trace.cell(2, final_column).protection.locked)
                self.assertFalse(trace.cell(3, final_column).protection.locked)
                self.assertEqual(
                    "销售商品、提供劳务收到的现金",
                    trace.cell(2, final_column).value,
                )
                self.assertIn("重要待复核事项", trace.cell(3, final_column).value)

                technical_headers = (
                    "人工改选基准项目(技术)",
                    "人工改选基准金额(技术)",
                    "人工改选目标金额(技术)",
                    "人工改选生效标志(技术)",
                )
                for header in technical_headers:
                    column = headers.index(header) + 1
                    self.assertTrue(_column_is_hidden(trace, column))
                base_column = headers.index("人工改选基准项目(技术)") + 1
                self.assertEqual(
                    "销售商品、提供劳务收到的现金",
                    trace.cell(2, base_column).value,
                )
                pending_base_formula = trace.cell(3, base_column).value
                self.assertIn("重要待复核事项", pending_base_formula)
                self.assertIn(model.rules.item_by_id["CFO-07"].name, pending_base_formula)
            finally:
                workbook.close()

    def test_trace_override_is_added_to_main_manual_adjustment_formula(self) -> None:
        base = workbook_model(0, 0)
        existing = ExistingStatementResult(
            values={},
            prior_values={},
            standardized_values={},
            custom_rows=(),
            unit_multiplier=1,
        )
        model = replace(
            base,
            comparison=compare_statement(existing, base.statement),
            trace_rows=(
                {
                    "业务组成编号(技术)": "DONE-1",
                    "本行分配现金变化": 100.0,
                    "最终决定项目": "销售商品、提供劳务收到的现金",
                },
            ),
        )
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "全量留痕改选联动正表.xlsx"
            build_output_workbook(model, path)
            workbook = load_workbook(path, data_only=False)
            try:
                trace = workbook["全量分类留痕"]
                trace_headers = [cell.value for cell in trace[1]]
                main = workbook["现金流量表正表"]
                main_headers = [cell.value for cell in main[3]]
                item_name = model.rules.item_by_id["CFO-01"].name
                item_row = next(
                    row
                    for row in range(4, main.max_row + 1)
                    if main.cell(row, main_headers.index("项目") + 1).value == item_name
                )
                formula = main.cell(
                    item_row, main_headers.index("人工调整") + 1
                ).value

                self.assertIn("'全量分类留痕'", formula)
                self.assertIn("-SUMIFS(", formula)
                for header in (
                    "人工改选基准项目(技术)",
                    "人工改选基准金额(技术)",
                    "人工改选目标金额(技术)",
                    "人工改选生效标志(技术)",
                    "最终决定项目",
                ):
                    column = get_column_letter(trace_headers.index(header) + 1)
                    self.assertIn(f"${column}$2:${column}$2", formula)

                comparison = workbook["正表核对报告"]
                self.assertEqual("='现金流量表正表'!E4", comparison["E2"].value)
                self.assertEqual("='现金流量表正表'!F4", comparison["F2"].value)
            finally:
                workbook.close()

    def test_reliable_group_sheet_and_completion_gate_are_absent(self) -> None:
        model = workbook_model(0, 0)
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "最终输出.xlsx"
            build_output_workbook(model, path)
            workbook = load_workbook(path, data_only=False)
            try:
                self.assertEqual(EXPECTED_SHEETS, workbook.sheetnames)
                status_formula = workbook["使用说明与状态"]["B3"].value
                self.assertNotIn("可靠同类组", str(status_formula))
            finally:
                workbook.close()

    def test_python_can_read_cached_values_when_customer_statement_has_blank_rows(self) -> None:
        model = workbook_model(0, 0)
        existing = ExistingStatementResult(
            values={},
            prior_values={},
            standardized_values={},
            custom_rows=(),
            unit_multiplier=1,
        )
        model = replace(
            model,
            comparison=compare_statement(existing, model.statement),
        )
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "空白客户行缓存可读.xlsx"
            build_output_workbook(model, path)
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                values = [
                    cell.value
                    for sheet in workbook.worksheets
                    for row in sheet.iter_rows()
                    for cell in row
                ]
                self.assertNotIn("None", values)
            finally:
                workbook.close()

    def test_review_sheet_contains_decision_facts_and_manual_record_fields(self) -> None:
        trace_row = {
            "来源文件": "匿名输入.xlsx",
            "来源工作表": "明细",
            "来源行号": "2",
            "来源单元格": "A2:R2",
            "日期": "2026-01-01",
            "凭证字": "记",
            "凭证号": "1",
            "本行摘要": "支付匿名工程款",
            "本行科目路径": "应付账款_工程款",
            "原始一级科目": "应付账款",
            "原始科目编码": "2202.01",
            "原始完整科目路径": "应付账款_工程款_匿名供应商",
            "本行完整对方科目路径": "在建工程_匿名项目",
            "标准一级科目": "应付账款",
            "中间层级": "工程款",
            "末级明细": "匿名供应商",
            "映射状态": "已确认",
            "现金账户路径": "银行存款_一般户",
            "现金账户范围状态": "现金及现金等价物范围内",
            "借方": 0.0,
            "贷方": 100.0,
            "流量金额（原币）": 100.0,
            "本行分配现金变化": -100.0,
            "现金方向依据": "借贷差额",
            "原现流项目": "购建固定资产、无形资产和其他长期资产支付的现金",
            "原项目标准化结果": "购建固定资产、无形资产和其他长期资产支付的现金",
            "系统候选项目": "支付其他与经营活动有关的现金",
            "判断理由": "工程事实与费用候选存在冲突，需要人工决定",
            "证据质量说明": "摘要中、路径中，两个来源独立但互相冲突，不形成总分",
            "证据得分": "来源冲突，无可用总分",
            "单笔金额": 100.0,
            "单笔重要性层级": "M2（实际执行至整体重要性）",
            "强制检查": "两个来源冲突",
            "异常": "两个来源冲突",
            "AI复核过程": "第一次AI复核后仍冲突",
            "当前决定形成过程": "系统识别冲突；AI复核后仍需人工决定",
            "业务组成编号(技术)": "RC-1",
        }
        model = replace(workbook_model(1, 0), trace_rows=(trace_row,))
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "人工复核事实完整.xlsx"
            build_output_workbook(model, path)
            workbook = load_workbook(path, data_only=False)
            try:
                sheet = workbook["重要待复核事项"]
                headers = [cell.value for cell in sheet[1]]
                required = (
                    "日期", "凭证字", "凭证号", "本行摘要",
                    "本行完整对方科目路径", "标准一级科目", "现金账户路径",
                    "借方", "贷方", "流量金额（原币）", "本行分配现金变化",
                    "原项目标准化结果", "系统候选项目", "判断理由",
                    "证据质量说明", "证据得分", "单笔金额",
                    "单笔重要性层级", "强制检查", "异常",
                    "同一业务序号", "人工确认项目",
                    "人工依据", "外部资料位置", "处理人", "处理时间",
                )
                for header in required:
                    self.assertIn(header, headers)
                self.assertEqual(
                    "支付匿名工程款",
                    sheet.cell(2, headers.index("本行摘要") + 1).value,
                )
                self.assertNotIn("原始科目编码", headers)
                self.assertNotIn("AI复核过程", headers)
                self.assertNotIn("已决定项目", headers)
                self.assertEqual(
                    100.0,
                    sheet.cell(2, headers.index("贷方") + 1).value,
                )
                self.assertNotIn("人工可选标准项目", headers)
                self.assertNotIn("人工处理状态", headers)
                money_format = sheet.cell(
                    2, headers.index("单笔金额") + 1
                ).number_format
                for header in ("借方", "贷方", "本行分配现金变化"):
                    self.assertEqual(
                        money_format,
                        sheet.cell(2, headers.index(header) + 1).number_format,
                    )
                helper_values = {
                    cell.value
                    for row in sheet.iter_rows()
                    for cell in row
                    if isinstance(cell.value, str)
                }
                self.assertNotIn("明确排除", helper_values)
            finally:
                workbook.close()

    def test_every_manual_dropdown_starts_with_the_system_recommendation(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "逐行人工下拉.xlsx"
            build_output_workbook(workbook_model(2, 0), path)
            workbook = load_workbook(path, data_only=False)
            try:
                sheet = workbook["重要待复核事项"]
                headers = [cell.value for cell in sheet[1]]
                manual_column = headers.index("人工确认项目") + 1
                self.assertNotIn("人工可选标准项目", headers)
                validations = tuple(sheet.data_validations.dataValidation)
                for row in (2, 3):
                    coordinate = sheet.cell(row, manual_column).coordinate
                    validation = next(
                        item
                        for item in validations
                        if any(coordinate in cell_range for cell_range in item.sqref.ranges)
                    )
                    range_ref = validation.formula1.split("!", 1)[1].replace("$", "")
                    start, end = range_ref.split(":", 1)
                    self.assertEqual(
                        USE_SYSTEM_RECOMMENDATION,
                        sheet[start].value,
                    )
                    self.assertNotEqual("明确排除", sheet[end].value)
            finally:
                workbook.close()

    def test_generic_money_columns_use_thousands_separators(self) -> None:
        model = replace(
            workbook_model(0, 0),
            reconciliation=ReconciliationResult(
                "现金流量表与货币资金变动的勾稽核对：相符",
                123_456_789,
                133_456_789,
                0,
                10_000_000,
                0,
            ),
            trace_rows=({"摘要": "匿名业务", "现金变化": 1_234_567.89},),
        )
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "金额格式.xlsx"
            build_output_workbook(model, path)
            workbook = load_workbook(path, data_only=False)
            try:
                cash = workbook["现金范围与现金流量表与货币资金变动的勾稽核对"]
                amount_column = [cell.value for cell in cash[1]].index("金额（元）") + 1
                opening_row = next(
                    row
                    for row in range(2, cash.max_row + 1)
                    if cash.cell(row, 1).value == "期初现金及现金等价物余额"
                )
                self.assertIn(
                    "#,##0.00", cash.cell(opening_row, amount_column).number_format
                )
                self.assertIn(
                    "#,##0.00", workbook["全量分类留痕"]["B2"].number_format
                )
            finally:
                workbook.close()

    def test_review_sheet_keeps_each_real_allocation_without_synthetic_total_row(self) -> None:
        trace_rows = (
            {
                "业务组成编号(技术)": "RC-1",
                "本行分配现金变化": -60.0,
                "组成明细": "银行存款_甲户：-60.00元",
            },
            {
                "业务组成编号(技术)": "RC-1",
                "本行分配现金变化": -40.0,
                "组成明细": "银行存款_乙户：-40.00元",
            },
        )
        model = replace(workbook_model(1, 0), trace_rows=trace_rows)
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "多现金账户组成.xlsx"
            build_output_workbook(model, path)
            workbook = load_workbook(path, data_only=False)
            try:
                sheet = workbook["重要待复核事项"]
                headers = [cell.value for cell in sheet[1]]
                allocated_column = headers.index("本行分配现金变化") + 1
                business_column = headers.index("同一业务序号") + 1
                real_rows = [
                    row
                    for row in range(2, sheet.max_row + 1)
                    if sheet.cell(row, business_column).value
                ]
                self.assertEqual(2, len(real_rows))
                self.assertEqual(
                    [-60.0, -40.0],
                    [sheet.cell(row, allocated_column).value for row in real_rows],
                )
                self.assertNotIn("组成明细", headers)
            finally:
                workbook.close()

    def test_original_auto_difference_sheet_keeps_rows_visible_and_read_only(self) -> None:
        difference_row = {
            "日期": "2026-01-01",
            "凭证字": "记",
            "凭证号": "1",
            "摘要": "匿名税费",
            "科目编码": "1002.01",
            "科目名称": "银行存款",
            "借方": None,
            "贷方": 100.0,
            "流量金额（原币）": 100.0,
            "对方科目": "营业外支出",
            "原项目标准化结果": "支付的各项税费",
            "审定现流表项目": "支付其他与经营活动有关的现金",
            "差异形成原因": "标准项目不一致",
            "打分逻辑描述及打分结果": "90分",
            "独立来源1": "缴纳税费；强（45分）",
            "独立来源2": "营业外支出；强（45分）",
            "来源文件": "匿名输入.xlsx",
            "来源工作表": "明细",
            "来源单元格": "A2:L2",
        }
        model = replace(workbook_model(0, 0), difference_rows=(difference_row,))
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "原表差异.xlsx"
            build_output_workbook(model, path)
            workbook = load_workbook(path, data_only=False)
            try:
                sheet = workbook["原表与系统决定差异"]
                self.assertEqual(
                    list(difference_row), [cell.value for cell in sheet[1]]
                )
                self.assertEqual("A2", sheet.freeze_panes)
                self.assertIsNotNone(sheet.auto_filter.ref)
                self.assertEqual("营业外支出", sheet["J2"].value)
                self.assertEqual(
                    "支付其他与经营活动有关的现金", sheet["L2"].value
                )
                self.assertIn("#,##0.00", sheet["H2"].number_format)
                self.assertIsNone(sheet["G2"].value)
                self.assertTrue(sheet["A2"].protection.locked)
                self.assertTrue(sheet.protection.sheet)
            finally:
                workbook.close()

    def test_empty_difference_sheet_remains_visible_with_explanation(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "无差异.xlsx"
            build_output_workbook(workbook_model(0, 0), path)
            workbook = load_workbook(path, data_only=False)
            try:
                sheet = workbook["原表与系统决定差异"]
                self.assertEqual("visible", sheet.sheet_state)
                self.assertIn("无差异", str(sheet["A2"].value))
            finally:
                workbook.close()

    def test_more_than_100000_difference_rows_are_rejected(self) -> None:
        row = {"原项目标准化结果": "支付的各项税费", "审定现流表项目": "不进入正表"}
        model = replace(workbook_model(0, 0), difference_rows=(row,) * 100_001)
        with tempfile.TemporaryDirectory() as tmp:
            with self.assertRaisesRegex(ValueError, "差异明细超过"):
                build_output_workbook(model, Path(tmp) / "超限.xlsx")

    def test_generated_workbook_uses_consistent_professional_base_format(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "统一格式.xlsx"
            build_output_workbook(workbook_model(1, 1), path)
            workbook = load_workbook(path, data_only=False)
            try:
                for sheet in workbook.worksheets:
                    with self.subTest(sheet=sheet.title):
                        self.assertFalse(sheet.sheet_view.showGridLines)
                        self.assertEqual(18, sheet.sheet_format.defaultRowHeight)
                        for row in sheet.iter_rows():
                            for cell in row:
                                if cell.value not in (None, ""):
                                    self.assertEqual("Times New Roman", cell.font.name)
                with zipfile.ZipFile(path) as package:
                    theme = package.read("xl/theme/theme1.xml").decode("utf-8")
                self.assertIn('script="Hans" typeface="宋体"', theme)
                self.assertNotIn("MS Gothic", theme)
            finally:
                workbook.close()

    def test_workbook_has_expected_visible_sheets_and_no_external_links(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "现金流量表正表及复核底稿.xlsx"
            model = workbook_model(review_batches=1, duplicate_groups=1)
            build_output_workbook(model, path)
            workbook = load_workbook(path, data_only=False, keep_links=True)
            try:
                self.assertEqual(EXPECTED_SHEETS, workbook.sheetnames)
                hidden = {
                    sheet.title
                    for sheet in workbook.worksheets
                    if sheet.sheet_state == "hidden"
                }
                self.assertEqual({"AI复核记录", "输入识别与字段映射"}, hidden)
                self.assertEqual([], workbook._external_links)
            finally:
                workbook.close()
            validation = validate_output_workbook(path, model)
            self.assertTrue(validation.valid, validation.errors)

    def test_manual_formula_references_all_authorized_adjustment_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "底稿.xlsx"
            base = workbook_model(review_batches=2, duplicate_groups=2)
            model = replace(
                base,
                trace_rows=(*base.trace_rows,
                    {
                        "业务组成编号(技术)": "DONE-1",
                        "本行分配现金变化": 100.0,
                        "最终决定项目": "销售商品、提供劳务收到的现金",
                    },
                ),
            )
            build_output_workbook(model, path)
            workbook = load_workbook(path, data_only=False)
            try:
                formulas = [
                    cell.value
                    for row in workbook["现金流量表正表"].iter_rows()
                    for cell in row
                    if isinstance(cell.value, str) and cell.value.startswith("=")
                ]
                self.assertTrue(any("重要待复核事项" in formula for formula in formulas))
                self.assertTrue(any("疑似重复事项" in formula for formula in formulas))
                self.assertTrue(any("全量分类留痕" in formula for formula in formulas))
                self.assertTrue(all("原表与自动判定差异" not in formula for formula in formulas))
                self.assertLess(len(formulas), 300)
                self.assertTrue(workbook["现金流量表正表"].freeze_panes)
                self.assertTrue(workbook["重要待复核事项"].data_validations.dataValidation)
            finally:
                workbook.close()

    def test_review_reclassification_and_duplicate_exclusion_adjust_once(self) -> None:
        base = workbook_model(review_batches=1, duplicate_groups=1)
        model = replace(
            base,
            review_batches=(
                replace(base.review_batches[0], baseline_item_code="CFO-07"),
            ),
        )
        adjustments = calculate_manual_adjustments(
            model,
            review_decisions={"REV-1": "CFI-09"},
            duplicate_decisions={"DUP-1": "exclude"},
        )
        self.assertEqual(-10_000, adjustments["CFO-07"])
        self.assertEqual(10_000, adjustments["CFI-09"])
        self.assertEqual(-20_000, adjustments["CFO-03"])
        formula = manual_adjustment_formula("支付其他与经营活动有关的现金", 2, 2)
        self.assertIn("重要待复核事项", formula)
        self.assertIn("疑似重复事项", formula)
        self.assertIn("支付其他与经营活动有关的现金", formula)
        self.assertIn(USE_SYSTEM_RECOMMENDATION, formula)
        self.assertNotIn("CFO-07", formula)

        recommended = calculate_manual_adjustments(
            model,
            review_decisions={"REV-1": USE_SYSTEM_RECOMMENDATION},
            duplicate_decisions={},
        )
        self.assertNotIn("CFO-07", recommended)

    def test_pending_duplicate_group_links_to_the_eventual_manual_item(self) -> None:
        base = workbook_model(review_batches=1, duplicate_groups=1)
        pending_group = replace(
            base.duplicate_groups[0],
            item_id="",
            blocks_manual_completion=True,
        )
        pending_review = replace(
            base.review_batches[0],
            component_ids=pending_group.component_ids,
            proposed_item_code="",
            alternative_item_codes=("CFO-03", "CFI-09"),
        )
        model = replace(
            base,
            review_batches=(pending_review,),
            duplicate_groups=(pending_group,),
            trace_rows=(*base.trace_rows,
                {
                    "业务组成编号(技术)": pending_group.component_ids[0],
                    "本行摘要": "待决定重复事项",
                    "本行分配现金变化": 200.0,
                },
                {
                    "业务组成编号(技术)": pending_group.component_ids[1],
                    "本行摘要": "待决定重复事项",
                    "本行分配现金变化": 200.0,
                },
            ),
        )

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "待决定项目的疑似重复事项.xlsx"
            build_output_workbook(model, path)
            workbook = load_workbook(path, data_only=False)
            try:
                duplicate_item = workbook["疑似重复事项"]["B2"].value
                self.assertIsInstance(duplicate_item, str)
                self.assertTrue(duplicate_item.startswith("="))
                self.assertIn("重要待复核事项", duplicate_item)
            finally:
                workbook.close()

    def test_invalid_pasted_review_text_is_neutral_and_status_only_reads_decision_cell(self) -> None:
        model = workbook_model(review_batches=1, duplicate_groups=0)
        self.assertEqual(
            {},
            calculate_manual_adjustments(model, {"REV-1": "随意粘贴的无效项目"}, {}),
        )
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "无效选择防护.xlsx"
            build_output_workbook(model, path)
            workbook = load_workbook(path, data_only=False)
            try:
                review_formulas = [
                    cell.value
                    for row in workbook["重要待复核事项"].iter_rows()
                    for cell in row
                    if isinstance(cell.value, str) and cell.value.startswith("=")
                ]
                self.assertFalse(any("无效选择" in formula for formula in review_formulas))
                headers = [cell.value for cell in workbook["重要待复核事项"][1]]
                status_formula = workbook["重要待复核事项"].cell(
                    2, REVIEW_HEADERS.index("人工处理状态") + 1
                ).value
                self.assertIn(get_column_letter(headers.index("人工确认项目") + 1), status_formula)
            finally:
                workbook.close()

    def test_reverse_direction_review_uses_signed_statement_amount(self) -> None:
        model = replace(
            workbook_model(0, 0),
            review_batches=(
                ReviewBatch(
                    "REV-REFUND",
                    ("C-REFUND",),
                    "CFO-04",
                    ("CFO-03",),
                    10_000,
                    "退款分类仍不确定",
                    baseline_statement_amount_cent=-10_000,
                    cash_delta_cent=10_000,
                ),
            ),
        )
        adjustments = calculate_manual_adjustments(
            model, {"REV-REFUND": "CFO-03"}, {}
        )
        self.assertNotIn("CFO-04", adjustments)
        self.assertEqual(10_000, adjustments["CFO-03"])

    def test_mandatory_batch_uses_hidden_item_list_and_range_dropdown(self) -> None:
        # 强制人工复核批次可改选任一标准项目，下拉使用隐藏区域。
        from cashflow_direct.classification import load_rule_pack
        rules = load_rule_pack(Path(__file__).resolve().parents[1])
        leaf_ids = tuple(item.item_id for item in rules.statement_items if item.is_leaf)
        batch = ReviewBatch(
            "REV-BIG",
            ("C-BIG",),
            "CFO-06",
            tuple(item_id for item_id in leaf_ids if item_id != "CFO-06"),
            100_000,
            "达到财务报表整体重要性，强制人工复核",
            baseline_statement_amount_cent=100_000,
            cash_delta_cent=-100_000,
            mandatory=True,
            baseline_item_code="CFO-06",
        )
        model = replace(
            workbook_model(0, 0),
            review_batches=(batch,),
            trace_rows=(
                {
                    "业务组成编号(技术)": "C-BIG",
                    "贷方": 1_000.0,
                    "本行分配现金变化": -1_000.0,
                    "单笔金额": "1000.0",
                },
            ),
        )
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "强制复核.xlsx"
            build_output_workbook(model, path)
            workbook = load_workbook(path, data_only=False)
            try:
                review = workbook["重要待复核事项"]
                headers = [cell.value for cell in review[1]]
                manual_column = headers.index("人工确认项目") + 1
                amount_column = headers.index("单笔金额") + 1
                credit_column = headers.index("贷方") + 1
                allocated_column = headers.index("本行分配现金变化") + 1
                status_column = REVIEW_HEADERS.index("人工处理状态") + 1
                adjustment_column = headers.index("系统项目调整(技术)") + 1
                helper_column = len(REVIEW_HEADERS) + 2
                helper_letter = get_column_letter(helper_column)
                manual_letter = get_column_letter(manual_column)
                outflow_names = tuple(
                    item.name
                    for item in rules.statement_items
                    if item.is_leaf and item.normal_direction == "outflow"
                )
                expected_options = (
                    USE_SYSTEM_RECOMMENDATION,
                    "支付的各项税费",
                    *(name for name in outflow_names if name != "支付的各项税费"),
                )
                written = tuple(
                    review.cell(row=row, column=helper_column).value
                    for row in range(1, len(expected_options) + 1)
                )
                self.assertEqual(
                    expected_options,
                    written,
                )
                self.assertNotIn("人工可选标准项目", headers)
                self.assertNotIn("人工处理状态", headers)
                self.assertEqual(1_000.0, review.cell(2, amount_column).value)
                self.assertEqual("n", review.cell(2, amount_column).data_type)
                self.assertEqual(
                    review.cell(2, credit_column).number_format,
                    review.cell(2, amount_column).number_format,
                )
                self.assertEqual(
                    review.cell(2, allocated_column).number_format,
                    review.cell(2, amount_column).number_format,
                )
                self.assertTrue(_column_is_hidden(review, helper_column))
                validation = review.data_validations.dataValidation[0].formula1
                status_formula = review.cell(2, status_column).value
                self.assertNotIn("COUNTIF", status_formula)
                # 区域末行必须与叶子项目数精确一致（行号算错也能过模糊断言）
                self.assertIn(
                    f"${helper_letter}$1:${helper_letter}${len(expected_options)}",
                    validation,
                )
                adjustment_formula = review.cell(2, adjustment_column).value
                self.assertIn(USE_SYSTEM_RECOMMENDATION, adjustment_formula)
                self.assertIn(get_column_letter(headers.index("原基线项目(技术)") + 1), adjustment_formula)
                self.assertIn(f"{manual_letter}2", status_formula)
                # 设计第五节：强制批次行必须有区域下拉——自检正路通过、被篡改为内联列表则报错
                ok = validate_output_workbook(path, model)
                self.assertTrue(ok.valid, ok.errors)
                review.data_validations.dataValidation[0].formula1 = '"支付的各项税费"'
                workbook.save(path)
            finally:
                workbook.close()
            check = validate_output_workbook(path, model)
            self.assertFalse(check.valid)
            self.assertTrue(any("区域" in error for error in check.errors))

    def test_mandatory_dropdown_validation_uses_expanded_detail_row_positions(self) -> None:
        first = ReviewBatch(
            "REV-FIRST",
            ("C-FIRST",),
            "CFO-07",
            ("CFO-04",),
            30_000,
            "第一项业务含三行真实明细",
            baseline_statement_amount_cent=30_000,
            cash_delta_cent=-30_000,
        )
        mandatory = ReviewBatch(
            "REV-MANDATORY-AFTER-DETAILS",
            ("C-MANDATORY",),
            "CFO-06",
            ("CFO-07",),
            100_000,
            "达到财务报表整体重要性，强制人工复核",
            baseline_statement_amount_cent=100_000,
            cash_delta_cent=-100_000,
            mandatory=True,
            baseline_item_code="CFO-06",
        )
        trace_rows = (
            {
                "业务组成编号(技术)": "C-FIRST",
                "贷方": 100.0,
                "本行分配现金变化": -100.0,
                "单笔金额": 100.0,
            },
            {
                "业务组成编号(技术)": "C-FIRST",
                "贷方": 200.0,
                "本行分配现金变化": -200.0,
                "单笔金额": 200.0,
            },
            {
                "业务组成编号(技术)": "C-FIRST",
                "贷方": 300.0,
                "本行分配现金变化": -300.0,
                "单笔金额": 300.0,
            },
            {
                "业务组成编号(技术)": "C-MANDATORY",
                "贷方": 1_000.0,
                "本行分配现金变化": -1_000.0,
                "单笔金额": 1_000.0,
            },
        )
        model = replace(
            workbook_model(0, 0),
            review_batches=(first, mandatory),
            trace_rows=trace_rows,
        )

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "展开明细后的强制复核.xlsx"
            build_output_workbook(model, path)
            check = validate_output_workbook(path, model)

        self.assertTrue(check.valid, check.errors)

    def test_manual_batch_without_system_choice_keeps_control_option_but_blocks_completion(self) -> None:
        base = workbook_model(1, 0)
        pending = replace(
            base.review_batches[0],
            proposed_item_code="",
            alternative_item_codes=("CFO-03", "CFI-09"),
        )
        model = replace(base, review_batches=(pending,))
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "无系统候选下拉.xlsx"
            build_output_workbook(model, path)
            workbook = load_workbook(path, data_only=False)
            try:
                review = workbook["重要待复核事项"]
                headers = [cell.value for cell in review[1]]
                status_column = REVIEW_HEADERS.index("人工处理状态") + 1
                validation = review.data_validations.dataValidation[0]
                helper_range = str(validation.formula1).split("!")[-1].replace("'", "")
                first_cell = helper_range.split(":", 1)[0].replace("$", "")
                self.assertEqual(USE_SYSTEM_RECOMMENDATION, review[first_cell].value)
                self.assertNotIn("人工可选标准项目", headers)
                self.assertNotIn("人工处理状态", headers)
                self.assertIn("系统没有首选项目，请改选", review.cell(2, status_column).value)
            finally:
                workbook.close()

    def test_zero_review_batches_show_clear_note_and_statement_still_builds(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "无重大事项.xlsx"
            build_output_workbook(workbook_model(0, 0), path)
            workbook = load_workbook(path, data_only=False)
            try:
                self.assertIn("无重大", workbook["重要待复核事项"]["A2"].value)
                self.assertEqual(35, workbook["现金流量表正表"].max_row - 3)
            finally:
                workbook.close()

    def test_manual_cells_are_narrowly_validated_and_formulas_are_protected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "受保护底稿.xlsx"
            model = workbook_model(1, 1)
            build_output_workbook(model, path)
            workbook = load_workbook(path, data_only=False)
            try:
                status_formula = workbook["使用说明与状态"]["B3"].value
                self.assertIn("待确认", status_formula)
                self.assertIn("无效选择", status_formula)
                review = workbook["重要待复核事项"]
                headers = [cell.value for cell in review[1]]
                manual_column = headers.index("人工确认项目") + 1
                manual_cell = review.cell(2, manual_column)
                helper_column = len(REVIEW_HEADERS) + 2
                helper_letter = get_column_letter(helper_column)
                self.assertIsNone(manual_cell.value)
                self.assertFalse(manual_cell.protection.locked)
                validation = review.data_validations.dataValidation[0].formula1
                self.assertIn(f"${helper_letter}$", validation)
                self.assertEqual(
                    USE_SYSTEM_RECOMMENDATION,
                    review.cell(1, helper_column).value,
                )
                self.assertEqual(
                    "支付其他与经营活动有关的现金",
                    review.cell(2, helper_column).value,
                )
                self.assertEqual(
                    "支付其他与投资活动有关的现金",
                    review.cell(3, helper_column).value,
                )
                self.assertIsNone(review.cell(4, helper_column).value)
                self.assertNotIn("CFI-09", validation)
                self.assertNotIn("CFO-01", validation)
                self.assertTrue(review.protection.sheet)
                self.assertTrue(workbook["现金流量表正表"].protection.sheet)
            finally:
                workbook.close()

    def test_comparison_and_reconciliation_follow_final_statement(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "动态核对底稿.xlsx"
            model = workbook_model(1, 0)
            existing = ExistingStatementResult(
                values=dict(model.statement.values),
                prior_values=dict(model.statement.prior_values),
                standardized_values=dict(model.statement.values),
                custom_rows=(),
                unit_multiplier=1,
            )
            model = replace(
                model,
                comparison=compare_statement(existing, model.statement),
                reconciliation=ReconciliationResult(
                    "现金流量表与货币资金变动的勾稽核对：相符", 100_000, 160_000, 0, 60_000, 0
                ),
            )
            build_output_workbook(model, path)
            workbook = load_workbook(path, data_only=False)
            try:
                self.assertIn("现金流量表正表", workbook["正表核对报告"]["E2"].value)
                headers = [cell.value for cell in workbook["现金范围与现金流量表与货币资金变动的勾稽核对"][1]]
                self.assertIn("金额（元）", headers)
                cash_sheet = workbook["现金范围与现金流量表与货币资金变动的勾稽核对"]
                net_row = next(
                    row for row in range(2, cash_sheet.max_row + 1)
                    if cash_sheet.cell(row, 1).value == "已分类现金流量表净额"
                )
                difference_row = next(
                    row for row in range(2, cash_sheet.max_row + 1)
                    if cash_sheet.cell(row, 1).value == "现金变动桥接差异"
                )
                self.assertIn("现金流量表正表", cash_sheet.cell(net_row, 3).value)
                self.assertTrue(str(cash_sheet.cell(difference_row, 3).value).startswith("="))
                self.assertIn("ROUND", cash_sheet.cell(difference_row, 3).value)
                self.assertIn("ROUND", workbook["正表核对报告"]["G2"].value)
                self.assertIn("现金范围与现金流量表与货币资金变动的勾稽核对", workbook["使用说明与状态"]["B3"].value)
                self.assertNotIn("101", str(workbook["全量分类留痕"].print_area))
            finally:
                workbook.close()


    def test_human_sheets_show_names_and_hide_machine_columns(self) -> None:
        base = workbook_model(1, 1)
        model = replace(
            base,
            review_batches=(replace(base.review_batches[0], component_ids=("C-1",)),),
            ai_records=(
                {
                    "阶段": "首次复核",
                    "task_id": "TASK-1",
                    "component_id": "COMP-1",
                    "item_id": "CFO-03",
                    "reason": "AI 与自动判断一致",
                    "confidence": "high",
                },
            ),
            trace_rows=(
                {
                    "记录类型": "现金流业务组成",
                    "摘要": "匿名业务",
                    "现金变化": 100.0,
                    "原现流项目": "支付其他与经营活动有关的现金",
                    "对方科目": "普通往来科目",
                    "自动判定现流项目": "支付其他与经营活动有关的现金",
                    "判断理由": "命中规则",
                    "证据强度": "高",
                    "证据得分": 55,
                    "异常": "",
                    "决策来源": "系统规则",
                    "方向依据": "借贷差额",
                    "来源文件": "匿名输入.xlsx",
                    "来源工作表": "匿名数据",
                    "来源单元格": "A1:H1",
                    "一致性复核状态": "重大一致性复核已取得一致决定",
                    "一致性复核理由": "整组业务实质一致",
                    "一致性重要性层级": "达到整体重要性",
                    "决策来源(技术)": "system",
                    "命中规则(技术)": "CFO-07-FALLBACK",
                    "业务组成编号(技术)": "C-1",
                    "来源占用键(技术)": "E-1",
                    "业务组编号(技术)": "CGR-1",
                },
            ),
        )
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "留痕分层.xlsx"
            build_output_workbook(model, path)
            workbook = load_workbook(path, data_only=False)
            try:
                headers = [cell.value for cell in workbook["全量分类留痕"][1]]
                self.assertEqual(
                    [
                        "记录类型", "摘要", "现金变化", "原现流项目", "对方科目", "自动判定现流项目",
                        "判断理由", "证据强度", "证据得分", "异常", "决策来源", "方向依据",
                        "来源文件", "来源工作表", "来源单元格", "一致性复核状态", "一致性复核理由",
                        "一致性重要性层级", "决策来源(技术)", "命中规则(技术)", "业务组成编号(技术)",
                        "来源占用键(技术)", "业务组编号(技术)",
                    ],
                    headers,
                )
                self.assertEqual(
                    "支付其他与经营活动有关的现金",
                    workbook["全量分类留痕"]["F2"].value,
                )
                trace = workbook["全量分类留痕"]
                for visible_header in ("决策来源", "一致性复核状态"):
                    column_index = headers.index(visible_header) + 1
                    self.assertFalse(
                        any(
                            dimension.hidden
                            and dimension.min <= column_index <= dimension.max
                            for dimension in trace.column_dimensions.values()
                        )
                    )
                for header in (
                    "决策来源(技术)",
                    "命中规则(技术)",
                    "业务组成编号(技术)",
                    "来源占用键(技术)",
                    "业务组编号(技术)",
                ):
                    column_index = headers.index(header) + 1
                    self.assertTrue(
                        any(
                            dimension.hidden
                            and dimension.min <= column_index <= dimension.max
                            for dimension in trace.column_dimensions.values()
                        )
                    )

                review = workbook["重要待复核事项"]
                review_headers = [cell.value for cell in review[1]]
                self.assertEqual(
                    "支付其他与经营活动有关的现金",
                    review.cell(2, review_headers.index("系统候选项目") + 1).value,
                )
                self.assertNotIn("来源文件", review_headers)
                for visible_header in ("本行摘要", "系统候选项目"):
                    self.assertFalse(
                        _column_is_hidden(
                            review, review_headers.index(visible_header) + 1
                        )
                    )
                self.assertNotIn("批次编号(技术)", review_headers)
                for technical_header in (
                    "系统项目(技术)",
                    "业务组成编号(技术)",
                ):
                    column_index = review_headers.index(technical_header) + 1
                    self.assertTrue(
                        any(
                            dimension.hidden
                            and dimension.min <= column_index <= dimension.max
                            for dimension in review.column_dimensions.values()
                        )
                    )
                self.assertTrue(
                    _column_is_hidden(trace, headers.index("证据得分") + 1)
                )
                status_formula = review.cell(
                    2, REVIEW_HEADERS.index("人工处理状态") + 1
                ).value
                self.assertNotIn("OR(FALSE)", status_formula)

                duplicate = workbook["疑似重复事项"]
                self.assertEqual("收到其他与经营活动有关的现金", duplicate["B2"].value)

                ai_sheet = workbook["AI复核记录"]
                ai_headers = [cell.value for cell in ai_sheet[1]]
                self.assertIn("现流项目", ai_headers)
                item_column = ai_headers.index("现流项目") + 1
                self.assertEqual(
                    "收到其他与经营活动有关的现金",
                    ai_sheet.cell(2, item_column).value,
                )
            finally:
                workbook.close()

    def test_comparison_uses_full_project_name_and_hides_support_column(self) -> None:
        model = workbook_model(0, 0)
        existing = ExistingStatementResult(
            values=dict(model.statement.values),
            prior_values=dict(model.statement.prior_values),
            standardized_values=dict(model.statement.values),
            custom_rows=(),
            unit_multiplier=1,
        )
        model = replace(model, comparison=compare_statement(existing, model.statement))
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "人类可读核对报告.xlsx"
            build_output_workbook(model, path)
            workbook = load_workbook(path, data_only=False)
            try:
                sheet = workbook["正表核对报告"]
                self.assertEqual(
                    (
                        "项目",
                        "客户金额",
                        "系统自动调整",
                        "自动基线",
                        "人工调整",
                        "最终金额",
                        "最终差异",
                        "明细重建金额",
                        "原表与明细勾稽差额",
                        "支持组成",
                    ),
                    tuple(cell.value for cell in sheet[1]),
                )
                self.assertEqual(model.rules.statement_items[0].name, sheet["A2"].value)
                self.assertTrue(sheet.column_dimensions["J"].hidden)
            finally:
                workbook.close()


    def test_new_workbook_never_contains_legacy_reconciliation_term(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "无旧词.xlsx"
            model = workbook_model(0, 0)
            build_output_workbook(model, path)
            workbook = load_workbook(path, data_only=False)
            try:
                hits = [
                    f"{sheet.title}!{cell.coordinate}"
                    for sheet in workbook.worksheets
                    for row in sheet.iter_rows()
                    for cell in row
                    if isinstance(cell.value, str) and "现金调节" in cell.value
                ]
                self.assertEqual([], hits)
            finally:
                workbook.close()


def test_trace_output_is_chinese(tmp_path):
    """A4 输出中文化：全量分类留痕可见单元格不含英文技术值，证据得分为数值（Task 14）。"""
    from dataclasses import replace

    base = workbook_model(0, 0)
    model = replace(
        base,
        trace_rows=(
            {
                "摘要": "测试",
                "证据强度": "中",
                "证据得分": 45,
                "异常": "内部划转",
                "决策来源": "系统规则",
                "命中规则(技术)": "CFO-XX",
            },
        ),
    )
    path = Path(tmp_path) / "out.xlsx"
    build_output_workbook(model, path)
    wb = load_workbook(path, data_only=False)
    sheet = wb["全量分类留痕"]
    hidden = {dim.min for dim in sheet.column_dimensions.values() if dim.hidden}
    english = ("high", "medium", "low", "internal_transfer", "system")
    for row in sheet.iter_rows():
        for cell in row:
            if cell.column in hidden or cell.value is None:
                continue
            text = str(cell.value)
            assert not any(token in text for token in english), f"可见列出现英文技术值：{text}"
    headers = [cell.value for cell in sheet[1]]
    score_column = headers.index("证据得分") + 1
    assert sheet.cell(2, score_column).value == 45
    wb.close()


if __name__ == "__main__":

    unittest.main()
