from __future__ import annotations

import posixpath
import zipfile
from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook


MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
DOC_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
HEADER_MARKERS = ("日期", "凭证", "摘要", "科目", "借方", "贷方", "金额", "现流", "现金流", "项目")


@dataclass(frozen=True, slots=True)
class SheetSnapshot:
    name: str
    rows: tuple[tuple[object, ...], ...]
    merged_ranges: tuple[str, ...]
    hidden_columns: tuple[int, ...]


@dataclass(frozen=True, slots=True)
class WorkbookSnapshot:
    path: Path
    sheets: tuple[SheetSnapshot, ...]


@dataclass(frozen=True, slots=True)
class HeaderBand:
    sheet_name: str
    row_start: int
    row_end: int
    score: int


def _xml_sheet_metadata(path: Path) -> dict[str, tuple[tuple[str, ...], tuple[int, ...]]]:
    metadata: dict[str, tuple[tuple[str, ...], tuple[int, ...]]] = {}
    with zipfile.ZipFile(path) as archive:
        workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        relationships = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        targets = {
            item.attrib["Id"]: item.attrib["Target"]
            for item in relationships.findall(f"{{{PKG_REL_NS}}}Relationship")
        }
        for sheet in workbook.findall(f".//{{{MAIN_NS}}}sheet"):
            name = sheet.attrib["name"]
            relation_id = sheet.attrib[f"{{{DOC_REL_NS}}}id"]
            target = targets[relation_id].lstrip("/")
            member = target if target.startswith("xl/") else posixpath.normpath(f"xl/{target}")
            root = ElementTree.fromstring(archive.read(member))
            merges = tuple(
                item.attrib["ref"] for item in root.findall(f".//{{{MAIN_NS}}}mergeCell")
            )
            hidden: set[int] = set()
            for item in root.findall(f".//{{{MAIN_NS}}}col"):
                if item.attrib.get("hidden") in {"1", "true"}:
                    hidden.update(range(int(item.attrib["min"]), int(item.attrib["max"]) + 1))
            metadata[name] = (merges, tuple(sorted(hidden)))
    return metadata


def scan_workbook(path: Path, sample_rows: int = 200) -> WorkbookSnapshot:
    """只读提取样例网格；合并信息直接从工作簿 XML 读取。"""
    source = Path(path)
    metadata = _xml_sheet_metadata(source)
    workbook = load_workbook(source, read_only=True, data_only=True, keep_vba=False)
    try:
        sheets: list[SheetSnapshot] = []
        for worksheet in workbook.worksheets:
            rows = tuple(
                tuple(row)
                for row in worksheet.iter_rows(min_row=1, max_row=sample_rows, values_only=True)
            )
            merges, hidden = metadata.get(worksheet.title, ((), ()))
            sheets.append(SheetSnapshot(worksheet.title, rows, merges, hidden))
        return WorkbookSnapshot(source.resolve(), tuple(sheets))
    finally:
        workbook.close()


def _semantic_cell_count(row: tuple[object, ...]) -> int:
    return sum(
        1
        for value in row
        if isinstance(value, str) and any(marker in value.replace(" ", "") for marker in HEADER_MARKERS)
    )


def _row_intersects_merge(row_number: int, merged_ranges: tuple[str, ...]) -> bool:
    from openpyxl.utils.cell import range_boundaries

    return any(
        min_row <= row_number <= max_row and (max_col > min_col or max_row > min_row)
        for min_col, min_row, max_col, max_row in map(range_boundaries, merged_ranges)
    )


def find_header_bands(snapshot: WorkbookSnapshot) -> tuple[HeaderBand, ...]:
    """按文本语义密度和合并边界寻找候选表头带。"""
    candidates: list[HeaderBand] = []
    for sheet in snapshot.sheets:
        for index, row in enumerate(sheet.rows, 1):
            score = _semantic_cell_count(row)
            if score < 3:
                continue
            start = index
            while start > 1 and _row_intersects_merge(start - 1, sheet.merged_ranges):
                start -= 1
            candidates.append(HeaderBand(sheet.name, start, index, score))
    return tuple(sorted(candidates, key=lambda item: (-item.score, item.row_start)))
