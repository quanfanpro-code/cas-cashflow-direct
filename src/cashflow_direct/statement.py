from __future__ import annotations

import re
from collections.abc import Sequence
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from cashflow_direct.classification import RulePack
from cashflow_direct.models import CashflowComponent, ClassificationDecision
from cashflow_direct.money import statement_amount_cent, yuan_to_cent
from cashflow_direct.semantic_mapping import ColumnMapping, MappingQuestion


@dataclass(frozen=True, slots=True)
class StatementResult:
    values: dict[str, int]
    prior_values: dict[str, int | None]
    support_component_ids: dict[str, tuple[str, ...]]


@dataclass(frozen=True, slots=True)
class ExistingCustomRow:
    name: str
    parent_item_id: str
    current_cent: int | None
    prior_cent: int | None
    source_cell: str


@dataclass(frozen=True, slots=True)
class ExistingStatementResult:
    values: dict[str, int | None]
    prior_values: dict[str, int | None]
    standardized_values: dict[str, int | None]
    custom_rows: tuple[ExistingCustomRow, ...]
    unit_multiplier: int
    sheet_name: str = ""


@dataclass(frozen=True, slots=True)
class ComparisonRow:
    item_id: str
    existing_cent: int | None
    computed_cent: int
    manual_adjustment_cent: int
    final_cent: int
    difference_cent: int | None
    support_component_ids: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class StatementComparison:
    rows: tuple[ComparisonRow, ...]


@dataclass(frozen=True, slots=True)
class ReconciliationResult:
    status: str
    opening_cent: int | None
    closing_cent: int | None
    fx_cent: int | None
    net_cash_cent: int | None
    difference_cent: int | None


def aggregate_statement(
    components: Sequence[CashflowComponent],
    decisions: Sequence[ClassificationDecision],
    rules: RulePack,
    *,
    opening_cent: int | None = None,
    fx_cent: int | None = None,
    prior_values: dict[str, int | None] | None = None,
) -> StatementResult:
    item_by_id = rules.item_by_id
    values = {item.item_id: 0 for item in rules.statement_items}
    support: dict[str, list[str]] = {item.item_id: [] for item in rules.statement_items}
    component_by_id = {item.component_id: item for item in components}
    for decision in decisions:
        if decision.excluded:
            continue
        component = component_by_id[decision.component_id]
        item = item_by_id[decision.system_item_id]
        if not item.is_leaf:
            raise ValueError(f"业务组成不能直接分类到汇总行：{item.item_id}")
        values[item.item_id] += statement_amount_cent(
            component.cash_delta_cent, item.normal_direction
        )
        support[item.item_id].append(component.component_id)

    if fx_cent is not None:
        values["FX"] = fx_cent
    if opening_cent is not None:
        values["CASH-OPENING"] = opening_cent
    for item in sorted(rules.statement_items, key=lambda value: value.display_order):
        if item.formula_components:
            values[item.item_id] = sum(
                values[source_id] * multiplier
                for source_id, multiplier in item.formula_components
            )
            support[item.item_id] = list(
                dict.fromkeys(
                    component_id
                    for source_id, _ in item.formula_components
                    for component_id in support[source_id]
                )
            )
    return StatementResult(
        values,
        {
            item.item_id: (prior_values or {}).get(item.item_id)
            for item in rules.statement_items
        },
        {item_id: tuple(component_ids) for item_id, component_ids in support.items()},
    )


def _normalize_item_name(value: str) -> str:
    text = re.sub(r"[\s，。；：、,.!！?？（）()《》\[\]【】]+", "", value)
    return text.replace("和", "").replace("其中", "")


def _mapping_question(
    role: str,
    text: str,
    row_number: int,
    sheet_name: str = "",
) -> MappingQuestion:
    candidate = ColumnMapping(
        role,
        1,
        "A",
        (text,),
        f"A{row_number}",
        0,
    )
    return MappingQuestion(role, candidate, (), (text, "", ""), sheet_name)


def _amount_to_cent(value: object, unit_multiplier: int) -> int | None:
    if value in (None, ""):
        return None
    return yuan_to_cent(value) * unit_multiplier


def _parse_statement_rows(
    rows: list[tuple[object, ...]],
    sheet_name: str,
    rules: RulePack,
) -> ExistingStatementResult | MappingQuestion | None:
    unit_text = "|".join(
        str(value)
        for row in rows[:20]
        for value in row
        if value not in (None, "")
    )
    unit_multiplier = 10_000 if "万元" in unit_text else 1
    header_row = None
    project_column = current_column = prior_column = None
    for row_index, row in enumerate(rows, 1):
        texts = [_normalize_item_name(str(value)) if value is not None else "" for value in row]
        project = next((index for index, text in enumerate(texts) if text == "项目"), None)
        current = next(
            (index for index, text in enumerate(texts) if "本期" in text or "本年" in text),
            None,
        )
        if project is not None and current is not None:
            header_row = row_index
            project_column = project
            current_column = current
            prior_column = next(
                (index for index, text in enumerate(texts) if "上期" in text or "上年" in text),
                None,
            )
            break
    if header_row is None or project_column is None or current_column is None:
        return None

    normalized_to_id = {
        _normalize_item_name(item.name): item.item_id for item in rules.statement_items
    }
    values: dict[str, int | None] = {}
    prior_values: dict[str, int | None] = {}
    custom_rows: list[ExistingCustomRow] = []
    last_standard_id = ""
    for row_number, row in enumerate(rows[header_row:], header_row + 1):
        if project_column >= len(row):
            continue
        raw_name = row[project_column]
        if raw_name in (None, ""):
            continue
        name = str(raw_name).strip()
        normalized = _normalize_item_name(name)
        item_id = normalized_to_id.get(normalized)
        current_value = row[current_column] if current_column < len(row) else None
        prior_value = row[prior_column] if prior_column is not None and prior_column < len(row) else None
        if item_id is not None:
            values[item_id] = _amount_to_cent(current_value, unit_multiplier)
            prior_values[item_id] = _amount_to_cent(prior_value, unit_multiplier)
            last_standard_id = item_id
            continue
        last_item = rules.item_by_id.get(last_standard_id)
        if last_standard_id and (
            name.startswith("其中")
            or (
                last_item is not None
                and last_item.is_leaf
                and not any(term in name for term in ("合计", "总额", "净额", "余额", "影响"))
            )
        ):
            custom_rows.append(
                ExistingCustomRow(
                    name,
                    last_standard_id,
                    _amount_to_cent(current_value, unit_multiplier),
                    _amount_to_cent(prior_value, unit_multiplier),
                    f"A{row_number}",
                )
            )
            continue
        return _mapping_question("statement_item", name, row_number, sheet_name)

    if len(values) != len(rules.statement_items):
        missing = next(
            item.item_id for item in rules.statement_items if item.item_id not in values
        )
        return _mapping_question(
            "statement_item",
            f"缺少标准项目 {missing}",
            header_row,
            sheet_name,
        )
    return ExistingStatementResult(
        values,
        prior_values,
        dict(values),
        tuple(custom_rows),
        unit_multiplier,
        sheet_name,
    )


def parse_existing_statement(
    path: Path,
    rules: RulePack,
) -> ExistingStatementResult | MappingQuestion:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    try:
        candidates = tuple(
            _parse_statement_rows(
                list(worksheet.iter_rows(values_only=True)),
                worksheet.title,
                rules,
            )
            for worksheet in workbook.worksheets
        )
        valid = tuple(item for item in candidates if isinstance(item, ExistingStatementResult))
        if len(valid) > 1:
            names = "、".join(item.sheet_name for item in valid)
            return _mapping_question(
                "statement_sheet",
                f"识别到多个现金流量表工作表：{names}",
                1,
            )
        if valid:
            return valid[0]
        question = next((item for item in candidates if isinstance(item, MappingQuestion)), None)
        return question or _mapping_question("statement_header", "未找到项目和本期或本年金额列", 1)
    finally:
        workbook.close()


def compare_statement(
    existing: ExistingStatementResult,
    computed: StatementResult,
) -> StatementComparison:
    rows: list[ComparisonRow] = []
    for item_id, computed_value in computed.values.items():
        existing_value = existing.standardized_values.get(item_id)
        rows.append(
            ComparisonRow(
                item_id,
                existing_value,
                computed_value,
                0,
                computed_value,
                None if existing_value is None else computed_value - existing_value,
                computed.support_component_ids[item_id],
            )
        )
    return StatementComparison(tuple(rows))


def reconcile_cash(
    statement: StatementResult,
    opening_cent: int | None,
    closing_cent: int | None,
    fx_cent: int | None,
) -> ReconciliationResult:
    if opening_cent is None or closing_cent is None or fx_cent is None:
        return ReconciliationResult(
            "现金调节未完成", opening_cent, closing_cent, fx_cent, None, None
        )
    net_cash = (
        statement.values["CFO-NET"]
        + statement.values["CFI-NET"]
        + statement.values["CFF-NET"]
        + fx_cent
    )
    difference = closing_cent - opening_cent - net_cash
    status = "现金调节完成" if difference == 0 else "现金调节存在差异"
    return ReconciliationResult(
        status, opening_cent, closing_cent, fx_cent, net_cash, difference
    )
