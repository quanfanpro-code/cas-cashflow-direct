from __future__ import annotations

import json
import re
from collections import defaultdict
from collections.abc import Mapping, Sequence
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

import xlsxwriter
from openpyxl import load_workbook
from xlsxwriter.utility import xl_col_to_name

from cashflow_direct.classification import RulePack
from cashflow_direct.decision_policy import DEFAULT_AUTOMATIC_CHANGE_SCORE
from cashflow_direct.duplicates import DuplicateGroup
from cashflow_direct.models import ReviewBatch
from cashflow_direct.money import statement_amount_cent
from cashflow_direct.statement import (
    ReconciliationResult,
    StatementComparison,
    StatementResult,
)


SHEET_NAMES = (
    "使用说明与状态",
    "现金流量表正表",
    "正表核对报告",
    "重要待复核事项",
    "低金额批量处理",
    "疑似重复事项",
    "AI复核记录",
    "原表与系统决定差异",
    "现金范围与现金流量表与货币资金变动的勾稽核对",
    "全量分类留痕",
    "科目语义词典",
    "同类检查",
    "输入识别与字段映射",
)

DIFFERENCE_HEADERS = (
    "日期",
    "凭证字",
    "凭证号",
    "摘要",
    "科目编码",
    "科目名称",
    "借方",
    "贷方",
    "流量金额（原币）",
    "对方科目",
    "原项目标准化结果",
    "审定现流表项目",
    "差异形成原因",
    "打分逻辑描述及打分结果",
    "独立来源1",
    "独立来源2",
    "来源文件",
    "来源工作表",
    "来源单元格",
)

REVIEW_HEADERS = (
    "日期",
    "凭证字",
    "凭证号",
    "本行摘要",
    "本行完整对方科目路径",
    "标准一级科目",
    "现金账户路径",
    "借方",
    "贷方",
    "流量金额（原币）",
    "本行分配现金变化",
    "现金方向依据",
    "原项目标准化结果",
    "系统候选项目",
    "判断理由",
    "摘要来源质量",
    "完整路径来源质量",
    "两个来源是否独立",
    "证据质量说明",
    "证据得分",
    "单笔金额",
    "单笔重要性层级",
    "强制检查",
    "唯一动作",
    "异常",
    "批次最不利影响金额",
    "批次现金变化金额",
    "行类型",
    "人工可选标准项目",
    "人工确认项目",
    "明确排除原因",
    "人工依据",
    "外部资料位置",
    "处理人",
    "处理时间",
    "人工处理状态",
    "批次编号(技术)",
    "系统项目(技术)",
    "系统基线金额(技术)",
    "系统项目调整(技术)",
    "目标项目金额(技术)",
    "包含笔数(技术)",
    "业务组成编号(技术)",
    "原基线项目(技术)",
)

USE_SYSTEM_RECOMMENDATION = "采用系统首选项目"
TRACE_MANUAL_HEADERS = (
    "人工改选基准项目(技术)",
    "人工改选基准金额(技术)",
    "人工改选目标金额(技术)",
    "人工改选生效标志(技术)",
)


def _display_value(header: object, value: object) -> object:
    """用户可见的日期列只显示日期，不携带午夜时间。"""
    if str(header).endswith("日期") or str(header) == "日期":
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, str):
            matched = re.match(r"^(\d{4}-\d{2}-\d{2})(?:[T ]00:00:00(?:\.0+)?(?:Z|[+-]\d{2}:?\d{2})?)?$", value.strip())
            if matched:
                return matched.group(1)
    return value


def _review_col(header: str, *, absolute: bool = False) -> str:
    name = xl_col_to_name(REVIEW_HEADERS.index(header))
    return f"${name}" if absolute else name


@dataclass(frozen=True, slots=True)
class WorkbookModel:
    statement: StatementResult
    rules: RulePack
    comparison: StatementComparison | None
    review_batches: tuple[ReviewBatch, ...]
    duplicate_groups: tuple[DuplicateGroup, ...]
    ai_records: tuple[Mapping[str, object], ...]
    cash_scope_rows: tuple[Mapping[str, object], ...]
    reconciliation: ReconciliationResult | None
    trace_rows: tuple[Mapping[str, object], ...]
    mapping_rows: tuple[Mapping[str, object], ...]
    overall_status: str
    automatic_change_threshold: int = DEFAULT_AUTOMATIC_CHANGE_SCORE
    difference_rows: tuple[Mapping[str, object], ...] = ()
    unconfirmed_statement: bool = False
    dictionary_rows: tuple[Mapping[str, object], ...] = ()
    consistency_rows: tuple[Mapping[str, object], ...] = ()
    manual_adjustments: Mapping[str, int] = field(default_factory=dict)
    low_amount_review_batches: tuple[ReviewBatch, ...] = ()


@dataclass(frozen=True, slots=True)
class WorkbookValidation:
    valid: bool
    errors: tuple[str, ...]


def manual_adjustment_formula(
    item_name: str,
    review_last_row: int,
    duplicate_last_row: int,
    base_amount: float = 0,
    low_amount_last_row: int = 2,
) -> str:
    review_end = max(2, review_last_row)
    duplicate_end = max(2, duplicate_last_row)
    low_end = max(2, low_amount_last_row)
    system_adjustment = _review_col("系统项目调整(技术)", absolute=True)
    baseline_item = _review_col("原基线项目(技术)", absolute=True)
    system_item = _review_col("系统项目(技术)", absolute=True)
    target_amount = _review_col("目标项目金额(技术)", absolute=True)
    manual_item = _review_col("人工确认项目", absolute=True)
    return (
        f'={base_amount}'
        f'+SUMIFS(\'重要待复核事项\'!{system_adjustment}$2:{system_adjustment}${review_end},\'重要待复核事项\'!{baseline_item}$2:{baseline_item}${review_end},"{item_name}")'
        f'+SUMIFS(\'重要待复核事项\'!{target_amount}$2:{target_amount}${review_end},\'重要待复核事项\'!{manual_item}$2:{manual_item}${review_end},"{item_name}")'
        f'+SUMIFS(\'重要待复核事项\'!{target_amount}$2:{target_amount}${review_end},\'重要待复核事项\'!{system_item}$2:{system_item}${review_end},"{item_name}",\'重要待复核事项\'!{manual_item}$2:{manual_item}${review_end},"{USE_SYSTEM_RECOMMENDATION}")'
        f'+SUMIFS(\'低金额批量处理\'!{system_adjustment}$2:{system_adjustment}${low_end},\'低金额批量处理\'!{baseline_item}$2:{baseline_item}${low_end},"{item_name}")'
        f'+SUMIFS(\'低金额批量处理\'!{target_amount}$2:{target_amount}${low_end},\'低金额批量处理\'!{manual_item}$2:{manual_item}${low_end},"{item_name}")'
        f'+SUMIFS(\'低金额批量处理\'!{target_amount}$2:{target_amount}${low_end},\'低金额批量处理\'!{system_item}$2:{system_item}${low_end},"{item_name}",\'低金额批量处理\'!{manual_item}$2:{manual_item}${low_end},"{USE_SYSTEM_RECOMMENDATION}")'
        f'+SUMIFS(\'疑似重复事项\'!$F$2:$F${duplicate_end},\'疑似重复事项\'!$B$2:$B${duplicate_end},"{item_name}")'
    )


def trace_manual_adjustment_terms(
    item_name: str,
    trace_last_row: int,
    trace_headers: Sequence[str],
) -> str:
    required = {*TRACE_MANUAL_HEADERS, "最终决定项目"}
    if not required.issubset(trace_headers):
        return ""

    def trace_col(header: str) -> str:
        return f"${xl_col_to_name(trace_headers.index(header))}"

    base_item = trace_col("人工改选基准项目(技术)")
    base_amount = trace_col("人工改选基准金额(技术)")
    target_amount = trace_col("人工改选目标金额(技术)")
    active = trace_col("人工改选生效标志(技术)")
    final_item = trace_col("最终决定项目")
    item = item_name.replace('"', '""')
    return (
        f'-SUMIFS(\'全量分类留痕\'!{base_amount}$2:{base_amount}${trace_last_row},'
        f'\'全量分类留痕\'!{base_item}$2:{base_item}${trace_last_row},"{item}",'
        f'\'全量分类留痕\'!{active}$2:{active}${trace_last_row},1)'
        f'+SUMIFS(\'全量分类留痕\'!{target_amount}$2:{target_amount}${trace_last_row},'
        f'\'全量分类留痕\'!{final_item}$2:{final_item}${trace_last_row},"{item}",'
        f'\'全量分类留痕\'!{active}$2:{active}${trace_last_row},1)'
    )


def calculate_manual_adjustments(
    model: WorkbookModel,
    review_decisions: Mapping[str, str],
    duplicate_decisions: Mapping[str, str],
) -> dict[str, int]:
    adjustments: defaultdict[str, int] = defaultdict(int)
    for batch in model.review_batches:
        selected = review_decisions.get(batch.batch_id)
        if selected == USE_SYSTEM_RECOMMENDATION:
            selected = batch.proposed_item_code
        selectable = tuple(
            item_id
            for item_id in (batch.proposed_item_code, *batch.alternative_item_codes)
            if item_id
        )
        if selected == "明确排除":
            if batch.baseline_item_code:
                adjustments[batch.baseline_item_code] -= batch.baseline_statement_amount_cent
        elif selected in selectable and selected != batch.baseline_item_code:
            if batch.baseline_item_code:
                adjustments[batch.baseline_item_code] -= batch.baseline_statement_amount_cent
            adjustments[selected] += statement_amount_cent(
                batch.cash_delta_cent,
                model.rules.item_by_id[selected].normal_direction,
            )
    for group in model.duplicate_groups:
        if duplicate_decisions.get(group.group_id, group.default_decision) in {"exclude", "剔除"}:
            adjustments[group.item_id] -= group.baseline_statement_amount_cent
    return dict(adjustments)


def _formats(workbook: xlsxwriter.Workbook) -> dict[str, xlsxwriter.format.Format]:
    base = {
        "font_name": "Times New Roman",
        "font_size": 11,
    }
    return {
        "title": workbook.add_format(
            {**base, "bold": True, "font_size": 16, "font_color": "#FFFFFF", "bg_color": "#17365D", "align": "center", "valign": "vcenter"}
        ),
        "header": workbook.add_format(
            {**base, "bold": True, "font_color": "#FFFFFF", "bg_color": "#1F4E78", "border": 1, "align": "center", "valign": "vcenter"}
        ),
        "section": workbook.add_format({**base, "bold": True, "bg_color": "#D9EAF7", "border": 1}),
        "text": workbook.add_format({**base, "border": 1, "valign": "top"}),
        "money": workbook.add_format({**base, "border": 1, "num_format": "#,##0.00;[Red](#,##0.00);-"}),
        "input": workbook.add_format({**base, "border": 1, "bg_color": "#DDEBF7", "font_color": "#0070C0", "locked": False}),
        "pending": workbook.add_format({**base, "border": 1, "bg_color": "#FFF2CC"}),
        "error": workbook.add_format({**base, "border": 1, "bg_color": "#F4CCCC", "font_color": "#9C0006"}),
        "note": workbook.add_format({**base, "italic": True, "font_color": "#666666"}),
        "link": workbook.add_format({**base, "font_color": "#0563C1", "underline": True}),
    }


def _configure_sheet(
    sheet: xlsxwriter.worksheet.Worksheet, columns: int, rows: int = 2
) -> None:
    sheet.set_default_row(18)
    sheet.hide_gridlines(2)
    sheet.freeze_panes(1, 0)
    sheet.set_landscape()
    sheet.fit_to_pages(1, 0)
    sheet.print_area(0, 0, max(1, rows - 1), max(0, columns - 1))


def _write_dict_rows(
    sheet: xlsxwriter.worksheet.Worksheet,
    rows: Sequence[Mapping[str, object]],
    formats: dict[str, xlsxwriter.format.Format],
    empty_note: str,
) -> None:
    if not rows:
        sheet.write(0, 0, "说明", formats["header"])
        sheet.write(1, 0, empty_note, formats["note"])
        sheet.set_column(0, 0, 60)
        _configure_sheet(sheet, 1)
        return
    headers = tuple(dict.fromkeys(key for row in rows for key in row.keys()))
    for column, header in enumerate(headers):
        sheet.write(0, column, header, formats["header"])
    for row_index, row in enumerate(rows, 1):
        for column, header in enumerate(headers):
            cell_format = (
                formats["money"]
                if header in {"金额（元）", "现金变化"}
                else formats["text"]
            )
            sheet.write(
                row_index,
                column,
                _display_value(header, row.get(header, "")),
                cell_format,
            )
    sheet.autofilter(0, 0, len(rows), len(headers) - 1)
    sheet.set_column(0, len(headers) - 1, 20)
    _configure_sheet(sheet, len(headers), len(rows) + 1)


def _write_difference_rows(
    sheet: xlsxwriter.worksheet.Worksheet,
    rows: Sequence[Mapping[str, object]],
    formats: dict[str, xlsxwriter.format.Format],
) -> None:
    money_headers = {"借方", "贷方", "流量金额（原币）"}
    for column, header in enumerate(DIFFERENCE_HEADERS):
        sheet.write(0, column, header, formats["header"])
    if rows:
        for row_index, row in enumerate(rows, 1):
            for column, header in enumerate(DIFFERENCE_HEADERS):
                value = _display_value(header, row.get(header))
                if header in money_headers:
                    if value is None:
                        sheet.write_blank(row_index, column, None, formats["money"])
                    else:
                        sheet.write_number(row_index, column, float(value), formats["money"])
                else:
                    sheet.write(row_index, column, value or "", formats["text"])
        sheet.autofilter(0, 0, len(rows), len(DIFFERENCE_HEADERS) - 1)
    else:
        sheet.write(1, 0, "原表项目与自动判定项目无差异。", formats["note"])
    sheet.set_column(0, 2, 14)
    sheet.set_column(3, 5, 24)
    sheet.set_column(6, 8, 16)
    sheet.set_column(9, 13, 38)
    sheet.set_column(14, 16, 24)
    _configure_sheet(sheet, len(DIFFERENCE_HEADERS), len(rows) + 1)


def _write_low_amount_review_sheet(
    sheet,
    model: WorkbookModel,
    formats: Mapping[str, object],
    item_name_by_id: Mapping[str, str],
) -> int:
    """低金额批次主行只判断一次，后续逐行展示现金分配明细。"""
    for column, header in enumerate(REVIEW_HEADERS):
        sheet.write(0, column, header, formats["header"])
    batches = model.low_amount_review_batches
    if not batches:
        sheet.write(1, 0, "本期无低金额人工批量事项。", formats["note"])
        _configure_sheet(sheet, len(REVIEW_HEADERS), 2)
        return 2

    trace_by_component: defaultdict[str, list[Mapping[str, object]]] = defaultdict(list)
    for trace_row in model.trace_rows:
        component_id = str(
            trace_row.get("业务组成编号(技术)")
            or trace_row.get("component_id")
            or ""
        )
        if component_id:
            trace_by_component[component_id].append(trace_row)

    helper_column = len(REVIEW_HEADERS) + 1
    helper_column_name = xl_col_to_name(helper_column)
    helper_row = 0
    current_row = 1
    numeric_headers = {
        "借方",
        "贷方",
        "流量金额（原币）",
        "本行分配现金变化",
        "单笔金额",
    }
    for batch in batches:
        master_row = current_row
        excel_master_row = master_row + 1
        selectable_ids = tuple(
            dict.fromkeys(
                item_id
                for item_id in (batch.proposed_item_code, *batch.alternative_item_codes)
                if item_id
            )
        )
        selectable_names = tuple(item_name_by_id[item_id] for item_id in selectable_ids)
        options = (USE_SYSTEM_RECOMMENDATION, *selectable_names)
        option_start = helper_row + 1
        for option in options:
            sheet.write(helper_row, helper_column, option, formats["text"])
            helper_row += 1
        list_range = f"${helper_column_name}${option_start}:${helper_column_name}${helper_row}"
        proposed_name = item_name_by_id.get(batch.proposed_item_code, "尚未形成系统候选")
        master_values = {
            "本行摘要": batch.representative_summary or "批次汇总",
            "本行完整对方科目路径": batch.counterpart_group or "见下方明细",
            "本行分配现金变化": batch.cash_delta_cent / 100,
            "单笔金额": abs(batch.cash_delta_cent) / 100,
            "批次最不利影响金额": batch.worst_case_impact_cent / 100,
            "批次现金变化金额": batch.cash_delta_cent / 100,
            "行类型": "批次判断",
            "系统候选项目": proposed_name,
            "判断理由": batch.reason,
            "人工可选标准项目": "、".join(selectable_names),
            "批次编号(技术)": batch.batch_id,
            "系统项目(技术)": proposed_name,
            "系统基线金额(技术)": batch.baseline_statement_amount_cent / 100,
            "包含笔数(技术)": len(batch.component_ids),
            "业务组成编号(技术)": "、" + "、".join(batch.component_ids) + "、",
            "原基线项目(技术)": item_name_by_id.get(batch.baseline_item_code, ""),
        }
        money_headers = numeric_headers | {
            "批次最不利影响金额",
            "批次现金变化金额",
            "系统基线金额(技术)",
        }
        for header, value in master_values.items():
            sheet.write(
                master_row,
                REVIEW_HEADERS.index(header),
                value,
                formats["money"] if header in money_headers else formats["text"],
            )
        sheet.write_blank(master_row, REVIEW_HEADERS.index("人工确认项目"), None, formats["input"])
        sheet.data_validation(
            master_row,
            REVIEW_HEADERS.index("人工确认项目"),
            master_row,
            REVIEW_HEADERS.index("人工确认项目"),
            {"validate": "list", "source": f"='低金额批量处理'!{list_range}"},
        )
        manual_col = _review_col("人工确认项目")
        system_col = _review_col("系统项目(技术)")
        baseline_col = _review_col("系统基线金额(技术)")
        baseline_item_col = _review_col("原基线项目(技术)")
        cash_change_col = _review_col("批次现金变化金额")
        effective_item = (
            f'IF({manual_col}{excel_master_row}="{USE_SYSTEM_RECOMMENDATION}",'
            f'{system_col}{excel_master_row},{manual_col}{excel_master_row})'
        )
        sheet.write_formula(
            master_row,
            REVIEW_HEADERS.index("系统项目调整(技术)"),
            f'=IF(OR({manual_col}{excel_master_row}="",{baseline_item_col}{excel_master_row}=""),0,'
            f'IF({effective_item}<>{baseline_item_col}{excel_master_row},-{baseline_col}{excel_master_row},0))',
            formats["money"],
            0,
        )
        target_formula = "0"
        for item_id, item_name in reversed(tuple(zip(selectable_ids, selectable_names, strict=True))):
            direction = model.rules.item_by_id[item_id].normal_direction
            amount = f"{cash_change_col}{excel_master_row}" if direction == "inflow" else f"-{cash_change_col}{excel_master_row}"
            target_formula = f'IF({manual_col}{excel_master_row}="{item_name}",{amount},{target_formula})'
        if batch.proposed_item_code:
            proposed_direction = model.rules.item_by_id[batch.proposed_item_code].normal_direction
            proposed_amount = f"{cash_change_col}{excel_master_row}" if proposed_direction == "inflow" else f"-{cash_change_col}{excel_master_row}"
            target_formula = f'IF({manual_col}{excel_master_row}="{USE_SYSTEM_RECOMMENDATION}",{proposed_amount},{target_formula})'
        sheet.write_formula(master_row, REVIEW_HEADERS.index("目标项目金额(技术)"), f"={target_formula}", formats["money"], 0)
        sheet.write_formula(
            master_row,
            REVIEW_HEADERS.index("人工处理状态"),
            f'=IF({manual_col}{excel_master_row}="","等待人工处理","人工处理完成")',
            formats["pending"],
            "等待人工处理",
        )
        current_row += 1

        for component_id in batch.component_ids:
            for trace_row in trace_by_component.get(component_id, ()):
                detail_values = {
                    header: trace_row.get(header, "")
                    for header in REVIEW_HEADERS
                    if header in trace_row
                }
                detail_values.update(
                    {
                        "行类型": "现金分配明细",
                        "批次编号(技术)": batch.batch_id,
                        "业务组成编号(技术)": component_id,
                    }
                )
                for header, value in detail_values.items():
                    if header in numeric_headers and not isinstance(value, (int, float)):
                        value = ""
                    sheet.write(
                        current_row,
                        REVIEW_HEADERS.index(header),
                        _display_value(header, value),
                        formats["money"] if header in numeric_headers and isinstance(value, (int, float)) else formats["text"],
                    )
                sheet.write_formula(
                    current_row,
                    REVIEW_HEADERS.index("人工确认项目"),
                    f"={manual_col}{excel_master_row}",
                    formats["text"],
                    "",
                )
                sheet.write(current_row, REVIEW_HEADERS.index("人工处理状态"), "明细随批次主行生效", formats["text"])
                current_row += 1

    sheet.autofilter(0, 0, current_row - 1, len(REVIEW_HEADERS) - 1)
    for column, header in enumerate(REVIEW_HEADERS):
        if header.endswith("(技术)"):
            sheet.set_column(column, column, 20, None, {"hidden": True})
    sheet.set_column(helper_column, helper_column, None, None, {"hidden": True})
    sheet.set_column("A:D", 20)
    sheet.set_column("E:AF", 22)
    _configure_sheet(sheet, len(REVIEW_HEADERS), current_row)
    sheet.protect("", {"autofilter": True, "sort": True, "select_unlocked_cells": True})
    return max(2, current_row)


def build_output_workbook(model: WorkbookModel, output_path: Path) -> Path:
    target = Path(output_path)
    if target.exists():
        raise FileExistsError(f"输出文件已存在，不会覆盖：{target}")
    if len(model.trace_rows) > 100_000:
        raise ValueError("全量分类留痕超过本版本 100,000 行验收范围")
    if len(model.difference_rows) > 100_000:
        raise ValueError("原表与系统决定差异明细超过本版本 100,000 行验收范围")
    if any(
        not batch.alternative_item_codes and not batch.mandatory
        for batch in model.review_batches
    ):
        raise ValueError("重要待复核事项存在没有备选现流项目的记录")
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = xlsxwriter.Workbook(str(target))
    workbook.set_calc_mode("auto")
    formats = _formats(workbook)
    sheets = {name: workbook.add_worksheet(name) for name in SHEET_NAMES}
    item_name_by_id = {
        item.item_id: item.name for item in model.rules.statement_items
    }
    cleaned_trace_rows = tuple(
        {key: value for key, value in row.items() if key != "人工决定"}
        for row in model.trace_rows
    )
    trace_supports_manual_choice = any(
        "最终决定项目" in row for row in cleaned_trace_rows
    )
    trace_rows = tuple(
        {
            **row,
            **(
                {header: "" for header in TRACE_MANUAL_HEADERS}
                if trace_supports_manual_choice
                else {}
            ),
        }
        for row in cleaned_trace_rows
    )
    trace_headers = tuple(
        dict.fromkeys(key for row in trace_rows for key in row.keys())
    )
    trace_last = max(2, len(trace_rows) + 1)
    try:
        low_amount_last = _write_low_amount_review_sheet(
            sheets["低金额批量处理"],
            model,
            formats,
            item_name_by_id,
        )
        status = sheets["使用说明与状态"]
        status.set_default_row(18)
        status.set_row(0, 24)
        status.hide_gridlines(2)
        status.merge_range("A1:D1", "直接法现金流量表正表编制与复核底稿", formats["title"])
        status.write("A3", "当前状态", formats["header"])
        reconciliation_complete = bool(
            model.reconciliation is not None
            and model.reconciliation.opening_cent is not None
            and model.reconciliation.closing_cent is not None
            and model.reconciliation.fx_cent is not None
        )
        hard_draft = (
            model.overall_status.startswith("诊断材料")
            or "输入存在未处理错误" in model.overall_status
            or (
            model.reconciliation is not None and not reconciliation_complete
            )
        )
        if model.unconfirmed_statement:
            status.write("B3", "草稿：存在未核对的疑似正表", formats["error"])
            status.write("A4", "提示", formats["header"])
            status.write(
                "B4",
                "检测到疑似客户现有正表但未确认纳入核对，请返回确认后再生成最终结果。",
                formats["note"],
            )
        elif hard_draft:
            status.write("B3", model.overall_status, formats["error"])
        elif (
            model.review_batches
            or model.low_amount_review_batches
            or model.duplicate_groups
            or reconciliation_complete
        ):
            review_end = max(2, len(model.review_batches) + 1)
            duplicate_end = max(2, len(model.duplicate_groups) + 1)
            pending_terms = []
            if model.review_batches:
                review_status_col = _review_col("人工处理状态")
                pending_terms.append(
                    f'COUNTIF(\'重要待复核事项\'!{review_status_col}2:{review_status_col}{review_end},"等待人工处理")'
                )
            if model.low_amount_review_batches:
                low_status_col = _review_col("人工处理状态")
                pending_terms.append(
                    f'COUNTIF(\'低金额批量处理\'!{low_status_col}2:{low_status_col}{low_amount_last},"等待人工处理")'
                )
            if model.duplicate_groups:
                pending_terms.append(
                    f'COUNTIF(\'疑似重复事项\'!H2:H{duplicate_end},"无效选择")'
                )
                pending_terms.append(
                    f'COUNTIFS(\'疑似重复事项\'!G2:G{duplicate_end},"是",'
                    f'\'疑似重复事项\'!H2:H{duplicate_end},"待确认")'
                )
            completed_value = '"最终可使用"'
            if reconciliation_complete:
                bridge_row = len(model.cash_scope_rows) + 7
                completed_value = (
                    f'IF(\'现金范围与现金流量表与货币资金变动的勾稽核对\'!B{bridge_row}="最终现金流量表勾稽成功",'
                    '"最终可使用","草稿：现金流量表尚待分类或现金变动桥接存在差异")'
                )
            condition = "+".join(pending_terms) or "0"
            status.write_formula(
                "B3",
                f'=IF({condition}=0,{completed_value},"待完成人工确认")',
                formats["pending"],
                model.overall_status,
            )
        else:
            status.write("B3", "最终可使用", formats["text"])
        status.write("A5", "使用说明", formats["header"])
        status.write("B5", "蓝色或黄色单元格为人工选择区；修改后结果会即时更新，无需再次运行本工具。", formats["text"])
        status.write("A6", "本次自动修改最低证据分", formats["header"])
        status.write(
            "B6",
            f"{model.automatic_change_threshold}分（客户选择；70为默认推荐）",
            formats["text"],
        )
        for index, name in enumerate(SHEET_NAMES[1:], 7):
            status.write_url(index - 1, 0, f"internal:'{name}'!A1", formats["link"], string=name)
        status.set_column("A:A", 24)
        status.set_column("B:D", 32)
        status.freeze_panes(2, 0)
        status.print_area("A1:D20")

        review = sheets["重要待复核事项"]
        review_detail_count = 0
        for column, header in enumerate(REVIEW_HEADERS):
            review.write(0, column, header, formats["header"])
        helper_column = len(REVIEW_HEADERS) + 1
        if model.review_batches:
            trace_by_component: defaultdict[str, list[Mapping[str, object]]] = defaultdict(list)
            for trace_row in model.trace_rows:
                component_id = str(
                    trace_row.get("业务组成编号(技术)")
                    or trace_row.get("component_id")
                    or ""
                )
                if component_id:
                    trace_by_component[component_id].append(trace_row)

            def review_fact(
                batch: ReviewBatch, header: str, default: object = "未记录"
            ) -> object:
                values = tuple(
                    dict.fromkeys(
                        row.get(header)
                        for component_id in batch.component_ids
                        for row in trace_by_component.get(component_id, ())
                        if row.get(header) not in (None, "")
                    )
                )
                if not values:
                    return default
                if len(values) == 1:
                    return values[0]
                return "、".join(str(value) for value in values)

            helper_column_name = xl_col_to_name(helper_column)
            helper_ranges: dict[tuple[str, ...], tuple[str, str]] = {}
            helper_row = 0
            review_row_by_component = {
                component_id: excel_row
                for excel_row, batch in enumerate(model.review_batches, 2)
                for component_id in batch.component_ids
            }
            for batch in model.review_batches:
                if (
                    batch.follows_component_id
                    and batch.follows_component_id not in review_row_by_component
                ):
                    raise ValueError(
                        "增值税附属复核行找不到基础项目复核行："
                        + batch.follows_component_id
                    )
            for batch in model.review_batches:
                selectable_ids = tuple(
                    item_id
                    for item_id in (
                        batch.proposed_item_code,
                        *batch.alternative_item_codes,
                    )
                    if item_id
                )
                selectable_names = tuple(item_name_by_id[item_id] for item_id in selectable_ids)
                recommendation_option = (USE_SYSTEM_RECOMMENDATION,)
                options = (*recommendation_option, *selectable_names)
                if options in helper_ranges:
                    continue
                start_row = helper_row + 1
                for option in options:
                    review.write(helper_row, helper_column, option, formats["text"])
                    helper_row += 1
                list_range = f"${helper_column_name}${start_row}:${helper_column_name}${helper_row}"
                helper_ranges[options] = (list_range, list_range)
            for row_index, batch in enumerate(model.review_batches, 1):
                proposed_name = item_name_by_id.get(
                    batch.proposed_item_code, "尚未形成系统候选"
                )
                selectable_ids = tuple(
                    item_id
                    for item_id in (
                        batch.proposed_item_code,
                        *batch.alternative_item_codes,
                    )
                    if item_id
                )
                selectable_names = tuple(item_name_by_id[item_id] for item_id in selectable_ids)
                recommendation_option = (USE_SYSTEM_RECOMMENDATION,)
                list_range, choice_range = helper_ranges[
                    (*recommendation_option, *selectable_names)
                ]
                source_fallbacks = tuple(
                    location.split("|", 2)
                    for location in batch.source_locations
                    if location
                )
                facts = {
                    "日期": review_fact(batch, "日期"),
                    "凭证字": review_fact(batch, "凭证字"),
                    "凭证号": review_fact(batch, "凭证号"),
                    "本行摘要": review_fact(batch, "本行摘要", batch.representative_summary or "未记录"),
                    "本行完整对方科目路径": review_fact(
                        batch,
                        "本行完整对方科目路径",
                        batch.counterpart_group or "未记录",
                    ),
                    "标准一级科目": review_fact(batch, "标准一级科目"),
                    "现金账户路径": review_fact(batch, "现金账户路径"),
                    "借方": review_fact(batch, "借方"),
                    "贷方": review_fact(batch, "贷方"),
                    "流量金额（原币）": review_fact(batch, "流量金额（原币）"),
                    "本行分配现金变化": batch.cash_delta_cent / 100,
                    "现金方向依据": review_fact(batch, "现金方向依据"),
                    "原项目标准化结果": review_fact(batch, "原项目标准化结果"),
                    "系统候选项目": review_fact(batch, "系统候选项目", proposed_name),
                    "判断理由": review_fact(batch, "判断理由", batch.reason),
                    "摘要来源质量": review_fact(batch, "摘要来源质量"),
                    "完整路径来源质量": review_fact(batch, "完整路径来源质量"),
                    "两个来源是否独立": review_fact(batch, "两个来源是否独立"),
                    "证据质量说明": review_fact(batch, "证据质量说明"),
                    "证据得分": review_fact(batch, "证据得分"),
                    "单笔金额": abs(batch.cash_delta_cent) / 100,
                    "单笔重要性层级": review_fact(batch, "单笔重要性层级"),
                    "强制检查": review_fact(batch, "强制检查"),
                    "唯一动作": review_fact(batch, "唯一动作"),
                    "异常": review_fact(batch, "异常"),
                    "行类型": "批次判断",
                    "人工可选标准项目": (
                        "随基础项目自动确定（无需重复选择）"
                        if batch.follows_component_id
                        else "、".join(selectable_names)
                    ),
                }
                for header, value in facts.items():
                    numeric_fact_headers = {
                        "借方",
                        "贷方",
                        "流量金额（原币）",
                        "本行分配现金变化",
                        "单笔金额",
                    }
                    if header in numeric_fact_headers and not isinstance(
                        value, (int, float)
                    ):
                        value = ""
                    cell_format = (
                        formats["money"]
                        if header in numeric_fact_headers
                        and isinstance(value, (int, float))
                        else formats["text"]
                    )
                    review.write(
                        row_index,
                        REVIEW_HEADERS.index(header),
                        _display_value(header, value),
                        cell_format,
                    )
                review.write_number(
                    row_index,
                    REVIEW_HEADERS.index("批次最不利影响金额"),
                    batch.worst_case_impact_cent / 100,
                    formats["money"],
                )
                review.write_number(
                    row_index,
                    REVIEW_HEADERS.index("批次现金变化金额"),
                    batch.cash_delta_cent / 100,
                    formats["money"],
                )
                for header in (
                    "人工确认项目",
                    "明确排除原因",
                    "人工依据",
                    "外部资料位置",
                    "处理人",
                    "处理时间",
                ):
                    if header == "人工确认项目" and batch.follows_component_id:
                        continue
                    review.write_blank(
                        row_index,
                        REVIEW_HEADERS.index(header),
                        None,
                        formats["input"],
                    )
                review.write(row_index, REVIEW_HEADERS.index("批次编号(技术)"), batch.batch_id, formats["text"])
                review.write(row_index, REVIEW_HEADERS.index("系统项目(技术)"), proposed_name, formats["text"])
                review.write_number(
                    row_index,
                    REVIEW_HEADERS.index("系统基线金额(技术)"),
                    batch.baseline_statement_amount_cent / 100,
                    formats["money"],
                )
                excel_row = row_index + 1
                manual_col = _review_col("人工确认项目")
                system_col = _review_col("系统项目(技术)")
                status_col = _review_col("人工处理状态")
                baseline_col = _review_col("系统基线金额(技术)")
                baseline_item_col = _review_col("原基线项目(技术)")
                cash_change_col = _review_col("批次现金变化金额")
                if batch.follows_component_id:
                    base_row = review_row_by_component[batch.follows_component_id]
                    base_choice = f"{manual_col}{base_row}"
                    base_system = f"{system_col}{base_row}"
                    review.write_formula(
                        row_index,
                        REVIEW_HEADERS.index("人工确认项目"),
                        f'=IF({base_choice}="","",IF({base_choice}="{USE_SYSTEM_RECOMMENDATION}",{base_system},{base_choice}))',
                        formats["text"],
                        "",
                    )
                effective_item = (
                    f'IF({manual_col}{excel_row}="{USE_SYSTEM_RECOMMENDATION}",'
                    f'{system_col}{excel_row},{manual_col}{excel_row})'
                )
                review.write_formula(
                    row_index,
                    REVIEW_HEADERS.index("系统项目调整(技术)"),
                    f'=IF(OR({manual_col}{excel_row}="",{baseline_item_col}{excel_row}=""),0,'
                    f'IF(OR({manual_col}{excel_row}="明确排除",'
                    f'{effective_item}<>{baseline_item_col}{excel_row}),-{baseline_col}{excel_row},0))',
                    formats["money"],
                    0,
                )
                target_formula = "0"
                for item_id, item_name in reversed(
                    tuple(zip(selectable_ids, selectable_names, strict=True))
                ):
                    direction = model.rules.item_by_id[item_id].normal_direction
                    amount = (
                        f'{cash_change_col}{excel_row}'
                        if direction == "inflow"
                        else f'-{cash_change_col}{excel_row}'
                    )
                    target_formula = (
                        f'IF({manual_col}{excel_row}="{item_name}",'
                        f'IF({baseline_item_col}{excel_row}="{item_name}",0,{amount}),'
                        f'{target_formula})'
                    )
                if batch.proposed_item_code:
                    proposed_direction = model.rules.item_by_id[
                        batch.proposed_item_code
                    ].normal_direction
                    proposed_amount = (
                        f'{cash_change_col}{excel_row}'
                        if proposed_direction == "inflow"
                        else f'-{cash_change_col}{excel_row}'
                    )
                    target_formula = (
                        f'IF({manual_col}{excel_row}="{USE_SYSTEM_RECOMMENDATION}",'
                        f'IF({baseline_item_col}{excel_row}={system_col}{excel_row},0,'
                        f'{proposed_amount}),{target_formula})'
                    )
                review.write_formula(
                    row_index,
                    REVIEW_HEADERS.index("目标项目金额(技术)"),
                    f'={target_formula}',
                    formats["money"],
                    0,
                )
                if batch.follows_component_id:
                    base_row = review_row_by_component[batch.follows_component_id]
                    review.write_formula(
                        row_index,
                        REVIEW_HEADERS.index("人工处理状态"),
                        f'=IF({status_col}{base_row}="人工处理完成","随基础项目完成","随基础项目待定")',
                        formats["pending"],
                        "随基础项目待定",
                    )
                else:
                    review.write_formula(
                        row_index,
                        REVIEW_HEADERS.index("人工处理状态"),
                        f'=IF({manual_col}{excel_row}="","等待人工处理",'
                        f'IF(AND({manual_col}{excel_row}="{USE_SYSTEM_RECOMMENDATION}",'
                        f'{system_col}{excel_row}="尚未形成系统候选"),'
                        '"系统没有首选项目，请改选","人工处理完成"))',
                        formats["pending"],
                        "等待人工处理",
                    )
                review.write_number(
                    row_index,
                    REVIEW_HEADERS.index("包含笔数(技术)"),
                    len(batch.component_ids),
                    formats["text"],
                )
                review.write(
                    row_index,
                    REVIEW_HEADERS.index("业务组成编号(技术)"),
                    "、" + "、".join(batch.component_ids) + "、",
                    formats["text"],
                )
                review.write(
                    row_index,
                    REVIEW_HEADERS.index("原基线项目(技术)"),
                    item_name_by_id.get(batch.baseline_item_code, ""),
                    formats["text"],
                )
                if not batch.follows_component_id:
                    review.data_validation(
                        row_index,
                        REVIEW_HEADERS.index("人工确认项目"),
                        row_index,
                        REVIEW_HEADERS.index("人工确认项目"),
                        {
                            "validate": "list",
                            "source": f"='重要待复核事项'!{list_range}",
                        },
                    )
            detail_row_index = len(model.review_batches) + 1
            detail_numeric_headers = {
                "借方",
                "贷方",
                "流量金额（原币）",
                "本行分配现金变化",
                "单笔金额",
            }
            for master_excel_row, batch in enumerate(model.review_batches, 2):
                for component_id in batch.component_ids:
                    for trace_row in trace_by_component.get(component_id, ()):
                        detail_values = {
                            header: trace_row.get(header, "")
                            for header in REVIEW_HEADERS
                            if header in trace_row
                        }
                        detail_values.update(
                            {
                                "行类型": "现金分配明细",
                                "批次编号(技术)": batch.batch_id,
                                "业务组成编号(技术)": component_id,
                            }
                        )
                        for header, value in detail_values.items():
                            if header in detail_numeric_headers and not isinstance(
                                value, (int, float)
                            ):
                                value = ""
                            review.write(
                                detail_row_index,
                                REVIEW_HEADERS.index(header),
                                _display_value(header, value),
                                formats["money"]
                                if header in detail_numeric_headers
                                and isinstance(value, (int, float))
                                else formats["text"],
                            )
                        review.write_formula(
                            detail_row_index,
                            REVIEW_HEADERS.index("人工确认项目"),
                            f"={_review_col('人工确认项目')}{master_excel_row}",
                            formats["text"],
                            "",
                        )
                        review.write(
                            detail_row_index,
                            REVIEW_HEADERS.index("人工处理状态"),
                            "明细随批次主行生效",
                            formats["text"],
                        )
                        detail_row_index += 1
                        review_detail_count += 1
            review.autofilter(
                0,
                0,
                len(model.review_batches) + review_detail_count,
                len(REVIEW_HEADERS) - 1,
            )
        else:
            review.write(1, 0, "本期无重大剩余不确定事项，无需人工复核。", formats["note"])
        review.set_column("A:D", 20)
        review.set_column("E:G", 12)
        review.set_column("H:Q", 28)
        review.set_column("R:AF", 22)
        review.set_column("AG:AL", 20)
        review.set_column("AM:AN", 18)
        hidden_review_headers = {
            "摘要来源质量",
            "完整路径来源质量",
            "两个来源是否独立",
            "证据质量说明",
            "证据得分",
            "单笔重要性层级",
            "强制检查",
            "唯一动作",
            "批次最不利影响金额",
            "批次现金变化金额",
            "明确排除原因",
            "人工依据",
            "外部资料位置",
            "处理人",
            "处理时间",
        }
        for column, header in enumerate(REVIEW_HEADERS):
            if header.endswith("(技术)") or header in hidden_review_headers:
                review.set_column(column, column, 20, None, {"hidden": True})
        review.set_column(helper_column, helper_column, None, None, {"hidden": True})
        _configure_sheet(
            review,
            len(REVIEW_HEADERS),
            len(model.review_batches) + review_detail_count + 1,
        )
        review.protect("", {"autofilter": True, "sort": True, "select_unlocked_cells": True})

        duplicate = sheets["疑似重复事项"]
        duplicate_headers = (
            "组编号",
            "标准项目",
            "人工决定",
            "最不利影响金额",
            "重复计入正表金额",
            "有效调整金额",
            "是否阻止完成",
            "状态",
        )
        for column, header in enumerate(duplicate_headers):
            duplicate.write(0, column, header, formats["header"])
        if model.duplicate_groups:
            for row_index, group in enumerate(model.duplicate_groups, 1):
                duplicate.write(row_index, 0, group.group_id, formats["text"])
                if group.item_id:
                    duplicate.write(
                        row_index,
                        1,
                        item_name_by_id[group.item_id],
                        formats["text"],
                    )
                else:
                    manual_column = xl_col_to_name(
                        REVIEW_HEADERS.index("人工确认项目") + 1
                    )
                    component_column = xl_col_to_name(
                        REVIEW_HEADERS.index("业务组成编号(技术)") + 1
                    )
                    duplicate.write_formula(
                        row_index,
                        1,
                        "=IFERROR(INDEX('重要待复核事项'!"
                        f"${manual_column}$2:${manual_column}${max(2, len(model.review_batches) + 1)},"
                        "MATCH(\"*"
                        f"{group.component_ids[0]}"
                        "*\",'重要待复核事项'!"
                        f"${component_column}$2:${component_column}${max(2, len(model.review_batches) + 1)},0)),"
                        '"待人工决定")',
                        formats["text"],
                        "待人工决定",
                    )
                duplicate.write_blank(row_index, 2, None, formats["input"])
                duplicate.write_number(row_index, 3, group.worst_case_impact_cent / 100, formats["money"])
                duplicate.write_number(
                    row_index,
                    4,
                    group.baseline_statement_amount_cent / 100,
                    formats["money"],
                )
                duplicate.write_formula(
                    row_index,
                    5,
                    f'=IF(C{row_index + 1}="剔除",-E{row_index + 1},0)',
                    formats["money"],
                    0,
                )
                duplicate.write(row_index, 6, "是" if group.blocks_manual_completion else "否", formats["pending"])
                duplicate.write_formula(
                    row_index,
                    7,
                    f'=IF(C{row_index + 1}="",IF(G{row_index + 1}="是","待确认","默认保留"),IF(C{row_index + 1}="保留","保留",IF(C{row_index + 1}="剔除","剔除","无效选择")))',
                    formats["pending"],
                    "待确认" if group.blocks_manual_completion else "默认保留",
                )
                duplicate.data_validation(row_index, 2, row_index, 2, {"validate": "list", "source": ["保留", "剔除"]})
            duplicate.autofilter(0, 0, len(model.duplicate_groups), len(duplicate_headers) - 1)
        else:
            duplicate.write(1, 0, "本期未发现跨文件疑似重复事项。", formats["note"])
        duplicate.set_column("A:C", 18)
        duplicate.set_column("D:H", 18)
        _configure_sheet(duplicate, len(duplicate_headers), len(model.duplicate_groups) + 1)
        duplicate.protect("", {"autofilter": True, "sort": True, "select_unlocked_cells": True})

        main = sheets["现金流量表正表"]
        main.set_default_row(18)
        main.set_row(0, 24)
        main.hide_gridlines(2)
        main.merge_range("A1:F1", "现金流量表", formats["title"])
        main.write("A2", "金额单位：人民币元", formats["note"])
        headers = ("项目编号", "项目", "上期金额", "自动基线", "人工调整", "最终金额")
        for column, header in enumerate(headers):
            main.write(2, column, header, formats["header"])
        ordered = sorted(model.rules.statement_items, key=lambda item: item.display_order)
        excel_row_by_id = {item.item_id: index + 4 for index, item in enumerate(ordered)}
        review_last = max(2, len(model.review_batches) + 1)
        duplicate_last = max(2, len(model.duplicate_groups) + 1)
        for index, item in enumerate(ordered):
            row = index + 3
            excel_row = row + 1
            row_format = formats["text"] if item.is_leaf else formats["section"]
            main.write(row, 0, item.item_id, row_format)
            main.write(row, 1, item.name, row_format)
            prior = model.statement.prior_values.get(item.item_id)
            if prior is not None:
                main.write_number(row, 2, prior / 100, formats["money"])
            else:
                main.write_blank(row, 2, None, formats["money"])
            main.write_number(row, 3, model.statement.values[item.item_id] / 100, formats["money"])
            if item.is_leaf:
                main.write_formula(
                    row,
                    4,
                    manual_adjustment_formula(
                        item.name,
                        review_last,
                        duplicate_last,
                        model.manual_adjustments.get(item.item_id, 0) / 100,
                        low_amount_last,
                    )
                    + trace_manual_adjustment_terms(
                        item.name,
                        trace_last,
                        trace_headers,
                    ),
                    formats["money"],
                    model.manual_adjustments.get(item.item_id, 0) / 100,
                )
            else:
                main.write_number(row, 4, 0, formats["money"])
            if item.formula_components:
                parts = [
                    f"F{excel_row_by_id[source_id]}*{multiplier}"
                    for source_id, multiplier in item.formula_components
                ]
                formula = "=" + "+".join(parts).replace("+-", "-")
            else:
                formula = f"=D{excel_row}+E{excel_row}"
            main.write_formula(
                row,
                5,
                formula,
                formats["money"],
                (
                    model.statement.values[item.item_id]
                    + model.manual_adjustments.get(item.item_id, 0)
                )
                / 100,
            )
        main.freeze_panes(3, 0)
        main.autofilter(2, 0, len(ordered) + 2, len(headers) - 1)
        main.set_column("A:A", 16)
        main.set_column("B:B", 50)
        main.set_column("C:F", 18)
        main.set_landscape()
        main.fit_to_pages(1, 1)
        main.print_area(0, 0, len(ordered) + 2, 5)
        main.protect("", {"autofilter": True, "sort": True})

        comparison_rows = () if model.comparison is None else tuple(
            {
                "项目": item_name_by_id[row.item_id],
                "客户金额": None if row.existing_cent is None else row.existing_cent / 100,
                "系统自动调整": row.system_adjustment_cent / 100,
                "自动基线": row.computed_cent / 100,
                "人工调整": row.manual_adjustment_cent / 100,
                "最终金额": row.final_cent / 100,
                "最终差异": None if row.difference_cent is None else row.difference_cent / 100,
                "明细重建金额": row.detail_reconstruction_cent / 100,
                "原表与明细勾稽差额": None if row.detail_gap_cent is None else row.detail_gap_cent / 100,
                "支持组成": "、".join(row.support_component_ids),
            }
            for row in model.comparison.rows
        )
        comparison_sheet = sheets["正表核对报告"]
        if comparison_rows:
            comparison_headers = (
                "项目",
                "客户金额",
                "系统自动调整",
                "自动基线",
                "人工调整",
                "最终金额",
                "最终差异",
                "明细重建金额",
                "原表与明细勾稽差额",
                "支持组成",
            )
            for column, header in enumerate(comparison_headers):
                comparison_sheet.write(0, column, header, formats["header"])
            for row_index, row in enumerate(comparison_rows, 1):
                main_row = row_index + 3
                comparison_sheet.write(row_index, 0, row["项目"], formats["text"])
                if row["客户金额"] is None:
                    comparison_sheet.write_blank(row_index, 1, None, formats["money"])
                else:
                    comparison_sheet.write_number(row_index, 1, row["客户金额"], formats["money"])
                comparison_sheet.write_number(row_index, 2, row["系统自动调整"], formats["money"])
                comparison_sheet.write_formula(row_index, 3, f"='现金流量表正表'!D{main_row}", formats["money"], row["自动基线"])
                comparison_sheet.write_formula(row_index, 4, f"='现金流量表正表'!E{main_row}", formats["money"], row["人工调整"])
                comparison_sheet.write_formula(row_index, 5, f"='现金流量表正表'!F{main_row}", formats["money"], row["最终金额"])
                comparison_sheet.write_formula(
                    row_index,
                    6,
                    f'=IF(B{row_index + 1}="","",ROUND(F{row_index + 1}-B{row_index + 1},2))',
                    formats["money"],
                    "" if row["最终差异"] is None else row["最终差异"],
                )
                comparison_sheet.write_number(row_index, 7, row["明细重建金额"], formats["money"])
                if row["原表与明细勾稽差额"] is None:
                    comparison_sheet.write_blank(row_index, 8, None, formats["money"])
                else:
                    comparison_sheet.write_number(row_index, 8, row["原表与明细勾稽差额"], formats["money"])
                comparison_sheet.write(row_index, 9, row["支持组成"], formats["text"])
            comparison_sheet.autofilter(0, 0, len(comparison_rows), len(comparison_headers) - 1)
            comparison_sheet.set_column("A:A", 48)
            comparison_sheet.set_column("B:I", 18)
            comparison_sheet.set_column("J:J", 40, None, {"hidden": True})
            _configure_sheet(comparison_sheet, len(comparison_headers), len(comparison_rows) + 1)
            comparison_sheet.protect("", {"autofilter": True, "sort": True})
        else:
            _write_dict_rows(comparison_sheet, (), formats, "本次为编制任务，未提供客户现有正表。")
        ai_display_rows = tuple(
            {
                ("现流项目" if key == "item_id" else key): (
                    item_name_by_id.get(str(value), value)
                    if key == "item_id"
                    else json.dumps(value, ensure_ascii=False)
                    if isinstance(value, (dict, list, tuple))
                    else value
                )
                for key, value in row.items()
            }
            for row in model.ai_records
        )
        _write_dict_rows(
            sheets["AI复核记录"],
            ai_display_rows,
            formats,
            "本期没有需要 AI 复核的事项。",
        )
        sheets["AI复核记录"].hide()

        _write_difference_rows(
            sheets["原表与系统决定差异"], model.difference_rows, formats
        )
        difference_sheet = sheets["原表与系统决定差异"]
        for header in (
            "来源文件",
            "来源工作表",
            "来源单元格",
        ):
            column = DIFFERENCE_HEADERS.index(header)
            difference_sheet.set_column(column, column, 20, None, {"hidden": True})
        difference_sheet.protect(
            "", {"autofilter": True, "sort": True}
        )

        cash_rows = list(model.cash_scope_rows)
        if model.reconciliation is not None:
            legacy_reconciliation = model.reconciliation.bridge_difference_cent is None
            classified_net_cent = (
                int(model.reconciliation.net_cash_cent or 0)
                - int(model.reconciliation.fx_cent or 0)
                if legacy_reconciliation
                else model.reconciliation.classified_net_cent
            )
            pending_net_cent = (
                0 if legacy_reconciliation else model.reconciliation.pending_net_cent
            )
            confirmed_adjustment_cent = (
                0
                if legacy_reconciliation
                else model.reconciliation.confirmed_adjustment_cent
            )
            bridge_difference_cent = (
                model.reconciliation.difference_cent
                if legacy_reconciliation
                else model.reconciliation.bridge_difference_cent
            )
            final_difference_cent = (
                model.reconciliation.difference_cent
                if legacy_reconciliation
                else model.reconciliation.final_difference_cent
            )
            for project, amount in (
                ("期初现金及现金等价物余额", model.reconciliation.opening_cent),
                ("已分类现金流量表净额", classified_net_cent),
                ("尚待分类现金净额", pending_net_cent),
                ("汇率变动影响", model.reconciliation.fx_cent),
                ("已确认调整", confirmed_adjustment_cent),
                ("现金变动桥接差异", bridge_difference_cent),
                ("现金流量表最终差异", final_difference_cent),
                ("期末现金及现金等价物余额", model.reconciliation.closing_cent),
            ):
                cash_rows.append(
                    {
                        "科目": project,
                        "决定": model.reconciliation.status,
                        "金额（元）": None if amount is None else amount / 100,
                    }
                )
        cash_sheet = sheets["现金范围与现金流量表与货币资金变动的勾稽核对"]
        _write_dict_rows(cash_sheet, tuple(cash_rows), formats, "现金范围尚未确认。")
        if reconciliation_complete:
            opening_row = len(model.cash_scope_rows) + 2
            classified_row = len(model.cash_scope_rows) + 3
            pending_row = len(model.cash_scope_rows) + 4
            fx_row = len(model.cash_scope_rows) + 5
            adjustment_row = len(model.cash_scope_rows) + 6
            bridge_row = len(model.cash_scope_rows) + 7
            final_row = len(model.cash_scope_rows) + 8
            closing_row = len(model.cash_scope_rows) + 9
            cash_sheet.write_formula(
                classified_row - 1,
                2,
                f"='现金流量表正表'!F{excel_row_by_id['NET-CASH']}-'现金流量表正表'!F{excel_row_by_id['FX']}",
                formats["money"],
                classified_net_cent / 100,
            )
            review_cash_change = _review_col("批次现金变化金额", absolute=True)
            review_status = _review_col("人工处理状态", absolute=True)
            review_baseline = _review_col("原基线项目(技术)", absolute=True)
            review_end = max(2, len(model.review_batches) + 1)
            low_end = max(2, low_amount_last)
            pending_formula = (
                f'=SUMIFS(\'重要待复核事项\'!{review_cash_change}$2:{review_cash_change}${review_end},'
                f'\'重要待复核事项\'!{review_baseline}$2:{review_baseline}${review_end},"",'
                f'\'重要待复核事项\'!{review_status}$2:{review_status}${review_end},"等待人工处理")'
                f'+SUMIFS(\'低金额批量处理\'!{review_cash_change}$2:{review_cash_change}${low_end},'
                f'\'低金额批量处理\'!{review_baseline}$2:{review_baseline}${low_end},"",'
                f'\'低金额批量处理\'!{review_status}$2:{review_status}${low_end},"等待人工处理")'
            )
            cash_sheet.write_formula(
                pending_row - 1,
                2,
                pending_formula,
                formats["money"],
                pending_net_cent / 100,
            )
            cash_sheet.write_formula(
                bridge_row - 1,
                2,
                f'=ROUND(C{closing_row}-C{opening_row}-C{classified_row}-C{pending_row}-C{fx_row}-C{adjustment_row},2)',
                formats["money"],
                int(bridge_difference_cent or 0) / 100,
            )
            cash_sheet.write_formula(
                final_row - 1,
                2,
                f'=ROUND(C{closing_row}-C{opening_row}-C{classified_row}-C{fx_row}-C{adjustment_row},2)',
                formats["money"],
                int(final_difference_cent or 0) / 100,
            )
            cash_sheet.write_formula(
                bridge_row - 1,
                1,
                f'=IF(C{bridge_row}<>0,"现金变动桥接存在无法解释差异",IF(C{final_row}=0,"最终现金流量表勾稽成功","现金变动桥接相符、现金流量表尚待分类"))',
                formats["pending"],
                model.reconciliation.status,
            )
        trace_sheet = sheets["全量分类留痕"]
        _write_dict_rows(trace_sheet, trace_rows, formats, "没有现金流业务组成。")
        if trace_rows and "最终决定项目" in trace_headers:
            final_column = trace_headers.index("最终决定项目")
            base_item_column = trace_headers.index("人工改选基准项目(技术)")
            base_amount_column = trace_headers.index("人工改选基准金额(技术)")
            target_amount_column = trace_headers.index("人工改选目标金额(技术)")
            active_column = trace_headers.index("人工改选生效标志(技术)")
            amount_column = (
                trace_headers.index("本行分配现金变化")
                if "本行分配现金变化" in trace_headers
                else None
            )
            helper_name_column = len(trace_headers) + 1
            helper_name = xl_col_to_name(helper_name_column)
            options = tuple(
                sorted(
                    (item for item in model.rules.statement_items if item.is_leaf),
                    key=lambda item: item.display_order,
                )
            )
            option_names = {item.name for item in options}
            for helper_index, item in enumerate(options):
                helper_column = helper_name_column + helper_index
                trace_sheet.write(0, helper_column, item.name, formats["text"])
                trace_sheet.write_number(
                    1,
                    helper_column,
                    1 if item.normal_direction == "inflow" else -1,
                    formats["text"],
                )
            helper_last_column = helper_name_column + len(options)
            helper_last = xl_col_to_name(helper_last_column)
            trace_sheet.write(0, helper_last_column, "明确排除", formats["text"])
            trace_sheet.write_number(1, helper_last_column, 0, formats["text"])
            option_range = f"${helper_name}$1:${helper_last}$1"
            lookup_range = f"${helper_name}$1:${helper_last}$2"
            trace_sheet.data_validation(
                1,
                final_column,
                len(trace_rows),
                final_column,
                {
                    "validate": "list",
                    "source": option_range,
                    "error_type": "stop",
                    "error_title": "项目无效",
                    "error_message": "请从下拉菜单选择现金流量表项目或明确排除。",
                    "input_title": "人工最终选择",
                    "input_message": "默认显示当前决定；需要调整时从下拉菜单改选。",
                },
            )
            trace_sheet.set_column(
                helper_name_column,
                helper_last_column,
                20,
                None,
                {"hidden": True},
            )
            review_choice = _review_col("人工确认项目", absolute=True)
            review_system_item = _review_col("系统项目(技术)", absolute=True)
            review_components = _review_col("业务组成编号(技术)", absolute=True)
            review_end = len(model.review_batches) + 1
            review_by_component = {
                component_id: (excel_row, batch)
                for excel_row, batch in enumerate(model.review_batches, 2)
                for component_id in batch.component_ids
            }
            for row_index, row in enumerate(trace_rows, 1):
                current_decision = str(row.get("最终决定项目") or "")
                component_id = str(row.get("业务组成编号(技术)") or "")
                review_match = review_by_component.get(component_id)
                if current_decision == "等待人工复核" and review_match is not None:
                    _, batch = review_match
                    safe_component_id = component_id.replace("~", "~~").replace("*", "~*").replace("?", "~?").replace('"', '""')
                    match_expression = (
                        f'MATCH("*、{safe_component_id}、*",'
                        f"'重要待复核事项'!{review_components}$2:{review_components}${review_end},0)"
                    )
                    selected_expression = (
                        f"INDEX('重要待复核事项'!{review_choice}$2:{review_choice}${review_end},"
                        f"{match_expression})"
                    )
                    system_expression = (
                        f"INDEX('重要待复核事项'!{review_system_item}$2:{review_system_item}${review_end},"
                        f"{match_expression})"
                    )
                    resolved_expression = (
                        f'IF({selected_expression}="{USE_SYSTEM_RECOMMENDATION}",'
                        f'{system_expression},{selected_expression})'
                    )
                    trace_sheet.write_formula(
                        row_index,
                        final_column,
                        f'=IFERROR(IF({selected_expression}="","等待人工复核",{resolved_expression}),"等待人工复核")',
                        formats["input"],
                        "等待人工复核",
                    )
                    baseline_name = item_name_by_id.get(batch.baseline_item_code, "")
                    safe_baseline = baseline_name.replace('"', '""')
                    trace_sheet.write_formula(
                        row_index,
                        base_item_column,
                        f'=IFERROR(IF({selected_expression}="","{safe_baseline}",{resolved_expression}),"{safe_baseline}")',
                        formats["text"],
                        baseline_name,
                    )
                else:
                    trace_sheet.write(
                        row_index,
                        final_column,
                        current_decision,
                        formats["input"],
                    )
                    fallback = str(row.get("原项目标准化结果") or "")
                    baseline_name = (
                        current_decision
                        if current_decision in option_names
                        or current_decision == "明确排除"
                        else fallback
                        if fallback in option_names
                        else ""
                    )
                    trace_sheet.write(
                        row_index,
                        base_item_column,
                        baseline_name,
                        formats["text"],
                    )
                excel_row = row_index + 1
                final_cell = f"{xl_col_to_name(final_column)}{excel_row}"
                base_item_cell = f"{xl_col_to_name(base_item_column)}{excel_row}"
                cash_cell = (
                    "0"
                    if amount_column is None
                    else f"{xl_col_to_name(amount_column)}{excel_row}"
                )
                trace_sheet.write_formula(
                    row_index,
                    base_amount_column,
                    f'=IFERROR({cash_cell}*HLOOKUP({base_item_cell},{lookup_range},2,FALSE),0)',
                    formats["money"],
                    0,
                )
                trace_sheet.write_formula(
                    row_index,
                    target_amount_column,
                    f'=IFERROR({cash_cell}*HLOOKUP({final_cell},{lookup_range},2,FALSE),0)',
                    formats["money"],
                    0,
                )
                trace_sheet.write_formula(
                    row_index,
                    active_column,
                    f'=IF(AND(COUNTIF({option_range},{final_cell})=1,{final_cell}<>{base_item_cell}),1,0)',
                    formats["text"],
                    0,
                )
        hidden_trace_headers = {
            "命中规则(技术)",
            "业务组成编号(技术)",
            "来源占用键(技术)",
            "业务组编号(技术)",
            "决策来源(技术)",
            "原始行编号(技术)",
            "来源文件",
            "来源工作表",
            "来源行号",
            "来源单元格",
            "本行完整对方科目路径",
            "中间层级",
            "末级明细",
            "映射状态",
            "一级科目映射候选",
            "一级科目映射依据",
            "现金方向依据",
            "原现流项目",
            "系统候选项目",
            "判断理由",
            "摘要来源质量",
            "完整路径来源质量",
            "两个来源是否独立",
            "证据质量说明",
            "证据得分",
            "单笔金额",
            "强制检查",
            "异常",
            "AI复核过程",
            "本行分配现金变化",
            "组成明细",
            "评分版本",
            "动作表版本",
        }
        for column, header in enumerate(trace_headers):
            if (
                header in hidden_trace_headers
                or header.endswith("(技术)")
                or header.endswith("（技术）")
            ):
                trace_sheet.set_column(column, column, 20, None, {"hidden": True})
        _write_dict_rows(
            sheets["科目语义词典"],
            model.dictionary_rows,
            formats,
            "本次没有需要展示的完整对方科目路径。",
        )
        _write_dict_rows(
            sheets["同类检查"],
            model.consistency_rows,
            formats,
            "本期未发现相同原始来源却形成不同项目的事项。",
        )
        _write_dict_rows(sheets["输入识别与字段映射"], model.mapping_rows, formats, "没有字段映射记录。")
        sheets["输入识别与字段映射"].hide()
    finally:
        workbook.close()
    return target


def validate_output_workbook(path: Path, model: WorkbookModel) -> WorkbookValidation:
    errors: list[str] = []
    workbook = load_workbook(path, data_only=False, keep_links=True)
    try:
        if tuple(workbook.sheetnames) != SHEET_NAMES:
            errors.append("工作表名称或顺序不正确")
        hidden_sheets = {
            sheet.title
            for sheet in workbook.worksheets
            if sheet.sheet_state == "hidden"
        }
        if hidden_sheets != {"AI复核记录", "输入识别与字段映射"}:
            errors.append("机器工作表的默认隐藏状态不正确")
        if workbook._external_links:
            errors.append("工作簿包含外部链接")
        difference = workbook["原表与系统决定差异"]
        if tuple(cell.value for cell in difference[1]) != DIFFERENCE_HEADERS:
            errors.append("原表与系统决定差异表头不正确")
        if difference.freeze_panes != "A2":
            errors.append("原表与系统决定差异冻结窗格不正确")
        if model.difference_rows and difference.auto_filter.ref is None:
            errors.append("原表与系统决定差异未设置筛选")
        if not difference.protection.sheet:
            errors.append("原表与系统决定差异未设置只读保护")
        main = workbook["现金流量表正表"]
        if main.freeze_panes != "A4":
            errors.append("正表冻结窗格不正确")
        if main.auto_filter.ref is None:
            errors.append("正表未设置筛选")
        formulas = [
            cell.value
            for row in main.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        ]
        if not any("重要待复核事项" in formula for formula in formulas):
            errors.append("正表未引用重要待复核事项调整层")
        if not any("疑似重复事项" in formula for formula in formulas):
            errors.append("正表未引用疑似重复事项调整层")
        if any("[" in formula for formula in formulas):
            errors.append("正表公式引用了外部工作簿")
        trace_supports_manual_choice = any(
            "最终决定项目" in row for row in model.trace_rows
        )
        if trace_supports_manual_choice and not any(
            "全量分类留痕" in formula for formula in formulas
        ):
            errors.append("正表未引用全量分类留痕人工改选层")
        if any("原表与系统决定差异" in formula for formula in formulas):
            errors.append("正表公式引用了原表与系统决定差异")
        status_formulas = [
            cell.value
            for row in workbook["使用说明与状态"].iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        ]
        if any("原表与系统决定差异" in formula for formula in status_formulas):
            errors.append("首页状态公式引用了原表与系统决定差异")
        for index, item in enumerate(sorted(model.rules.statement_items, key=lambda value: value.display_order), 4):
            actual = main.cell(index, 4).value
            expected = model.statement.values[item.item_id] / 100
            if actual != expected:
                errors.append(f"自动基线不一致：{item.item_id}")
                break
        manual_batches = tuple(
            (index + 2, batch)
            for index, batch in enumerate(model.review_batches)
            if not batch.follows_component_id
        )
        if manual_batches and not workbook["重要待复核事项"].data_validations.dataValidation:
            errors.append("重要待复核事项缺少下拉选择")
        review_sheet = workbook["重要待复核事项"]
        if tuple(cell.value for cell in review_sheet[1][: len(REVIEW_HEADERS)]) != REVIEW_HEADERS:
            errors.append("重要待复核事项表头不完整或顺序不正确")
        # 设计第五节：强制人工复核批次的行必须使用区域引用下拉（内联列表会超 255 字符上限）
        mandatory_excel_rows = [
            index + 2
            for index, batch in enumerate(model.review_batches)
            if batch.mandatory and not batch.follows_component_id
        ]
        if mandatory_excel_rows:
            validations = workbook["重要待复核事项"].data_validations.dataValidation
            choice_column = _review_col("人工确认项目")
            for excel_row in mandatory_excel_rows:
                if not any(
                    "$" in str(validation.formula1 or "")
                    and f"{choice_column}{excel_row}" in str(validation.sqref)
                    for validation in validations
                ):
                    errors.append(f"强制人工复核批次第 {excel_row} 行未使用区域引用下拉")
                    break
        if model.duplicate_groups and not workbook["疑似重复事项"].data_validations.dataValidation:
            errors.append("疑似重复事项缺少下拉选择")
        if manual_batches:
            choice_column = _review_col("人工确认项目")
            first_manual_row = manual_batches[0][0]
            fill = workbook["重要待复核事项"][f"{choice_column}{first_manual_row}"].fill.fgColor.rgb
            if fill not in {"FFDDEBF7", "DDEBF7"}:
                errors.append("人工输入单元格未使用蓝色标识")
    finally:
        workbook.close()
    try:
        cached_workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in cached_workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        _ = cell.value
        finally:
            cached_workbook.close()
    except Exception as error:
        errors.append(f"工作簿公式缓存无法读取：{error}")
    return WorkbookValidation(not errors, tuple(errors))
