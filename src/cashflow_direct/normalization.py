from __future__ import annotations

import re
from collections import Counter
from collections.abc import Sequence
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from cashflow_direct.models import EvidenceProfile, NormalizedEntry, SourceLocator
from cashflow_direct.money import stable_id, yuan_to_cent
from cashflow_direct.semantic_mapping import DatasetMapping
from cashflow_direct.workbook_structure import open_workbook_robust


CASH_ACCOUNT_TERMS = ("库存现金", "银行存款", "其他货币资金", "现金等价物")


@dataclass(frozen=True, slots=True)
class RowExclusion:
    source: SourceLocator
    discard_reason: str


@dataclass(frozen=True, slots=True)
class RowError:
    source: SourceLocator
    message: str


@dataclass(frozen=True, slots=True)
class NormalizationResult:
    entries: tuple[NormalizedEntry, ...]
    profile: EvidenceProfile
    exclusions: tuple[RowExclusion, ...]
    errors: tuple[RowError, ...]
    rows_read: int


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _date_text(value: object) -> str:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return _text(value)


def _cell_value(row: tuple[object, ...], mapping: DatasetMapping, role: str) -> object:
    column = mapping.role_to_column.get(role)
    if column is None or column.column_index > len(row):
        return None
    return row[column.column_index - 1]


def _row_locator(file_id: str, sheet_name: str, row_number: int, columns: list[int]) -> SourceLocator:
    first = min(columns) if columns else 1
    last = max(columns) if columns else 1
    return SourceLocator(
        file_id,
        sheet_name,
        row_number,
        row_number,
        f"{get_column_letter(first)}{row_number}:{get_column_letter(last)}{row_number}",
    )


def _error_locator(file_id: str, sheet_name: str, row_number: int, column: int) -> SourceLocator:
    cell = f"{get_column_letter(column)}{row_number}"
    return SourceLocator(file_id, sheet_name, row_number, row_number, cell)


def _is_repeated_header(row: tuple[object, ...], mapping: DatasetMapping) -> bool:
    matches = 0
    for column in mapping.role_to_column.values():
        if column.column_index <= len(row) and column.header_path:
            if _text(row[column.column_index - 1]).replace(" ", "") == column.header_path[-1].replace(" ", ""):
                matches += 1
    return matches >= 3


def _is_total_row(summary: str, first_cell: object) -> bool:
    totals = {"合计", "总计", "本页合计", "本月合计", "本年累计"}
    normalized = lambda value: re.sub(r"[\s：:]+", "", _text(value))
    return normalized(summary) in totals or normalized(first_cell) in totals


def _retained_side(account_name: str) -> str:
    if not account_name:
        return "unknown"
    return "cash" if any(term in account_name for term in CASH_ACCOUNT_TERMS) else "counterpart"


def infer_evidence_profile(entries: Sequence[NormalizedEntry]) -> EvidenceProfile:
    voucher_counts = Counter(entry.voucher_key for entry in entries)
    has_accounts = any(entry.account_name for entry in entries)
    matched = any(entry.counterpart_name for entry in entries)
    has_flow_item = any(entry.original_flow_item for entry in entries)
    label_sides = frozenset(entry.label_side for entry in entries if entry.label_side != "unknown")
    retained_sides = frozenset(
        entry.retained_side for entry in entries if entry.retained_side != "unknown"
    )
    return EvidenceProfile(
        full_voucher=has_accounts and any(count > 1 for count in voucher_counts.values()),
        matched_counterparty=matched,
        has_flow_item=has_flow_item,
        label_sides=label_sides,
        retained_side_values=retained_sides,
        has_flow_amount=any(entry.flow_amount_cent != 0 for entry in entries),
        summary_only=bool(entries) and all(
            not entry.account_name and not entry.counterpart_name for entry in entries
        ),
        split_duplication_risk=matched and has_flow_item,
    )


def normalize_dataset(path: Path, file_id: str, mapping: DatasetMapping) -> NormalizationResult:
    """把不同证据形态统一为逐行分录，并完整记录排除和错误。"""
    workbook = open_workbook_robust(path)
    entries: list[NormalizedEntry] = []
    exclusions: list[RowExclusion] = []
    errors: list[RowError] = []
    rows_read = 0
    try:
        worksheet = workbook[mapping.sheet_name]
        columns = [item.column_index for item in mapping.role_to_column.values()]
        merged_values: dict[tuple[int, int], object] = {}
        for merged in mapping.merged_ranges:
            min_col, min_row, max_col, max_row = range_boundaries(merged)
            if max_row <= mapping.header_row_end:
                continue
            anchor = worksheet.cell(min_row, min_col).value
            for row_index in range(max(min_row, mapping.header_row_end + 1), max_row + 1):
                for column_index in range(min_col, max_col + 1):
                    merged_values[(row_index, column_index)] = anchor
        previous_voucher_date = previous_voucher_no = ""
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=mapping.header_row_end + 1, values_only=True),
            mapping.header_row_end + 1,
        ):
            logical_row = list(row)
            for column_index in columns:
                if column_index > len(logical_row):
                    logical_row.extend([None] * (column_index - len(logical_row)))
                if logical_row[column_index - 1] in (None, ""):
                    logical_row[column_index - 1] = merged_values.get((row_number, column_index))
            row = tuple(logical_row)
            rows_read += 1
            source = _row_locator(file_id, worksheet.title, row_number, columns)
            if all(value in (None, "") for value in row):
                exclusions.append(RowExclusion(source, "blank_structure"))
                continue
            if (
                "voucher_date" in mapping.role_to_column
                and "voucher_no" in mapping.role_to_column
                and _cell_value(tuple(row), mapping, "voucher_date") in (None, "")
                and _cell_value(tuple(row), mapping, "voucher_no") in (None, "")
            ):
                # 明细表中无日期且无凭证号的行视为小计/汇总行，剔除并留痕
                exclusions.append(RowExclusion(source, "subtotal_row"))
                continue
            if _is_repeated_header(tuple(row), mapping):
                exclusions.append(RowExclusion(source, "repeated_header"))
                continue
            summary = _text(_cell_value(tuple(row), mapping, "summary"))
            if _is_total_row(summary, row[0] if row else None):
                exclusions.append(RowExclusion(source, "total_row"))
                continue

            money: dict[str, int] = {}
            source_money: dict[str, int | None] = {}
            failed = False
            for role in ("debit", "credit", "flow_amount"):
                column = mapping.role_to_column.get(role)
                value = _cell_value(tuple(row), mapping, role)
                if column is None or value in (None, ""):
                    money[role] = 0
                    source_money[role] = None
                    continue
                try:
                    money[role] = yuan_to_cent(value)
                    source_money[role] = money[role]
                except ValueError:
                    errors.append(
                        RowError(
                            _error_locator(file_id, worksheet.title, row_number, column.column_index),
                            f"金额无法识别：{value!r}",
                        )
                    )
                    failed = True
                    break
            if failed:
                continue

            direction = _text(_cell_value(tuple(row), mapping, "direction"))
            if "direction" in mapping.role_to_column and "flow_amount" in mapping.role_to_column:
                if direction.startswith("借"):
                    money["debit"] = money["flow_amount"]
                elif direction.startswith("贷"):
                    money["credit"] = money["flow_amount"]
                else:
                    column = mapping.role_to_column["direction"].column_index
                    errors.append(
                        RowError(
                            _error_locator(file_id, worksheet.title, row_number, column),
                            f"借贷方向无法识别：{direction!r}",
                        )
                    )
                    continue

            account = _text(_cell_value(tuple(row), mapping, "account_name"))
            account_code = _text(_cell_value(tuple(row), mapping, "account_code"))
            counterpart = _text(_cell_value(tuple(row), mapping, "counterpart_name"))
            retained_side = _retained_side(account)
            original_item = _text(_cell_value(tuple(row), mapping, "flow_item"))
            label_side = retained_side if original_item else "unknown"
            voucher_date = _date_text(_cell_value(tuple(row), mapping, "voucher_date"))
            voucher_word = _text(_cell_value(tuple(row), mapping, "voucher_word"))
            voucher_no = _text(_cell_value(tuple(row), mapping, "voucher_no"))
            if "voucher_date" in mapping.role_to_column and not voucher_date:
                voucher_date = previous_voucher_date
            if "voucher_no" in mapping.role_to_column and not voucher_no:
                voucher_no = previous_voucher_no
            if voucher_date:
                previous_voucher_date = voucher_date
            if voucher_no:
                previous_voucher_no = voucher_no
            voucher_key = stable_id(
                "VCH", file_id, worksheet.title, voucher_date, voucher_no or row_number
            )
            entries.append(
                NormalizedEntry(
                    entry_id=stable_id("ENT", file_id, worksheet.title, row_number),
                    source=source,
                    voucher_key=voucher_key,
                    voucher_date=voucher_date,
                    voucher_no=voucher_no,
                    summary=summary,
                    account_name=account,
                    counterpart_name=counterpart,
                    debit_cent=money["debit"],
                    credit_cent=money["credit"],
                    flow_amount_cent=money["flow_amount"],
                    original_flow_item=original_item,
                    label_side=label_side,
                    retained_side=retained_side,
                    voucher_word=voucher_word,
                    account_code=account_code,
                    source_debit_cent=source_money["debit"],
                    source_credit_cent=source_money["credit"],
                    source_flow_amount_cent=source_money["flow_amount"],
                )
            )
    finally:
        workbook.close()

    frozen_entries = tuple(entries)
    return NormalizationResult(
        frozen_entries,
        infer_evidence_profile(frozen_entries),
        tuple(exclusions),
        tuple(errors),
        rows_read,
    )


def subtotal_exclusion_warning(result: NormalizationResult) -> dict[str, object] | None:
    """小计剔除占比超10%时返回核对报告提示，否则 None。"""
    if not result.rows_read:
        return None
    count = sum(1 for item in result.exclusions if item.discard_reason == "subtotal_row")
    if count / result.rows_read <= 0.10:
        return None
    return {
        "kind": "提示",
        "message": f"小计行剔除占比异常：{count}/{result.rows_read} 行被按小计剔除，请人工确认未误删明细",
    }
