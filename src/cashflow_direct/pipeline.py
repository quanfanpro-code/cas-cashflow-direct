from __future__ import annotations

import hashlib
import json
import re
from collections.abc import Mapping, Sequence
from dataclasses import asdict, dataclass, replace
from datetime import datetime, timezone
from pathlib import Path
from time import sleep

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string

from cashflow_direct.account_mapping import (
    AccountMappingRecord,
    build_account_mappings,
    load_standard_accounts,
    resolve_account_mappings,
    standardize_entries,
)
from cashflow_direct.account_dictionary import (
    AccountDictionary,
    AccountNodeConcept,
    AccountPathRelation,
    AccountPathSemanticResult,
    AccountPathSlot,
    AccountSemanticEntry,
    analyze_account_path,
    build_account_agent_task,
    load_account_semantic_rules,
    load_common_dictionary,
    merge_account_agent_concepts,
    merge_dictionaries,
    split_account_levels,
)
from cashflow_direct.ai_review import (
    build_blind_ai_tasks,
    company_note_applies,
    company_note_is_active,
    chunk_ai_tasks,
    merge_structured_ai_results,
    resolve_structured_ai_results,
    review_text_pattern as _review_text_pattern,
    structured_ai_result_from_mapping,
    validate_basis_text,
    write_ai_tasks_jsonl,
)
from cashflow_direct.classification import (
    classify_all,
    load_rule_pack,
    route_classification_decisions,
)
from cashflow_direct.component_structure_ai import (
    StructureAIResult,
    StructureAITask,
    build_structure_ai_tasks,
    resolve_structure_ai_request,
    validate_structure_ai_results,
)
from cashflow_direct.components import (
    CashScope,
    ComponentSourceAllocation,
    InternalTransferLeg,
    build_cashflow_components,
    compute_rough_reconciliation,
    confirm_cash_scope as make_cash_scope,
    discover_cash_scope,
    find_cash_row_cleanup_requests,
    flow_direction_source,
)
from cashflow_direct.consistency import (
    apply_consistency_forced_checks,
    find_consistency_groups,
)
from cashflow_direct.decision_policy import (
    DEFAULT_AUTOMATIC_CHANGE_SCORE,
    EvidenceQuality,
    materiality_level,
    validate_automatic_change_threshold,
)
from cashflow_direct.duplicates import assign_duplicate_items, find_suspected_duplicates
from cashflow_direct.differences import build_original_auto_differences
from cashflow_direct.excel_recalculation import recalculate_workbook_with_excel
from cashflow_direct.intake import register_inputs, validate_materiality
from cashflow_direct.models import (
    CashflowComponent,
    ClassificationDecision,
    AITask,
    MaterialityAmounts,
    NormalizedEntry,
    EvidenceProfile,
    SourceLocator,
    UnresolvedDecision,
)
from cashflow_direct.materiality import build_review_batches
from cashflow_direct.money import stable_id, statement_amount_cent, yuan_to_cent
from cashflow_direct.normalization import (
    infer_evidence_profile,
    normalize_dataset,
    subtotal_exclusion_warning,
)
from cashflow_direct.semantic_mapping import (
    DatasetMapping,
    MappingQuestion,
    infer_dataset_mappings,
)
from cashflow_direct.statement import (
    ExistingStatementResult,
    aggregate_statement,
    build_statement_layers,
    compare_statement,
    detect_statement_sheets,
    internal_transfer_statement_adjustments,
    parse_existing_statement,
    reconcile_cash,
)
from cashflow_direct.summary_semantics import (
    SummarySemanticResult,
    SummarySpan,
    analyze_summary,
    build_summary_agent_task,
    load_summary_rules,
    merge_summary_agent_slots,
    validate_summary_batch,
)
from cashflow_direct.trace_output import build_trace_rows
from cashflow_direct.storage import RunStore
from cashflow_direct.validation import (
    validate_classification,
    validate_final_readiness,
    validate_final_output,
    validate_input_hashes,
    validate_statement,
)
from cashflow_direct.versions import assert_current_versions, current_versions
from cashflow_direct.workbook_output import WorkbookModel, build_output_workbook
from cashflow_direct.workbook_structure import open_workbook_robust, scan_workbook


def _dictionary_display_result(
    confirmed: Mapping[str, Mapping[str, object]],
    standard_path: str,
    raw_path: str,
) -> Mapping[str, object]:
    """科目语义工作表按已确认的完整路径取值，不用末级短名串行。"""
    return confirmed.get(standard_path) or confirmed.get(raw_path) or {}


PROJECT_ROOT = Path(__file__).resolve().parents[2]
TRACE_DIR_NAME = "计算留痕数据"
STATE_FILE_NAME = "运行状态.json"
DB_FILE_NAME = "计算留痕.sqlite3"
# A4 输出中文化：人类可读列一律中文大白话（技术英文值只进隐藏技术列）
_EVIDENCE_TIER_TEXT = {"high": "高", "medium": "中", "low": "低"}
_ANOMALY_TEXT = {
    "internal_transfer": "内部划转",
    "non_cash": "非现金事项",
    "accrual_with_cash_leg": "权责发生制事项（含现金腿）",
    "netting_suspect": "疑似净额结算",
    "voucher_unbalanced": "凭证借贷不平衡",
    "unallocated_cash": "现金未分配",
    "cash_allocation_mismatch": "现金分配不符",
}
_DECISION_SOURCE_TEXT = {
    "system": "系统规则",
    "ai_agreement": "AI复核一致",
    "ai_adjudication": "AI裁决",
    "ai_conflict": "AI仍有分歧",
    "consistency_review": "一致性复核",
    "consistency_adjudication": "一致性裁决",
    "manual": "人工确认",
}
@dataclass(frozen=True, slots=True)
class PreflightResult:
    run_id: str
    run_dir: Path
    status: str
    recommended_cash_decisions: dict[str, str]
    mapping_question_count: int
    source_entry_count: int


@dataclass(frozen=True, slots=True)
class StageResult:
    run_id: str
    run_dir: Path
    stage: str
    status: str
    next_action: str


@dataclass(frozen=True, slots=True)
class ClassificationStageResult:
    run_id: str
    run_dir: Path
    component_count: int
    component_hash: str
    source_entry_count: int
    cash_delta_cent: int
    ai_tasks_missing: int
    status: str


@dataclass(frozen=True, slots=True)
class AIStageResult:
    run_id: str
    run_dir: Path
    valid_count: int
    missing_count: int
    status: str


@dataclass(frozen=True, slots=True)
class FinalizeResult:
    run_id: str
    run_dir: Path
    workbook_path: Path
    overall_status: str


def _trace_dir(run_dir: Path) -> Path:
    return Path(run_dir) / TRACE_DIR_NAME


def _state_path(run_dir: Path) -> Path:
    return _trace_dir(run_dir) / STATE_FILE_NAME


def _store(run_dir: Path) -> RunStore:
    return RunStore(_trace_dir(run_dir) / DB_FILE_NAME)


def _load_state(run_dir: Path) -> dict[str, object]:
    path = _state_path(run_dir)
    if not path.is_file():
        raise RuntimeError("未找到运行状态，请先执行资料预检")
    with path.open("r", encoding="utf-8-sig") as source:
        state = json.load(source)
    assert_current_versions(state.get("versions"), PROJECT_ROOT)
    return state


def _replace_with_windows_retry(source: Path, target: Path) -> Path:
    for attempt in range(10):
        try:
            return source.replace(target)
        except PermissionError:
            if attempt == 9:
                raise
            sleep(min(0.05 * (2**attempt), 2.0))
    raise AssertionError("文件替换重试次数计算错误")


def _save_state(run_dir: Path, state: Mapping[str, object]) -> None:
    target = _state_path(run_dir)
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_suffix(".json.tmp")
    with temporary.open("w", encoding="utf-8-sig", newline="\n") as output:
        json.dump(state, output, ensure_ascii=False, separators=(",", ":"))
    _replace_with_windows_retry(temporary, target)


def _write_trace_jsonl(run_dir: Path, filename: str, records: Sequence[object]) -> None:
    target = _trace_dir(run_dir) / filename
    with target.open("w", encoding="utf-8-sig", newline="\n") as output:
        for record in records:
            payload = asdict(record) if hasattr(record, "__dataclass_fields__") else record
            output.write(json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + "\n")


def _write_trace_json(run_dir: Path, filename: str, payload: object) -> None:
    target = _trace_dir(run_dir) / filename
    temporary = target.with_suffix(f"{target.suffix}.tmp")
    with temporary.open("w", encoding="utf-8-sig", newline="\n") as output:
        json.dump(payload, output, ensure_ascii=False, indent=2)
    _replace_with_windows_retry(temporary, target)


def _entry_from_dict(payload: Mapping[str, object]) -> NormalizedEntry:
    data = dict(payload)
    data["source"] = SourceLocator(**data["source"])
    return NormalizedEntry(**data)


def _component_from_dict(payload: Mapping[str, object]) -> CashflowComponent:
    data = dict(payload)
    for key in (
        "counterpart_accounts",
        "source_keys",
        "anomalies",
        "source_file_ids",
        "original_counterpart_accounts",
    ):
        data[key] = tuple(data.get(key, ()))
    return CashflowComponent(**data)


def _decision_from_dict(payload: Mapping[str, object]) -> ClassificationDecision:
    data = dict(payload)
    data["excluded_conflict_rule_ids"] = tuple(data.get("excluded_conflict_rule_ids", ()))
    raw_score = data.get("evidence_score", 0)
    data["evidence_score"] = None if raw_score is None else int(raw_score)
    data["evidence_sources"] = tuple(data.get("evidence_sources", ()))
    data["candidate_item_ids"] = tuple(data.get("candidate_item_ids", ()))
    return ClassificationDecision(**data)


def _evidence_assessment_payload(
    decision: ClassificationDecision,
) -> dict[str, object]:
    """保存评分结论及所有会改变路由的强制检查事实。"""
    return {
        "component_id": decision.component_id,
        "candidate_item_ids": decision.candidate_item_ids,
        "summary_quality": decision.summary_quality,
        "account_path_quality": decision.account_path_quality,
        "sources_independent": decision.sources_independent,
        "source_conflict": decision.source_conflict,
        "business_conflict": decision.business_conflict,
        "company_rule_conflict": decision.company_rule_conflict,
        "vat_base_missing": decision.vat_base_missing,
        "net_item_facts_missing": decision.net_item_facts_missing,
        "direction_status": decision.direction_status,
        "evidence_score": decision.evidence_score,
    }


def _ai_records_from_state(
    state: Mapping[str, object],
) -> tuple[dict[str, object], ...]:
    """汇总有效结果和技术失败终态，供最终工作簿完整留痕。"""
    groups = (
        (
            "分类AI有效结果",
            state.get("structured_ai_validation", {}).get("valid_results", ()),
        ),
        ("分类AI技术失败", state.get("ai_technical_failure_log", ())),
        (
            "业务组成结构AI有效结果",
            state.get("component_structure_ai_results", ()),
        ),
        (
            "业务组成结构AI技术失败",
            state.get("component_structure_ai_technical_failure_log", ()),
        ),
    )
    return tuple(
        {"阶段": stage_name, **dict(item)}
        for stage_name, records in groups
        for item in records
    )


def _ai_task_from_dict(payload: Mapping[str, object]) -> AITask:
    data = dict(payload)
    data["candidate_item_ids"] = tuple(data.get("candidate_item_ids", ()))
    return AITask(**data)


def _scope_from_dict(payload: Mapping[str, object]) -> CashScope:
    return CashScope(
        frozenset(payload["included_keys"]),
        frozenset(payload["excluded_keys"]),
        tuple((item[0], tuple(item[1])) for item in payload["account_names_by_key"]),
        str(payload["scope_hash"]),
    )


def _materiality_from_state(state: Mapping[str, object]) -> MaterialityAmounts:
    return MaterialityAmounts(**state["materiality"])


def _prepare_account_mapping_state(
    state: dict[str, object], entries: Sequence[NormalizedEntry]
) -> tuple[AccountMappingRecord, ...]:
    paths = tuple(
        path
        for entry in entries
        for path in (entry.account_name, entry.counterpart_name)
        if path.strip()
    )
    records = build_account_mappings(paths, load_standard_accounts(PROJECT_ROOT))
    state["account_mapping_records"] = [asdict(item) for item in records]
    state["account_mapping_questions"] = [
        {
            "original_level1": item.original_level1,
            "candidate_standard_names": list(item.candidate_standard_names),
            "basis": item.basis,
        }
        for item in records
        if item.status != "confirmed"
    ]
    return records


def _account_mapping_records_from_state(
    state: Mapping[str, object],
) -> tuple[AccountMappingRecord, ...]:
    records = tuple(
        AccountMappingRecord(**item)
        for item in state.get("account_mapping_records", ())
    )
    if not records or any(item.status != "confirmed" for item in records):
        raise RuntimeError("一级科目映射未全部确认，请先完成一级科目确认")
    return records


def _standardized_entries_from_state(
    state: Mapping[str, object],
) -> tuple[NormalizedEntry, ...]:
    raw_entries = tuple(_entry_from_dict(item) for item in state["entries"])
    records = _account_mapping_records_from_state(state)
    return standardize_entries(
        raw_entries,
        {item.original_level1: item for item in records},
    )


def _prepare_cash_scope_state(state: dict[str, object]) -> dict[str, str]:
    proposal = discover_cash_scope(_standardized_entries_from_state(state))
    recommended = {
        candidate.account_key: candidate.system_suggestion
        for candidate in proposal.candidates
        if candidate.system_suggestion in {"include", "exclude"}
    }
    state["cash_scope_proposal"] = asdict(proposal)
    state["recommended_cash_decisions"] = recommended
    return recommended


def _persist_ai_results(
    run_dir: Path,
    stage_name: str,
    results: Sequence[object],
) -> None:
    with _store(run_dir).stage(f"ai_result_{stage_name}") as connection:
        connection.executemany(
            "INSERT OR REPLACE INTO ai_result(record_id, payload_json) VALUES (?, ?)",
            (
                (
                    f"{stage_name}:{item.task_id}",
                    json.dumps({"阶段": stage_name, **asdict(item)}, ensure_ascii=False),
                )
                for item in results
            ),
        )


def _assert_inputs_unchanged(state: Mapping[str, object]) -> None:
    result = validate_input_hashes(state["files"])
    if not result.valid:
        raise RuntimeError("输入文件已被修改，请建立新运行目录后重新处理")


def _profile_to_dict(profile: EvidenceProfile) -> dict[str, object]:
    return {
        "full_voucher": profile.full_voucher,
        "matched_counterparty": profile.matched_counterparty,
        "has_flow_item": profile.has_flow_item,
        "label_sides": sorted(profile.label_sides),
        "retained_side_values": sorted(profile.retained_side_values),
        "has_flow_amount": profile.has_flow_amount,
        "summary_only": profile.summary_only,
        "split_duplication_risk": profile.split_duplication_risk,
    }


def _profile_from_dict(payload: Mapping[str, object]) -> EvidenceProfile:
    return EvidenceProfile(
        full_voucher=bool(payload["full_voucher"]),
        matched_counterparty=bool(payload["matched_counterparty"]),
        has_flow_item=bool(payload["has_flow_item"]),
        label_sides=frozenset(str(item) for item in payload["label_sides"]),
        retained_side_values=frozenset(str(item) for item in payload["retained_side_values"]),
        has_flow_amount=bool(payload["has_flow_amount"]),
        summary_only=bool(payload["summary_only"]),
        split_duplication_risk=bool(payload["split_duplication_risk"]),
    )


def _balances_from_existing(result: ExistingStatementResult) -> dict[str, int | None]:
    return {
        "opening_cent": result.values.get("CASH-OPENING"),
        "closing_cent": result.values.get("CASH-CLOSING"),
        "fx_cent": result.values.get("FX"),
    }


def _read_cash_balances(path: Path) -> dict[str, tuple[int, int]]:
    found: dict[str, tuple[int, int]] = {}
    workbook = open_workbook_robust(path)
    try:
        for sheet in workbook.worksheets:
            priority = 2 if "余额" in sheet.title else 1
            unit_text = "|".join(
                str(value)
                for row in sheet.iter_rows(min_row=1, max_row=20, max_col=10, values_only=True)
                for value in row
                if value not in (None, "")
            )
            multiplier = 10_000 if "万元" in unit_text else 1
            for row in sheet.iter_rows(min_row=1, max_row=100, max_col=10, values_only=True):
                for index, value in enumerate(row[:-1]):
                    if not isinstance(value, str) or row[index + 1] in (None, ""):
                        continue
                    key = None
                    if "期初现金及现金等价物余额" in value:
                        key = "opening_cent"
                    elif "期末现金及现金等价物余额" in value:
                        key = "closing_cent"
                    elif "汇率变动对现金及现金等价物的影响" in value:
                        key = "fx_cent"
                    if key is not None:
                        try:
                            amount = yuan_to_cent(row[index + 1]) * multiplier
                            if key not in found or priority > found[key][0]:
                                found[key] = (priority, amount)
                        except ValueError:
                            continue
    finally:
        workbook.close()
    return found


def _result_from_classification(state: Mapping[str, object], run_dir: Path) -> ClassificationStageResult:
    data = state["classification_summary"]
    return ClassificationStageResult(
        str(state["run_id"]),
        Path(run_dir),
        int(data["component_count"]),
        str(data["component_hash"]),
        int(data["source_entry_count"]),
        int(data["cash_delta_cent"]),
        int(data["ai_tasks_missing"]),
        str(data["status"]),
    )


def run_preflight(
    inputs: Sequence[Path],
    materiality: tuple[object, object, object],
    output_parent: Path | None = None,
    statement_path: Path | None = None,
    notes: str | None = None,
    automatic_change_threshold: int = DEFAULT_AUTOMATIC_CHANGE_SCORE,
) -> PreflightResult:
    automatic_change_threshold = validate_automatic_change_threshold(
        automatic_change_threshold
    )
    amounts = validate_materiality(*materiality)
    intake = register_inputs(inputs, output_parent=output_parent)
    run_dir = intake.run_dir
    trace_dir = _trace_dir(run_dir)
    trace_dir.mkdir(parents=True, exist_ok=True)
    store = _store(run_dir)
    store.initialize()
    rules = load_rule_pack(PROJECT_ROOT)

    entries: list[NormalizedEntry] = []
    mappings: list[dict[str, object]] = []
    questions: list[dict[str, object]] = []
    normalization_issues: list[dict[str, object]] = []
    sheet_structures: list[dict[str, object]] = []
    existing_statement_path: str | None = None
    profiles: dict[str, dict[str, object]] = {}
    designated_target = statement_path.resolve() if statement_path is not None else None
    designated_hit = False
    balance_candidates: dict[str, tuple[int, int]] = {}
    statement_candidates: list[dict[str, object]] = []
    mapped_dataset_sheets_by_file: dict[str, frozenset[str]] = {}
    for registered in intake.active_files:
        for key, candidate in _read_cash_balances(registered.path).items():
            if key not in balance_candidates or candidate[0] > balance_candidates[key][0]:
                balance_candidates[key] = candidate
        is_designated = designated_target is not None and registered.path.resolve() == designated_target
        # 第一遍：只收集余额、工作表结构、字段映射与明细归一化（疑似正表识别放第二遍，需明细日期年份）
        snapshot = scan_workbook(registered.path)
        sheet_structures.extend(
            {
                "file_id": registered.file_id,
                "sheet": sheet.name,
                "sample_rows": len(sheet.rows),
                "merged_ranges": sheet.merged_ranges,
                "hidden_columns": sheet.hidden_columns,
            }
            for sheet in snapshot.sheets
        )
        exclude_sheets = frozenset()
        if is_designated:
            # designated 文件归一化时需排除正表工作表；仅需工作表名，无需年份列选择，先用无年份识别取表名
            exclude_sheets = frozenset(
                name
                for name, result in detect_statement_sheets(registered.path, rules).items()
                if isinstance(result, ExistingStatementResult)
            )
        detected = infer_dataset_mappings(snapshot, exclude_sheets=exclude_sheets)
        mapped_dataset_sheets_by_file[registered.file_id] = frozenset(
            mapping.sheet_name for mapping in detected if isinstance(mapping, DatasetMapping)
        )
        for mapping in detected:
            if isinstance(mapping, DatasetMapping):
                normalized = normalize_dataset(registered.path, registered.file_id, mapping)
                entries.extend(normalized.entries)
                profiles[str(registered.file_id)] = _profile_to_dict(normalized.profile)
                mappings.append(
                    {
                        "file_id": registered.file_id,
                        "file": registered.path.name,
                        "sheet": mapping.sheet_name,
                        "header_rows": f"{mapping.header_row_start}:{mapping.header_row_end}",
                        "roles": {role: column.column_letter for role, column in mapping.role_to_column.items()},
                    }
                )
                normalization_issues.extend(
                    {
                        "file_id": registered.file_id,
                        "file": registered.path.name,
                        "sheet": issue.source.sheet_name,
                        "cell": issue.source.cell_range,
                        "kind": "错误" if hasattr(issue, "message") else "排除",
                        "message": getattr(issue, "message", getattr(issue, "discard_reason", "")),
                    }
                    for issue in (*normalized.errors, *normalized.exclusions)
                )
                normalization_issues.extend(
                    {
                        "file_id": registered.file_id,
                        "file": registered.path.name,
                        "sheet": issue.source.sheet_name,
                        "cell": issue.source.cell_range,
                        "kind": "警告",
                        "message": issue.message,
                    }
                    for issue in normalized.warnings
                )
                warning = subtotal_exclusion_warning(normalized)
                if warning is not None:
                    normalization_issues.append(
                        {
                            "file_id": registered.file_id,
                            "file": registered.path.name,
                            "sheet": mapping.sheet_name,
                            "cell": "",
                            **warning,
                        }
                    )
            else:
                questions.append(
                    {
                        "file_id": registered.file_id,
                        "file": registered.path.name,
                        "sheet": mapping.sheet_name,
                        "role": mapping.role,
                        "recommended": mapping.recommended.column_letter,
                    }
                )

    # 第二遍：entries 收齐后，由明细日期区间推断本期年份，再逐文件识别正表（A3 多时间列选列）
    reference_years = frozenset(
        int(year.group(1))
        for entry in entries
        if (year := re.match(r"(\d{4})", entry.voucher_date))
    )
    for registered in intake.active_files:
        is_designated = designated_target is not None and registered.path.resolve() == designated_target
        statement_by_sheet = detect_statement_sheets(registered.path, rules, reference_years)
        statement_hits = {
            name: result
            for name, result in statement_by_sheet.items()
            if isinstance(result, ExistingStatementResult)
        }
        if is_designated:
            designated_hit = True
            if not statement_hits:
                question = next(
                    (item for item in statement_by_sheet.values() if isinstance(item, MappingQuestion)),
                    None,
                )
                raise ValueError(
                    f"指定的正表文件识别失败：{registered.path.name}（{question.sample_values[0] if question is not None else '未找到项目列或可用金额列'}）"
                )
            if len(statement_hits) > 1:
                raise ValueError(
                    f"指定的正表文件识别到多个现金流量表工作表：{registered.path.name}"
                )
            existing_statement_path = str(registered.path)
            hit = next(iter(statement_hits.values()))
            for key, value in _balances_from_existing(hit).items():
                if value is not None:
                    balance_candidates[key] = (3, value)
        else:
            auto_hits = {
                name: result
                for name, result in statement_hits.items()
                if name not in mapped_dataset_sheets_by_file.get(registered.file_id, frozenset())
            }
            if len(auto_hits) == 1:
                hit = next(iter(auto_hits.values()))
                questions.append(
                    {
                        "file_id": registered.file_id,
                        "file": registered.path.name,
                        "sheet": hit.sheet_name,
                        "role": "statement_sheet",
                        "recommended": hit.sheet_name,
                        "message": f"检测到疑似正表工作表《{hit.sheet_name}》，请确认是否作为客户现有正表核对",
                        "kind": "statement",
                    }
                )
                statement_candidates.append(
                    {
                        "file_id": registered.file_id,
                        "file": registered.path.name,
                        "sheet": hit.sheet_name,
                    }
                )
            elif len(auto_hits) > 1:
                questions.append(
                    {
                        "file_id": registered.file_id,
                        "file": registered.path.name,
                        "sheet": "",
                        "role": "statement_sheet",
                        "recommended": "",
                        "message": f"识别到多个现金流量表工作表：{'、'.join(auto_hits)}",
                        "kind": "statement",
                    }
                )
    if designated_target is not None and not designated_hit:
        raise ValueError(f"指定的正表文件不在已选输入中：{statement_path}")

    recommended: dict[str, str] = {}
    run_id = stable_id("RUN", run_dir.name, *(item.sha256 for item in intake.files))
    versions = current_versions(PROJECT_ROOT)
    state: dict[str, object] = {
        "schema_version": versions["schema"],
        "versions": versions,
        "run_id": run_id,
        "stage": "preflight",
        "files": [
            {
                "file_id": item.file_id,
                "path": str(item.path),
                "sha256": item.sha256,
                "duplicate_of": item.duplicate_of,
                "is_macro_workbook": item.is_macro_workbook,
            }
            for item in intake.files
        ],
        "materiality": asdict(amounts),
        "automatic_change_threshold": automatic_change_threshold,
        "entries": [asdict(entry) for entry in entries],
        "mappings": mappings,
        "mapping_questions": questions,
        "existing_statement_path": existing_statement_path,
        "statement_candidates": statement_candidates,
        "statement_confirmations": {},
        "cash_balances": {key: value for key, (_, value) in balance_candidates.items()},
        "evidence_profiles": profiles,
        "normalization_issues": normalization_issues,
    }
    if notes:
        state["company_notes_raw"] = notes
    account_mapping_records = (
        _prepare_account_mapping_state(state, entries) if not questions else ()
    )
    _assert_inputs_unchanged(state)
    with store.stage("preflight") as connection:
        connection.execute(
            "INSERT INTO run_manifest(record_id, payload_json) VALUES (?, ?)",
            (
                run_id,
                json.dumps(
                    {
                        "materiality": asdict(amounts),
                        "automatic_change_threshold": automatic_change_threshold,
                        "versions": versions,
                    },
                    ensure_ascii=False,
                ),
            ),
        )
        connection.executemany(
            "INSERT INTO run_version(record_id, payload_json) VALUES (?, ?)",
            (
                (name, json.dumps({"value": value}, ensure_ascii=False))
                for name, value in versions.items()
            ),
        )
        connection.executemany(
            "INSERT INTO source_entry(record_id, payload_json) VALUES (?, ?)",
            ((entry.entry_id, json.dumps(asdict(entry), ensure_ascii=False)) for entry in entries),
        )
        connection.executemany(
            "INSERT INTO source_file(record_id, payload_json) VALUES (?, ?)",
            (
                (item.file_id, json.dumps(asdict(item), ensure_ascii=False, default=str))
                for item in intake.files
            ),
        )
        connection.executemany(
            "INSERT INTO sheet_structure(record_id, payload_json) VALUES (?, ?)",
            (
                (
                    stable_id("SHT", item["file_id"], item["sheet"]),
                    json.dumps(item, ensure_ascii=False),
                )
                for item in sheet_structures
            ),
        )
        connection.executemany(
            "INSERT INTO field_mapping(record_id, payload_json) VALUES (?, ?)",
            (
                (
                    stable_id("MAP", item["file_id"], item["sheet"]),
                    json.dumps(item, ensure_ascii=False),
                )
                for item in mappings
            ),
        )
    if questions:
        status = "waiting_mapping"
    elif any(item.status != "confirmed" for item in account_mapping_records):
        state["stage"] = "waiting_account_mapping"
        status = "waiting_account_mapping"
    else:
        recommended = _prepare_cash_scope_state(state)
        state["stage"] = "waiting_cash_scope"
        status = "waiting_cash_scope"
    _save_state(run_dir, state)
    return PreflightResult(run_id, run_dir, status, recommended, len(questions), len(entries))


def confirm_mapping(run_dir: Path, decisions: Mapping[str, str]) -> StageResult:
    state = _load_state(run_dir)
    _assert_inputs_unchanged(state)
    confirmations = dict(state.get("mapping_confirmations", {}))
    confirmations.update(decisions)
    files_by_id = {str(item["file_id"]): Path(str(item["path"])) for item in state["files"]}

    # ---- 疑似正表确认（取值 use / ignore）----
    statement_candidates = list(state.get("statement_candidates", ()))
    statement_confirmations = dict(state.get("statement_confirmations", {}))
    for candidate in statement_candidates:
        key = f"{candidate['file_id']}:statement:{candidate['sheet']}"
        choice = confirmations.get(key)
        if choice not in {"use", "ignore"}:
            raise RuntimeError(f"等待疑似正表确认：{key}（取值 use 或 ignore）")
        statement_confirmations[key] = choice
    candidate_keys = {
        f"{c['file_id']}:statement:{c['sheet']}" for c in statement_candidates
    }
    ambiguous_statements = [
        question for question in state.get("mapping_questions", ())
        if question.get("kind") == "statement"
        and f"{question['file_id']}:statement:{question.get('sheet', '')}" not in candidate_keys
    ]
    if ambiguous_statements:
        messages = "；".join(
            str(question.get("message") or "存在多个现金流量表工作表")
            for question in ambiguous_statements
        )
        raise RuntimeError(f"{messages}；请合并工作表或明确目标正表后重新预检")

    # ---- 字段映射确认 ----
    pending_questions = [
        question for question in state.get("mapping_questions", ())
        if question.get("kind") != "statement"
    ]
    if not pending_questions and not statement_candidates:
        state["mapping_confirmations"] = confirmations
        records = _prepare_account_mapping_state(
            state, tuple(_entry_from_dict(item) for item in state["entries"])
        )
        waiting_account = any(item.status != "confirmed" for item in records)
        if not waiting_account:
            _prepare_cash_scope_state(state)
        state["stage"] = (
            "waiting_account_mapping" if waiting_account else "waiting_cash_scope"
        )
        _save_state(run_dir, state)
        return StageResult(
            str(state["run_id"]),
            Path(run_dir),
            "mapping",
            "waiting" if waiting_account else "completed",
            "确认客户一级科目" if waiting_account else "确认现金范围",
        )

    pending_by_file: dict[str, list[dict[str, object]]] = {}
    for question in pending_questions:
        pending_by_file.setdefault(str(question["file_id"]), []).append(question)

    new_questions: list[dict[str, object]] = []
    new_entries: list[NormalizedEntry] = []
    profiles: dict[str, dict[str, object]] = {
        str(file_id): dict(payload)
        for file_id, payload in state.get("evidence_profiles", {}).items()
    }
    new_mappings: list[dict[str, object]] = []
    new_issues: list[dict[str, object]] = []
    mapped_sheets = {
        (str(item["file_id"]), str(item["sheet"])) for item in state.get("mappings", ())
    }
    for file_id, questions in pending_by_file.items():
        # 把所有已确认的字段映射一并作为列覆盖传回，避免只带"当前待确认项"
        # 导致先确认过的字段在下一轮重新变回待确认、形成两字段来回切换死循环。
        overrides_by_sheet: dict[str, dict[str, int]] = {}
        for record_key, record_choice in confirmations.items():
            # 只取属于本文件、且为"工作表:字段"形式的角色键
            prefix = f"{file_id}:"
            if not record_key.startswith(prefix):
                continue
            remainder = record_key[len(prefix):]
            if ":" not in remainder:
                continue
            sheet_part, _, role_part = remainder.partition(":")
            if not sheet_part or not role_part or sheet_part == "statement":
                continue
            try:
                col = column_index_from_string(str(record_choice).strip().upper())
            except ValueError:
                continue  # 不是字段列映射（如正表 use/ignore），跳过
            overrides_by_sheet.setdefault(sheet_part, {})[role_part] = col
        for question in questions:
            role = str(question["role"])
            sheet_name = str(question.get("sheet", ""))
            full_key = f"{file_id}:{sheet_name}:{role}"
            legacy_key = f"{file_id}:{role}"
            choice = confirmations.get(full_key, confirmations.get(legacy_key))
            if not choice:
                raise ValueError(f"等待字段确认：缺少 {full_key}")
            try:
                overrides_by_sheet.setdefault(sheet_name, {})[role] = column_index_from_string(
                    str(choice).strip().upper()
                )
            except ValueError as error:
                raise ValueError(f"字段 {full_key} 的确认值必须是 Excel 列字母") from error
        path = files_by_id[file_id]
        for mapping in infer_dataset_mappings(scan_workbook(path), overrides_by_sheet):
            if isinstance(mapping, MappingQuestion):
                new_questions.append(
                    {
                        "file_id": file_id,
                        "file": path.name,
                        "sheet": mapping.sheet_name,
                        "role": mapping.role,
                        "recommended": mapping.recommended.column_letter,
                    }
                )
                continue
            if (file_id, mapping.sheet_name) in mapped_sheets:
                continue
            normalized = normalize_dataset(path, file_id, mapping)
            new_entries.extend(normalized.entries)
            profiles[file_id] = _profile_to_dict(normalized.profile)
            new_mappings.append(
                {
                    "file_id": file_id,
                    "file": path.name,
                    "sheet": mapping.sheet_name,
                    "header_rows": f"{mapping.header_row_start}:{mapping.header_row_end}",
                    "roles": {role: column.column_letter for role, column in mapping.role_to_column.items()},
                }
            )
            new_issues.extend(
                {
                    "file_id": file_id,
                    "file": path.name,
                    "sheet": issue.source.sheet_name,
                    "cell": issue.source.cell_range,
                    "kind": "错误" if hasattr(issue, "message") else "排除",
                    "message": getattr(issue, "message", getattr(issue, "discard_reason", "")),
                }
                for issue in (*normalized.errors, *normalized.exclusions)
            )

    entries = [_entry_from_dict(item) for item in state["entries"]] + new_entries
    state["entries"] = [asdict(item) for item in entries]
    state["mappings"] = [*state.get("mappings", ()), *new_mappings]
    state["normalization_issues"] = [*state.get("normalization_issues", ()), *new_issues]
    state["mapping_questions"] = new_questions
    state["mapping_confirmations"] = confirmations
    state["statement_confirmations"] = statement_confirmations
    state["evidence_profiles"] = profiles
    use_paths = {
        str(files_by_id[str(candidate["file_id"])])
        for candidate in statement_candidates
        if statement_confirmations.get(f"{candidate['file_id']}:statement:{candidate['sheet']}") == "use"
    }
    if len(use_paths) > 1:
        raise RuntimeError("多个文件被确认为客户现有正表，请只保留一个")
    if use_paths:
        state["existing_statement_path"] = next(iter(use_paths))
    elif statement_candidates:
        # 存在待确认的疑似正表但均未确认纳入核对时，清空正表登记
        state["existing_statement_path"] = None
    # 无待确认疑似正表时，保留 preflight 通过 --statement-path 指定的正表
    if state["existing_statement_path"] is not None:
        parsed = parse_existing_statement(
            Path(str(state["existing_statement_path"])), load_rule_pack(PROJECT_ROOT)
        )
        if isinstance(parsed, ExistingStatementResult):
            balances = dict(state.get("cash_balances", {}))
            balances.update(
                {
                    key: value
                    for key, value in _balances_from_existing(parsed).items()
                    if value is not None
                }
            )
            state["cash_balances"] = balances
    account_records = () if new_questions else _prepare_account_mapping_state(state, entries)
    waiting_account = bool(account_records) and any(
        item.status != "confirmed" for item in account_records
    )
    if not new_questions and not waiting_account:
        _prepare_cash_scope_state(state)
    state["stage"] = (
        "waiting_mapping"
        if new_questions
        else "waiting_account_mapping"
        if waiting_account
        else "waiting_cash_scope"
    )
    if new_entries:
        with _store(run_dir).stage("mapping") as connection:
            connection.executemany(
                "INSERT OR REPLACE INTO source_entry(record_id, payload_json) VALUES (?, ?)",
                ((entry.entry_id, json.dumps(asdict(entry), ensure_ascii=False)) for entry in new_entries),
            )
            connection.executemany(
                "INSERT OR REPLACE INTO field_mapping(record_id, payload_json) VALUES (?, ?)",
                (
                    (
                        stable_id("MAP", item["file_id"], item["sheet"]),
                        json.dumps(item, ensure_ascii=False),
                    )
                    for item in new_mappings
                ),
            )
    _save_state(run_dir, state)
    return StageResult(
        str(state["run_id"]),
        Path(run_dir),
        "mapping",
        "waiting"
        if new_questions or state["stage"] == "waiting_account_mapping"
        else "completed",
        "继续确认剩余字段"
        if new_questions
        else "确认客户一级科目"
        if state["stage"] == "waiting_account_mapping"
        else "确认现金范围",
    )


def confirm_account_mapping(
    run_dir: Path,
    decisions: Mapping[str, str],
) -> StageResult:
    state = _load_state(run_dir)
    _assert_inputs_unchanged(state)
    if state.get("mapping_questions"):
        raise RuntimeError("字段映射仍有待确认项，请先完成字段确认")
    raw_records = state.get("account_mapping_records")
    records = (
        tuple(AccountMappingRecord(**item) for item in raw_records)
        if isinstance(raw_records, list)
        else _prepare_account_mapping_state(
            state, tuple(_entry_from_dict(item) for item in state["entries"])
        )
    )
    resolved = resolve_account_mappings(
        records, decisions, load_standard_accounts(PROJECT_ROOT)
    )
    before = {
        item.original_level1: (item.standard_level1, item.status)
        for item in records
    }
    after = {
        item.original_level1: (item.standard_level1, item.status)
        for item in resolved
    }
    changed = before != after
    downstream_started = any(
        key in state
        for key in ("cash_scope", "account_dictionary", "classification_summary")
    )
    if downstream_started and changed:
        raise RuntimeError(
            "一级科目映射已被现金范围或分类使用；映射变化后必须新建运行目录，旧运行不得继续复用"
        )
    if downstream_started:
        return StageResult(
            str(state["run_id"]),
            Path(run_dir),
            "account_mapping",
            "completed",
            "映射未变化，无需重复确认",
        )
    state["account_mapping_records"] = [asdict(item) for item in resolved]
    state["account_mapping_questions"] = []
    state["account_mapping_confirmations"] = dict(decisions)
    _prepare_cash_scope_state(state)
    state["stage"] = "waiting_cash_scope"
    _save_state(run_dir, state)
    _write_trace_jsonl(run_dir, "一级科目映射留痕.jsonl", resolved)
    return StageResult(
        str(state["run_id"]),
        Path(run_dir),
        "account_mapping",
        "completed",
        "确认现金范围",
    )


def confirm_cash_scope(
    run_dir: Path,
    decisions: Mapping[str, object],
) -> StageResult:
    state = _load_state(run_dir)
    _assert_inputs_unchanged(state)
    if state.get("mapping_questions"):
        raise RuntimeError("字段映射仍有待确认项，请先完成字段确认")
    if state.get("account_mapping_questions"):
        raise RuntimeError("客户一级科目映射仍有待确认项，请先完成一级科目确认")
    entries = _standardized_entries_from_state(state)
    proposal = discover_cash_scope(entries)
    scope = make_cash_scope(proposal, decisions)
    state["cash_scope_proposal"] = asdict(proposal)
    state["recommended_cash_decisions"] = {
        candidate.account_key: candidate.system_suggestion
        for candidate in proposal.candidates
        if candidate.system_suggestion in {"include", "exclude"}
    }
    store = _store(run_dir)
    with store.stage("cash_scope") as connection:
        connection.execute(
            "INSERT OR REPLACE INTO cash_scope(record_id, payload_json) VALUES (?, ?)",
            (scope.scope_hash, json.dumps(asdict(scope), ensure_ascii=False, default=list)),
        )
    state["cash_scope"] = {
        "included_keys": sorted(scope.included_keys),
        "excluded_keys": sorted(scope.excluded_keys),
        "account_names_by_key": scope.account_names_by_key,
        "scope_hash": scope.scope_hash,
    }
    state["stage"] = "cash_scope_confirmed"
    _save_state(run_dir, state)
    return StageResult(str(state["run_id"]), Path(run_dir), "cash_scope", "completed", "执行自动分类")


def supplement_cash_balances(
    run_dir: Path,
    opening: object,
    closing: object,
    fx: object,
    source_note: str,
) -> StageResult:
    """补充缺失的现金余额和汇率影响，保留来源说明且不做倒挤。"""
    state = _load_state(run_dir)
    _assert_inputs_unchanged(state)
    if not source_note.strip():
        raise ValueError("补充现金余额时必须填写资料来源说明")
    state["cash_balances"] = {
        "opening_cent": yuan_to_cent(opening),
        "closing_cent": yuan_to_cent(closing),
        "fx_cent": yuan_to_cent(fx),
    }
    state["cash_balance_source"] = source_note.strip()
    state.pop("overall_status", None)
    state.pop("workbook_path", None)
    _save_state(run_dir, state)
    return StageResult(
        str(state["run_id"]), Path(run_dir), "cash_balances", "completed", "生成最终工作簿"
    )


def _store_consistency_resolution(
    state: dict[str, object],
    resolution: object,
) -> None:
    state["decisions"] = [asdict(item) for item in resolution.decisions]
    state["consistency_resolution"] = {
        "statuses": [
            {
                "group_id": group_id,
                "status": status,
                "reason": reason,
                "tier": tier,
            }
            for group_id, status, reason, tier in resolution.statuses
        ],
        "unresolved": [asdict(item) for item in resolution.unresolved],
    }


def _prepare_consistency_stage(
    state: dict[str, object],
    run_dir: Path,
) -> tuple[str, int]:
    if "consistency_groups" in state:
        missing = int(state["classification_summary"].get("ai_tasks_missing", 0))
        return (
            "待完成人工决定"
            if state.get("stage") == "waiting_human"
            else "AI 已完成",
            missing,
        )
    components = tuple(_component_from_dict(item) for item in state["components"])
    decisions = tuple(_decision_from_dict(item) for item in state["decisions"])
    groups = find_consistency_groups(
        components, decisions, _materiality_from_state(state)
    )
    state["consistency_schema_version"] = 1
    state["consistency_groups"] = [asdict(item) for item in groups]
    rule_pack = load_rule_pack(PROJECT_ROOT)
    resolution = apply_consistency_forced_checks(
        groups,
        decisions,
        {item.item_id: item.name for item in rule_pack.statement_items},
        {item.item_id: item.normal_direction for item in rule_pack.statement_items},
    )
    _store_consistency_resolution(state, resolution)
    state["classification_summary"]["ai_tasks_missing"] = 0
    if resolution.unresolved:
        state["classification_summary"]["status"] = "waiting_human"
        state["stage"] = "waiting_human"
        return "待完成人工决定", 0
    state["classification_summary"]["status"] = "consistency_completed"
    state["stage"] = "consistency_completed"
    return "AI 已完成", 0


def _account_path_result_to_dict(
    result: AccountPathSemanticResult,
) -> dict[str, object]:
    return {
        "account": result.account,
        "status": result.status,
        "concepts": [asdict(item) for item in result.concepts],
        "candidate_item_ids": list(result.candidate_item_ids),
        "inflow_candidate_item_ids": list(result.inflow_candidate_item_ids),
        "outflow_candidate_item_ids": list(result.outflow_candidate_item_ids),
        "quality_score": result.quality.value,
        "semantic": result.semantic,
        "basis": result.basis,
        "unresolved_slots": [asdict(item) for item in result.unresolved_slots],
        "matched_rule_ids": list(result.matched_rule_ids),
        "relations": [asdict(item) for item in result.relations],
    }


def _account_path_result_from_dict(
    payload: Mapping[str, object],
) -> AccountPathSemanticResult:
    return AccountPathSemanticResult(
        account=str(payload["account"]),
        status=str(payload.get("status", "未识别")),
        concepts=tuple(
            AccountNodeConcept(
                int(item["level_index"]),
                str(item["node_text"]),
                str(item["concept"]),
                str(item["source_text"]),
                str(item.get("source", "direct")),
            )
            for item in payload.get("concepts", ())
        ),
        candidate_item_ids=tuple(str(value) for value in payload.get("candidate_item_ids", ())),
        inflow_candidate_item_ids=tuple(
            str(value) for value in payload.get("inflow_candidate_item_ids", ())
        ),
        outflow_candidate_item_ids=tuple(
            str(value) for value in payload.get("outflow_candidate_item_ids", ())
        ),
        quality=EvidenceQuality(int(payload.get("quality_score", 0))),
        semantic=str(payload.get("semantic", "")),
        basis=str(payload.get("basis", "")),
        unresolved_slots=tuple(
            AccountPathSlot(
                int(item["level_index"]),
                str(item["node_text"]),
                tuple(str(value) for value in item.get("allowed_concepts", ())),
                tuple(str(value) for value in item.get("allowed_relations", ())),
            )
            for item in payload.get("unresolved_slots", ())
        ),
        matched_rule_ids=tuple(str(value) for value in payload.get("matched_rule_ids", ())),
        relations=tuple(
            AccountPathRelation(
                int(item["parent_level_index"]),
                int(item["child_level_index"]),
                str(item["relation"]),
                str(item.get("source", "agent")),
            )
            for item in payload.get("relations", ())
        ),
    )


def _dictionary_from_state(state: Mapping[str, object]) -> AccountDictionary:
    """由运行状态中的企业专属自定义条目 + 内置通用词典，合并出本次分类所用的语义词典。"""
    common = load_common_dictionary(PROJECT_ROOT)
    dictionary_state = state.get("account_dictionary", {})
    if dictionary_state and dictionary_state.get("schema_version") != "2.0.0":
        raise RuntimeError("旧版科目路径语义结果不能用于当前规则，请重新执行科目扫描")
    valid = dictionary_state.get("valid_results", ())
    custom_entries = tuple(
        AccountSemanticEntry(
            account=str(item["account"]),
            semantic=str(item.get("semantic", "")),
            item_id=(
                str(item.get("candidate_item_ids", ())[0])
                if len(item.get("candidate_item_ids", ())) == 1
                else ""
            ),
            basis=str(item.get("basis", "")),
            confidence="",
            layer="runtime",
            note_id=str(item.get("note_id", "")),
            inflow_item_id=(
                str(item.get("inflow_candidate_item_ids", ())[0])
                if len(item.get("inflow_candidate_item_ids", ())) == 1
                else ""
            ),
            outflow_item_id=(
                str(item.get("outflow_candidate_item_ids", ())[0])
                if len(item.get("outflow_candidate_item_ids", ())) == 1
                else ""
            ),
            classification_facts=tuple(
                str(value) for value in item.get("classification_facts", ())
            ),
            candidate_item_ids=tuple(
                str(value) for value in item.get("candidate_item_ids", ())
            ),
            inflow_candidate_item_ids=tuple(
                str(value) for value in item.get("inflow_candidate_item_ids", ())
            ),
            outflow_candidate_item_ids=tuple(
                str(value) for value in item.get("outflow_candidate_item_ids", ())
            ),
            quality_score=int(item.get("quality_score", 0)),
        )
        for item in valid
    )
    return merge_dictionaries(common, AccountDictionary(custom_entries))


def _summary_result_to_dict(result: SummarySemanticResult) -> dict[str, object]:
    return {
        "summary": result.summary,
        "status": result.status,
        "spans": [asdict(span) for span in result.spans],
        "candidate_item_ids": list(result.candidate_item_ids),
        "quality": result.quality.value,
        "reason": result.reason,
        "unresolved_slots": list(result.unresolved_slots),
    }


def _summary_result_from_dict(payload: Mapping[str, object]) -> SummarySemanticResult:
    return SummarySemanticResult(
        summary=str(payload["summary"]),
        status=str(payload["status"]),
        spans=tuple(
            SummarySpan(
                slot=str(span["slot"]),
                text=str(span["text"]),
                start=int(span["start"]),
                end=int(span["end"]),
                source=str(span.get("source", "rule")),
            )
            for span in payload.get("spans", ())
        ),
        candidate_item_ids=tuple(
            str(value) for value in payload.get("candidate_item_ids", ())
        ),
        quality=EvidenceQuality(int(payload.get("quality", 0))),
        reason=str(payload.get("reason", "")),
        unresolved_slots=tuple(
            str(value) for value in payload.get("unresolved_slots", ())
        ),
    )


def _summary_semantics_from_state(
    state: Mapping[str, object],
) -> dict[str, SummarySemanticResult]:
    payload = state.get("summary_semantics", {})
    return {
        result.summary: result
        for result in (
            _summary_result_from_dict(item) for item in payload.get("results", ())
        )
    }


def _write_dictionary_batches(run_dir: Path, tasks: Sequence[dict[str, object]]) -> None:
    for index in range(0, len(tasks), 25):
        _write_trace_jsonl(
            run_dir,
            f"科目语义待判断_第{index // 25 + 1:02d}批.jsonl",
            tasks[index : index + 25],
        )


def confirm_company_notes(
    run_dir: Path,
    entries: Sequence[Mapping[str, object]],
) -> StageResult:
    """登记经用户确认的公司特殊规则清单（B9）。

    无 --notes 文本时也可直接口述登记。只有采用中的规则生效；
    规则停用、替代或变更时，已依赖它的词典和分类结果立即失效。
    """
    state = _load_state(run_dir)
    _assert_inputs_unchanged(state)
    previous_active = {
        str(note.get("note_id", "")): dict(note)
        for note in state.get("company_notes", ())
    }
    validated: list[dict[str, object]] = []
    seen_note_ids: set[str] = set()
    for index, item in enumerate(entries):
        text = str(item.get("内容", "")).strip()
        if not text:
            raise ValueError(f"公司特殊规则第 {index + 1} 条缺少内容")
        note_id = str(item.get("note_id") or f"NOTE-{index + 1:02d}").strip()
        if re.fullmatch(r"NOTE-\d{2}", note_id) is None:
            raise ValueError(f"公司特殊规则编号必须使用唯一的 NOTE-xx 格式：{note_id}")
        if note_id in seen_note_ids:
            raise ValueError(f"公司特殊规则编号重复：{note_id}")
        seen_note_ids.add(note_id)
        previous = previous_active.get(note_id, {})
        status = str(item.get("状态", "") or previous.get("状态", "采用"))
        if status not in {
            "采用",
            "长期采用",
            "仅本次采用",
            "冲突未采用",
            "已停用",
            "已被替代",
        }:
            raise ValueError(
                f"公司特殊规则第 {index + 1} 条状态非法：{status}"
            )
        def value_or_previous(key: str, default: object) -> object:
            return item[key] if key in item else previous.get(key, default)

        terms = (
            item.get("涉及科目或词")
            if "涉及科目或词" in item
            else item.get("涉及科目")
            if "涉及科目" in item
            else previous.get("涉及科目或词", ())
        )

        def list_value(value: object) -> list[object]:
            if value in (None, ""):
                return []
            return list(value) if isinstance(value, (list, tuple)) else [value]

        record = {
            "note_id": note_id,
            "内容": text,
            "涉及科目或词": list(terms)
            if isinstance(terms, (list, tuple))
            else [str(terms)],
            "建议处理": str(value_or_previous("建议处理", "")),
            "依据": str(value_or_previous("依据", "")),
            "规则类型": str(value_or_previous("规则类型", "")),
            "状态": status,
            "规则版本": int(value_or_previous("规则版本", 1)),
            "运行编号": str(value_or_previous("运行编号", state["run_id"])),
            "适用主体": str(value_or_previous("适用主体", "本次运行主体")),
            "适用期间": str(value_or_previous("适用期间", "本次运行期间")),
            "原始证据示例": str(value_or_previous("原始证据示例", "")),
            "判断理由": str(
                value_or_previous("判断理由", item.get("依据", ""))
            ),
            "确认人": str(value_or_previous("确认人", "用户确认")),
            "确认时间": str(
                value_or_previous(
                    "确认时间", datetime.now(timezone.utc).isoformat()
                )
            ),
            "替代规则编号": str(value_or_previous("替代规则编号", "")),
            "影响业务组成": list_value(
                value_or_previous("影响业务组成", ())
            ),
            "影响金额分": int(value_or_previous("影响金额分", 0)),
        }
        for key in (
            "适用完整路径",
            "适用标准一级科目",
            "适用中间层级",
            "适用末级明细",
            "适用摘要词",
            "适用公司别名",
        ):
            values = value_or_previous(key, ())
            record[key] = [str(value) for value in list_value(values)]
        if previous and "规则版本" not in item:
            ignored_for_version = {"规则版本", "确认人", "确认时间"}
            changed = any(
                record.get(key) != previous.get(key)
                for key in set(record) | set(previous)
                if key not in ignored_for_version
            )
            record["规则版本"] = int(previous.get("规则版本", 1)) + int(changed)
            if changed:
                record["确认人"] = str(item.get("确认人", "用户确认"))
                record["确认时间"] = str(
                    item.get("确认时间", datetime.now(timezone.utc).isoformat())
                )
        validated.append(record)
    current_active = {
        str(note.get("note_id", "")): note
        for note in validated
    }
    changed_note_ids = sorted(
        note_id
        for note_id in set(previous_active) | set(current_active)
        if previous_active.get(note_id) != current_active.get(note_id)
    )
    dependency_keys = (
        "account_dictionary",
        "classification_summary",
        "components",
        "decisions",
        "source_allocations",
        "materiality_assessments",
        "ai_tasks",
        "structured_ai_validation",
        "human_decisions",
        "consistency_groups",
        "consistency_resolution",
        "reconciliation",
        "rough_reconciliation",
        "internal_transfers",
        "final_readiness",
        "overall_status",
        "workbook_path",
        "standardized_evidence_profiles",
        "semantic_account_paths",
    )
    invalidated = [key for key in dependency_keys if key in state]
    if changed_note_ids and invalidated:
        for key in invalidated:
            state.pop(key, None)
        state["account_dictionary_completed"] = False
        state["note_dependency_rebuild"] = {
            "changed_note_ids": changed_note_ids,
            "invalidated_results": invalidated,
            "reason": "公司特殊规则变更，依赖结果已失效，必须重建科目语义和分类",
        }
        state["stage"] = (
            "cash_scope_confirmed" if "cash_scope" in state else "waiting_cash_scope"
        )
    state["company_notes"] = validated
    _save_state(run_dir, state)
    _write_trace_json(run_dir, "公司规则登记.json", validated)
    return StageResult(
        str(state["run_id"]), Path(run_dir), "company_notes", "completed", "执行科目语义确认"
    )


def _write_summary_batches(run_dir: Path, tasks: Sequence[dict[str, object]]) -> None:
    for index in range(0, len(tasks), 25):
        _write_trace_jsonl(
            run_dir,
            f"摘要语义待判断_第{index // 25 + 1:02d}批.jsonl",
            tasks[index : index + 25],
        )


def _standard_basis_matches_items(
    text: str,
    item_ids: Sequence[str],
    item_by_id: Mapping[str, object],
) -> bool:
    """准则依据必须可追查，并且逐项对应本次选择的正表项目。"""
    basis = (text or "").strip()
    selected = tuple(dict.fromkeys(item_id for item_id in item_ids if item_id))
    if not selected or validate_basis_text(basis) is not None:
        return not selected
    if "准则" not in basis and "应用指南" not in basis:
        return False
    for item_id in selected:
        item = item_by_id.get(item_id)
        item_name = str(getattr(item, "name", ""))
        if item_id not in basis and (not item_name or item_name not in basis):
            return False
        if "应用指南" in basis:
            continue
        expected_clause = {
            "CFO": "第十条",
            "CFI": "第十三条",
            "CFF": "第十五条",
        }.get(item_id.split("-", 1)[0], "")
        lease_clause = (
            item_id == "CFF-06"
            and "企业会计准则第21号" in basis
            and "第五十三条" in basis
        )
        if expected_clause not in basis and not lease_clause:
            return False
    return True


def _write_dictionary_doc(
    run_dir: Path,
    valid: Sequence[dict[str, object]],
    company_notes: Sequence[Mapping[str, object]] = (),
    coverage: Mapping[str, object] | None = None,
) -> None:
    lines = [
        "# 科目语义词典说明",
        "",
        "本文件按完整父路径展示本次运行的固定语义结果；覆盖统计只用于完善规则，不参与分类、评分、重要性或人工门禁。",
        "",
    ]
    if coverage:
        path_counts = dict(coverage.get("path_counts", {}))
        component_counts = dict(coverage.get("component_counts", {}))
        component_amounts = dict(coverage.get("component_absolute_amount_cent", {}))
        lines.extend(
            [
                "## 规则覆盖情况",
                "",
                f"- 完整路径总数：{coverage.get('path_total', 0)}",
                f"- 业务组成总数：{coverage.get('component_total', 0)}",
                "",
                "| 状态 | 路径数 | 业务组成笔数 | 绝对金额（元） |",
                "|---|---:|---:|---:|",
            ]
        )
        for status in ("固定规则完整解释", "Agent补充", "部分解释", "冲突", "未识别"):
            lines.append(
                f"| {status} | {path_counts.get(status, 0)} | "
                f"{component_counts.get(status, 0)} | "
                f"{int(component_amounts.get(status, 0)) / 100:,.2f} |"
            )
        pending_paths = tuple(coverage.get("pending_paths", ()))
        if pending_paths:
            lines.extend(
                [
                    "",
                    "### 高频或高金额待补路径",
                    "",
                    "| 完整路径 | 状态 | 业务组成笔数 | 绝对金额（元） |",
                    "|---|---|---:|---:|",
                ]
            )
            for item in pending_paths:
                lines.append(
                    f"| {item.get('account', '')} | {item.get('status', '')} | "
                    f"{item.get('component_count', 0)} | "
                    f"{int(item.get('absolute_amount_cent', 0)) / 100:,.2f} |"
                )
        lines.extend(["", "## 完整路径明细", ""])
    lines.append(
        "| 客户原完整路径 | 客户一级科目 | 标准一级科目 | 中间层级 | 末级明细 | "
        "规范化路径 | 科目语义 | 疑似现金流项目 | 固定质量 | 节点解释 | 未识别节点 | "
        "事实依据 | 适用NOTE | 解释状态 |"
    )
    lines.append("|---|---|---|---|---|---|---|---|---|---|---|---|---|---|")
    source_names = {
        "direct": "直接识别",
        "parent_inheritance": "继承父属性",
        "neutral": "中性限定",
        "agent": "Agent补充",
    }
    for item in valid:
        levels = split_account_levels(str(item["account"]))
        originals = item.get("original_paths", ()) or (item["account"],)
        node_explanations = "；".join(
            f"第{int(concept.get('level_index', 0)) + 1}层“{concept.get('node_text', '')}”="
            f"{concept.get('concept', '')}（{source_names.get(str(concept.get('source', '')), concept.get('source', ''))}）"
            for concept in item.get("concepts", ())
        ) or "无直接、继承或中性解释"
        unresolved_nodes = "；".join(
            f"第{int(slot.get('level_index', 0)) + 1}层“{slot.get('node_text', '')}”"
            for slot in item.get("unresolved_slots", ())
        ) or "无"
        if item.get("candidate_item_ids"):
            item_display = "候选：" + "、".join(item["candidate_item_ids"])
        elif any(
            item.get(key)
            for key in (
                "inflow_candidate_item_ids",
                "outflow_candidate_item_ids",
            )
        ):
            inflow = "、".join(item.get("inflow_candidate_item_ids", ()))
            outflow = "、".join(item.get("outflow_candidate_item_ids", ()))
            item_display = f"流入：{inflow}；流出：{outflow}"
        else:
            item_display = "已识别但不指向特定项目"
        lines.append(
            f"| {'、'.join(str(value) for value in originals)} | "
            f"{item.get('customer_level1', levels[0] if levels else '')} | "
            f"{item.get('standard_level1', levels[0] if levels else '')} | "
            f"{' / '.join(levels[1:-1])} | {levels[-1] if levels else ''} | {item['account']} | "
            f"{item.get('semantic', '')} | "
            f"{item_display} | "
            f"{item.get('quality_score', 0)} | {node_explanations} | {unresolved_nodes} | "
            f"{item.get('basis', '')} | "
            f"{item.get('note_id', '')} | {item.get('status', '')} |"
        )
    if company_notes:
        # 复核修复：公司特殊规则分"已采用"与"冲突未采用（仅说明，不生效）"两节列示
        adopted = [note for note in company_notes if company_note_is_active(note)]
        declined = [note for note in company_notes if not company_note_is_active(note)]
        lines.append("")
        lines.append("## 公司特殊规则")
        for heading, notes in (("### 已采用", adopted), ("### 冲突未采用（仅说明，不生效）", declined)):
            if notes:
                lines.append(heading)
                for note in notes:
                    lines.append(
                        f"- {note.get('note_id', '')}：{note.get('内容', '')}（涉及科目或词："
                        f"{'、'.join(note.get('涉及科目或词', ()) or ())}；建议处理：{note.get('建议处理', '')}）"
                    )
    (Path(run_dir) / "科目语义词典说明.md").write_text(
        "\n".join(lines) + "\n", encoding="utf-8-sig"
    )


def _semantic_components_after_component_build(
    state: Mapping[str, object],
    entries: Sequence[NormalizedEntry],
) -> tuple[CashflowComponent, ...]:
    """独立扫描入口按正式现金范围和结构选择重建业务组成。"""
    scope = _scope_from_dict(state["cash_scope"])
    if find_cash_row_cleanup_requests(entries, scope):
        raise RuntimeError("现金分录仍无法可靠识别，请先执行分类并完成清洗门禁")
    entries_by_file: dict[str, list[NormalizedEntry]] = {}
    for entry in entries:
        entries_by_file.setdefault(entry.source.file_id, []).append(entry)
    single_sided_file_ids = frozenset(
        file_id
        for file_id, file_entries in entries_by_file.items()
        if (
            (profile := infer_evidence_profile(file_entries)).has_flow_amount
            and profile.has_flow_item
        )
    )
    build = build_cashflow_components(
        entries,
        scope,
        single_sided_file_ids=single_sided_file_ids,
        structure_selections={
            str(voucher_key): tuple(str(value) for value in selected)
            for voucher_key, selected in dict(
                state.get("component_structure_selections", {})
            ).items()
        },
        structure_selection_basis={
            str(voucher_key): str(value)
            for voucher_key, value in dict(
                state.get("component_structure_selection_basis", {})
            ).items()
        },
    )
    if build.structure_requests:
        raise RuntimeError("业务组成结构尚未确定，请先执行分类并完成结构确认")
    return tuple(build.components)


def _semantic_account_paths_after_component_build(
    state: Mapping[str, object],
    entries: Sequence[NormalizedEntry],
) -> tuple[str, ...]:
    components = _semantic_components_after_component_build(state, entries)
    return tuple(
        sorted(
            {
                path
                for component in components
                for path in component.counterpart_accounts
                if path.strip()
            }
        )
    )


def _account_dictionary_coverage(
    components: Sequence[CashflowComponent],
    results: Sequence[AccountPathSemanticResult],
) -> dict[str, object]:
    """覆盖统计只用于补规则，不参与分类、评分、重要性或门禁。"""
    result_by_path = {item.account: item for item in results}
    statuses = (
        "固定规则完整解释",
        "Agent补充",
        "部分解释",
        "冲突",
        "未识别",
    )
    path_counts = {
        status: sum(item.status == status for item in results) for status in statuses
    }
    rank = {status: index for index, status in enumerate(statuses)}
    component_counts = {status: 0 for status in statuses}
    component_amounts = {status: 0 for status in statuses}
    path_usage: dict[str, list[int]] = {}
    for component in components:
        component_statuses = [
            result_by_path[path].status
            for path in component.counterpart_accounts
            if path in result_by_path
        ]
        status = (
            max(component_statuses, key=lambda value: rank.get(value, len(rank)))
            if component_statuses
            else "未识别"
        )
        component_counts[status] += 1
        component_amounts[status] += abs(component.cash_delta_cent)
        for path in dict.fromkeys(component.counterpart_accounts):
            usage = path_usage.setdefault(path, [0, 0])
            usage[0] += 1
            usage[1] += abs(component.cash_delta_cent)
    pending = [
        {
            "account": path,
            "status": result_by_path[path].status,
            "component_count": values[0],
            "absolute_amount_cent": values[1],
        }
        for path, values in path_usage.items()
        if path in result_by_path
        and result_by_path[path].status not in {"固定规则完整解释", "Agent补充"}
    ]
    pending.sort(
        key=lambda item: (-int(item["absolute_amount_cent"]), -int(item["component_count"]), str(item["account"]))
    )
    return {
        "path_total": len(results),
        "path_counts": path_counts,
        "component_total": len(components),
        "component_counts": component_counts,
        "component_absolute_amount_cent": component_amounts,
        "pending_paths": pending[:20],
    }


def scan_accounts(run_dir: Path) -> dict[str, object]:
    """扫描本次运行全部完整对方科目路径，生成一般及企业专属语义任务。"""
    state = _load_state(run_dir)
    _assert_inputs_unchanged(state)
    if "cash_scope" not in state:
        raise RuntimeError("请确认现金范围后继续")
    if state.get("company_notes_raw") and not state.get("company_notes"):
        return {"status": "待确认公司特殊规则", "missing": 0}
    entries = _standardized_entries_from_state(state)
    components = _semantic_components_after_component_build(state, entries)
    all_paths = tuple(
        sorted(
            {
                path
                for component in components
                for path in component.counterpart_accounts
                if path.strip()
            }
        )
    )
    state["semantic_account_paths"] = list(all_paths)
    effective_names: list[tuple[NormalizedEntry, str]] = []
    for entry in entries:
        if entry.counterpart_name in all_paths:
            effective_names.append((entry, entry.counterpart_name))
        if entry.account_name in all_paths:
            effective_names.append((entry, entry.account_name))
    rules = load_account_semantic_rules(PROJECT_ROOT)
    analyses = tuple(analyze_account_path(path, rules) for path in all_paths)
    valid: list[dict[str, object]] = []
    tasks: list[dict[str, object]] = []
    base_results: list[dict[str, object]] = []
    adopted_notes = [
        note for note in state.get("company_notes", ()) if company_note_is_active(note)
    ]
    for result in analyses:
        account_path = result.account
        levels = split_account_levels(account_path)
        original_paths = tuple(
            dict.fromkeys(
                (
                    entry.original_counterpart_name
                    or entry.original_account_name
                    or name
                )
                for entry, name in effective_names
                if account_path == name
            )
        )
        metadata = {
            "original_path": account_path,
            "standard_level1": levels[0] if levels else "",
            "original_paths": list(original_paths),
            "customer_level1": (
                split_account_levels(original_paths[0])[0]
                if original_paths and split_account_levels(original_paths[0])
                else (levels[0] if levels else "")
            ),
        }
        record = {**_account_path_result_to_dict(result), **metadata}
        base_results.append(record)
        if not result.unresolved_slots:
            valid.append(record)
            continue
        task = {
            "task_id": stable_id("ACC", account_path),
            **build_account_agent_task(result),
            **metadata,
        }
        if adopted_notes:
            relevant = [
                f"{note.get('note_id', '')}：{note.get('内容', '')}"
                for note in adopted_notes
                if company_note_applies(note, "", (account_path,))
            ]
            if relevant:
                task["company_notes"] = relevant
        tasks.append(task)
    state["account_dictionary"] = {
        "schema_version": "2.0.0",
        "tasks": tasks,
        "base_results": base_results,
        "valid_results": valid,
        "missing_ids": [task["task_id"] for task in tasks],
        "coverage": _account_dictionary_coverage(components, analyses),
    }
    state["account_dictionary_completed"] = not tasks
    if tasks:
        _write_dictionary_batches(run_dir, tasks)
        state["stage"] = "waiting_dictionary"
    else:
        state["stage"] = "dictionary_completed"
        _write_dictionary_doc(
            run_dir,
            valid,
            state.get("company_notes", ()),
            state["account_dictionary"]["coverage"],
        )
    _save_state(run_dir, state)
    return {
        "status": "待科目语义确认" if tasks else "科目语义已齐备",
        "missing": len(tasks),
    }


def _import_dictionary_results_legacy(run_dir: Path, result_path: Path) -> dict[str, object]:
    """校验并导入完整路径科目语义结果；全部有效后标记齐备并生成人读说明文档。"""
    state = _load_state(run_dir)
    _assert_inputs_unchanged(state)
    pending = state.get("account_dictionary")
    if not pending or not pending.get("tasks"):
        raise RuntimeError("尚未生成科目语义待判断任务，请先执行 scan-accounts")
    expected_by_id = {task["task_id"]: task for task in pending["tasks"]}
    # 复核修复：NOTE 编号必须指向已登记且"采用"的公司特殊规则
    adopted_note_ids = {
        str(note.get("note_id", ""))
        for note in state.get("company_notes", ())
        if company_note_is_active(note)
    }
    statement_items = tuple(
        item for item in load_rule_pack(PROJECT_ROOT).statement_items if item.is_leaf
    )
    leaf_ids = {item.item_id for item in statement_items}
    item_by_id = {item.item_id: item for item in statement_items}
    valid: list[dict[str, object]] = []
    missing: list[str] = []
    seen: set[str] = set()
    with Path(result_path).open("r", encoding="utf-8-sig") as source:
        for raw in source:
            line = raw.strip()
            if not line:
                continue
            record = json.loads(line)
            task_id = str(record.get("task_id", ""))
            account = str(record.get("account", ""))
            item_id = str(record.get("item_id", ""))
            candidate_item_ids = tuple(
                dict.fromkeys(
                    str(value) for value in record.get("candidate_item_ids", ())
                )
            )
            inflow_candidate_item_ids = tuple(
                dict.fromkeys(
                    str(value)
                    for value in record.get("inflow_candidate_item_ids", ())
                )
            )
            outflow_candidate_item_ids = tuple(
                dict.fromkeys(
                    str(value)
                    for value in record.get("outflow_candidate_item_ids", ())
                )
            )
            inflow_item_id = str(record.get("inflow_item_id", ""))
            outflow_item_id = str(record.get("outflow_item_id", ""))
            confidence = str(record.get("confidence", ""))
            basis = str(record.get("basis", ""))
            standard_basis = str(record.get("standard_basis", ""))
            semantic = str(record.get("semantic", ""))
            note_id = str(record.get("note_id", ""))
            facts = tuple(str(value) for value in record.get("classification_facts", ()))
            if task_id not in expected_by_id or task_id in seen:
                if task_id:
                    missing.append(task_id)
                continue
            # 复核修复：科目段必须与任务一致，防止结果错行、张冠李戴
            if account != expected_by_id[task_id]["account"]:
                missing.append(task_id)
                continue
            if any(
                value and value not in leaf_ids
                for value in (
                    item_id,
                    inflow_item_id,
                    outflow_item_id,
                    *candidate_item_ids,
                    *inflow_candidate_item_ids,
                    *outflow_candidate_item_ids,
                )
            ):
                missing.append(task_id)
                continue
            if (
                confidence not in {"high", "medium", "low"}
                or (
                    (
                        item_id
                        or inflow_item_id
                        or outflow_item_id
                        or candidate_item_ids
                        or inflow_candidate_item_ids
                        or outflow_candidate_item_ids
                    )
                    and not facts
                )
                or (
                    max(
                        len(candidate_item_ids),
                        len(inflow_candidate_item_ids),
                        len(outflow_candidate_item_ids),
                    )
                    > 1
                    and confidence != "low"
                )
                or (
                    candidate_item_ids
                    and (
                        item_id
                        or inflow_item_id
                        or outflow_item_id
                        or inflow_candidate_item_ids
                        or outflow_candidate_item_ids
                    )
                )
                or (inflow_item_id and inflow_candidate_item_ids)
                or (outflow_item_id and outflow_candidate_item_ids)
                or not _standard_basis_matches_items(
                    standard_basis,
                    (
                        item_id,
                        inflow_item_id,
                        outflow_item_id,
                        *candidate_item_ids,
                        *inflow_candidate_item_ids,
                        *outflow_candidate_item_ids,
                    ),
                    item_by_id,
                )
                or (
                    validate_basis_text(basis) is not None
                    and not path_basis_is_traceable(account, basis)
                )
                or refuses_general_semantic_judgment(
                    semantic,
                    basis,
                    item_id
                    or inflow_item_id
                    or outflow_item_id
                    or next(
                        iter(
                            candidate_item_ids
                            or inflow_candidate_item_ids
                            or outflow_candidate_item_ids
                        ),
                        "",
                    ),
                )
            ):
                missing.append(task_id)
                continue
            if note_id and note_id not in adopted_note_ids:
                missing.append(task_id)
                continue
            seen.add(task_id)
            valid.append(
                {
                    "task_id": task_id,
                    "account": account,
                    "semantic": semantic,
                    "item_id": item_id,
                    "inflow_item_id": inflow_item_id,
                    "outflow_item_id": outflow_item_id,
                    "candidate_item_ids": list(candidate_item_ids),
                    "inflow_candidate_item_ids": list(inflow_candidate_item_ids),
                    "outflow_candidate_item_ids": list(outflow_candidate_item_ids),
                    "confidence": confidence,
                    "basis": basis.strip(),
                    "standard_basis": standard_basis.strip(),
                    "note_id": note_id,
                    "classification_facts": list(facts),
                    "original_paths": list(expected_by_id[task_id].get("original_paths", ())),
                    "customer_level1": str(expected_by_id[task_id].get("customer_level1", "")),
                    "standard_level1": str(expected_by_id[task_id].get("standard_level1", "")),
                }
            )
    missing.extend(task_id for task_id in expected_by_id if task_id not in seen)
    if missing:
        return {"status": "AI 未完成", "missing_ids": sorted(set(missing))}
    pending["valid_results"] = valid
    pending["missing_ids"] = []
    state["account_dictionary"] = pending
    state["account_dictionary_completed"] = True
    state["stage"] = "dictionary_completed"
    _write_dictionary_doc(run_dir, valid, state.get("company_notes", ()))
    _save_state(run_dir, state)
    return {"status": "科目语义已导入", "count": len(valid)}


def import_dictionary_results(run_dir: Path, result_path: Path) -> dict[str, object]:
    """导入受限节点概念；候选和质量始终由固定程序重新计算。"""
    state = _load_state(run_dir)
    _assert_inputs_unchanged(state)
    pending = state.get("account_dictionary")
    if not pending or pending.get("schema_version") != "2.0.0":
        raise RuntimeError("旧版科目路径语义答案不兼容，请重新执行 scan-accounts")
    tasks = tuple(pending.get("tasks", ()))
    if not tasks:
        raise RuntimeError("尚未生成科目路径待判断任务，请先执行 scan-accounts")
    expected_by_id = {str(task["task_id"]): task for task in tasks}
    base_by_account = {
        str(item["account"]): item for item in pending.get("base_results", ())
    }
    rules = load_account_semantic_rules(PROJECT_ROOT)
    valid = list(pending.get("valid_results", ()))
    seen: set[str] = set()
    missing: list[str] = []
    allowed_fields = {"task_id", "account", "node_concepts", "relations"}
    with Path(result_path).open("r", encoding="utf-8-sig") as source:
        for raw in source:
            if not raw.strip():
                continue
            record = json.loads(raw)
            task_id = str(record.get("task_id", ""))
            task = expected_by_id.get(task_id)
            if task is None or task_id in seen:
                if task_id:
                    missing.append(task_id)
                continue
            if set(record).difference(allowed_fields):
                forbidden = set(record).difference(allowed_fields)
                if forbidden.intersection(
                    {
                        "item_id",
                        "candidate_item_ids",
                        "inflow_item_id",
                        "outflow_item_id",
                        "confidence",
                        "quality",
                        "score",
                        "materiality",
                        "action",
                    }
                ):
                    raise ValueError("科目路径Agent不得返回项目、质量或分数")
                missing.append(task_id)
                continue
            account = str(record.get("account", ""))
            if account != str(task["account"]) or account not in base_by_account:
                missing.append(task_id)
                continue
            base = _account_path_result_from_dict(base_by_account[account])
            merged = merge_account_agent_concepts(base, record, rules)
            metadata = {
                key: value
                for key, value in base_by_account[account].items()
                if key
                in {
                    "original_path",
                    "original_paths",
                    "customer_level1",
                    "standard_level1",
                }
            }
            valid.append({**_account_path_result_to_dict(merged), **metadata})
            seen.add(task_id)
    missing.extend(task_id for task_id in expected_by_id if task_id not in seen)
    if missing:
        return {"status": "AI 未完成", "missing_ids": sorted(set(missing))}
    entries = _standardized_entries_from_state(state)
    components = _semantic_components_after_component_build(state, entries)
    final_results = tuple(_account_path_result_from_dict(item) for item in valid)
    coverage = _account_dictionary_coverage(components, final_results)
    pending["valid_results"] = valid
    pending["missing_ids"] = []
    pending["coverage"] = coverage
    state["account_dictionary"] = pending
    state["account_dictionary_completed"] = True
    state["stage"] = "dictionary_completed"
    _write_dictionary_doc(
        run_dir,
        valid,
        state.get("company_notes", ()),
        coverage,
    )
    _save_state(run_dir, state)
    return {"status": "科目语义已导入", "count": len(valid)}


def import_summary_results(run_dir: Path, result_path: Path) -> dict[str, object]:
    """导入受限Agent补充的原文槽位，候选和质量始终由固定程序重算。"""
    state = _load_state(run_dir)
    _assert_inputs_unchanged(state)
    pending = state.get("summary_semantics")
    if not pending or not pending.get("tasks"):
        raise RuntimeError("尚未生成摘要语义待判断任务，请先执行 classify")
    expected = {task["task_id"]: task for task in pending["tasks"]}
    rules = load_summary_rules(PROJECT_ROOT)
    results = {
        result.summary: result
        for result in (
            _summary_result_from_dict(item) for item in pending.get("results", ())
        )
    }
    seen: set[str] = set()
    with Path(result_path).open("r", encoding="utf-8-sig") as source:
        for raw in source:
            if not raw.strip():
                continue
            record = json.loads(raw)
            task_id = str(record.get("task_id", ""))
            task = expected.get(task_id)
            if task is None:
                raise ValueError(f"Agent结果任务编号不属于当前运行：{task_id}")
            if task_id in seen:
                raise ValueError(f"同一导入文件包含重复摘要任务：{task_id}")
            seen.add(task_id)
            base = analyze_summary(str(task["summary"]), rules)
            merged = merge_summary_agent_slots(base, record, rules)
            previous = results.get(merged.summary)
            if previous is not None and previous.status == "agent_complete" and previous != merged:
                raise ValueError(f"摘要任务存在相互冲突的重复结果：{task_id}")
            results[merged.summary] = merged

    missing = sorted(
        task_id
        for task_id, task in expected.items()
        if results.get(str(task["summary"]), SummarySemanticResult(
            str(task["summary"]), "needs_agent", (), (), EvidenceQuality.INVALID, ""
        )).status
        != "agent_complete"
    )
    pending["results"] = [
        _summary_result_to_dict(result)
        for result in sorted(results.values(), key=lambda item: item.summary)
    ]
    pending["missing_ids"] = missing
    state["summary_semantics"] = pending
    if missing:
        _save_state(run_dir, state)
        return {"status": "AI 未完成", "missing_ids": missing}

    ordered = tuple(sorted(results.values(), key=lambda item: item.summary))
    validate_summary_batch(
        ordered,
        tuple(result.summary for result in ordered),
    )
    state["summary_semantics_completed"] = True
    state["summary_semantics_version"] = str(rules["schema_version"])
    state["stage"] = "summary_semantics_completed"
    _save_state(run_dir, state)
    return {"status": "摘要语义已导入", "count": len(ordered)}


def _structure_ai_task_from_dict(payload: Mapping[str, object]) -> StructureAITask:
    return StructureAITask(
        task_id=str(payload["task_id"]),
        voucher_key=str(payload["voucher_key"]),
        review_round=str(payload["review_round"]),
        candidate_entry_id_combinations=tuple(
            tuple(str(value) for value in combination)
            for combination in payload["candidate_entry_id_combinations"]
        ),
        context=str(payload["context"]),
    )


def import_component_structure_ai_results(
    run_dir: Path,
    result_path: Path,
) -> dict[str, object]:
    state = _load_state(run_dir)
    _assert_inputs_unchanged(state)
    tasks = tuple(
        _structure_ai_task_from_dict(item)
        for item in state.get("component_structure_ai_tasks", ())
    )
    if not tasks:
        raise RuntimeError("当前没有待完成的业务组成AI任务")
    payloads = []
    with Path(result_path).open("r", encoding="utf-8-sig") as source:
        for line in source:
            if line.strip():
                payloads.append(json.loads(line))
    prior_payloads = list(state.get("component_structure_ai_results", ()))
    validation = validate_structure_ai_results(
        tasks,
        (*prior_payloads, *payloads),
    )
    terminal = _register_ai_technical_attempts(
        state,
        payloads,
        validation.invalid_ids,
        validation.duplicate_ids,
        validation.missing_ids,
        state_prefix="component_structure_ai",
    )
    pending = tuple(
        task_id for task_id in validation.missing_ids if task_id not in terminal
    )
    blocking_invalid = tuple(
        task_id for task_id in validation.invalid_ids if task_id not in terminal
    )
    blocking_duplicates = tuple(
        task_id for task_id in validation.duplicate_ids if task_id not in terminal
    )
    state["component_structure_ai_results"] = [
        asdict(item) for item in validation.valid_results
    ]
    state["component_structure_ai_validation"] = {
        "missing_ids": pending,
        "invalid_ids": blocking_invalid,
        "duplicate_ids": blocking_duplicates,
        "terminal_failure_ids": tuple(sorted(terminal)),
    }
    if pending or blocking_invalid or blocking_duplicates:
        state["stage"] = "waiting_component_structure_ai"
        _save_state(run_dir, state)
        return {
            "status": "AI 未完成",
            "missing_ids": pending,
            "invalid_ids": blocking_invalid,
            "terminal_failure_ids": tuple(sorted(terminal)),
        }

    tasks_by_voucher: dict[str, list[StructureAITask]] = {}
    results_by_voucher: dict[str, list[StructureAIResult]] = {}
    for task in tasks:
        tasks_by_voucher.setdefault(task.voucher_key, []).append(task)
    for result in validation.valid_results:
        results_by_voucher.setdefault(result.voucher_key, []).append(result)
    selected = {
        str(key): list(value)
        for key, value in dict(
            state.get("component_structure_selections", {})
        ).items()
    }
    basis = {
        str(key): str(value)
        for key, value in dict(
            state.get("component_structure_selection_basis", {})
        ).items()
    }
    followups: list[StructureAITask] = []
    user_requests: list[dict[str, object]] = []
    for request in state.get("component_structure_requests", ()):
        voucher_key = str(request["voucher_key"])
        level = str(request["materiality_level"])
        voucher_tasks = tuple(tasks_by_voucher.get(voucher_key, ()))
        voucher_results = tuple(results_by_voucher.get(voucher_key, ()))
        resolution = resolve_structure_ai_request(
            request,
            level,
            voucher_tasks,
            voucher_results,
            terminal,
        )
        if resolution.status == "selected":
            selected[voucher_key] = list(resolution.selected_entry_ids)
            basis[voucher_key] = resolution.basis_type
        elif resolution.status in {"needs_second", "needs_c"}:
            review_round = "second" if resolution.status == "needs_second" else "C"
            new_task = build_structure_ai_tasks(
                request, level, (review_round,)
            )[0]
            if review_round == "C":
                history = [
                    {
                        "轮次": result.review_round,
                        "所选组合": result.selected_entry_ids,
                        "把握": result.confidence,
                        "理由": result.reason,
                    }
                    for result in voucher_results
                    if result.review_round in {"A", "B"}
                ]
                new_task = replace(
                    new_task,
                    context=(
                        new_task.context
                        + "；既有互盲意见："
                        + json.dumps(history, ensure_ascii=False, separators=(",", ":"))
                        + "；只许整理和比较上述既有候选及理由，不得增加事实。"
                    ),
                )
            followups.append(new_task)
        else:
            enriched = dict(request)
            enriched["ai_review_history"] = [
                asdict(result) for result in voucher_results
            ]
            user_requests.append(enriched)
    if followups:
        all_tasks = (*tasks, *followups)
        state["component_structure_ai_tasks"] = [
            asdict(task) for task in all_tasks
        ]
        state["component_structure_selections"] = selected
        state["component_structure_selection_basis"] = basis
        state["stage"] = "waiting_component_structure_ai"
        for batch_number in range(0, len(followups), 25):
            _write_trace_jsonl(
                run_dir,
                f"业务组成AI后续请求_第{batch_number // 25 + 1:02d}批.jsonl",
                followups[batch_number : batch_number + 25],
            )
        _save_state(run_dir, state)
        return {
            "status": "待AI继续判断业务组成",
            "missing": len(followups),
        }
    state["component_structure_selections"] = selected
    state["component_structure_selection_basis"] = basis
    state["component_structure_requests"] = user_requests
    state["stage"] = (
        "waiting_component_structure_confirmation"
        if user_requests
        else "component_structure_confirmed"
    )
    _save_state(run_dir, state)
    return {
        "status": "待确认业务组成" if user_requests else "业务组成AI判断已完成",
        "pending_user_count": len(user_requests),
    }


def confirm_component_structure(
    run_dir: Path,
    selections: Mapping[str, object],
) -> StageResult:
    """确认金额均可闭合时的业务组成组合；只能从系统列出的组合中选择。"""
    state = _load_state(run_dir)
    _assert_inputs_unchanged(state)
    requests = tuple(state.get("component_structure_requests", ()))
    if not requests:
        raise RuntimeError("当前没有待确认的业务组成")
    request_by_voucher = {
        str(item["voucher_key"]): item for item in requests
    }
    if set(selections) != set(request_by_voucher):
        raise ValueError("必须逐项确认全部待定业务组成，不能漏选或增加凭证")
    confirmed: dict[str, list[str]] = {}
    basis_by_voucher: dict[str, str] = {}
    for voucher_key, selection_payload in selections.items():
        if isinstance(selection_payload, Mapping):
            selected_values = selection_payload.get("entry_ids", ())
            basis_type = str(
                selection_payload.get("basis_type", "existing_evidence")
            )
        else:
            selected_values = selection_payload
            basis_type = "existing_evidence"
        if basis_type not in {"existing_evidence", "independent_external"}:
            raise ValueError(f"业务组成确认依据类型非法：{voucher_key}")
        if not isinstance(selected_values, (list, tuple)):
            raise ValueError(f"业务组成确认必须提供来源行编号数组：{voucher_key}")
        selected = tuple(str(value) for value in selected_values)
        candidates = {
            tuple(str(value) for value in combination)
            for combination in request_by_voucher[voucher_key][
                "candidate_entry_id_combinations"
            ]
        }
        if selected not in candidates:
            raise ValueError(f"业务组成确认不属于既有候选组合：{voucher_key}")
        confirmed[voucher_key] = list(selected)
        basis_by_voucher[voucher_key] = basis_type
    existing_confirmed = dict(state.get("component_structure_selections", {}))
    existing_confirmed.update(confirmed)
    existing_basis = dict(state.get("component_structure_selection_basis", {}))
    existing_basis.update(basis_by_voucher)
    state["component_structure_selections"] = existing_confirmed
    state["component_structure_selection_basis"] = existing_basis
    state.pop("component_structure_requests", None)
    state["stage"] = "component_structure_confirmed"
    _save_state(run_dir, state)
    return StageResult(
        str(state["run_id"]),
        Path(run_dir),
        "component_structure",
        "completed",
        "继续构造业务组成和结构化语义",
    )


def run_classification(run_dir: Path) -> ClassificationStageResult:
    state = _load_state(run_dir)
    _assert_inputs_unchanged(state)
    entries = _standardized_entries_from_state(state)
    if "classification_summary" in state:
        return _result_from_classification(state, run_dir)
    if "cash_scope" not in state:
        raise RuntimeError("请确认现金范围后继续")
    balances = state.get("cash_balances", {})
    if not all(balances.get(key) is not None for key in ("opening_cent", "closing_cent", "fx_cent")):
        raise RuntimeError("请先补充期初、期末现金余额和汇率影响后再分类")
    scope = _scope_from_dict(state["cash_scope"])
    entries_by_file: dict[str, list[NormalizedEntry]] = {}
    for entry in entries:
        entries_by_file.setdefault(entry.source.file_id, []).append(entry)
    profiles = {
        file_id: infer_evidence_profile(file_entries)
        for file_id, file_entries in entries_by_file.items()
    }
    state["standardized_evidence_profiles"] = {
        file_id: _profile_to_dict(profile) for file_id, profile in profiles.items()
    }
    single_sided_file_ids = frozenset(
        file_id
        for file_id, profile in profiles.items()
        if profile.has_flow_amount and profile.has_flow_item
    )
    rough = compute_rough_reconciliation(
        entries,
        profiles,
        int(balances["opening_cent"]),
        int(balances["closing_cent"]),
        int(balances["fx_cent"]),
    )
    state["rough_reconciliation"] = asdict(rough)
    state["single_sided_file_ids"] = sorted(single_sided_file_ids)
    _write_trace_jsonl(run_dir, "粗勾稽留痕.jsonl", (asdict(rough),))
    cleanup_requests = find_cash_row_cleanup_requests(entries, scope)
    if cleanup_requests:
        entry_by_id = {entry.entry_id: entry for entry in entries}
        file_name_by_id = {
            str(item["file_id"]): Path(str(item["path"])).name
            for item in state["files"]
        }
        visible_requests: list[dict[str, object]] = []
        for request in cleanup_requests:
            for entry_id in request.entry_ids:
                entry = entry_by_id[entry_id]
                visible_requests.append(
                    {
                        "文件": file_name_by_id.get(entry.source.file_id, entry.source.file_id),
                        "工作表": entry.source.sheet_name,
                        "凭证": entry.voucher_key,
                        "来源行": entry.source.row_start,
                        "来源单元格": entry.source.cell_range,
                        "摘要": entry.summary,
                        "科目": entry.original_account_name or entry.account_name,
                        "原因": request.reason,
                    }
                )
        state["cash_row_cleanup_requests"] = visible_requests
        state["stage"] = "waiting_cash_row_cleanup"
        _save_state(run_dir, state)
        request_path = Path(run_dir) / "现金分录清洗请求.md"
        lines = [
            "# 现金分录清洗请求",
            "",
            "以下位置无法可靠区分现金分录和非现金分录。请清洗输入或明确现金行后重新开始本次运行。",
            "",
        ]
        for index, item in enumerate(visible_requests, start=1):
            lines.append(
                f"{index}. {item['文件']}｜{item['工作表']}｜第{item['来源行']}行｜"
                f"凭证{item['凭证']}｜科目：{item['科目']}｜原因：{item['原因']}"
            )
        request_path.write_text("\n".join(lines) + "\n", encoding="utf-8-sig")
        return ClassificationStageResult(
            str(state["run_id"]),
            Path(run_dir),
            0,
            "",
            len(entries),
            0,
            0,
            "待用户清洗现金分录",
        )
    _save_state(run_dir, state)
    build = build_cashflow_components(
        entries,
        scope,
        single_sided_file_ids=single_sided_file_ids,
        structure_selections={
            str(voucher_key): tuple(str(value) for value in selected)
            for voucher_key, selected in dict(
                state.get("component_structure_selections", {})
            ).items()
        },
        structure_selection_basis={
            str(voucher_key): str(value)
            for voucher_key, value in dict(
                state.get("component_structure_selection_basis", {})
            ).items()
        },
    )
    if build.structure_requests:
        entry_by_id = {entry.entry_id: entry for entry in entries}
        thresholds = _materiality_from_state(state)
        requests = []
        structure_tasks: list[StructureAITask] = []
        for item in build.structure_requests:
            request = asdict(item)
            level = materiality_level(item.cash_delta_cent, thresholds).value
            request["materiality_level"] = level
            request["candidate_details"] = [
                [
                    {
                        "entry_id": entry_id,
                        "摘要": entry_by_id[entry_id].summary,
                        "科目": entry_by_id[entry_id].account_name,
                        "借方金额分": entry_by_id[entry_id].debit_cent,
                        "贷方金额分": entry_by_id[entry_id].credit_cent,
                        "流量金额分": entry_by_id[entry_id].flow_amount_cent,
                        "原现流项目": entry_by_id[entry_id].original_flow_item,
                    }
                    for entry_id in combination
                ]
                for combination in item.candidate_entry_id_combinations
            ]
            requests.append(request)
            rounds = (
                ("single",)
                if level in {"M0", "M1"}
                else ("A", "B")
                if level == "M2"
                else ()
            )
            structure_tasks.extend(build_structure_ai_tasks(request, level, rounds))
        state["component_structure_requests"] = requests
        state["component_structure_ai_tasks"] = [
            asdict(task) for task in structure_tasks
        ]
        state["component_structure_ai_results"] = []
        state["stage"] = (
            "waiting_component_structure_ai"
            if structure_tasks
            else "waiting_component_structure_confirmation"
        )
        _write_trace_jsonl(run_dir, "业务组成待确认.jsonl", build.structure_requests)
        for batch_number in range(0, len(structure_tasks), 25):
            _write_trace_jsonl(
                run_dir,
                f"业务组成AI请求_第{batch_number // 25 + 1:02d}批.jsonl",
                structure_tasks[batch_number : batch_number + 25],
            )
        _save_state(run_dir, state)
        return ClassificationStageResult(
            str(state["run_id"]),
            Path(run_dir),
            0,
            "",
            len(entries),
            0,
            len(structure_tasks) if structure_tasks else len(requests),
            "待AI判断业务组成" if structure_tasks else "待确认业务组成",
        )
    state.pop("component_structure_requests", None)
    state["semantic_account_paths"] = sorted(
        {
            path
            for component in build.components
            for path in component.counterpart_accounts
            if path.strip()
        }
    )
    _save_state(run_dir, state)
    if not state.get("account_dictionary_completed"):
        scan = scan_accounts(run_dir)
        state = _load_state(run_dir)
        if scan.get("status") == "待确认公司特殊规则":
            return ClassificationStageResult(
                str(state["run_id"]), Path(run_dir), 0, "", 0, 0, int(scan.get("missing", 0)),
                "待确认公司特殊规则",
            )
        if not state.get("account_dictionary_completed"):
            return ClassificationStageResult(
                str(state["run_id"]),
                Path(run_dir),
                0,
                "",
                len(entries),
                0,
                int(scan["missing"]),
                "待科目语义确认",
            )
    entry_by_id = {entry.entry_id: entry for entry in entries}
    raw_components = tuple(
        replace(
            component,
            voucher_date=next((entry_by_id[key].voucher_date for key in component.source_keys if key in entry_by_id), ""),
            voucher_no=next((entry_by_id[key].voucher_no for key in component.source_keys if key in entry_by_id), ""),
            source_file_ids=tuple(
                sorted({entry_by_id[key].source.file_id for key in component.source_keys if key in entry_by_id})
            ),
        )
        for component in build.components
    )
    entries_by_voucher: dict[str, list[NormalizedEntry]] = {}
    for entry in entries:
        entries_by_voucher.setdefault(entry.voucher_key, []).append(entry)
    components = tuple(
        replace(
            component,
            original_counterpart_accounts=tuple(
                dict.fromkeys(
                    entry.original_account_name or entry.account_name
                    for entry in entries_by_voucher.get(component.voucher_key, ())
                    if entry.account_name in component.counterpart_accounts
                )
            ),
            account_mapping_status="confirmed",
        )
        for component in raw_components
    )
    if not state.get("summary_semantics_completed"):
        summary_rules = load_summary_rules(PROJECT_ROOT)
        semantic_results = tuple(
            analyze_summary(summary, summary_rules)
            for summary in sorted(
                {
                    component.summary.strip()
                    for component in components
                    if component.summary.strip()
                }
            )
        )
        tasks = tuple(
            task
            for task in (build_summary_agent_task(result) for result in semantic_results)
            if task is not None
        )
        state["summary_semantics"] = {
            "tasks": list(tasks),
            "results": [_summary_result_to_dict(result) for result in semantic_results],
            "missing_ids": [task["task_id"] for task in tasks],
        }
        state["summary_semantics_version"] = str(summary_rules["schema_version"])
        state["summary_semantics_completed"] = not tasks
        if tasks:
            _write_summary_batches(run_dir, tasks)
            state["stage"] = "waiting_summary_semantics"
            _save_state(run_dir, state)
            return ClassificationStageResult(
                str(state["run_id"]),
                Path(run_dir),
                len(components),
                "",
                len(entries),
                sum(item.cash_delta_cent for item in components),
                len(tasks),
                "待摘要语义确认",
            )
        validate_summary_batch(
            semantic_results,
            tuple(result.summary for result in semantic_results),
        )
    rules = load_rule_pack(PROJECT_ROOT)
    candidate_decisions = classify_all(
        components,
        rules,
        _dictionary_from_state(state),
        _summary_semantics_from_state(state),
    )
    routing = route_classification_decisions(
        components,
        candidate_decisions,
        _materiality_from_state(state),
        company_notes=state.get("company_notes", ()),
        automatic_change_threshold=int(state["automatic_change_threshold"]),
    )
    decisions = routing.decisions
    checked = validate_classification(components, decisions)
    if not checked.valid:
        raise RuntimeError("自动分类不变量失败：" + "；".join(checked.errors))
    tasks = routing.ai_tasks
    serialized_components = [asdict(item) for item in components]
    serialized_decisions = [asdict(item) for item in decisions]
    digest_source = json.dumps(serialized_components, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    component_hash = hashlib.sha256(digest_source.encode("utf-8")).hexdigest()
    has_pending_human = any(
        not item.resolved
        and not item.excluded
        and item.decision_action
        not in {"ai_review", "double_ai_review"}
        for item in decisions
    )
    classification_status = (
        "waiting_ai"
        if tasks
        else "waiting_human"
        if has_pending_human
        else "classification_completed"
    )
    summary = {
        "component_count": len(components),
        "component_hash": component_hash,
        "source_entry_count": len(entries),
        "cash_delta_cent": sum(item.cash_delta_cent for item in components),
        "ai_tasks_missing": len(tasks),
        "status": classification_status,
    }
    store = _store(run_dir)
    with store.stage("classification") as connection:
        vouchers: dict[str, list[str]] = {}
        for entry in entries:
            vouchers.setdefault(entry.voucher_key, []).append(entry.entry_id)
        connection.executemany(
            "INSERT OR REPLACE INTO voucher(record_id, payload_json) VALUES (?, ?)",
            (
                (
                    voucher_key,
                    json.dumps({"entry_ids": entry_ids}, ensure_ascii=False),
                )
                for voucher_key, entry_ids in vouchers.items()
            ),
        )
        connection.executemany(
            "INSERT INTO cashflow_component(record_id, payload_json) VALUES (?, ?)",
            ((item.component_id, json.dumps(asdict(item), ensure_ascii=False)) for item in components),
        )
        connection.executemany(
            "INSERT INTO classification_decision(record_id, payload_json) VALUES (?, ?)",
            ((item.component_id, json.dumps(asdict(item), ensure_ascii=False)) for item in decisions),
        )
        connection.executemany(
            "INSERT INTO source_allocation(record_id, payload_json) VALUES (?, ?)",
            (
                (
                    stable_id("ALLOC", item.component_id, item.entry_id),
                    json.dumps(asdict(item), ensure_ascii=False),
                )
                for item in build.source_allocations
            ),
        )
        connection.executemany(
            "INSERT INTO evidence_assessment(record_id, payload_json) VALUES (?, ?)",
            (
                (
                    item.component_id,
                    json.dumps(
                        _evidence_assessment_payload(item),
                        ensure_ascii=False,
                    ),
                )
                for item in decisions
            ),
        )
        connection.executemany(
            "INSERT INTO materiality_assessment(record_id, payload_json) VALUES (?, ?)",
            (
                (item.record_id, json.dumps(asdict(item), ensure_ascii=False))
                for item in routing.materiality_assessments
            ),
        )
        connection.executemany(
            "INSERT INTO decision_route(record_id, payload_json) VALUES (?, ?)",
            (
                (
                    item.component_id,
                    json.dumps(
                        {
                            "component_id": item.component_id,
                            "action": item.decision_action,
                            "materiality_level": item.materiality_level,
                            "resolved": item.resolved,
                            "decision_source": item.decision_source,
                        },
                        ensure_ascii=False,
                    ),
                )
                for item in decisions
            ),
        )
        connection.executemany(
            "INSERT INTO ai_task(record_id, payload_json) VALUES (?, ?)",
            ((item.task_id, json.dumps(asdict(item), ensure_ascii=False)) for item in tasks),
        )
        # 同一凭证、同一现金腿、同一冲抵金额可能对应多条对等内部划转腿，
        # 其 stable_id 会撞车。这里在基础编号后追加出现序号，保证每行 record_id 唯一，
        # 既保留全量内部划转留痕，又避免 UNIQUE 约束导致整单失败。
        transfer_tuples = [
            (item.voucher_key, item.entry_id, item.matched_cent)
            for item in build.excluded_internal_transfers
        ]
        transfer_ids: list[str] = []
        seen_transfer_ids: dict[str, int] = {}
        for item, tup in zip(build.excluded_internal_transfers, transfer_tuples):
            base = stable_id("ITR", *tup)
            occurrence = seen_transfer_ids.get(base, 0)
            seen_transfer_ids[base] = occurrence + 1
            transfer_ids.append(base if occurrence == 0 else f"{base}_{occurrence+1:02d}")
        connection.executemany(
            "INSERT INTO internal_transfer(record_id, payload_json) VALUES (?, ?)",
            (
                (
                    transfer_id,
                    json.dumps(asdict(item), ensure_ascii=False),
                )
                for transfer_id, item in zip(transfer_ids, build.excluded_internal_transfers)
            ),
        )
    state["components"] = serialized_components
    state["decisions"] = serialized_decisions
    state["ai_tasks"] = [asdict(item) for item in tasks]
    state["ai_schema_version"] = 3
    state["materiality_assessments"] = [
        asdict(item) for item in routing.materiality_assessments
    ]
    state["source_allocations"] = [
        asdict(item) for item in build.source_allocations
    ]
    state["internal_transfers"] = [asdict(item) for item in build.excluded_internal_transfers]
    state["classification_summary"] = summary
    state["stage"] = summary["status"]
    for batch_number, batch in enumerate(chunk_ai_tasks(tasks), 1):
        write_ai_tasks_jsonl(
            _trace_dir(run_dir) / f"AI复核请求_第{batch_number:02d}批.jsonl",
            batch,
        )
    _write_trace_jsonl(run_dir, "内部划转排除.jsonl", build.excluded_internal_transfers)
    if not tasks and not has_pending_human:
        _prepare_consistency_stage(state, run_dir)
    _save_state(run_dir, state)
    return _result_from_classification(state, run_dir)


def _build_pending_ai_followups(
    decisions: Sequence[ClassificationDecision],
    components: Sequence[CashflowComponent],
    existing_tasks: Sequence[AITask],
    company_notes: Sequence[Mapping[str, object]],
) -> tuple[AITask, ...]:
    """根据尚未完成的动作生成下一轮任务，且不把既有AI意见传给互盲复核。"""
    component_by_id = {item.component_id: item for item in components}
    tasks_by_component: dict[str, list[AITask]] = {}
    for task in existing_tasks:
        tasks_by_component.setdefault(task.component_id, []).append(task)
    followups: list[AITask] = []
    for decision in decisions:
        component = component_by_id.get(decision.component_id)
        if component is None:
            raise RuntimeError(f"AI后续复核缺少业务组成：{decision.component_id}")
        current_tasks = tasks_by_component.get(decision.component_id, ())
        if decision.decision_action in {
            "double_ai_review",
            "ai_double_followup_review",
        }:
            existing_slots = {
                slot
                for slot in ("A", "B")
                if any(f"；独立复核{slot}：" in task.context for task in current_tasks)
            }
            missing_slots = tuple(slot for slot in ("A", "B") if slot not in existing_slots)
            if missing_slots:
                followups.extend(
                    build_blind_ai_tasks(
                        component,
                        decision,
                        missing_slots,
                        company_notes,
                    )
                )
        elif decision.decision_action == "ai_third_review" and not any(
            "；独立复核C：" in task.context for task in current_tasks
        ):
            followups.extend(
                build_blind_ai_tasks(
                    component,
                    decision,
                    ("C",),
                    company_notes,
                )
            )
    return tuple(followups)


def _register_ai_technical_attempts(
    state: dict[str, object],
    payloads: Sequence[Mapping[str, object]],
    invalid_ids: Sequence[str],
    duplicate_ids: Sequence[str],
    missing_ids: Sequence[str] = (),
    *,
    state_prefix: str = "ai",
) -> set[str]:
    """记录明确提交但校验失败的任务；第三次失败后形成终态。"""
    attempts = {
        str(task_id): int(count)
        for task_id, count in dict(
            state.get(f"{state_prefix}_task_attempts", {})
        ).items()
    }
    terminal = {
        str(task_id)
        for task_id in state.get(f"{state_prefix}_terminal_failure_ids", ())
    }
    submitted = {
        str(payload.get("task_id", "")) for payload in payloads if payload.get("task_id")
    }
    failed = submitted.intersection({*invalid_ids, *duplicate_ids}) | set(missing_ids)
    failure_log = list(state.get(f"{state_prefix}_technical_failure_log", ()))
    for task_id in sorted(failed):
        if task_id in terminal:
            continue
        attempts[task_id] = attempts.get(task_id, 0) + 1
        is_terminal = attempts[task_id] >= 3
        if is_terminal:
            terminal.add(task_id)
        failure_log.append(
            {
                "task_id": task_id,
                "attempt": attempts[task_id],
                "status": "无有效结果" if is_terminal else "技术失败，可重试",
            }
        )
    state[f"{state_prefix}_task_attempts"] = attempts
    state[f"{state_prefix}_terminal_failure_ids"] = sorted(terminal)
    state[f"{state_prefix}_technical_failure_log"] = failure_log
    return terminal


def import_ai_results(run_dir: Path, result_path: Path) -> AIStageResult:
    state = _load_state(run_dir)
    _assert_inputs_unchanged(state)
    payloads = []
    with Path(result_path).open("r", encoding="utf-8-sig") as source:
        for line in source:
            if line.strip():
                payloads.append(json.loads(line))
    leaf_ids = {
        item.item_id for item in load_rule_pack(PROJECT_ROOT).statement_items if item.is_leaf
    }
    if int(state.get("ai_schema_version", 0)) < 2:
        raise RuntimeError("旧版AI结果格式不可复用，请新建运行目录")
    if int(state.get("ai_schema_version", 0)) >= 2:
        tasks = tuple(
            _ai_task_from_dict(item) for item in state.get("ai_tasks", ())
        )
        prior_results = tuple(
            structured_ai_result_from_mapping(item)
            for item in state.get("structured_ai_validation", {}).get(
                "valid_results", ()
            )
        )
        validation = merge_structured_ai_results(
            tasks,
            prior_results,
            payloads,
            leaf_ids,
        )
        terminal_failure_ids = _register_ai_technical_attempts(
            state,
            payloads,
            validation.invalid_ids,
            validation.duplicate_ids,
            validation.missing_ids,
        )
        pending_ids = tuple(
            task_id
            for task_id in validation.missing_ids
            if task_id not in terminal_failure_ids
        )
        blocking_invalid_ids = tuple(
            task_id
            for task_id in validation.invalid_ids
            if task_id not in terminal_failure_ids
        )
        blocking_duplicate_ids = tuple(
            task_id
            for task_id in validation.duplicate_ids
            if task_id not in terminal_failure_ids
        )
        ai_round_complete = not (
            pending_ids or blocking_invalid_ids or blocking_duplicate_ids
        )
        state["structured_ai_validation"] = {
            "valid_results": [asdict(item) for item in validation.valid_results],
            "missing_ids": validation.missing_ids,
            "duplicate_ids": validation.duplicate_ids,
            "invalid_ids": validation.invalid_ids,
            "terminal_failure_ids": tuple(sorted(terminal_failure_ids)),
            "status": "AI 已完成" if ai_round_complete else "AI 未完成",
        }
        _persist_ai_results(run_dir, "结构化复核", validation.valid_results)
        _write_trace_jsonl(run_dir, "AI结构化复核结果.jsonl", validation.valid_results)
        state["classification_summary"]["ai_tasks_missing"] = len(pending_ids)
        if ai_round_complete:
            rules = load_rule_pack(PROJECT_ROOT)
            item_by_id = rules.item_by_id
            decisions = tuple(
                _decision_from_dict(item) for item in state["decisions"]
            )
            resolved = resolve_structured_ai_results(
                decisions,
                tasks,
                validation.valid_results,
                {key: item.name for key, item in item_by_id.items()},
                {
                    key: item.normal_direction
                    for key, item in item_by_id.items()
                },
                failed_task_ids=terminal_failure_ids,
                automatic_change_threshold=int(
                    state["automatic_change_threshold"]
                ),
            )
            state["decisions"] = [asdict(item) for item in resolved]
            with _store(run_dir).stage("structured_ai_resolution") as connection:
                connection.executemany(
                    "INSERT OR REPLACE INTO classification_decision(record_id, payload_json) VALUES (?, ?)",
                    (
                        (
                            item.component_id,
                            json.dumps(asdict(item), ensure_ascii=False),
                        )
                        for item in resolved
                    ),
                )
                connection.executemany(
                    "INSERT OR REPLACE INTO evidence_assessment(record_id, payload_json) VALUES (?, ?)",
                    (
                        (
                            item.component_id,
                            json.dumps(
                                _evidence_assessment_payload(item),
                                ensure_ascii=False,
                            ),
                        )
                        for item in resolved
                    ),
                )
                connection.executemany(
                    "INSERT OR REPLACE INTO decision_route(record_id, payload_json) VALUES (?, ?)",
                    (
                        (
                            item.component_id,
                            json.dumps(
                                {
                                    "component_id": item.component_id,
                                    "action": item.decision_action,
                                    "materiality_level": item.materiality_level,
                                    "resolved": item.resolved,
                                    "decision_source": item.decision_source,
                                },
                                ensure_ascii=False,
                            ),
                        )
                        for item in resolved
                    ),
                )
            followups = _build_pending_ai_followups(
                resolved,
                tuple(
                    _component_from_dict(payload)
                    for payload in state.get("components", ())
                ),
                tasks,
                state.get("company_notes", ()),
            )
            if followups:
                state["ai_tasks"] = [*state.get("ai_tasks", ()), *[asdict(item) for item in followups]]
                state["classification_summary"]["ai_tasks_missing"] = len(followups)
                state["classification_summary"]["status"] = "waiting_ai"
                state["stage"] = "waiting_ai"
                for batch_number, batch in enumerate(chunk_ai_tasks(followups), 1):
                    write_ai_tasks_jsonl(
                        _trace_dir(run_dir)
                        / f"AI后续复核请求_第{batch_number:02d}批.jsonl",
                        batch,
                    )
                status = "待AI后续复核"
                missing_count = len(followups)
            elif all(item.resolved or item.excluded for item in resolved):
                status, missing_count = _prepare_consistency_stage(state, run_dir)
            else:
                state["classification_summary"]["ai_tasks_missing"] = 0
                state["classification_summary"]["status"] = "waiting_human"
                state["stage"] = "waiting_human"
                status = "待完成人工决定"
                missing_count = 0
        else:
            state["classification_summary"]["status"] = "waiting_ai"
            state["stage"] = "waiting_ai"
            status = "AI 未完成"
            missing_count = len(pending_ids)
        _save_state(run_dir, state)
        return AIStageResult(
            str(state["run_id"]),
            Path(run_dir),
            len(validation.valid_results),
            missing_count,
            status,
        )


def _requires_overall_manual_override(decision: ClassificationDecision) -> bool:
    """整体重要性通常转人工；仅保留已确认的空白项目90分AI补列例外。"""
    blank_score_90_ai_fill = bool(
        decision.original_item_state in {"blank", "unstandardizable"}
        and decision.evidence_score == 90
        and decision.decision_action == "automatic_fill"
        and decision.decision_source.startswith("ai_")
    )
    return decision.decision_source != "manual" and not blank_score_90_ai_fill


def _needs_overall_manual_listing(
    decision: ClassificationDecision,
    amount_cent: int,
    overall_cent: int,
) -> bool:
    return (
        abs(amount_cent) >= overall_cent
        and _requires_overall_manual_override(decision)
    )


def confirm_manual_decisions(
    run_dir: Path,
    entries: Sequence[Mapping[str, object]],
) -> StageResult:
    state = _load_state(run_dir)
    _assert_inputs_unchanged(state)
    if "classification_summary" not in state:
        raise RuntimeError("请先完成分类并形成待人工决定事项")
    rules = load_rule_pack(PROJECT_ROOT)
    leaf_items = {
        item.item_id: item for item in rules.statement_items if item.is_leaf
    }
    decisions = tuple(
        _decision_from_dict(item) for item in state.get("decisions", ())
    )
    decision_by_id = {item.component_id: item for item in decisions}
    seen: set[str] = set()
    replacements: dict[str, ClassificationDecision] = {}
    records: list[dict[str, object]] = []
    decided_at = datetime.now(timezone.utc).isoformat()
    for index, entry in enumerate(entries, 1):
        component_id = str(entry.get("component_id", "")).strip()
        if component_id not in decision_by_id:
            raise ValueError(f"第 {index} 条人工决定找不到对应业务：{component_id}")
        if component_id in seen:
            raise ValueError(f"人工决定重复：{component_id}")
        seen.add(component_id)
        original = decision_by_id[component_id]
        if original.resolved or original.excluded:
            raise ValueError(f"业务已经取得决定，不能重复覆盖：{component_id}")
        basis = str(entry.get("basis", "")).strip()
        operator = str(entry.get("operator", "")).strip()
        excluded = bool(entry.get("exclude", False))
        item_id = str(entry.get("item_id", "")).strip()
        if excluded == bool(item_id):
            raise ValueError("人工决定必须且只能选择一个正表项目或明确排除")
        if item_id and item_id not in leaf_items:
            raise ValueError(f"人工选择的正表项目无效：{item_id}")
        if excluded:
            current = replace(
                original,
                system_item_id="",
                system_item_name="",
                normal_direction="net",
                reason=(
                    f"{original.reason}；人工明确排除"
                    + (f"：{basis}" if basis else "")
                ),
                resolved=True,
                excluded=True,
                decision_source="manual",
                decision_action="manual_exclude",
            )
        else:
            item = leaf_items[item_id]
            current = replace(
                original,
                system_item_id=item_id,
                system_item_name=item.name,
                normal_direction=item.normal_direction,
                reason=(
                    f"{original.reason}；人工确认"
                    + (f"：{basis}" if basis else "")
                ),
                resolved=True,
                excluded=False,
                decision_source="manual",
                decision_action="manual_decision",
            )
        replacements[component_id] = current
        records.append(
            {
                "component_id": component_id,
                "item_id": item_id,
                "excluded": excluded,
                "basis": basis,
                "external_source": str(entry.get("external_source", "")).strip(),
                "operator": operator,
                "decided_at": str(entry.get("decided_at", "")).strip() or decided_at,
                "original_evidence_score": original.evidence_score,
                "amount_impact_cent": next(
                    int(item["cash_delta_cent"])
                    for item in state.get("components", ())
                    if item["component_id"] == component_id
                ),
                "suggest_new_rule": bool(entry.get("suggest_new_rule", False)),
            }
        )
    updated = tuple(
        replacements.get(item.component_id, item) for item in decisions
    )
    state["decisions"] = [asdict(item) for item in updated]
    state["human_decisions"] = [*state.get("human_decisions", ()), *records]
    pending = [
        item.component_id for item in updated if not item.resolved and not item.excluded
    ]
    state["classification_summary"]["ai_tasks_missing"] = 0
    state["classification_summary"]["status"] = (
        "waiting_human" if pending else "manual_decisions_completed"
    )
    state["stage"] = state["classification_summary"]["status"]
    for key in (
        "overall_status",
        "workbook_path",
        "final_readiness",
        "consistency_groups",
        "consistency_resolution",
    ):
        state.pop(key, None)
    if not pending:
        _prepare_consistency_stage(state, run_dir)
    with _store(run_dir).stage("manual_decision") as connection:
        connection.executemany(
            "INSERT OR REPLACE INTO human_decision(record_id, payload_json) VALUES (?, ?)",
            (
                (
                    record["component_id"],
                    json.dumps(record, ensure_ascii=False),
                )
                for record in records
            ),
        )
        connection.executemany(
            "INSERT OR REPLACE INTO classification_decision(record_id, payload_json) VALUES (?, ?)",
            (
                (item.component_id, json.dumps(asdict(item), ensure_ascii=False))
                for item in updated
            ),
        )
        connection.executemany(
            "INSERT OR REPLACE INTO decision_route(record_id, payload_json) VALUES (?, ?)",
            (
                (
                    item.component_id,
                    json.dumps(
                        {
                            "component_id": item.component_id,
                            "action": item.decision_action,
                            "materiality_level": item.materiality_level,
                            "resolved": item.resolved,
                            "decision_source": item.decision_source,
                        },
                        ensure_ascii=False,
                    ),
                )
                for item in updated
            ),
        )
    _write_trace_jsonl(run_dir, "人工决定记录.jsonl", records)
    _save_state(run_dir, state)
    return StageResult(
        str(state["run_id"]),
        Path(run_dir),
        "manual_decision",
        "人工决定已记录",
        (
            "仍有待人工决定事项"
            if state.get("stage") == "waiting_human"
            else "可生成最终工作簿"
        ),
    )


def _assert_agent_gates_closed(state: Mapping[str, object]) -> None:
    if state.get("component_structure_requests"):
        raise RuntimeError("业务组成结构仍待Agent或用户确认，不能生成最终工作簿")


def finalize_run(run_dir: Path) -> FinalizeResult:
    state = _load_state(run_dir)
    _assert_inputs_unchanged(state)
    _assert_agent_gates_closed(state)
    if state.get("overall_status") and state.get("workbook_path"):
        existing_output = Path(str(state["workbook_path"]))
        if existing_output.is_file():
            try:
                workbook = load_workbook(existing_output, read_only=True, data_only=False)
                workbook.close()
                recalculation = state.get("excel_recalculation", {})
                existing_hash = hashlib.sha256(existing_output.read_bytes()).hexdigest()
                final_recalculation_valid = (
                    state.get("overall_status") != "最终可使用"
                    or (
                        isinstance(recalculation, Mapping)
                        and recalculation.get("status") == "completed"
                        and recalculation.get("workbook_sha256") == existing_hash
                    )
                )
                if final_recalculation_valid:
                    return FinalizeResult(
                        str(state["run_id"]),
                        Path(run_dir),
                        existing_output,
                        str(state["overall_status"]),
                    )
                state["recovery_note"] = (
                    "原最终工作簿缺少有效的Excel完整重算记录，已改为重建"
                )
            except Exception as error:
                state["recovery_note"] = f"原结果文件无法打开，已改为重建：{error}"
    if "classification_summary" not in state:
        raise RuntimeError("请先完成自动分类")
    if int(state["classification_summary"]["ai_tasks_missing"]) > 0:
        raise RuntimeError("AI 复核尚未逐编号完成，不能生成最终结果")

    components = tuple(_component_from_dict(item) for item in state["components"])
    decisions = tuple(_decision_from_dict(item) for item in state["decisions"])
    component_by_id = {item.component_id: item for item in components}
    overall_cent = _materiality_from_state(state).overall_cent
    decisions = tuple(
        replace(
            decision,
            resolved=False,
            decision_action="human_decision",
        )
        if (
            decision.resolved
            and not decision.excluded
            and _needs_overall_manual_listing(
                decision,
                component_by_id[decision.component_id].cash_delta_cent,
                overall_cent,
            )
        )
        else decision
        for decision in decisions
    )
    state["decisions"] = [asdict(item) for item in decisions]
    with _store(run_dir).stage("finalize_decision_routes") as connection:
        connection.executemany(
            "INSERT OR REPLACE INTO classification_decision(record_id, payload_json) VALUES (?, ?)",
            (
                (item.component_id, json.dumps(asdict(item), ensure_ascii=False))
                for item in decisions
            ),
        )
        connection.executemany(
            "INSERT OR REPLACE INTO decision_route(record_id, payload_json) VALUES (?, ?)",
            (
                (
                    item.component_id,
                    json.dumps(
                        {
                            "component_id": item.component_id,
                            "action": item.decision_action,
                            "materiality_level": item.materiality_level,
                            "resolved": item.resolved,
                            "decision_source": item.decision_source,
                        },
                        ensure_ascii=False,
                    ),
                )
                for item in decisions
            ),
        )
    _save_state(run_dir, state)
    unfinished_ai = tuple(
        decision.component_id
        for decision in decisions
        if not decision.resolved
        and not decision.excluded
        and decision.decision_action
        in {
            "ai_review",
            "dual_ai_review",
            "double_ai_review",
            "ai_double_followup_review",
            "ai_third_review",
        }
    )
    if unfinished_ai:
        raise RuntimeError(
            "AI复核流程尚未完成，不能生成最终工作簿："
            + "、".join(unfinished_ai)
        )
    source_allocations = tuple(
        ComponentSourceAllocation(**item)
        for item in state.get("source_allocations", ())
    )
    readiness = validate_final_readiness(
        components,
        decisions,
        source_allocations,
        ai_tasks_missing=int(state["classification_summary"]["ai_tasks_missing"]),
        mapping_complete=(
            not state.get("mapping_questions")
            and not state.get("account_mapping_questions")
            and bool(state.get("account_mapping_records"))
            and all(
                item.get("status") == "confirmed"
                for item in state.get("account_mapping_records", ())
            )
        ),
        versions_consistent=True,
    )
    state["final_readiness"] = asdict(readiness)
    _save_state(run_dir, state)
    entries = tuple(_entry_from_dict(item) for item in state["entries"])
    entry_by_id = {entry.entry_id: entry for entry in entries}
    rules = load_rule_pack(PROJECT_ROOT)
    leaf_item_ids = tuple(
        item.item_id for item in rules.statement_items if item.is_leaf
    )

    def original_statement_amount(
        decision: ClassificationDecision,
        component: CashflowComponent,
    ) -> int:
        item = rules.item_by_id.get(decision.original_standard_item_id)
        if item is None:
            return 0
        return statement_amount_cent(
            component.cash_delta_cent,
            item.normal_direction,
        )
    comparison = None
    existing = None
    existing_path = state.get("existing_statement_path")
    if existing_path:
        reference_years = frozenset(
            int(year.group(1))
            for entry in entries
            if (year := re.match(r"(\d{4})", entry.voucher_date))
        )
        existing = parse_existing_statement(Path(str(existing_path)), rules, reference_years)
        if isinstance(existing, MappingQuestion):
            raise RuntimeError("客户现有正表仍有无法映射的项目，请先确认")

    balances = state.get("cash_balances", {})
    layers = build_statement_layers(
        components,
        decisions,
        rules,
        existing=existing,
        opening_cent=balances.get("opening_cent"),
        fx_cent=balances.get("fx_cent"),
        additional_system_adjustments=internal_transfer_statement_adjustments(
            entries,
            tuple(
                InternalTransferLeg(**item)
                for item in state.get("internal_transfers", ())
            ),
            rules,
        ),
    )
    statement = layers.automatic_baseline
    statement_check = validate_statement(statement)
    if not statement_check.valid:
        raise RuntimeError("正表金额勾稽失败：" + "；".join(statement_check.errors))
    if existing is not None:
        comparison = compare_statement(
            existing,
            statement,
            system_adjustments=layers.system_adjustments,
            manual_adjustments=layers.manual_adjustments,
            detail_reconstruction=layers.detail_reconstruction,
        )
    reconciliation = reconcile_cash(
        layers.final_statement,
        balances.get("opening_cent"),
        balances.get("closing_cent"),
        balances.get("fx_cent"),
    )
    duplicate_groups = assign_duplicate_items(
        find_suspected_duplicates(
            components, _materiality_from_state(state).performance_cent
        ),
        decisions,
    )
    component_by_id = {item.component_id: item for item in components}
    file_name_by_id = {
        str(item["file_id"]): Path(str(item["path"])).name for item in state["files"]
    }
    consistency_unresolved = tuple(
        state.get("consistency_resolution", {}).get("unresolved", ())
    )
    consistency_unresolved_component_ids = {
        str(component_id)
        for payload in consistency_unresolved
        for component_id in payload.get("component_ids", ())
    }
    unresolved_list = [
        UnresolvedDecision(
            component_id=decision.component_id,
            cash_delta_cent=component_by_id[decision.component_id].cash_delta_cent,
            cash_direction=(
                "inflow" if component_by_id[decision.component_id].cash_delta_cent > 0 else "outflow"
            ),
            original_item=component_by_id[decision.component_id].original_item_text,
            system_item_id=(
                decision.system_item_id or decision.original_standard_item_id
            ),
            review_status="统一动作表要求人工决定",
            counterpart_group=_review_text_pattern(
                "、".join(component_by_id[decision.component_id].counterpart_accounts)
            ),
            summary_pattern=_review_text_pattern(
                component_by_id[decision.component_id].summary
            ),
            alternative_item_ids=tuple(
                dict.fromkeys(
                    item
                    for item in decision.candidate_item_ids
                    if item
                    and item
                    != (decision.system_item_id or decision.original_standard_item_id)
                )
            ) or tuple(
                item_id
                for item_id in leaf_item_ids
                if item_id
                != (decision.system_item_id or decision.original_standard_item_id)
            ),
            reason=decision.reason,
            system_statement_amount_cent=original_statement_amount(
                decision, component_by_id[decision.component_id]
            ),
            source_locations=tuple(
                dict.fromkeys(
                    f"{file_name_by_id.get(entry_by_id[key].source.file_id, entry_by_id[key].source.file_id)}|{entry_by_id[key].source.sheet_name}|{entry_by_id[key].source.cell_range}"
                    for key in component_by_id[decision.component_id].source_keys
                    if key in entry_by_id
                )
            ),
            mandatory=(
                abs(component_by_id[decision.component_id].cash_delta_cent)
                >= _materiality_from_state(state).overall_cent
            ),
            baseline_item_code=decision.original_standard_item_id,
        )
        for decision in decisions
        if (
            not decision.resolved
            and not decision.excluded
            and decision.component_id not in consistency_unresolved_component_ids
        )
    ]
    for payload in consistency_unresolved:
        candidate_by_component = {
            str(item[0]): tuple(str(value) for value in item[1])
            for item in payload.get("candidate_item_ids", ())
        }
        for component_id in payload.get("component_ids", ()):
            component_id = str(component_id)
            component = component_by_id[component_id]
            decision = next(
                item for item in decisions if item.component_id == component_id
            )
            unresolved_list.append(
                UnresolvedDecision(
                    component_id=component_id,
                    cash_delta_cent=component.cash_delta_cent,
                    cash_direction=(
                        "inflow" if component.cash_delta_cent > 0 else "outflow"
                    ),
                    original_item=component.original_item_text,
                    system_item_id=(
                        decision.system_item_id or decision.original_standard_item_id
                    ),
                    review_status=(
                        "同一业务组仍待人工决定："
                        + str(payload["group_id"])
                    ),
                    counterpart_group=_review_text_pattern(
                        "、".join(component.counterpart_accounts)
                    ),
                    summary_pattern=_review_text_pattern(component.summary),
                    alternative_item_ids=tuple(
                        item_id
                        for item_id in candidate_by_component.get(component_id, ())
                        if item_id
                        != (decision.system_item_id or decision.original_standard_item_id)
                    ),
                    reason=str(payload["reason"]),
                    system_statement_amount_cent=original_statement_amount(
                        decision, component
                    ),
                    source_locations=tuple(
                        dict.fromkeys(
                            f"{file_name_by_id.get(entry_by_id[key].source.file_id, entry_by_id[key].source.file_id)}|{entry_by_id[key].source.sheet_name}|{entry_by_id[key].source.cell_range}"
                            for key in component.source_keys
                            if key in entry_by_id
                        )
                    ),
                    group_impact_cent=int(payload["gross_cent"]),
                    mandatory=abs(component.cash_delta_cent) >= _materiality_from_state(state).overall_cent,
                    baseline_item_code=decision.original_standard_item_id,
                )
            )
    unresolved = tuple(unresolved_list)
    # 大额自动判断即使已确定也须进入人工表；已经由人工决定的事项不得重复确认。
    listed_ids = {item.component_id for item in unresolved_list}
    for decision in decisions:
        if (
            decision.excluded
            or decision.decision_source == "manual"
            or decision.component_id in listed_ids
        ):
            continue
        component = component_by_id[decision.component_id]
        if not _needs_overall_manual_listing(
            decision,
            component.cash_delta_cent,
            _materiality_from_state(state).overall_cent,
        ):
            continue
        unresolved_list.append(
            UnresolvedDecision(
                component_id=component.component_id,
                cash_delta_cent=component.cash_delta_cent,
                cash_direction=(
                    "inflow" if component.cash_delta_cent > 0 else "outflow"
                ),
                original_item=component.original_item_text,
                system_item_id=decision.system_item_id,
                review_status="达到财务报表整体重要性，强制人工复核",
                counterpart_group=_review_text_pattern(
                    "、".join(component.counterpart_accounts)
                ),
                summary_pattern=_review_text_pattern(component.summary),
                alternative_item_ids=(),
                reason=decision.reason,
                system_statement_amount_cent=original_statement_amount(
                    decision, component
                ),
                source_locations=tuple(
                    dict.fromkeys(
                        f"{file_name_by_id.get(entry_by_id[key].source.file_id, entry_by_id[key].source.file_id)}|{entry_by_id[key].source.sheet_name}|{entry_by_id[key].source.cell_range}"
                        for key in component.source_keys
                        if key in entry_by_id
                    )
                ),
                mandatory=True,
                baseline_item_code=decision.original_standard_item_id,
            )
        )
    unresolved = tuple(unresolved_list)
    review_batches = build_review_batches(
        unresolved,
        _materiality_from_state(state).performance_cent,
        all_leaf_item_ids=leaf_item_ids,
    )
    state["review_batches"] = [asdict(item) for item in review_batches]
    consistency_group_by_component: dict[str, Mapping[str, object]] = {}
    for group in state.get("consistency_groups", ()):
        for component_id in group.get("component_ids", ()):
            consistency_group_by_component[str(component_id)] = group
    consistency_status_by_group = {
        str(item["group_id"]): item
        for item in state.get("consistency_resolution", {}).get("statuses", ())
    }
    tier_names = {
        "M0": "低于明显微小错报临界值",
        "M1": "达到明显微小错报临界值但低于实际执行重要性",
        "M2": "达到实际执行重要性但低于整体重要性",
        "M3": "达到整体重要性",
    }
    trace_rows_list: list[dict[str, object]] = []
    for component, decision in zip(components, decisions, strict=True):
        source_entries = tuple(
            entry_by_id[key] for key in component.source_keys if key in entry_by_id
        )
        consistency_group = consistency_group_by_component.get(
            component.component_id, {}
        )
        consistency_status = consistency_status_by_group.get(
            str(consistency_group.get("group_id", "")), {}
        )
        trace_rows_list.append(
            {
                "记录类型": "现金流业务组成",
                "摘要": component.summary,
                "现金变化": component.cash_delta_cent / 100,
                "原现流项目": component.original_item_text,
                "对方科目": "、".join(component.counterpart_accounts),
                "自动判定现流项目": decision.system_item_name or "不进入正表",
                "判断理由": decision.reason,
                "证据强度": _EVIDENCE_TIER_TEXT.get(decision.evidence_level, decision.evidence_level),
                "证据得分": decision.evidence_score,
                "异常": "、".join(_ANOMALY_TEXT.get(anomaly, anomaly) for anomaly in component.anomalies),
                "决策来源": _DECISION_SOURCE_TEXT.get(decision.decision_source, decision.decision_source),
                "方向依据": "、".join(
                    dict.fromkeys(flow_direction_source(entry) for entry in source_entries)
                ),
                "来源文件": "、".join(
                    dict.fromkeys(
                        file_name_by_id.get(entry.source.file_id, entry.source.file_id)
                        for entry in source_entries
                    )
                ),
                "来源工作表": "、".join(
                    dict.fromkeys(entry.source.sheet_name for entry in source_entries)
                ),
                "来源单元格": "、".join(
                    dict.fromkeys(entry.source.cell_range for entry in source_entries)
                ),
                "一致性复核状态": consistency_status.get("status", ""),
                "一致性复核理由": consistency_status.get("reason", ""),
                "一致性重要性层级": tier_names.get(
                    str(consistency_group.get("tier", "")), ""
                ),
                "决策来源(技术)": decision.decision_source,
                "命中规则(技术)": decision.matched_rule_id,
                "业务组成编号(技术)": component.component_id,
                "来源占用键(技术)": "、".join(component.source_keys),
                "业务组编号(技术)": consistency_group.get("group_id", ""),
            }
        )
    for transfer in state.get("internal_transfers", ()):
        entry = entry_by_id.get(str(transfer["entry_id"]))
        trace_rows_list.append(
            {
                "记录类型": "内部划转排除",
                "摘要": "" if entry is None else entry.summary,
                "现金变化": int(transfer["matched_cent"]) / 100,
                "原现流项目": "" if entry is None else entry.original_flow_item,
                "对方科目": "" if entry is None else entry.counterpart_name,
                "自动判定现流项目": "不进入正表",
                "判断理由": "现金及现金等价物内部划转",
                "证据强度": "高",
                "异常": "内部划转已排除",
                "方向依据": "内部划转",
                "来源文件": "" if entry is None else file_name_by_id.get(entry.source.file_id, entry.source.file_id),
                "来源工作表": "" if entry is None else entry.source.sheet_name,
                "来源单元格": "" if entry is None else entry.source.cell_range,
                "一致性复核状态": "",
                "一致性复核理由": "",
                "一致性重要性层级": "",
                "决策来源(技术)": "system",
                "命中规则(技术)": "INTERNAL-TRANSFER",
                "业务组成编号(技术)": transfer["entry_id"],
                "来源占用键(技术)": transfer["entry_id"],
                "业务组编号(技术)": "",
            }
        )
    trace_rows = tuple(trace_rows_list)
    trace_rows = build_trace_rows(
        entries,
        components,
        decisions,
        source_allocations,
        state.get("materiality_assessments", ()),
        rules,
        state,
        file_name_by_id,
    )
    difference_rows = build_original_auto_differences(
        entries,
        components,
        decisions,
        frozenset(
            str(item["entry_id"])
            for item in state.get("internal_transfers", ())
        ),
        rules,
        file_name_by_id,
    )
    mapping_rows = tuple(
        {
            "文件": item["file"],
            "工作表": item["sheet"],
            "表头行": item["header_rows"],
            "字段映射": json.dumps(item["roles"], ensure_ascii=False),
        }
        for item in state.get("mappings", ())
    ) + tuple(
        {
            "文件": item["file"],
            "工作表": item["sheet"],
            "表头行": item["cell"],
            "字段映射": f"{item['kind']}：{item['message']}",
        }
        for item in state.get("normalization_issues", ())
    )
    statement_unconfirmed = any(
        state.get("statement_confirmations", {}).get(
            f"{item['file_id']}:statement:{item['sheet']}"
        ) != "use"
        for item in state.get("statement_candidates", ())
    )
    blocking_readiness_errors = tuple(
        error
        for error in readiness.errors
        if not error.startswith("仍待人工决定：")
    )
    isolated_invalid_input = any(
        decision.decision_action == "isolate_invalid_input"
        for decision in decisions
    )
    normalization_errors = tuple(
        item
        for item in state.get("normalization_issues", ())
        if item.get("kind") == "错误"
    )
    isolated_input_errors = bool(
        isolated_invalid_input
        and normalization_errors
        and all(
            "摘要为空" in str(item.get("message", ""))
            or "对方科目" in str(item.get("message", ""))
            for item in normalization_errors
        )
    )
    if blocking_readiness_errors:
        status = "诊断材料，不可作为最终表"
    elif statement_unconfirmed:
        status = "草稿：存在未核对的疑似正表"
    elif (
        normalization_errors
        and not isolated_input_errors
    ):
        status = "诊断材料，不可作为最终表"
    elif (
        review_batches
        or any(group.blocks_manual_completion for group in duplicate_groups)
    ):
        status = "待完成人工确认"
    elif reconciliation.status != "现金流量表与货币资金变动的勾稽核对：相符":
        status = "草稿：现金流量表与货币资金变动的勾稽核对未完成或存在差异"
    else:
        status = "最终可使用"
    dictionary_valid = {
        str(item.get("account", "")): item
        for item in state.get("account_dictionary", {}).get("valid_results", ())
    }
    level1_mapping_by_original = {
        str(item.get("original_level1", "")): item
        for item in state.get("account_mapping_records", ())
    }
    counterpart_paths = tuple(
        sorted(
            {
                (raw, standardized, component.account_mapping_status)
                for component in components
                for raw, standardized in zip(
                    component.original_counterpart_accounts
                    or component.counterpart_accounts,
                    component.counterpart_accounts,
                )
                if raw.strip()
            }
        )
    )
    dictionary_rows = tuple(
        {
            "客户原路径": raw_path,
            "客户一级科目": raw_levels[0] if raw_levels else "未识别",
            "标准一级科目": standard_levels[0] if mapping_status == "confirmed" and standard_levels else "待确认",
            "中间层级": "_".join(standard_levels[1:-1]) if len(standard_levels) > 2 else "",
            "末级明细": standard_levels[-1] if standard_levels else "",
            "规范化路径": "_".join(standard_levels) if mapping_status == "confirmed" and standard_levels else raw_path,
            "科目语义": str(
                _dictionary_display_result(dictionary_valid, standard_path, raw_path).get(
                    "semantic", "未记录"
                )
            ),
            "疑似项目": str(
                _dictionary_display_result(dictionary_valid, standard_path, raw_path).get(
                    "item_id", "未记录"
                )
            ),
            "证据状况": "已形成完整路径",
            "依据": str(
                _dictionary_display_result(dictionary_valid, standard_path, raw_path).get(
                    "basis", "内置规则或本行业务来源"
                )
            ),
            "适用NOTE": str(
                _dictionary_display_result(dictionary_valid, standard_path, raw_path).get(
                    "note_id", ""
                )
            ),
            "映射状态": "已确认" if mapping_status == "confirmed" else "待确认（流程已停止）",
            "一级科目映射候选": "、".join(
                str(item)
                for item in level1_mapping_by_original.get(
                    raw_levels[0] if raw_levels else "", {}
                ).get("candidate_standard_names", ())
            ) or "无自动候选",
            "一级科目映射依据": str(
                level1_mapping_by_original.get(
                    raw_levels[0] if raw_levels else "", {}
                ).get("basis", "未记录")
            ),
        }
        for raw_path, standard_path, mapping_status, raw_levels, standard_levels in (
            (
                raw_path,
                standard_path,
                mapping_status,
                split_account_levels(raw_path),
                split_account_levels(standard_path),
            )
            for raw_path, standard_path, mapping_status in counterpart_paths
        )
    )
    consistency_status = {
        str(item.get("group_id", "")): item
        for item in state.get("consistency_resolution", {}).get("statuses", ())
    }
    consistency_rows = tuple(
        {
            "同一凭证内不同分类的处理结果": str(
                consistency_status.get(str(group.get("group_id", "")), {}).get(
                    "status", "等待人工复核"
                )
            ),
            "为什么需要整组检查以及检查依据": str(
                consistency_status.get(str(group.get("group_id", "")), {}).get(
                    "reason", "相同原始来源形成不同项目"
                )
            ),
            "该组金额与采用的复核要求": (
                f"{int(group.get('gross_cent', 0)) / 100:,.2f}元；"
                f"有效重要性为{tier_names.get(str(group.get('materiality_level', '')), str(group.get('materiality_level', '')))}"
            ),
            "业务组成编号": "、".join(group.get("component_ids", ())),
        }
        for group in state.get("consistency_groups", ())
    )
    cash_account_names = {
        str(key): tuple(str(name) for name in names)
        for key, names in state["cash_scope"].get("account_names_by_key", ())
    }

    def cash_account_display_name(key: str) -> str:
        names = cash_account_names.get(key, ())
        return "、".join(names) if names else key

    model = WorkbookModel(
        statement=statement,
        rules=rules,
        comparison=comparison,
        review_batches=review_batches,
        duplicate_groups=duplicate_groups,
        ai_records=_ai_records_from_state(state),
        cash_scope_rows=(
            tuple(
                {"科目": cash_account_display_name(str(key)), "决定": "纳入"}
                for key in state["cash_scope"]["included_keys"]
            )
            + tuple(
                {"科目": cash_account_display_name(str(key)), "决定": "不纳入"}
                for key in state["cash_scope"]["excluded_keys"]
            )
        ),
        reconciliation=reconciliation,
        trace_rows=trace_rows,
        difference_rows=difference_rows,
        mapping_rows=mapping_rows,
        overall_status=status,
        automatic_change_threshold=int(state["automatic_change_threshold"]),
        unconfirmed_statement=statement_unconfirmed,
        dictionary_rows=dictionary_rows,
        consistency_rows=consistency_rows,
        manual_adjustments=layers.manual_adjustments,
    )
    sequence = 1
    while True:
        suffix = "" if sequence == 1 else f"_重建{sequence}"
        workbook_path = Path(run_dir) / f"现金流量表正表及复核底稿{suffix}.xlsx"
        temporary_path = workbook_path.with_name(f"{workbook_path.stem}_生成中.xlsx")
        if not workbook_path.exists() and not temporary_path.exists():
            break
        sequence += 1
    build_output_workbook(model, temporary_path)
    if status == "最终可使用":
        recalculate_workbook_with_excel(temporary_path)
    output_check = validate_final_output(temporary_path, model)
    if not output_check.valid:
        raise RuntimeError("输出工作簿验收失败：" + "；".join(output_check.errors))
    _replace_with_windows_retry(temporary_path, workbook_path)
    store = _store(run_dir)
    with store.stage("finalize") as connection:
        connection.executemany(
            "INSERT OR REPLACE INTO classification_decision(record_id, payload_json) VALUES (?, ?)",
            ((item.component_id, json.dumps(asdict(item), ensure_ascii=False)) for item in decisions),
        )
        connection.executemany(
            "INSERT OR REPLACE INTO statement_value(record_id, payload_json) VALUES (?, ?)",
            (
                (item_id, json.dumps({"amount_cent": amount}, ensure_ascii=False))
                for item_id, amount in statement.values.items()
            ),
        )
        connection.executemany(
            "INSERT OR REPLACE INTO review_batch(record_id, payload_json) VALUES (?, ?)",
            ((item.batch_id, json.dumps(asdict(item), ensure_ascii=False)) for item in review_batches),
        )
        connection.executemany(
            "INSERT OR REPLACE INTO duplicate_group(record_id, payload_json) VALUES (?, ?)",
            ((item.group_id, json.dumps(asdict(item), ensure_ascii=False)) for item in duplicate_groups),
        )
        if comparison is not None:
            connection.executemany(
                "INSERT OR REPLACE INTO statement_comparison(record_id, payload_json) VALUES (?, ?)",
                ((item.item_id, json.dumps(asdict(item), ensure_ascii=False)) for item in comparison.rows),
            )
        connection.execute(
            "INSERT OR REPLACE INTO reconciliation(record_id, payload_json) VALUES (?, ?)",
            ("cash_reconciliation", json.dumps(asdict(reconciliation), ensure_ascii=False)),
        )
    state["reconciliation"] = asdict(reconciliation)
    state["workbook_path"] = str(workbook_path)
    state["overall_status"] = status
    state["stage"] = "finalized"
    if status == "最终可使用":
        state["excel_recalculation"] = {
            "status": "completed",
            "recalculated_at": datetime.now(timezone.utc).isoformat(),
            "workbook_sha256": hashlib.sha256(workbook_path.read_bytes()).hexdigest(),
        }
    else:
        state.pop("excel_recalculation", None)
    _save_state(run_dir, state)
    return FinalizeResult(str(state["run_id"]), Path(run_dir), workbook_path, status)
