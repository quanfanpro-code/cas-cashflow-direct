from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from collections.abc import Mapping

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from cashflow_direct.workbook_structure import HeaderBand, SheetSnapshot, WorkbookSnapshot, find_header_bands
from cashflow_direct.rule_registry import default_rule_registry


ALWAYS_REQUIRED_ROLES = ("summary",)


@dataclass(frozen=True, slots=True)
class ColumnMapping:
    role: str
    column_index: int
    column_letter: str
    header_path: tuple[str, ...]
    header_cell: str
    score: int


@dataclass(frozen=True, slots=True)
class DatasetMapping:
    sheet_name: str
    header_row_start: int
    header_row_end: int
    role_to_column: dict[str, ColumnMapping]
    merged_ranges: tuple[str, ...] = ()


@dataclass(frozen=True, slots=True)
class MappingQuestion:
    role: str
    recommended: ColumnMapping
    alternatives: tuple[ColumnMapping, ...]
    sample_values: tuple[str, str, str]
    sheet_name: str = ""


def _dictionary() -> dict[str, dict[str, list[str]]]:
    return default_rule_registry().field_semantics["roles"]


def _merged_value(sheet: SheetSnapshot, row: int, column: int) -> object:
    value = sheet.rows[row - 1][column - 1] if row <= len(sheet.rows) and column <= len(sheet.rows[row - 1]) else None
    if value not in (None, ""):
        return value
    for merged in sheet.merged_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(merged)
        if min_row <= row <= max_row and min_col <= column <= max_col:
            if min_row <= len(sheet.rows) and min_col <= len(sheet.rows[min_row - 1]):
                return sheet.rows[min_row - 1][min_col - 1]
    return None


def _header_path(sheet: SheetSnapshot, band: HeaderBand, column: int) -> tuple[str, ...]:
    values: list[str] = []
    for row in range(band.row_start, band.row_end + 1):
        value = _merged_value(sheet, row, column)
        if value not in (None, ""):
            text = str(value).strip()
            if text and (not values or values[-1] != text):
                values.append(text)
    return tuple(values)


def _sample_values(sheet: SheetSnapshot, band: HeaderBand, column: int) -> tuple[object, ...]:
    values: list[object] = []
    for row in sheet.rows[band.row_end :]:
        value = row[column - 1] if column <= len(row) else None
        if value not in (None, ""):
            values.append(value)
        if len(values) == 8:
            break
    return tuple(values)


_DATE_RE = re.compile(
    r"^\d{4}[-/.]\d{1,2}[-/.]\d{1,2}([ T]\d{1,2}:\d{2}(:\d{2})?)?$|^\d{4}年\d{1,2}月\d{1,2}日$"
)


def _is_strict_date(value: object) -> bool:
    if isinstance(value, (date, datetime)):
        return True
    if isinstance(value, str):
        return bool(_DATE_RE.match(value.strip()))
    return False


def _strict_date_share(values: tuple[object, ...]) -> float:
    if not values:
        return 0.0
    return sum(1 for value in values if _is_strict_date(value)) / len(values)


def _type_bonus(values: tuple[object, ...], expected: list[str]) -> int:
    if not values:
        return 0
    if "date" in expected:
        hits = sum(1 for value in values if _is_strict_date(value))
    elif "money" in expected:
        hits = sum(isinstance(value, (int, float)) and not isinstance(value, bool) for value in values)
    elif "integer" in expected:
        hits = sum(isinstance(value, int) and not isinstance(value, bool) for value in values)
    else:
        hits = sum(isinstance(value, str) for value in values)
    return 2 if hits / len(values) >= 0.6 else 0


def _account_path_bonus(path: tuple[str, ...], values: tuple[object, ...]) -> int:
    """一级科目与完整路径并存时，只在样本确有层级结构时明确优先完整路径。"""
    header_bonus = 1 if any(
        item.replace(" ", "") in {"科目名称", "会计科目", "完整科目名称", "科目全称"}
        for item in path
    ) else 0
    text_values = tuple(str(value).strip() for value in values if str(value).strip())
    hierarchical = sum(
        bool(re.search(r"[_/\\>|：:]", value)) for value in text_values
    )
    sample_bonus = 2 if text_values and hierarchical / len(text_values) >= 0.6 else 0
    return header_bonus + sample_bonus


def _score(role: str, path: tuple[str, ...], values: tuple[object, ...], spec: dict[str, list[str]]) -> int:
    joined = "".join(path).replace(" ", "")
    best = 0
    for term in spec["terms"]:
        normalized = term.replace(" ", "")
        if any(normalized == item.replace(" ", "") for item in path):
            best = max(best, 10)
        elif len(normalized) > 2 and normalized in joined:
            best = max(best, 7)
    score = best + _type_bonus(values, spec["value_types"])
    if role == "account_name":
        score += _account_path_bonus(path, values)
    return score


def _map_band(
    sheet: SheetSnapshot,
    band: HeaderBand,
    overrides: Mapping[str, int],
) -> DatasetMapping | MappingQuestion | None:
    roles = _dictionary()
    max_columns = max((len(row) for row in sheet.rows), default=0)
    candidates: dict[str, list[ColumnMapping]] = {role: [] for role in roles}
    for column in range(1, max_columns + 1):
        path = _header_path(sheet, band, column)
        values = _sample_values(sheet, band, column)
        for role, spec in roles.items():
            score = _score(role, path, values, spec)
            if role == "voucher_date" and _strict_date_share(values) >= 0.6 and score < 9:
                score = 9
            if score >= 7:
                candidates[role].append(
                    ColumnMapping(
                        role,
                        column,
                        get_column_letter(column),
                        path,
                        f"{get_column_letter(column)}{band.row_end}",
                        score,
                    )
                )

    has_debit_credit = bool(candidates["debit"] and candidates["credit"])
    has_direction_amount = bool(candidates["direction"] and candidates["flow_amount"])
    if not has_debit_credit and not has_direction_amount:
        return None
    if any(not candidates[role] for role in ALWAYS_REQUIRED_ROLES):
        return None

    chosen: dict[str, ColumnMapping] = {}
    for role in roles:
        ranked = sorted(candidates[role], key=lambda item: (-item.score, item.column_index))
        if not ranked:
            continue
        if role in overrides:
            selected = next(
                (item for item in ranked if item.column_index == overrides[role]),
                None,
            )
            if selected is None:
                raise ValueError(f"字段 {role} 的确认列不在候选范围内")
            chosen[role] = selected
            continue
        if len(ranked) > 1 and ranked[0].score - ranked[1].score <= 1:
            samples = [str(value) for value in _sample_values(sheet, band, ranked[0].column_index)[:3]]
            samples.extend([""] * (3 - len(samples)))
            return MappingQuestion(
                role, ranked[0], tuple(ranked[1:]), tuple(samples[:3]), sheet.name
            )
        chosen[role] = ranked[0]
    return DatasetMapping(sheet.name, band.row_start, band.row_end, chosen, sheet.merged_ranges)


def infer_dataset_mappings(
    snapshot: WorkbookSnapshot,
    overrides: Mapping[str, Mapping[str, int]] | None = None,
    exclude_sheets: frozenset[str] = frozenset(),
) -> tuple[DatasetMapping | MappingQuestion, ...]:
    """逐工作表识别数据集，避免按月或按账户分表被静默跳过。

    exclude_sheets 用于把已确认为客户现有正表的 sheet 排除出明细推断，
    支持"明细与正表同文件"双角色读取。
    """
    sheets = {sheet.name: sheet for sheet in snapshot.sheets}
    results: list[DatasetMapping | MappingQuestion] = []
    completed: set[str] = set()
    for band in find_header_bands(snapshot):
        if band.sheet_name in completed or band.sheet_name in exclude_sheets:
            continue
        result = _map_band(
            sheets[band.sheet_name],
            band,
            (overrides or {}).get(band.sheet_name, {}),
        )
        if result is not None:
            results.append(result)
            completed.add(band.sheet_name)
    return tuple(results)


def infer_dataset_mapping(
    snapshot: WorkbookSnapshot,
    overrides: Mapping[str, int] | None = None,
) -> DatasetMapping | MappingQuestion:
    """综合表头路径和列值类型映射字段；接近候选必须显式提问。"""
    results = infer_dataset_mappings(
        snapshot,
        {sheet.name: dict(overrides or {}) for sheet in snapshot.sheets},
    )
    if results:
        return results[0]
    placeholder = ColumnMapping("unknown", 1, "A", (), "A1", 0)
    return MappingQuestion("dataset", placeholder, (), ("", "", ""))
