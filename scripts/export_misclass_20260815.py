# -*- coding: utf-8 -*-
"""一次性脚本：按 2026-08-15 人工复核结论，为《现流表明细.xlsx》标注疑似错分。

不进 skill 的 src/，仅本次交付物使用。输入只读，输出到桌面。
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

INPUT = Path(r"C:\Users\27651\Desktop\现流表明细.xlsx")
OUTPUT = Path(r"C:\Users\27651\Desktop\现流表明细_错分标注.xlsx")

# 科目级编码（取编码点号前一段）
def _base_code(code: object) -> str:
    return (str(code or "").split(".")[0]).strip()


def _has(text: str, terms: tuple[str, ...]) -> bool:
    return any(term in text for term in terms)


def annotate(row: tuple[object, ...]):
    """按判定规则表逐行标注，返回 (是否疑似错分, 建议调整至, 理由, 涉及凭证) 或 None。"""
    summary = str(row[4] or "")
    base_code = _base_code(row[5])
    account_name = str(row[6] or "")
    flow = row[10]
    project = str(row[12] or "")
    voucher = f"{row[1]}{row[2]}"
    flow_negative = isinstance(flow, (int, float)) and flow < 0

    # 规则1：罚款、滞纳金误挂"支付的各项税费"
    if "营业外支出_罚款、滞纳金" in account_name and "支付的各项税费" in project:
        return ("是", "支付其他与经营活动有关的现金", "滞纳金/罚款不是税", voucher)
    # 规则2：罚款、滞纳金误挂"支付给职工"
    if "营业外支出_罚款、滞纳金" in account_name and "支付给职工" in project:
        return ("是", "支付其他与经营活动有关的现金", "滞纳金/罚款不是职工薪酬", voucher)
    # 规则3：记231 整凭证票据背书抵应付账款（非现金）
    if voucher == "记231":
        return ("是", "不进现金流量表", "票据背书属非现金交易；一正一负虽轧平但明细不该出现", "记231")
    # 规则4：自家电费挂到收款项目（含“矿业公司X月电费”“矿山收电费”等写法，
    # 判定本质：科目是制造费用_电费或进项税、挂在收款项目、金额为负）
    if (
        "电费" in summary
        and (base_code == "5101" or "进项税" in account_name)
        and "收到其他与经营" in project
        and flow_negative
    ):
        return ("是", "购买商品、接受劳务支付的现金", "自家电费是付款，不得在收款项目内轧差", voucher)
    # 规则5：收款挂到付款项目
    if "收水电气、房租" in summary and flow_negative and "支付其他与经营" in project:
        return ("是", "收到其他与经营活动有关的现金", "收款挂到了付款项目", voucher)
    # 规则6：收款/退回挂错方向
    if (
        ("收重庆中环工程款" in summary or "付矿业公司850中段排水通道费用" in summary)
        and "购买商品" in project
        and flow_negative
    ):
        return ("是", "收到其他与经营活动有关的现金", "收款/退回挂错方向", voucher)
    # 规则7：内部向部门收费不是对外销售
    if "茶叶费" in summary and flow_negative and "销售商品" in project:
        return ("是", "收到其他与经营活动有关的现金", "内部向部门收费不是对外销售", voucher)
    # 规则8：资本性支出误挂经营项目
    if "购建固定资产、无形资产和其他长期资产支付的现金" in _capital_target(
        base_code, summary
    ):
        return ("是", _capital_target(base_code, summary), "资本性支出", voucher)
    # 规则9：结构性存款属投资活动
    if "结构性存款" in summary and "收到其他与经营" in project:
        return ("是", "取得投资收益收到的现金", "结构性存款属投资活动", voucher)
    # 规则10：车船使用税误挂经营
    if "车船使用税" in account_name and "支付其他与经营" in project:
        return ("是", "支付的各项税费", "随保险缴纳的车船税是税款", voucher)
    # 规则11：电费基金误挂税费
    if "电费基金" in summary and "支付的各项税费" in project:
        return ("是", "购买商品、接受劳务支付的现金", "电费基金是电费组成，不是税", voucher)
    # 规则12：代扣个税应随工资
    if "个人所得税" in summary and base_code == "2241" and "支付的各项税费" in project:
        return ("是", "支付给职工以及为职工支付的现金", "代扣个税应随工资", voucher)
    # 规则13：劳务派遣费不是职工薪酬
    if "聚佳" in summary and "支付给职工" in project:
        return ("是", "购买商品、接受劳务支付的现金", "付给劳务公司的派遣费不属职工；且摘要写计提需核实", voucher)
    # 规则14：口径提示（不算错分）
    if base_code == "2211" and _has(summary, ("独生子女费", "慰问金", "护理费")):
        return ("否", "", "口径不一致，建议统一", voucher)
    if base_code == "6603" and flow_negative:
        return ("否", "", "口径不一致，建议统一", voucher)
    return None


def _capital_target(base_code: str, summary: str) -> str:
    """规则8：资本性支出判定，命中返回目标项目名，否则返回空串。"""
    target = "购建固定资产、无形资产和其他长期资产支付的现金"
    # 在建工程食堂/多功能厅；长期待摊装修/食堂/厨房；固定资产食堂计费系统；预付食堂工程款
    if base_code == "1604" and _has(summary, ("食堂", "多功能厅")):
        return target
    if base_code == "1801" and _has(summary, ("厕所", "食堂", "厨房")):
        return target
    if base_code == "1601" and "食堂计费系统" in summary:
        return target
    if base_code == "1123" and "付总部食堂建设工程费用" in summary:
        return target
    return ""


def main() -> None:
    wb = load_workbook(INPUT)
    ws = wb["明细"]
    start_col = ws.max_column + 1  # 新增列从 N 开始
    new_headers = ("是否疑似错分", "建议调整至", "理由", "涉及凭证")
    for index, header in enumerate(new_headers):
        ws.cell(1, start_col + index, header)

    summary_rows: dict[str, list[float]] = {}
    for row_index in range(2, ws.max_row + 1):
        row = tuple(ws.cell(row_index, col).value for col in range(1, ws.max_column + 1))
        label = annotate(row)
        if label is None:
            continue
        is_mis, target, reason, voucher = label
        ws.cell(row_index, start_col, is_mis)
        ws.cell(row_index, start_col + 1, target)
        ws.cell(row_index, start_col + 2, reason)
        ws.cell(row_index, start_col + 3, voucher)
        flow = row[10] if isinstance(row[10], (int, float)) else 0
        summary_rows.setdefault(reason, []).append(abs(flow))

    # 错分汇总页
    total = wb.create_sheet("错分汇总")
    total.append(["类别", "笔数", "金额合计（元）", "理由"])
    for index, (reason, amounts) in enumerate(sorted(summary_rows.items()), 2):
        total.append([f"第{index - 1}类", len(amounts), round(sum(amounts), 2), reason])

    wb.save(OUTPUT)
    print(f"已输出：{OUTPUT}")
    print(f"标注行数：{sum(len(v) for v in summary_rows.values())}")
    for reason, amounts in summary_rows.items():
        print(f"  {len(amounts):>3} 笔  {round(sum(amounts), 2):>15,.2f}  {reason}")


if __name__ == "__main__":
    main()